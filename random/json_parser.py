#!/usr/bin/env python3
"""
Software Inventory JSON → Excel Parser

Reads a (potentially large) JSON file where each record has:
    client     : string  — client/tenant name
    output     : string  — software list separated by \\n, entries may
                           include "[installed on <date>]" suffix
    start_date : string  — audit/scan date for this record

Produces a 3-tab Excel report:
    Tab 1 — Software by Client & Date  : name, client, occurrences, start_date
    Tab 2 — Software Totals            : name, total occurrences (one count per
                                         client regardless of how many start_dates
                                         that client has)
    Tab 3 — Installation Dates         : client, software, installed_on

Uses a streaming bracket-counter parser so the 1 GB file is never fully loaded.
Handles: JSON arrays, NDJSON, MySQL INTO OUTFILE (raw newlines inside strings),
multi-line objects, and concatenated objects.
"""

import re
import os
import sys
import json
import argparse
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is required.  Run:  pip install openpyxl")


# ── regex to find "[installed on ...]" anywhere in a software line ────────────
_INSTALLED_RE = re.compile(r'\[installed on ([^\]]+)\]', re.I)


def parse_line(raw: str) -> tuple[str, str]:
    """Return (software_name, installed_date).  installed_date may be empty."""
    raw = raw.strip()
    m = _INSTALLED_RE.search(raw)
    installed = m.group(1).strip() if m else ""
    name = _INSTALLED_RE.sub("", raw).strip()
    return name, installed


# ── JSON streaming ────────────────────────────────────────────────────────────
#
# MySQL INTO OUTFILE writes raw control bytes (newline 0x0A, tab 0x09, etc.)
# inside string field values without JSON-escaping them.  Standard parsers
# (ijson, json.loads) reject this as invalid JSON.
#
# _iter_records() uses a bracket-counter that:
#   1. Reads the file in 64 KB chunks — memory is bounded to one object at a time
#   2. Tracks string/escape state character-by-character
#   3. Replaces any bare control character inside a JSON string with its
#      two-char JSON escape sequence  (\n  \t  etc.)
#   4. Emits each complete {…} object to json.loads once the closing } is found
#
# This handles: JSON arrays, NDJSON, MySQL OUTFILE, multi-line objects,
# and concatenated objects — without needing ijson.

_CTRL = {'\n': '\\n', '\r': '\\r', '\t': '\\t', '\b': '\\b', '\f': '\\f'}

# Excel hard row limit (row 1 is header; max data rows = 1_048_575)
EXCEL_MAX_ROWS = 1_048_575


def _try_parse(s: str) -> dict | None:
    """Try json.loads, then retry after undoing MySQL OUTFILE backslash-doubling."""
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError:
        pass
    # MySQL INTO OUTFILE doubles every backslash (\ → \\), which turns JSON's
    # \" (escaped quote) into \\" (escaped backslash + raw quote = end-of-string).
    # Halving all backslash-pairs restores the original JSON text.
    try:
        obj = json.loads(s.replace('\\\\', '\\'))
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError as e:
        print(f"  [!] Skipping object: {e}", file=sys.stderr)
        return None


def _iter_records(file_path: str):
    """Yield one dict per JSON object from any supported format."""
    CHUNK = 65_536  # 64 KB per read
    buf: list[str] = []
    depth = 0
    in_str = False
    esc_next = False
    parsed = skipped = 0

    with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as fh:
        while True:
            chunk = fh.read(CHUNK)
            if not chunk:
                break
            for ch in chunk:
                if esc_next:
                    if depth:
                        buf.append(ch)
                    esc_next = False
                    continue
                if ch == "\\" and in_str:
                    if depth:
                        buf.append(ch)
                    esc_next = True
                    continue
                if ch == '"':
                    in_str = not in_str
                    if depth:
                        buf.append(ch)
                    continue
                if in_str:
                    if depth:
                        if ord(ch) < 0x20:
                            # Sanitize bare control chars MySQL OUTFILE leaves unescaped
                            buf.append(_CTRL.get(ch, f"\\u{ord(ch):04x}"))
                        else:
                            buf.append(ch)
                    continue
                # Outside strings
                if ch == "{":
                    depth += 1
                    buf.append(ch)
                elif ch == "}":
                    if depth:
                        buf.append(ch)
                        depth -= 1
                        if depth == 0:
                            s = "".join(buf)
                            buf.clear()
                            obj = _try_parse(s)
                            if obj is not None:
                                yield obj
                                parsed += 1
                            else:
                                skipped += 1
                elif depth:
                    buf.append(ch)

    print(f"[*] Parsed {parsed:,} records ({skipped:,} skipped)", file=sys.stderr)


# ── CSV parser ────────────────────────────────────────────────────────────────

def _iter_csv_records(file_path: str):
    """
    Yield dicts with keys client/output/start_date from a CSV file.
    The output column may contain newlines (quoted field) or literal \\n sequences.
    Column order is detected from the header row; column names are case-insensitive.
    """
    import csv
    parsed = 0
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        reader = csv.DictReader(fh)
        # Normalise header names to lowercase for flexible matching
        if reader.fieldnames is None:
            print("[!] CSV has no header row", file=sys.stderr)
            return
        norm = {f.strip().lower(): f for f in reader.fieldnames}
        client_col     = norm.get("client")
        output_col     = norm.get("output")
        start_date_col = norm.get("start_date") or norm.get("startdate") or norm.get("start date")
        if not all([client_col, output_col, start_date_col]):
            print(f"[!] CSV missing required columns. Found: {list(reader.fieldnames)}", file=sys.stderr)
            return
        for row in reader:
            yield {
                "client":     (row.get(client_col)     or "").strip(),
                "output":     (row.get(output_col)     or ""),
                "start_date": (row.get(start_date_col) or "").strip(),
            }
            parsed += 1
    print(f"[*] CSV: read {parsed:,} rows", file=sys.stderr)


# ── Aggregation ───────────────────────────────────────────────────────────────

def _accumulate(records, tab1, tab2, tab2_seen, tab3):
    for record in records:
        client     = str(record.get("client",     "")).strip()
        start_date = str(record.get("start_date", "")).strip()
        output     = str(record.get("output",     ""))
        for raw in re.split(r'\\n|\n', output):
            name, installed = parse_line(raw)
            if not name:
                continue
            tab1[(client, start_date, name)] += 1
            pair = (client, name)
            if pair not in tab2_seen:
                tab2_seen.add(pair)
                tab2[name] += 1
            if installed:
                tab3.append((client, name, installed))


def aggregate(input_path: str) -> tuple[dict, dict, list]:
    tab1: dict[tuple[str, str, str], int] = defaultdict(int)
    tab2: dict[str, int] = defaultdict(int)
    tab2_seen: set[tuple[str, str]] = set()
    tab3: list[tuple[str, str, str]] = []

    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".csv":
        print(f"[*] Format: CSV", file=sys.stderr)
        _accumulate(_iter_csv_records(input_path), tab1, tab2, tab2_seen, tab3)
    else:
        with open(input_path, "rb") as fh:
            probe = fh.read(256).decode("utf-8", errors="replace")
        print(f"[*] File probe: {probe[:120]!r}", file=sys.stderr)
        _accumulate(_iter_records(input_path), tab1, tab2, tab2_seen, tab3)

    return tab1, tab2, tab3


# ── Styling helpers ───────────────────────────────────────────────────────────

HDR_BG   = "1B3A5C"
HDR_FG   = "FFFFFF"
ALT_ROW  = "F5F5F5"

_thin    = Side(style="thin", color="CCCCCC")
THIN     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _align(h="left") -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=False)

def _hdr_row(ws, headers: list[str], row: int = 1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, color=HDR_FG)
        c.fill      = _fill(HDR_BG)
        c.alignment = _align("center")
        c.border    = THIN
    ws.row_dimensions[row].height = 20

def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(tab1: dict, tab2: dict, tab3: list, out_path: str):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Tab 1: Software by Client & Date ──────────────────────────────────────
    ws1 = wb.create_sheet("By Client & Date")
    ws1.sheet_view.showGridLines = False
    headers1 = ["Software Name", "Client", "Occurrences", "Start Date"]
    _hdr_row(ws1, headers1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}1"

    # Sort by start_date, client, software for readability
    rows1 = sorted(
        ((k[0], k[1], k[2], v) for k, v in tab1.items()),
        key=lambda x: (x[1], x[0], x[3])   # client, start_date, software
    )
    if len(rows1) > EXCEL_MAX_ROWS:
        print(f"[!] Tab1: {len(rows1):,} rows exceeds Excel limit; truncating to {EXCEL_MAX_ROWS:,}", file=sys.stderr)
        rows1 = rows1[:EXCEL_MAX_ROWS]

    for i, (client, start_date, name, count) in enumerate(rows1, 2):
        rb = ALT_ROW if i % 2 == 0 else None
        for col, val in enumerate([name, client, count, start_date], 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font      = _font()
            c.alignment = _align("center" if col == 3 else "left")
            c.border    = THIN
            if rb:
                c.fill = _fill(rb)

    _set_widths(ws1, [55, 30, 14, 18])

    # ── Tab 2: Software Totals ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Software Totals")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Software Name", "Occurrences"]
    _hdr_row(ws2, headers2)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

    # Sort by count descending, then name
    rows2 = sorted(tab2.items(), key=lambda x: (-x[1], x[0]))
    if len(rows2) > EXCEL_MAX_ROWS:
        print(f"[!] Tab2: {len(rows2):,} rows exceeds Excel limit; truncating to {EXCEL_MAX_ROWS:,}", file=sys.stderr)
        rows2 = rows2[:EXCEL_MAX_ROWS]

    for i, (name, count) in enumerate(rows2, 2):
        rb = ALT_ROW if i % 2 == 0 else None
        for col, val in enumerate([name, count], 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font      = _font()
            c.alignment = _align("center" if col == 2 else "left")
            c.border    = THIN
            if rb:
                c.fill = _fill(rb)

    _set_widths(ws2, [55, 14])

    # ── Tab 3: Installation Dates ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Installation Dates")
    ws3.sheet_view.showGridLines = False
    headers3 = ["Client", "Software Name", "Installed On"]
    _hdr_row(ws3, headers3)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}1"

    # Sort by client, software name
    rows3 = sorted(tab3, key=lambda x: (x[0], x[1]))
    if len(rows3) > EXCEL_MAX_ROWS:
        print(f"[!] Tab3: {len(rows3):,} rows exceeds Excel limit; truncating to {EXCEL_MAX_ROWS:,}", file=sys.stderr)
        rows3 = rows3[:EXCEL_MAX_ROWS]

    for i, (client, name, installed) in enumerate(rows3, 2):
        rb = ALT_ROW if i % 2 == 0 else None
        for col, val in enumerate([client, name, installed], 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.font      = _font()
            c.alignment = _align()
            c.border    = THIN
            if rb:
                c.fill = _fill(rb)

    _set_widths(ws3, [30, 55, 22])

    wb.save(out_path)


# ── HTML builder ──────────────────────────────────────────────────────────────

_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Software Inventory Report</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Calibri,'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#222;font-size:14px}
header{background:#1b3a5c;color:#fff;padding:0 24px;display:flex;align-items:center;height:50px;gap:8px;position:sticky;top:0;z-index:100;box-shadow:0 2px 6px rgba(0,0,0,.3)}
.logo{font-size:1.05em;font-weight:700;margin-right:16px;white-space:nowrap}
nav{display:flex;gap:4px}
nav button{background:none;border:none;color:rgba(255,255,255,.75);font:inherit;font-size:.9em;padding:6px 14px;border-radius:4px;cursor:pointer;transition:background .15s,color .15s}
nav button:hover,nav button.active{background:rgba(255,255,255,.18);color:#fff}
main{max-width:1400px;margin:20px auto;padding:0 16px}
.card{background:#fff;border-radius:6px;box-shadow:0 1px 4px rgba(0,0,0,.12);padding:20px 24px;margin-bottom:20px}
.view{display:none}.view.active{display:block}
.vh{display:flex;align-items:center;gap:12px;margin-bottom:14px;flex-wrap:wrap}
.vh h2{font-size:1.15em;color:#1b3a5c;flex:1 1 auto}
.sbox{display:flex;align-items:center;border:1px solid #ccd6e0;border-radius:4px;background:#fff;overflow:hidden}
.sbox input{border:none;padding:6px 10px;font:inherit;font-size:.9em;width:260px;outline:none}
.sbox button{border:none;background:none;padding:5px 8px;cursor:pointer;color:#aaa;font-size:1em}
.sbox button:hover{color:#555}
.stats{font-size:.84em;color:#888;margin-bottom:8px}
.warn{background:#fff3cd;border-left:4px solid #f0ad4e;padding:8px 12px;border-radius:0 4px 4px 0;margin-bottom:10px;font-size:.85em;color:#856404}
table{width:100%;border-collapse:collapse;font-size:.88em}
thead th{background:#1b3a5c;color:#fff;padding:8px 12px;text-align:left;white-space:nowrap;user-select:none}
thead th.sort{cursor:pointer}
thead th.sort:hover{background:#254e7e}
thead th .si{margin-left:4px;opacity:.5;font-size:.78em}
thead th.asc .si::after{content:'▲';opacity:1}
thead th.desc .si::after{content:'▼';opacity:1}
thead th:not(.asc):not(.desc) .si::after{content:'⇅'}
tbody tr:nth-child(even){background:#f5f8fb}
tbody tr:hover{background:#dce8f5}
td{padding:6px 12px;border-bottom:1px solid #e8edf2;vertical-align:top}
td.r{text-align:right;font-variant-numeric:tabular-nums}
.lnk{color:#1a6eb5;cursor:pointer;text-decoration:none}
.lnk:hover{text-decoration:underline}
.bc{font-size:.87em;color:#888;margin-bottom:12px}
.bc a{color:#1a6eb5;cursor:pointer;text-decoration:none}.bc a:hover{text-decoration:underline}
.bc span{margin:0 5px}
.dtitle{font-size:1.2em;font-weight:700;color:#1b3a5c;margin-bottom:4px}
.dsub{font-size:.86em;color:#777;margin-bottom:14px}
.pg{display:flex;gap:8px;align-items:center;margin-top:10px;font-size:.86em;color:#666}
.pg button{background:#1b3a5c;color:#fff;border:none;padding:4px 12px;border-radius:3px;cursor:pointer;font:inherit}
.pg button:disabled{background:#bbb;cursor:default}
</style>
</head>
<body>
<header>
  <span class="logo">Software Inventory</span>
  <nav>
    <button id="nb-totals" onclick="A.go('totals')">Software Totals</button>
    <button id="nb-bydate" onclick="A.go('bydate')">By Client &amp; Date</button>
    <button id="nb-installs" onclick="A.go('installs')">Installation Dates</button>
  </nav>
</header>
<main>

<div id="v-totals" class="view card">
  <div class="vh"><h2>Software Totals</h2>
    <div class="sbox"><input id="f-totals" type="search" placeholder="Filter software..." oninput="A.filt('totals')"><button onclick="A.clr('totals')">&#x2715;</button></div>
  </div>
  <div id="w-totals"></div>
  <div id="s-totals" class="stats"></div>
  <table><thead><tr>
    <th class="sort" onclick="A.sort('totals','s')">Software Name<span class="si"></span></th>
    <th class="sort" style="width:130px" onclick="A.sort('totals','n')">Occurrences<span class="si"></span></th>
  </tr></thead><tbody id="b-totals"></tbody></table>
  <div id="p-totals" class="pg"></div>
</div>

<div id="v-bydate" class="view card">
  <div class="vh"><h2>By Client &amp; Date</h2>
    <div class="sbox"><input id="f-bydate" type="search" placeholder="Filter..." oninput="A.filt('bydate')"><button onclick="A.clr('bydate')">&#x2715;</button></div>
  </div>
  <div id="w-bydate"></div>
  <div id="s-bydate" class="stats"></div>
  <table><thead><tr>
    <th class="sort" onclick="A.sort('bydate','s')">Software Name<span class="si"></span></th>
    <th class="sort" onclick="A.sort('bydate','c')">Client<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('bydate','n')">Occurrences<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('bydate','d')">Start Date<span class="si"></span></th>
  </tr></thead><tbody id="b-bydate"></tbody></table>
  <div id="p-bydate" class="pg"></div>
</div>

<div id="v-installs" class="view card">
  <div class="vh"><h2>Installation Dates</h2>
    <div class="sbox"><input id="f-installs" type="search" placeholder="Filter..." oninput="A.filt('installs')"><button onclick="A.clr('installs')">&#x2715;</button></div>
  </div>
  <div id="w-installs"></div>
  <div id="s-installs" class="stats"></div>
  <table><thead><tr>
    <th class="sort" onclick="A.sort('installs','c')">Client<span class="si"></span></th>
    <th class="sort" onclick="A.sort('installs','s')">Software Name<span class="si"></span></th>
    <th class="sort" style="width:160px" onclick="A.sort('installs','i')">Installed On<span class="si"></span></th>
  </tr></thead><tbody id="b-installs"></tbody></table>
  <div id="p-installs" class="pg"></div>
</div>

<div id="v-software" class="view card">
  <div class="bc"><a onclick="A.go('totals')">Software Totals</a><span>&#x203A;</span><span id="bc-s"></span></div>
  <div id="dt-s" class="dtitle"></div>
  <div id="ds-s" class="dsub"></div>
  <div class="vh">
    <div class="sbox"><input id="f-software" type="search" placeholder="Filter clients..." oninput="A.filt('software')"><button onclick="A.clr('software')">&#x2715;</button></div>
  </div>
  <div id="s-software" class="stats"></div>
  <table><thead><tr>
    <th class="sort" onclick="A.sort('software','c')">Client<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('software','n')">Occurrences<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('software','d')">Start Date<span class="si"></span></th>
    <th class="sort" style="width:160px" onclick="A.sort('software','i')">Installed On<span class="si"></span></th>
  </tr></thead><tbody id="b-software"></tbody></table>
  <div id="p-software" class="pg"></div>
</div>

<div id="v-client" class="view card">
  <div class="bc"><a onclick="A.go('bydate')">By Client &amp; Date</a><span>&#x203A;</span><span id="bc-c"></span></div>
  <div id="dt-c" class="dtitle"></div>
  <div id="ds-c" class="dsub"></div>
  <div class="vh">
    <div class="sbox"><input id="f-client" type="search" placeholder="Filter software..." oninput="A.filt('client')"><button onclick="A.clr('client')">&#x2715;</button></div>
  </div>
  <div id="s-client" class="stats"></div>
  <table><thead><tr>
    <th class="sort" onclick="A.sort('client','s')">Software Name<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('client','n')">Occurrences<span class="si"></span></th>
    <th class="sort" style="width:120px" onclick="A.sort('client','d')">Start Date<span class="si"></span></th>
    <th class="sort" style="width:160px" onclick="A.sort('client','i')">Installed On<span class="si"></span></th>
  </tr></thead><tbody id="b-client"></tbody></table>
  <div id="p-client" class="pg"></div>
</div>

</main>
<script>
const D1=__DATA1__;
const D2=__DATA2__;
const D3=__DATA3__;
const TRUNC={t1:__TR1__,t2:__TR2__,t3:__TR3__,max:__MAX__};

// Build lookup indices from D1 (by-client-date data)
const swRows={}, clRows={};
D1.forEach(r=>{
  (swRows[r.s]=swRows[r.s]||[]).push({c:r.c,n:r.n,d:r.d});
  (clRows[r.c]=clRows[r.c]||[]).push({s:r.s,n:r.n,d:r.d});
});
// Install-date lookup from D3
const instMap={};
D3.forEach(r=>{ instMap[r.c+'|'+r.s]=r.i; });

const A=(()=>{
  const PS=500;
  const st={};

  function initSt(v,data){
    st[v]={data:data.slice(),filt:data.slice(),sc:null,sa:true,pg:0};
  }
  initSt('totals',D2);
  initSt('bydate',D1);
  initSt('installs',D3);
  initSt('software',[]);
  initSt('client',[]);

  // Show truncation warnings on main views
  const warnKeys={totals:'t2',bydate:'t1',installs:'t3'};
  Object.entries(warnKeys).forEach(([v,k])=>{
    if(TRUNC[k]){
      const el=document.getElementById('w-'+v);
      if(el){
        el.className='warn';
        el.textContent='Showing first '+TRUNC.max.toLocaleString()+' rows — use filters to narrow results.';
      }
    }
  });

  function go(v){
    document.querySelectorAll('.view').forEach(e=>e.classList.remove('active'));
    document.querySelectorAll('nav button').forEach(b=>b.classList.remove('active'));
    document.getElementById('v-'+v).classList.add('active');
    const nb=document.getElementById('nb-'+v);
    if(nb) nb.classList.add('active');
    render(v);
  }

  function showSoftware(name){
    const rows=(swRows[name]||[]).map(r=>({c:r.c,n:r.n,d:r.d,i:instMap[r.c+'|'+name]||''}));
    initSt('software',rows);
    document.getElementById('bc-s').textContent=name;
    document.getElementById('dt-s').textContent=name;
    const tot=rows.reduce((s,r)=>s+r.n,0);
    document.getElementById('ds-s').textContent=
      rows.length+' client(s) · '+tot.toLocaleString()+' total occurrences';
    document.getElementById('f-software').value='';
    go('software');
  }

  function showClient(name){
    const rows=(clRows[name]||[]).map(r=>({s:r.s,n:r.n,d:r.d,i:instMap[name+'|'+r.s]||''}));
    initSt('client',rows);
    document.getElementById('bc-c').textContent=name;
    document.getElementById('dt-c').textContent=name;
    document.getElementById('ds-c').textContent=rows.length+' software title(s)';
    document.getElementById('f-client').value='';
    go('client');
  }

  function filt(v){
    const q=(document.getElementById('f-'+v)||{}).value||'';
    const term=q.toLowerCase();
    const s=st[v];
    s.filt=term
      ? s.data.filter(r=>Object.values(r).some(x=>String(x).toLowerCase().includes(term)))
      : s.data.slice();
    s.pg=0;
    render(v);
  }

  function clr(v){
    const el=document.getElementById('f-'+v);
    if(el) el.value='';
    filt(v);
  }

  function sort(v,col){
    const s=st[v];
    s.sa=(s.sc===col)?!s.sa:true;
    s.sc=col;
    s.pg=0;
    const asc=s.sa;
    s.filt.sort((a,b)=>{
      const av=a[col]??'', bv=b[col]??'';
      if(typeof av==='number'&&typeof bv==='number') return asc?av-bv:bv-av;
      return asc?String(av).localeCompare(String(bv)):String(bv).localeCompare(String(av));
    });
    render(v);
    updSort(v,col,asc);
  }

  function updSort(v,col,asc){
    const tbl=document.querySelector('#v-'+v+' table');
    if(!tbl) return;
    tbl.querySelectorAll('thead th').forEach(th=>th.classList.remove('asc','desc'));
    const colMap={totals:['s','n'],bydate:['s','c','n','d'],installs:['c','s','i'],
                  software:['c','n','d','i'],client:['s','n','d','i']};
    const idx=(colMap[v]||[]).indexOf(col);
    const ths=tbl.querySelectorAll('thead th.sort');
    if(idx>=0&&ths[idx]) ths[idx].classList.add(asc?'asc':'desc');
  }

  function render(v){
    const s=st[v];
    const all=s.filt;
    const start=s.pg*PS;
    const slice=all.slice(start,start+PS);
    const tbody=document.getElementById('b-'+v);
    if(!tbody) return;
    tbody.innerHTML='';
    const frag=document.createDocumentFragment();
    const rfn={totals:rTotals,bydate:rBydate,installs:rInstalls,
               software:rSwDetail,client:rClDetail};
    slice.forEach(r=>frag.appendChild((rfn[v]||rTotals)(r)));
    tbody.appendChild(frag);

    const total=s.data.length;
    const shown=Math.min(all.length-start,PS);
    const el=document.getElementById('s-'+v);
    if(el){
      if(shown===0){
        el.textContent='No matching rows.';
      } else if(all.length<total){
        el.textContent='Showing '+(start+1).toLocaleString()+'–'+(start+shown).toLocaleString()+
          ' of '+all.length.toLocaleString()+' matching ('+total.toLocaleString()+' total)';
      } else {
        el.textContent='Showing '+(start+1).toLocaleString()+'–'+(start+shown).toLocaleString()+
          ' of '+total.toLocaleString()+' rows';
      }
    }
    renderPg(v,all.length);
  }

  function renderPg(v,tot){
    const s=st[v];
    const pages=Math.ceil(tot/PS);
    const el=document.getElementById('p-'+v);
    if(!el) return;
    el.innerHTML='';
    if(pages<=1) return;
    const prev=document.createElement('button');
    prev.textContent='← Prev'; prev.disabled=s.pg===0;
    prev.onclick=()=>{ s.pg--; render(v); };
    el.appendChild(prev);
    const info=document.createElement('span');
    info.textContent='Page '+(s.pg+1)+' of '+pages;
    el.appendChild(info);
    const next=document.createElement('button');
    next.textContent='Next →'; next.disabled=s.pg>=pages-1;
    next.onclick=()=>{ s.pg++; render(v); };
    el.appendChild(next);
  }

  // ── Row renderers ────────────────────────────────────────────────────────
  function mkTd(txt,cls){ const e=document.createElement('td'); if(cls)e.className=cls; e.textContent=txt??''; return e; }
  function mkLnk(txt,fn){ const e=document.createElement('td'); const a=document.createElement('a'); a.className='lnk'; a.textContent=txt??''; a.onclick=fn; e.appendChild(a); return e; }
  function mkNum(n){ const e=document.createElement('td'); e.className='r'; e.textContent=Number(n).toLocaleString(); return e; }
  function mkTr(){ return document.createElement('tr'); }

  function rTotals(r)   { const t=mkTr(); t.appendChild(mkLnk(r.s,()=>showSoftware(r.s))); t.appendChild(mkNum(r.n)); return t; }
  function rBydate(r)   { const t=mkTr(); t.appendChild(mkLnk(r.s,()=>showSoftware(r.s))); t.appendChild(mkLnk(r.c,()=>showClient(r.c))); t.appendChild(mkNum(r.n)); t.appendChild(mkTd(r.d)); return t; }
  function rInstalls(r) { const t=mkTr(); t.appendChild(mkLnk(r.c,()=>showClient(r.c))); t.appendChild(mkLnk(r.s,()=>showSoftware(r.s))); t.appendChild(mkTd(r.i)); return t; }
  function rSwDetail(r) { const t=mkTr(); t.appendChild(mkLnk(r.c,()=>showClient(r.c))); t.appendChild(mkNum(r.n)); t.appendChild(mkTd(r.d)); t.appendChild(mkTd(r.i)); return t; }
  function rClDetail(r) { const t=mkTr(); t.appendChild(mkLnk(r.s,()=>showSoftware(r.s))); t.appendChild(mkNum(r.n)); t.appendChild(mkTd(r.d)); t.appendChild(mkTd(r.i)); return t; }

  go('totals');
  return {go,filt,clr,sort};
})();
</script>
</body>
</html>
"""


def build_html(tab1: dict, tab2: dict, tab3: list, out_path: str, max_rows: int = 500_000):
    rows1 = sorted(tab1.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2]))
    rows2 = sorted(tab2.items(), key=lambda kv: (-kv[1], kv[0]))
    rows3 = sorted(tab3, key=lambda x: (x[0], x[1]))

    tr1 = len(rows1) > max_rows
    tr2 = len(rows2) > max_rows
    tr3 = len(rows3) > max_rows
    if tr1: print(f"[!] HTML by-date: truncating {len(rows1):,} → {max_rows:,} rows", file=sys.stderr)
    if tr2: print(f"[!] HTML totals:  truncating {len(rows2):,} → {max_rows:,} rows", file=sys.stderr)
    if tr3: print(f"[!] HTML installs:truncating {len(rows3):,} → {max_rows:,} rows", file=sys.stderr)
    rows1 = rows1[:max_rows]
    rows2 = rows2[:max_rows]
    rows3 = rows3[:max_rows]

    sep = (',', ':')
    d1 = json.dumps([{"s": k[2], "c": k[0], "n": v, "d": k[1]} for k, v in rows1], separators=sep)
    d2 = json.dumps([{"s": s, "n": n} for s, n in rows2], separators=sep)
    d3 = json.dumps([{"c": c, "s": s, "i": i} for c, s, i in rows3], separators=sep)

    html = (_HTML
            .replace('__DATA1__', d1)
            .replace('__DATA2__', d2)
            .replace('__DATA3__', d3)
            .replace('__TR1__', 'true' if tr1 else 'false')
            .replace('__TR2__', 'true' if tr2 else 'false')
            .replace('__TR3__', 'true' if tr3 else 'false')
            .replace('__MAX__', str(max_rows)))

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Convert software inventory JSON or CSV to an HTML or Excel report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python json_parser.py inventory.json              # → HTML report (default)
  python json_parser.py inventory.csv               # → HTML report from CSV
  python json_parser.py inventory.json -o out.xlsx  # → Excel workbook
  python json_parser.py inventory.csv  -o out.html  # → explicit HTML

CSV columns : client, output, start_date
JSON keys   : client, output, start_date  (MySQL INTO OUTFILE handled automatically)
""")
    ap.add_argument("input",  help="Input JSON or CSV file")
    ap.add_argument("-o", "--output", default=None,
                    help="Output file (.html default, or .xlsx for Excel)")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"File not found: {args.input}")

    input_dir = os.path.dirname(os.path.abspath(args.input))
    if not args.output:
        stem = os.path.splitext(os.path.basename(args.input))[0]
        args.output = os.path.join(input_dir, f"{stem}_report.html")
    elif not os.path.dirname(args.output):
        args.output = os.path.join(input_dir, args.output)

    print(f"[*] Streaming: {args.input}")
    tab1, tab2, tab3 = aggregate(args.input)

    print(f"[*] Aggregated:")
    print(f"      By client+date rows  : {len(tab1):,}")
    print(f"      Unique software names: {len(tab2):,}")
    print(f"      Install-date rows    : {len(tab3):,}")

    ext = os.path.splitext(args.output)[1].lower()
    if ext == '.xlsx':
        print("[*] Writing Excel...")
        build_excel(tab1, tab2, tab3, args.output)
    else:
        print("[*] Writing HTML report...")
        build_html(tab1, tab2, tab3, args.output)
    print(f"[+] Saved: {os.path.abspath(args.output)}")


if __name__ == "__main__":
    main()
