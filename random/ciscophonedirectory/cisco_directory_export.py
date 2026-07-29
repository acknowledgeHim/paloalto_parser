#!/usr/bin/env python3
"""
Export the Cisco UCM (CallManager) IP Phone corporate directory to an XLSX file.

Queries the CCMCIP directory servlet:
    https://<cucm-host>:8443/ccmcip/xmldirectorylist.jsp

That endpoint returns a CiscoIPPhoneDirectory XML document like:

    <CiscoIPPhoneDirectory>
      <Title>Corporate Directory</Title>
      <DirectoryEntry>
        <Name>Doe, John</Name>
        <Telephone>1234</Telephone>
      </DirectoryEntry>
      ...
    </CiscoIPPhoneDirectory>

The servlet is a *search*, not a full dump -- it expects f= (first name)
and/or l= (last name) query params and returns entries whose name starts
with what you supply. To pull the whole directory this script can crawl
every combination of a-z0-9 as the last-name prefix (--crawl, the
default) and merge/dedupe the results, or you can pass one specific query
via --first/--last for a single request.
"""

import argparse
import string
import sys
import xml.etree.ElementTree as ET
from getpass import getpass

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def parse_directory_xml(xml_bytes: bytes) -> list[tuple[str, str]]:
    """Return list of (name, extension) tuples from a CiscoIPPhoneDirectory XML blob."""
    entries = []
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return entries

    for entry in root.findall(".//DirectoryEntry"):
        name_el = entry.find("Name")
        tel_el = entry.find("Telephone")
        name = (name_el.text or "").strip() if name_el is not None else ""
        tel = (tel_el.text or "").strip() if tel_el is not None else ""
        if name or tel:
            entries.append((name, tel))
    return entries


def fetch(session: requests.Session, url: str, params: dict, verify: bool, timeout: int):
    resp = session.get(url, params=params, verify=verify, timeout=timeout)
    resp.raise_for_status()
    return resp.content


def build_session(username: str | None, password: str | None) -> requests.Session:
    session = requests.Session()
    if username:
        session.auth = (username, password or "")
    return session


def crawl_directory(session, base_url, verify, timeout, prefixes, quiet=False) -> dict[tuple[str, str], None]:
    """Query the directory once per last-name prefix and merge results (dedup)."""
    seen: dict[tuple[str, str], None] = {}
    for i, prefix in enumerate(prefixes, 1):
        try:
            data = fetch(session, base_url, {"l": prefix}, verify, timeout)
        except requests.RequestException as exc:
            print(f"  [!] prefix '{prefix}' failed: {exc}", file=sys.stderr)
            continue
        entries = parse_directory_xml(data)
        for name, tel in entries:
            seen[(name, tel)] = None
        if not quiet:
            print(f"  [{i}/{len(prefixes)}] l={prefix!r}: {len(entries)} entries (running total {len(seen)})")
    return seen


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--url", default="https://ip:8443/ccmcip/xmldirectorylist.jsp",
                         help="CCMCIP directory URL, e.g. https://<cucm-host>:8443/ccmcip/xmldirectorylist.jsp")
    parser.add_argument("--username", help="Basic auth username (end user), if the CCMCIP servlet requires auth")
    parser.add_argument("--password", help="Basic auth password. If --username is given and this is omitted, you'll be prompted.")
    parser.add_argument("--insecure", action="store_true", help="Skip TLS certificate verification (common with self-signed CUCM certs)")
    parser.add_argument("--timeout", type=int, default=15, help="Per-request timeout in seconds (default: 15)")
    parser.add_argument("--output", "-o", default="cisco_phone_directory.xlsx", help="Output XLSX path")

    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--first", help="Single query: first-name filter (combine with --last if desired)")
    mode.add_argument("--last", help="Single query: fetch once using this last-name prefix instead of crawling a-z0-9")
    parser.add_argument("--no-crawl", action="store_true",
                         help="Don't crawl a-z0-9 prefixes; just hit --url with no search params (some CUCM builds return everything this way)")

    args = parser.parse_args()

    if args.insecure:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    password = args.password
    if args.username and password is None:
        password = getpass(f"Password for {args.username}: ")

    session = build_session(args.username, password)
    verify = not args.insecure

    if args.first or args.last:
        params = {}
        if args.first:
            params["f"] = args.first
        if args.last:
            params["l"] = args.last
        print(f"Querying {args.url} with {params} ...")
        try:
            data = fetch(session, args.url, params, verify, args.timeout)
        except requests.RequestException as exc:
            print(f"Request failed: {exc}", file=sys.stderr)
            sys.exit(1)
        entries = dict.fromkeys(parse_directory_xml(data))
    elif args.no_crawl:
        print(f"Querying {args.url} with no search params ...")
        try:
            data = fetch(session, args.url, {}, verify, args.timeout)
        except requests.RequestException as exc:
            print(f"Request failed: {exc}", file=sys.stderr)
            sys.exit(1)
        entries = dict.fromkeys(parse_directory_xml(data))
    else:
        prefixes = list(string.ascii_lowercase) + list(string.digits)
        print(f"Crawling {args.url} across {len(prefixes)} last-name prefixes ...")
        entries = crawl_directory(session, args.url, verify, args.timeout, prefixes)

    if not entries:
        print("No directory entries were returned. Check the URL/auth/search params.", file=sys.stderr)
        sys.exit(1)

    rows = sorted(entries.keys(), key=lambda nt: nt[0].lower())

    wb = Workbook()
    ws = wb.active
    ws.title = "Directory"
    ws.append(["Name", "Extension"])
    for name, tel in rows:
        ws.append([name, tel])

    for col_idx, header in enumerate(["Name", "Extension"], start=1):
        max_len = max([len(header)] + [len(str(r[col_idx - 1])) for r in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4
    ws.freeze_panes = "A2"

    wb.save(args.output)
    print(f"Saved {len(rows)} entries to {args.output}")


if __name__ == "__main__":
    main()
