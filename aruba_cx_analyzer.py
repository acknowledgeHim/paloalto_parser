#!/usr/bin/env python3
"""
Aruba CX Switch Configuration Analyzer
Parses AOS-CX running-config (show running-config) and produces an Excel
security audit report with CIS Controls v8 mapping.

Designed for AOS-CX 10.x on CX 6300 / 6400 / 8xxx series.

Usage:
    python aruba_cx_analyzer.py running-config.txt
    python aruba_cx_analyzer.py running-config.txt -o audit.xlsx
"""

import re
import os
import sys
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── CIS Controls v8 mapping ───────────────────────────────────────────────────
CIS_CTRL_DESC = {
    "3.10":  "Encrypt Sensitive Data in Transit",
    "4.2":   "Secure Configuration Process for Network Infrastructure",
    "4.7":   "Manage Default Accounts on Enterprise Assets and Software",
    "4.8":   "Disable Unnecessary Services on Enterprise Assets and Software",
    "4.9":   "Configure Trusted DNS Servers on Enterprise Assets",
    "5.2":   "Use Unique Passwords",
    "5.4":   "Restrict Administrator Privileges to Dedicated Accounts",
    "6.3":   "Require MFA for Externally-Exposed Applications",
    "6.5":   "Require MFA for Administrative Access",
    "6.7":   "Centralize Access Control",
    "6.8":   "Define and Maintain Role-Based Access Control",
    "8.2":   "Collect Audit Logs",
    "8.4":   "Standardize Time Synchronization",
    "8.9":   "Centralize Audit Logs",
    "12.2":  "Establish and Maintain a Secure Network Architecture",
    "12.3":  "Securely Manage Network Infrastructure",
    "12.6":  "Use Secure Network Management and Communication Protocols",
    "13.4":  "Perform Traffic Filtering Between Network Segments",
    "13.9":  "Deploy Port-Level Access Control",
}

CIS_CONTROL_MAP: dict[str, list[str]] = {
    # ── Credentials / accounts ────────────────────────────────────────────────
    "Plaintext User Password":              ["5.2"],
    "Account Without Password":             ["5.2"],
    "Plaintext RADIUS Key":                 ["5.2", "12.6"],
    "Plaintext TACACS+ Key":                ["5.2", "12.6"],
    "Default/Weak SNMP Community String":   ["4.7", "12.3"],
    "Excessive Privileged Accounts":        ["5.4", "6.8"],
    # ── Authentication / access ───────────────────────────────────────────────
    "No RADIUS or TACACS+ Configured":      ["6.5", "6.7"],
    "Local Auth Only on Console":           ["6.5"],
    "No MFA / AAA for Management":          ["6.3", "6.5"],
    "SSH Not Enabled":                      ["4.2", "12.6"],
    "Telnet Enabled":                       ["4.2", "12.6"],
    "No Management Access Restriction":     ["12.3", "6.7"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1 Enabled":                       ["4.2", "12.3", "12.6"],
    "SNMPv2c in Use":                       ["4.2", "12.3"],
    "SNMP Read-Write Community":            ["12.3"],
    "SNMP Without ACL Restriction":         ["12.3", "6.7"],
    # ── Layer 2 / switching security ─────────────────────────────────────────
    "DHCP Snooping Not Enabled":            ["12.2", "13.4"],
    "Dynamic ARP Inspection Not Enabled":   ["12.2", "13.4"],
    "BPDU Guard Not on Access Port":        ["12.2", "13.9"],
    "STP Root Guard Not Configured":        ["12.2", "13.9"],
    "IP Source Guard Not Enabled":          ["12.2", "13.4"],
    "No Port Security / 802.1X":           ["13.9", "6.7"],
    "VLAN 1 in Use on Access Port":         ["12.2"],
    "Trunk Native VLAN Is VLAN 1":          ["12.2"],
    "Spanning Tree Not Enabled":            ["12.2", "13.9"],
    # ── Logging / time ────────────────────────────────────────────────────────
    "No Syslog Servers Configured":         ["8.2", "8.9"],
    "NTP Not Configured":                   ["8.4"],
    "Only One NTP Server":                  ["8.4"],
    "NTP Authentication Not Configured":    ["8.4", "12.6"],
    # ── Password & session policy ─────────────────────────────────────────────
    "Password Complexity Not Enforced":     ["5.2"],
    "Weak Password Minimum Length":         ["5.2"],
    "No Account Lockout Policy":            ["5.2", "6.5"],
    "Management Session Timeout Not Set":   ["12.3"],
    # ── ACL protocol checks ───────────────────────────────────────────────────
    "ACL Permits Insecure Protocol":        ["4.8", "12.6"],
    "ACL Missing Default Deny":             ["12.2", "13.4"],
    "Password Expiry Not Configured":       ["5.2"],
    "SSH v1 Enabled":                       ["3.10", "12.6"],
    "Syslog Not Using TLS":                 ["8.9"],
    # ── Configuration hygiene ─────────────────────────────────────────────────
    "No Login/MOTD Banner":                 ["4.2"],
    "No System Location":                   ["4.2"],
    "No System Contact":                    ["4.2"],
    "Interface Missing Description":        ["4.2"],
    "Active Interface Not Shut Down":       ["4.2", "4.8"],
    "DNS Not Configured":                   ["4.9"],
}


def _cis_label(ctrl_ids: list[str]) -> str:
    return " · ".join(f"CIS {c}" for c in ctrl_ids)


# ── PCI DSS v4.0 mapping ─────────────────────────────────────────────────────
PCI_DSS_DESC = {
    "1.2.4":  "All traffic between trusted/untrusted networks is explicitly controlled",
    "1.2.7":  "Unused network access points are disabled",
    "1.3.1":  "Inbound traffic to the CDE is restricted to what is necessary",
    "2.2.1":  "Configuration standards are defined for all system components",
    "2.2.4":  "Only necessary services, protocols, and functions are enabled",
    "2.2.7":  "All non-console administrative access is encrypted",
    "4.2.1":  "Strong cryptography is used to safeguard PAN during transmission",
    "7.2.1":  "All user access is appropriate and assigned by business need",
    "8.2.1":  "All user IDs and authentication credentials are managed securely",
    "8.2.8":  "Session idle time-out is 15 minutes or less",
    "8.3.4":  "Invalid authentication attempts are limited",
    "8.3.6":  "Passwords meet minimum length and complexity requirements",
    "8.3.9":  "Passwords are changed at least once every 90 days",
    "8.4.1":  "MFA is implemented for all non-console administrative access",
    "10.5.4": "Audit log files are protected (written to external log servers)",
    "10.6.1": "System clocks are synchronized using time-synchronization technology",
}

PCI_DSS_MAP: dict[str, list[str]] = {
    # ── Credentials ───────────────────────────────────────────────────────────
    "Plaintext User Password":            ["8.2.1"],
    "Account Without Password":           ["8.2.1"],
    "Plaintext RADIUS Key":               ["8.2.1"],
    "Plaintext TACACS+ Key":              ["8.2.1"],
    "Default/Weak SNMP Community String": ["2.2.4"],
    "Excessive Privileged Accounts":      ["7.2.1"],
    # ── Authentication / access ───────────────────────────────────────────────
    "No RADIUS or TACACS+ Configured":    ["8.4.1"],
    "Local Auth Only on Console":         ["8.4.1"],
    "No MFA / AAA for Management":        ["8.4.1"],
    "SSH Not Enabled":                    ["2.2.7"],
    "Telnet Enabled":                     ["2.2.4", "2.2.7"],
    "No Management Access Restriction":   ["1.2.4"],
    "SSH v1 Enabled":                     ["4.2.1"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1 Enabled":                     ["2.2.4"],
    "SNMPv2c in Use":                     ["2.2.4"],
    "SNMP Read-Write Community":          ["2.2.4"],
    "SNMP Without ACL Restriction":       ["1.2.4"],
    # ── Layer 2 ───────────────────────────────────────────────────────────────
    "DHCP Snooping Not Enabled":          ["1.2.4"],
    "Dynamic ARP Inspection Not Enabled": ["1.2.4"],
    "BPDU Guard Not on Access Port":      ["1.2.4"],
    "STP Root Guard Not Configured":      ["1.2.4"],
    "IP Source Guard Not Enabled":        ["1.2.4"],
    "No Port Security / 802.1X":          ["1.2.7"],
    "VLAN 1 in Use on Access Port":       ["1.2.4"],
    "Trunk Native VLAN Is VLAN 1":        ["1.2.4"],
    "Spanning Tree Not Enabled":          ["1.2.4"],
    "Active Interface Not Shut Down":     ["1.2.7"],
    # ── Logging / time ────────────────────────────────────────────────────────
    "No Syslog Servers Configured":       ["10.5.4"],
    "Syslog Not Using TLS":               ["10.5.4"],
    "NTP Not Configured":                 ["10.6.1"],
    "Only One NTP Server":                ["10.6.1"],
    "NTP Authentication Not Configured":  ["10.6.1"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Password Complexity Not Enforced":   ["8.3.6"],
    "Weak Password Minimum Length":       ["8.3.6"],
    "No Account Lockout Policy":          ["8.3.4"],
    "Management Session Timeout Not Set": ["8.2.8"],
    "Password Expiry Not Configured":     ["8.3.9"],
    # ── ACL ───────────────────────────────────────────────────────────────────
    "ACL Permits Insecure Protocol":      ["2.2.4", "4.2.1"],
    "ACL Missing Default Deny":           ["1.2.4"],
    # ── Config hygiene ────────────────────────────────────────────────────────
    "No Login/MOTD Banner":               ["2.2.1"],
}


def _pci_label(req_ids: list[str]) -> str:
    return " · ".join(f"PCI {r}" for r in req_ids)


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "hdr_bg":   "1B3A5C", "hdr_fg":   "FFFFFF",
    "critical": "C00000", "critical_l": "FFB3B3",
    "high":     "FF0000", "high_l":     "FFD9B3",
    "medium":   "FF8C00", "medium_l":   "FFF2CC",
    "low":      "0070C0", "low_l":      "BDD7EE",
    "info":     "595959", "info_l":     "F2F2F2",
    "ok":       "375623", "ok_l":       "E2EFDA",
    "disabled": "A6A6A6",
    "alt_row":  "F5F5F5",
}

_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=size)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Config parser ─────────────────────────────────────────────────────────────
class ArubaCXParser:

    # Interfaces we treat as "physical access" candidates for L2 checks
    _PHYSICAL_RE = re.compile(r"^\d+/\d+/\d+$")

    def __init__(self, config_file: str):
        self.config_file = config_file
        self.lines: list[str] = []

        # Parsed data
        self.system: dict         = {}
        self.vlans: list[dict]    = []
        self.interfaces: list[dict] = []
        self.acls: list[dict]     = []
        self.aaa: dict            = {}
        self.radius_servers: list[dict] = []
        self.tacacs_servers: list[dict] = []
        self.snmp: dict           = {}
        self.ntp_servers: list[str] = []
        self.syslog_servers: list[dict] = []
        self.users: list[dict]    = []
        self.dhcp_snooping: dict  = {}
        self.arp_inspection: dict = {}
        self.spanning_tree: dict  = {}
        self.password_policy: dict = {}
        self.issues: list[dict]   = []

    # ── Entry point ───────────────────────────────────────────────────────────
    def parse(self):
        try:
            with open(self.config_file, encoding="utf-8", errors="replace") as fh:
                self.lines = fh.readlines()
        except FileNotFoundError:
            sys.exit(f"File not found: {self.config_file}")

        blocks = self._split_blocks()
        self._parse_system(blocks)
        self._parse_vlans(blocks)
        self._parse_interfaces(blocks)
        self._parse_acls(blocks)
        self._parse_aaa(blocks)
        self._parse_radius(blocks)
        self._parse_tacacs(blocks)
        self._parse_snmp(blocks)
        self._parse_ntp(blocks)
        self._parse_syslog(blocks)
        self._parse_users(blocks)
        self._parse_dhcp_snooping(blocks)
        self._parse_arp_inspection(blocks)
        self._parse_spanning_tree(blocks)
        self._parse_password_policy(blocks)
        self._run_checks()

    # ── Block splitter ────────────────────────────────────────────────────────
    def _split_blocks(self) -> list[dict]:
        """Split config into a list of {'header', 'sub', 'lineno'} dicts.

        Top-level (non-indented, non-empty, non-comment) lines start a new
        block.  Indented lines are collected as sub-commands of the current
        block.  '!' lines and blank lines flush the current block.
        """
        blocks: list[dict] = []
        cur_header: str | None = None
        cur_sub: list[tuple[int, str]] = []
        cur_lineno: int = 0

        def _flush():
            nonlocal cur_header, cur_sub, cur_lineno
            if cur_header is not None:
                blocks.append({
                    "header":  cur_header,
                    "sub":     cur_sub,
                    "lineno":  cur_lineno,
                })
            cur_header = None
            cur_sub = []
            cur_lineno = 0

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()

            if not stripped or stripped.startswith("!"):
                _flush()
                continue

            if line[0] in (" ", "\t"):
                # Sub-command
                if cur_header is not None:
                    cur_sub.append((lineno, stripped))
                # else: orphaned indented line — ignore
            else:
                # New top-level command
                _flush()
                cur_header = stripped
                cur_lineno = lineno

        _flush()
        return blocks

    # ── Helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _sub_val(sub: list[tuple[int, str]], pattern: str,
                 group: int = 1, default: str = "") -> str:
        """Return the first regex match in sub-commands."""
        rx = re.compile(pattern, re.IGNORECASE)
        for _, line in sub:
            m = rx.search(line)
            if m:
                return m.group(group).strip()
        return default

    @staticmethod
    def _sub_has(sub: list[tuple[int, str]], pattern: str) -> bool:
        rx = re.compile(pattern, re.IGNORECASE)
        return any(rx.search(line) for _, line in sub)

    @staticmethod
    def _sub_all(sub: list[tuple[int, str]], pattern: str) -> list[str]:
        rx = re.compile(pattern, re.IGNORECASE)
        return [m.group(0) for _, line in sub for m in [rx.search(line)] if m]

    # ── Issue helper ──────────────────────────────────────────────────────────
    def _issue(self, severity: str, category: str, object_name: str,
               description: str, recommendation: str,
               details: str = "", line: int | str = ""):
        cis_ids = CIS_CONTROL_MAP.get(category, [])
        pci_ids = PCI_DSS_MAP.get(category, [])
        self.issues.append({
            "severity":       severity,
            "category":       category,
            "object":         object_name,
            "line":           str(line) if line else "",
            "description":    description,
            "recommendation": recommendation,
            "details":        details,
            "cis_controls":   _cis_label(cis_ids),
            "cis_ids":        cis_ids,
            "pci_dss":        _pci_label(pci_ids),
            "pci_ids":        pci_ids,
        })

    # ── System / global settings ──────────────────────────────────────────────
    def _parse_system(self, blocks: list[dict]):
        sys: dict = {
            "hostname": "", "location": "", "contact": "",
            "banner": "", "dns_servers": [],
            "version": "", "ssh_vrfs": [],
            "telnet_enabled": False,
            "ssh_v1_disabled": False,
            "ssh_line": "", "ssh_v1_line": "",
        }

        for blk in blocks:
            hdr = blk["header"]

            m = re.match(r"^hostname\s+(.+)$", hdr, re.I)
            if m:
                sys["hostname"] = m.group(1).strip().strip('"')
                continue

            m = re.match(r"^system-location\s+(.+)$", hdr, re.I)
            if m:
                sys["location"] = m.group(1).strip().strip('"')
                continue

            m = re.match(r"^system-contact\s+(.+)$", hdr, re.I)
            if m:
                sys["contact"] = m.group(1).strip().strip('"')
                continue

            if re.match(r"^banner\s+(motd|login)", hdr, re.I):
                sys["banner"] = hdr
                continue

            m = re.match(r"^ip\s+dns\s+server-address\s+(.+)$", hdr, re.I)
            if m:
                sys["dns_servers"].append(m.group(1).strip())
                continue

            m = re.match(r"^ssh\s+server\s+vrf\s+(\S+)", hdr, re.I)
            if m:
                sys["ssh_vrfs"].append(m.group(1))
                if not sys["ssh_line"]:
                    sys["ssh_line"] = str(blk["lineno"])
                continue

            if re.match(r"^telnet\s+server", hdr, re.I):
                sys["telnet_enabled"] = True
                continue

            # "no ssh server v1" or "ssh server version 2" → SSHv1 disabled
            if re.match(r"^no\s+ssh\s+server\s+v1", hdr, re.I) or \
               re.match(r"^ssh\s+server\s+version\s+2", hdr, re.I):
                sys["ssh_v1_disabled"] = True
                sys["ssh_v1_line"] = str(blk["lineno"])
                continue

            # Version line at top of file
            m = re.match(r"^!?\s*Version\s+(\S+)", hdr, re.I)
            if m:
                sys["version"] = m.group(1)
                continue

        self.system = sys

    # ── VLANs ─────────────────────────────────────────────────────────────────
    def _parse_vlans(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^vlan\s+(\d+)\s*$", blk["header"], re.I)
            if not m:
                continue
            vid = int(m.group(1))
            name = self._sub_val(blk["sub"], r"^name\s+(\S+)")
            desc = self._sub_val(blk["sub"], r"^description\s+(.+)")
            self.vlans.append({
                "id":          vid,
                "name":        name,
                "description": desc.strip('"'),
                "line":        blk["lineno"],
            })

    # ── Interfaces ────────────────────────────────────────────────────────────
    def _parse_interfaces(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^interface\s+(.+)$", blk["header"], re.I)
            if not m:
                continue
            name = m.group(1).strip()
            sub  = blk["sub"]

            # Determine interface type
            if re.match(r"^\d+/\d+/\d+", name):
                itype = "physical"
            elif re.match(r"^vlan\s*\d+", name, re.I):
                itype = "svi"
            elif re.match(r"^lag\s*\d+", name, re.I):
                itype = "lag"
            elif re.match(r"^loopback\s*\d+", name, re.I):
                itype = "loopback"
            elif re.match(r"^mgmt", name, re.I):
                itype = "mgmt"
            else:
                itype = "other"

            # Shutdown / admin state
            shutdown = self._sub_has(sub, r"^\s*shutdown\s*$") and \
                       not self._sub_has(sub, r"^\s*no\s+shutdown")
            # AOS-CX: default is shutdown on physical, no shutdown on SVI
            # "no shutdown" explicitly enables
            no_shutdown = self._sub_has(sub, r"^no\s+shutdown")
            admin_up = no_shutdown or (itype in ("svi", "loopback") and not shutdown)

            # Description
            desc = self._sub_val(sub, r"^description\s+(.+)")
            desc = desc.strip('"')

            # VLAN configuration
            vlan_mode = ""
            access_vlan = ""
            native_vlan = ""
            trunk_vlans = ""
            if self._sub_has(sub, r"^vlan\s+access"):
                vlan_mode = "access"
                access_vlan = self._sub_val(sub, r"^vlan\s+access\s+(\d+)")
            elif self._sub_has(sub, r"^vlan\s+trunk"):
                vlan_mode = "trunk"
                native_vlan = self._sub_val(sub, r"^vlan\s+trunk\s+native\s+(\d+)")
                trunk_vlans = self._sub_val(sub, r"^vlan\s+trunk\s+allowed\s+(.+)")
            elif self._sub_has(sub, r"^routing") or self._sub_has(sub, r"^ip\s+address"):
                vlan_mode = "routed"

            # IP addresses
            ip_addrs = []
            for _, ln in sub:
                im = re.match(r"^ip\s+address\s+(\S+)", ln, re.I)
                if im:
                    ip_addrs.append(im.group(1))

            # ACLs applied
            acl_in  = self._sub_val(sub, r"^ip\s+access-group\s+(\S+)\s+in")
            acl_out = self._sub_val(sub, r"^ip\s+access-group\s+(\S+)\s+out")
            acl6_in  = self._sub_val(sub, r"^ipv6\s+access-group\s+(\S+)\s+in")
            acl6_out = self._sub_val(sub, r"^ipv6\s+access-group\s+(\S+)\s+out")

            # Spanning tree per-port
            bpduguard      = self._sub_has(sub, r"spanning-tree\s+bpduguard\s+enable")
            root_guard     = self._sub_has(sub, r"spanning-tree\s+root-guard")
            port_type      = self._sub_val(sub, r"spanning-tree\s+port-type\s+(\S+)")
            loop_guard     = self._sub_has(sub, r"spanning-tree\s+loop-guard")

            # L2 security
            dhcp_trust     = self._sub_has(sub, r"dhcp-snooping\s+trust")
            arp_trust      = self._sub_has(sub, r"arp\s+inspection\s+trust")
            ip_src_guard   = self._sub_has(sub, r"ip\s+source-guard")
            dot1x          = self._sub_has(sub, r"aaa\s+authentication\s+port-access\s+dot1x")
            mac_auth       = self._sub_has(sub, r"aaa\s+authentication\s+port-access\s+mac-auth")
            port_sec       = self._sub_has(sub, r"port-security")

            # Storm control
            storm_ctrl     = self._sub_has(sub, r"storm-control")

            # LLDP
            lldp_tx        = not self._sub_has(sub, r"no\s+lldp\s+transmit")
            lldp_rx        = not self._sub_has(sub, r"no\s+lldp\s+receive")

            # MTU
            mtu = self._sub_val(sub, r"^mtu\s+(\d+)")

            self.interfaces.append({
                "name":        name,
                "type":        itype,
                "line":        blk["lineno"],
                "admin_up":    admin_up,
                "description": desc,
                "vlan_mode":   vlan_mode,
                "access_vlan": access_vlan,
                "native_vlan": native_vlan,
                "trunk_vlans": trunk_vlans,
                "ip_addresses": ", ".join(ip_addrs),
                "acl_in":      acl_in,
                "acl_out":     acl_out,
                "acl6_in":     acl6_in,
                "acl6_out":    acl6_out,
                "bpduguard":   bpduguard,
                "root_guard":  root_guard,
                "port_type":   port_type,
                "loop_guard":  loop_guard,
                "dhcp_trust":  dhcp_trust,
                "arp_trust":   arp_trust,
                "ip_src_guard": ip_src_guard,
                "dot1x":       dot1x,
                "mac_auth":    mac_auth,
                "port_security": port_sec,
                "storm_control": storm_ctrl,
                "lldp_tx":     lldp_tx,
                "lldp_rx":     lldp_rx,
                "mtu":         mtu,
            })

    # ── ACLs ──────────────────────────────────────────────────────────────────
    def _parse_acls(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^(ip|ipv6)\s+access-list\s+(\S+)", blk["header"], re.I)
            if not m:
                continue
            proto = m.group(1).lower()
            name  = m.group(2)
            entries = []
            for lineno, ln in blk["sub"]:
                am = re.match(
                    r"^(\d+)\s+(permit|deny)\s+(.+)$", ln, re.I)
                if am:
                    entries.append({
                        "seq":    am.group(1),
                        "action": am.group(2).lower(),
                        "match":  am.group(3).strip(),
                        "line":   lineno,
                    })
            self.acls.append({
                "name":    name,
                "type":    proto,
                "entries": entries,
                "line":    blk["lineno"],
            })

    # ── AAA ───────────────────────────────────────────────────────────────────
    def _parse_aaa(self, blocks: list[dict]):
        aaa: dict = {
            "login_default":   [],
            "login_console":   [],
            "enable_default":  [],
            "accounting":      False,
            "dot1x_global":    False,
        }
        for blk in blocks:
            hdr = blk["header"]
            m = re.match(r"^aaa\s+authentication\s+login\s+default\s+(.+)$", hdr, re.I)
            if m:
                aaa["login_default"] = m.group(1).split()
                continue
            m = re.match(r"^aaa\s+authentication\s+login\s+console\s+(.+)$", hdr, re.I)
            if m:
                aaa["login_console"] = m.group(1).split()
                continue
            m = re.match(r"^aaa\s+authentication\s+enable\s+default\s+(.+)$", hdr, re.I)
            if m:
                aaa["enable_default"] = m.group(1).split()
                continue
            if re.match(r"^aaa\s+accounting", hdr, re.I):
                aaa["accounting"] = True
                continue
            if re.match(r"^aaa\s+authentication\s+port-access\s+dot1x\s+authenticator", hdr, re.I):
                aaa["dot1x_global"] = True
                continue
        self.aaa = aaa

    # ── RADIUS ────────────────────────────────────────────────────────────────
    def _parse_radius(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^radius-server\s+host\s+(\S+)", blk["header"], re.I)
            if not m:
                continue
            host = m.group(1)
            key_type = ""
            key_val  = ""
            km = re.match(r"^radius-server\s+host\s+\S+\s+key\s+(plaintext|ciphertext)\s+(\S+)",
                          blk["header"], re.I)
            if km:
                key_type = km.group(1).lower()
                key_val  = km.group(2)
            else:
                kt = self._sub_val(blk["sub"], r"^key\s+(plaintext|ciphertext)\s+(\S+)", group=1)
                key_type = kt.lower() if kt else ""

            vrf = self._sub_val(blk["sub"], r"^vrf\s+(\S+)")
            self.radius_servers.append({
                "host":      host,
                "key_type":  key_type,
                "vrf":       vrf,
                "line":      blk["lineno"],
            })

    # ── TACACS+ ───────────────────────────────────────────────────────────────
    def _parse_tacacs(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^tacacs-server\s+host\s+(\S+)", blk["header"], re.I)
            if not m:
                continue
            host = m.group(1)
            key_type = self._sub_val(blk["sub"], r"^key\s+(plaintext|ciphertext)", group=1)
            vrf      = self._sub_val(blk["sub"], r"^vrf\s+(\S+)")
            self.tacacs_servers.append({
                "host":     host,
                "key_type": key_type.lower() if key_type else "",
                "vrf":      vrf,
                "line":     blk["lineno"],
            })

    # ── SNMP ──────────────────────────────────────────────────────────────────
    def _parse_snmp(self, blocks: list[dict]):
        snmp: dict = {
            "communities": [],   # {name, access, acl, line}
            "v3_users":    [],
            "trap_hosts":  [],
            "contact":     "",
            "location":    "",
        }
        for blk in blocks:
            hdr = blk["header"]

            # snmp-server community <name> [ro|rw] [acl]
            m = re.match(r"^snmp-server\s+community\s+(\S+)(?:\s+(ro|rw))?(?:\s+(\S+))?",
                         hdr, re.I)
            if m:
                snmp["communities"].append({
                    "name":   m.group(1),
                    "access": (m.group(2) or "ro").lower(),
                    "acl":    m.group(3) or "",
                    "line":   blk["lineno"],
                })
                continue

            # snmp-server vrf block — communities may be inside
            if re.match(r"^snmp-server\s+vrf", hdr, re.I):
                for _, ln in blk["sub"]:
                    cm = re.match(r"^community\s+(\S+)(?:\s+(ro|rw))?", ln, re.I)
                    if cm:
                        snmp["communities"].append({
                            "name":   cm.group(1),
                            "access": (cm.group(2) or "ro").lower(),
                            "acl":    "",
                            "line":   blk["lineno"],
                        })
                continue

            # snmp-server host (trap receiver)
            m = re.match(r"^snmp-server\s+host\s+(\S+)\s+version\s+(\S+)\s+(\S+)",
                         hdr, re.I)
            if m:
                snmp["trap_hosts"].append({
                    "host":      m.group(1),
                    "version":   m.group(2),
                    "community": m.group(3),
                    "line":      blk["lineno"],
                })
                continue

            # SNMPv3 user
            m = re.match(r"^snmp-server\s+user\s+(\S+)", hdr, re.I)
            if m:
                snmp["v3_users"].append({
                    "name": m.group(1),
                    "line": blk["lineno"],
                })
                continue

            m = re.match(r"^snmp-server\s+contact\s+(.+)$", hdr, re.I)
            if m:
                snmp["contact"] = m.group(1).strip().strip('"')
                continue

            m = re.match(r"^snmp-server\s+location\s+(.+)$", hdr, re.I)
            if m:
                snmp["location"] = m.group(1).strip().strip('"')
                continue

        self.snmp = snmp

    # ── NTP ───────────────────────────────────────────────────────────────────
    def _parse_ntp(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^ntp\s+server\s+(\S+)", blk["header"], re.I)
            if m:
                self.ntp_servers.append(m.group(1))

    # ── Syslog ────────────────────────────────────────────────────────────────
    def _parse_syslog(self, blocks: list[dict]):
        for blk in blocks:
            hdr = blk["header"]
            # logging remote <ip> [port <n>] [tls] or sub-command "transport tls"
            m = re.match(r"^logging\s+remote\s+(\S+)", hdr, re.I)
            if m:
                tls_inline = bool(re.search(r"\btls\b", hdr, re.I))
                tls_sub    = self._sub_has(blk["sub"], r"^transport\s+tls")
                self.syslog_servers.append({
                    "host":     m.group(1),
                    "severity": self._sub_val(blk["sub"], r"^severity\s+(\S+)"),
                    "vrf":      self._sub_val(blk["sub"], r"^vrf\s+(\S+)"),
                    "tls":      tls_inline or tls_sub,
                    "line":     blk["lineno"],
                })
                continue
            # syslog-server / syslog remote variations
            m = re.match(r"^syslog\s+(?:server|remote)\s+(\S+)", hdr, re.I)
            if m:
                self.syslog_servers.append({
                    "host": m.group(1), "severity": "", "vrf": "",
                    "tls": False, "line": blk["lineno"],
                })
                continue
            # Plain "logging <ip-or-hostname>" (AOS-CX native form)
            m = re.match(r"^logging\s+(\d[\d.]+|\w[\w.-]+)\s*$", hdr, re.I)
            if m and not re.match(r"^logging\s+severity", hdr, re.I):
                self.syslog_servers.append({
                    "host": m.group(1), "severity": "", "vrf": "",
                    "tls": False, "line": blk["lineno"],
                })

    # ── Local users ───────────────────────────────────────────────────────────
    def _parse_users(self, blocks: list[dict]):
        for blk in blocks:
            m = re.match(r"^user\s+(\S+)", blk["header"], re.I)
            if not m:
                continue
            name  = m.group(1)
            group = self._sub_val(blk["sub"], r"^group\s+(\S+)")
            # password plaintext / ciphertext / sha256 / ...
            pm    = None
            for _, ln in blk["sub"]:
                pm = re.match(r"^password\s+(plaintext|ciphertext|sha256|sha1|md5)\s*(\S*)", ln, re.I)
                if pm:
                    break
            pwd_type = pm.group(1).lower() if pm else "none"
            self.users.append({
                "name":      name,
                "group":     group,
                "pwd_type":  pwd_type,
                "line":      blk["lineno"],
            })

    # ── DHCP Snooping ─────────────────────────────────────────────────────────
    def _parse_dhcp_snooping(self, blocks: list[dict]):
        ds: dict = {"enabled": False, "vlans": [], "line": 0}
        for blk in blocks:
            hdr = blk["header"]
            # Match both "dhcp-snooping" and "ip dhcp snooping" forms
            if re.match(r"^(?:ip\s+)?dhcp[\s-]snooping$", hdr, re.I):
                ds["enabled"] = True
                ds["line"]    = blk["lineno"]
                continue
            m = re.match(r"^(?:ip\s+)?dhcp[\s-]snooping\s+vlan\s+(.+)$", hdr, re.I)
            if m:
                ds["enabled"] = True
                ds["vlans"]  += self._expand_vlans(m.group(1))
        self.dhcp_snooping = ds

    # ── Dynamic ARP Inspection ────────────────────────────────────────────────
    def _parse_arp_inspection(self, blocks: list[dict]):
        dai: dict = {"enabled": False, "vlans": [], "line": 0}
        for blk in blocks:
            hdr = blk["header"]
            # Match both "arp inspection" and "ip arp inspection" forms
            if re.match(r"^(?:ip\s+)?arp\s+inspection$", hdr, re.I):
                dai["enabled"] = True
                dai["line"]    = blk["lineno"]
                continue
            m = re.match(r"^(?:ip\s+)?arp\s+inspection\s+vlan\s+(.+)$", hdr, re.I)
            if m:
                dai["enabled"] = True
                dai["vlans"]  += self._expand_vlans(m.group(1))
        self.arp_inspection = dai

    # ── Spanning Tree ─────────────────────────────────────────────────────────
    def _parse_spanning_tree(self, blocks: list[dict]):
        st: dict = {
            "enabled":  False,
            "mode":     "",
            "priority": "",
            "line":     0,
        }
        for blk in blocks:
            hdr = blk["header"]
            if re.match(r"^spanning-tree$", hdr, re.I):
                st["enabled"] = True
                st["line"]    = blk["lineno"]
                st["mode"]    = self._sub_val(blk["sub"], r"^spanning-tree\s+mode\s+(\S+)") or \
                                self._sub_val(blk["sub"], r"^mode\s+(\S+)")
                st["priority"] = self._sub_val(blk["sub"],
                                              r"^(?:spanning-tree\s+)?priority\s+(\d+)")
                continue
            # Some AOS-CX versions have these as top-level
            m = re.match(r"^spanning-tree\s+mode\s+(\S+)", hdr, re.I)
            if m:
                st["enabled"] = True
                st["mode"]    = m.group(1)
                continue
            m = re.match(r"^spanning-tree\s+priority\s+(\d+)", hdr, re.I)
            if m:
                st["priority"] = m.group(1)
        self.spanning_tree = st

    def _parse_password_policy(self, blocks: list[dict]):
        """Parse password policy, account lockout, session timeout, and NTP auth settings."""
        pp: dict = {
            "login_attempts":     0,  "lockout_time":    0,
            "min_length":         0,
            "complexity_upper":   0,  "complexity_lower":   0,
            "complexity_numeric": 0,  "complexity_special": 0,
            "password_age":       0,
            "session_timeout":    0,
            "ntp_auth_enabled":   False,
            "ntp_auth_keys":      [],
            "line_lockout":       None,
            "line_length":        None,
            "line_session":       None,
            "line_age":           None,
        }
        for blk in blocks:
            hdr = blk["header"]
            m = re.match(r"^aaa\s+authentication\s+login-attempts\s+(\d+)", hdr, re.I)
            if m:
                pp["login_attempts"] = int(m.group(1))
                pp["line_lockout"]   = blk["lineno"]
                continue
            m = re.match(r"^aaa\s+authentication\s+login-lockout-time\s+(\d+)", hdr, re.I)
            if m:
                pp["lockout_time"] = int(m.group(1))
                continue
            m = re.match(r"^password\s+min-length\s+(\d+)", hdr, re.I)
            if m:
                pp["min_length"]  = int(m.group(1))
                pp["line_length"] = blk["lineno"]
                continue
            m = re.match(r"^password\s+complexity\s+min-uppercase\s+(\d+)", hdr, re.I)
            if m:
                pp["complexity_upper"] = int(m.group(1))
                continue
            m = re.match(r"^password\s+complexity\s+min-lowercase\s+(\d+)", hdr, re.I)
            if m:
                pp["complexity_lower"] = int(m.group(1))
                continue
            m = re.match(r"^password\s+complexity\s+min-numeric\s+(\d+)", hdr, re.I)
            if m:
                pp["complexity_numeric"] = int(m.group(1))
                continue
            m = re.match(r"^password\s+complexity\s+min-special-char\s+(\d+)", hdr, re.I)
            if m:
                pp["complexity_special"] = int(m.group(1))
                continue
            m = re.match(r"^password\s+age\s+(\d+)", hdr, re.I)
            if m:
                pp["password_age"] = int(m.group(1))
                pp["line_age"] = blk["lineno"]
                continue
            m = re.match(r"^(?:cli\s+)?session-timeout\s+(\d+)", hdr, re.I)
            if m:
                pp["session_timeout"] = int(m.group(1))
                pp["line_session"]    = blk["lineno"]
                continue
            if re.match(r"^ntp\s+authentication$", hdr, re.I):
                pp["ntp_auth_enabled"] = True
                continue
            m = re.match(r"^ntp\s+authentication-key\s+(\d+)\s+(md5|sha1|sha256)\s+\S+", hdr, re.I)
            if m:
                pp["ntp_auth_keys"].append(
                    {"id": m.group(1), "type": m.group(2), "line": blk["lineno"]})
                continue
        self.password_policy = pp

    # ── VLAN range expander ───────────────────────────────────────────────────
    @staticmethod
    def _expand_vlans(vlan_str: str) -> list[int]:
        result = []
        for part in vlan_str.split(","):
            part = part.strip()
            if "-" in part:
                try:
                    lo, hi = part.split("-", 1)
                    result.extend(range(int(lo), int(hi) + 1))
                except ValueError:
                    pass
            else:
                try:
                    result.append(int(part))
                except ValueError:
                    pass
        return result

    # ── Security checks ───────────────────────────────────────────────────────
    def _run_checks(self):
        self._chk_credentials()
        self._chk_aaa()
        self._chk_ssh_telnet()
        self._chk_snmp()
        self._chk_dhcp_snooping()
        self._chk_arp_inspection()
        self._chk_spanning_tree()
        self._chk_interfaces()
        self._chk_ntp()
        self._chk_syslog()
        self._chk_system()
        self._chk_password_policy()
        self._chk_ssh_version()
        self._chk_acl_insecure_protocols()
        self._chk_acl_default_deny()
        self._chk_syslog_tls()

    # Check: plaintext credentials ────────────────────────────────────────────
    def _chk_credentials(self):
        for u in self.users:
            if u["pwd_type"] == "plaintext":
                self._issue(
                    "CRITICAL", "Plaintext User Password", f"user: {u['name']}",
                    f"User '{u['name']}' has a plaintext password stored in the config.",
                    "Use ciphertext (hashed) password storage. Change the password immediately.",
                    line=u["line"])
            elif u["pwd_type"] == "none":
                self._issue(
                    "HIGH", "Account Without Password", f"user: {u['name']}",
                    f"User '{u['name']}' has no password configured.",
                    "Set a strong password with ciphertext storage.",
                    line=u["line"])

        for srv in self.radius_servers:
            if srv["key_type"] == "plaintext":
                self._issue(
                    "CRITICAL", "Plaintext RADIUS Key", f"RADIUS: {srv['host']}",
                    f"RADIUS server {srv['host']} has a plaintext shared secret in the config.",
                    "Change to ciphertext storage or use a secrets manager. Rotate the key.",
                    line=srv["line"])

        for srv in self.tacacs_servers:
            if srv["key_type"] == "plaintext":
                self._issue(
                    "CRITICAL", "Plaintext TACACS+ Key", f"TACACS+: {srv['host']}",
                    f"TACACS+ server {srv['host']} has a plaintext key in the config.",
                    "Change to ciphertext storage. Rotate the key.",
                    line=srv["line"])

    # Check: AAA ──────────────────────────────────────────────────────────────
    def _chk_aaa(self):
        has_radius  = bool(self.radius_servers)
        has_tacacs  = bool(self.tacacs_servers)
        has_ext_aaa = has_radius or has_tacacs

        login_methods = self.aaa.get("login_default", [])
        console_methods = self.aaa.get("login_console", [])

        if not has_ext_aaa:
            self._issue(
                "HIGH", "No RADIUS or TACACS+ Configured", "AAA",
                "No external authentication servers (RADIUS or TACACS+) are configured.",
                "Configure RADIUS or TACACS+ for centralized authentication, authorisation, "
                "and accounting of all management access.")

        # Console using only local auth while external is configured
        if has_ext_aaa and console_methods and \
                all(m.lower() == "local" for m in console_methods):
            self._issue(
                "MEDIUM", "Local Auth Only on Console", "AAA / Console",
                "Console login uses local authentication only even though an AAA server is configured.",
                "Add RADIUS/TACACS+ to the console authentication method list as primary method.")

        # No accounting
        if not self.aaa.get("accounting"):
            self._issue(
                "MEDIUM", "No MFA / AAA for Management", "AAA Accounting",
                "AAA accounting is not configured; management sessions are not centrally logged.",
                "Configure 'aaa accounting' for commands and exec sessions to the AAA server.")

    # Check: SSH / Telnet ─────────────────────────────────────────────────────
    def _chk_ssh_telnet(self):
        if not self.system.get("ssh_vrfs"):
            self._issue(
                "HIGH", "SSH Not Enabled", "Management",
                "SSH server is not enabled on any VRF.",
                "Enable SSH on the management VRF: 'ssh server vrf mgmt'.")

        if self.system.get("telnet_enabled"):
            self._issue(
                "HIGH", "Telnet Enabled", "Management",
                "Telnet server is enabled — credentials and session data transmitted in cleartext.",
                "Disable the Telnet server and use SSH exclusively.")

    # Check: SNMP ─────────────────────────────────────────────────────────────
    def _chk_snmp(self):
        DEFAULT_COMMUNITIES = {"public", "private", "cisco", "community", "snmp", "admin"}
        communities = self.snmp.get("communities", [])

        for c in communities:
            name_lc = c["name"].lower()
            if name_lc in DEFAULT_COMMUNITIES:
                self._issue(
                    "CRITICAL", "Default/Weak SNMP Community String",
                    f"SNMP community: {c['name']}",
                    f"SNMP community '{c['name']}' is a well-known default value.",
                    "Replace with a long random community string. Migrate to SNMPv3 authPriv.",
                    line=c["line"])

            if c["access"] == "rw":
                self._issue(
                    "HIGH", "SNMP Read-Write Community",
                    f"SNMP community: {c['name']}",
                    f"Community '{c['name']}' has read-write (RW) access, allowing config changes via SNMP.",
                    "Remove RW community or restrict to NMS host ACL. Prefer SNMPv3 authPriv.",
                    line=c["line"])

            if not c.get("acl"):
                self._issue(
                    "MEDIUM", "SNMP Without ACL Restriction",
                    f"SNMP community: {c['name']}",
                    f"Community '{c['name']}' has no source ACL restriction.",
                    "Apply an ACL to limit SNMP access to authorised NMS hosts only.",
                    line=c["line"])

        # v2c vs v3
        has_v3 = bool(self.snmp.get("v3_users"))
        if communities and not has_v3:
            self._issue(
                "MEDIUM", "SNMPv2c in Use", "SNMP",
                "SNMPv2c communities are in use with no SNMPv3 users configured.",
                "Migrate to SNMPv3 with authPriv (AES-128+, SHA-256) for encrypted, "
                "authenticated SNMP management.")

        # Check trap hosts using v1/v2c
        for th in self.snmp.get("trap_hosts", []):
            if th["version"].lower() in ("1", "2c"):
                self._issue(
                    "MEDIUM", "SNMPv2c in Use",
                    f"SNMP trap host: {th['host']}",
                    f"SNMP traps to {th['host']} use version {th['version']} (cleartext).",
                    "Migrate trap delivery to SNMPv3 informs.",
                    line=th["line"])

    # Check: DHCP Snooping ────────────────────────────────────────────────────
    def _chk_dhcp_snooping(self):
        if not self.dhcp_snooping.get("enabled"):
            self._issue(
                "HIGH", "DHCP Snooping Not Enabled", "DHCP Snooping",
                "DHCP snooping is not enabled. Rogue DHCP servers can redirect client traffic.",
                "Enable DHCP snooping globally and per user VLAN: "
                "'dhcp-snooping' and 'dhcp-snooping vlan <id>'.",
                line=self.dhcp_snooping.get("line", ""))

    # Check: Dynamic ARP Inspection ───────────────────────────────────────────
    def _chk_arp_inspection(self):
        if not self.arp_inspection.get("enabled"):
            self._issue(
                "HIGH", "Dynamic ARP Inspection Not Enabled", "ARP Inspection",
                "Dynamic ARP Inspection (DAI) is not enabled. ARP spoofing / MITM attacks possible.",
                "Enable DAI globally and per user VLAN: "
                "'arp inspection' and 'arp inspection vlan <id>'. "
                "Requires DHCP snooping or static ARP ACLs.",
                line=self.arp_inspection.get("line", ""))

    # Check: Spanning Tree ────────────────────────────────────────────────────
    def _chk_spanning_tree(self):
        if not self.spanning_tree.get("enabled"):
            self._issue(
                "HIGH", "Spanning Tree Not Enabled", "Spanning Tree",
                "Spanning Tree Protocol is not explicitly enabled. "
                "Layer 2 loops could bring down the network.",
                "Enable STP: 'spanning-tree' with mode 'mstp' or 'rapid-pvst'.")

    # Check: per-interface ────────────────────────────────────────────────────
    def _chk_interfaces(self):
        ds_vlans  = set(self.dhcp_snooping.get("vlans", []))
        dai_vlans = set(self.arp_inspection.get("vlans", []))

        for iface in self.interfaces:
            name = iface["name"]
            itype = iface["type"]
            is_physical = itype == "physical"
            is_access   = iface["vlan_mode"] == "access" and is_physical
            is_trunk    = iface["vlan_mode"] == "trunk"  and is_physical
            is_up       = iface["admin_up"]

            # Missing descriptions on physical/LAG ports
            if is_physical and not iface["description"]:
                self._issue(
                    "LOW", "Interface Missing Description", f"interface {name}",
                    f"Interface {name} has no description.",
                    "Add a description identifying the connected device or purpose.",
                    line=iface["line"])

            # Physical ports that appear unused (no VLAN, no IP) but are admin up
            if is_physical and is_up and not iface["vlan_mode"] and \
                    not iface["ip_addresses"]:
                self._issue(
                    "MEDIUM", "Active Interface Not Shut Down", f"interface {name}",
                    f"Interface {name} is administratively up but has no VLAN or IP configured.",
                    "Shut down unused ports: 'shutdown' under the interface.",
                    line=iface["line"])

            if is_access and is_up:
                vlan_id = int(iface["access_vlan"]) if iface["access_vlan"].isdigit() else 0

                # VLAN 1 on access port
                if vlan_id == 1:
                    self._issue(
                        "MEDIUM", "VLAN 1 in Use on Access Port", f"interface {name}",
                        f"Access port {name} is on VLAN 1 (default VLAN). "
                        "VLAN 1 is often used for management and should not carry user traffic.",
                        "Move user devices to a dedicated VLAN; reserve VLAN 1 for management "
                        "or disable it on access ports.",
                        line=iface["line"])

                # BPDU guard on access ports
                if not iface["bpduguard"]:
                    self._issue(
                        "MEDIUM", "BPDU Guard Not on Access Port", f"interface {name}",
                        f"Access port {name} does not have BPDU guard enabled.",
                        "Enable BPDU guard on all access (edge) ports: "
                        "'spanning-tree bpduguard enable'.",
                        line=iface["line"])

                # IP source guard for snooped VLANs
                if vlan_id and vlan_id in ds_vlans and not iface["ip_src_guard"]:
                    self._issue(
                        "MEDIUM", "IP Source Guard Not Enabled", f"interface {name}",
                        f"Interface {name} is on a DHCP-snooped VLAN but has no IP source guard.",
                        "Enable IP source guard: 'ip source-guard' on interfaces "
                        "where DHCP snooping is active.",
                        line=iface["line"])

                # No 802.1X or MAC auth on access ports
                if not iface["dot1x"] and not iface["mac_auth"] and not iface["port_security"]:
                    self._issue(
                        "LOW", "No Port Security / 802.1X", f"interface {name}",
                        f"Access port {name} has no 802.1X, MAC authentication, or port security.",
                        "Implement 802.1X ('aaa authentication port-access dot1x') "
                        "or MAC auth for NAC on all access ports.",
                        line=iface["line"])

            if is_trunk and is_up:
                # Native VLAN 1 on trunk
                if iface["native_vlan"] == "1" or iface["native_vlan"] == "":
                    self._issue(
                        "MEDIUM", "Trunk Native VLAN Is VLAN 1", f"interface {name}",
                        f"Trunk port {name} uses VLAN 1 as native VLAN (default). "
                        "Native VLAN on a trunk carries untagged frames and can be exploited "
                        "for VLAN hopping.",
                        "Set native VLAN to an unused dedicated VLAN: "
                        "'vlan trunk native <unused-vlan>'.",
                        line=iface["line"])

                # Root guard on trunk ports (uplinks)
                if not iface["root_guard"]:
                    self._issue(
                        "LOW", "STP Root Guard Not Configured", f"interface {name}",
                        f"Trunk/uplink port {name} does not have STP root guard enabled.",
                        "Enable root guard on uplink/distribution ports to prevent a rogue "
                        "switch from becoming the STP root: 'spanning-tree root-guard'.",
                        line=iface["line"])

    # Check: NTP ──────────────────────────────────────────────────────────────
    def _chk_ntp(self):
        if not self.ntp_servers:
            self._issue(
                "MEDIUM", "NTP Not Configured", "NTP",
                "No NTP servers are configured. Inaccurate timestamps break log correlation, "
                "certificate validation, and RADIUS/TACACS+ accounting.",
                "Configure at least two NTP servers: 'ntp server <ip>'.")
            return

        if len(self.ntp_servers) == 1:
            self._issue("LOW", "Only One NTP Server", "NTP",
                "Only one NTP server is configured. Loss of this server leaves the switch without time sync.",
                "Add a second NTP server: 'ntp server <ip>'.",
                f"Current: {self.ntp_servers[0]}")

    # Check: Syslog ───────────────────────────────────────────────────────────
    def _chk_syslog(self):
        if not self.syslog_servers:
            self._issue(
                "MEDIUM", "No Syslog Servers Configured", "Logging",
                "No remote syslog servers are configured. Logs reside on the device only "
                "and will be lost on reboot or if the device is compromised.",
                "Configure at least one remote syslog server: 'logging remote <ip>'.")

    # Check: system hygiene ───────────────────────────────────────────────────
    def _chk_system(self):
        if not self.system.get("location"):
            self._issue("LOW", "No System Location", "System",
                "No system location is configured.",
                "Set 'system-location \"<location>\"' for asset tracking and SNMP MIB.")

        if not self.system.get("contact"):
            self._issue("LOW", "No System Contact", "System",
                "No system contact is configured.",
                "Set 'system-contact \"<email>\"' for accountability and SNMP MIB.")

        if not self.system.get("banner"):
            self._issue("MEDIUM", "No Login/MOTD Banner", "System",
                "No login or MOTD banner is configured.",
                "Configure a legal warning banner: 'banner motd <delim>\\n...\\n<delim>'.")

        if not self.system.get("dns_servers"):
            self._issue("LOW", "DNS Not Configured", "System",
                "No DNS server is configured on the switch.",
                "Configure 'ip dns server-address <ip>' for name resolution in ACLs and logging.")

    def _chk_password_policy(self):
        pp = self.password_policy
        if not pp:
            return

        if not pp.get("login_attempts"):
            self._issue("HIGH", "No Account Lockout Policy", "Password Policy",
                "No login-attempt limit is configured — brute-force attacks face no lockout.",
                "Set 'aaa authentication login-attempts 3' (or ≤ 5) and "
                "'aaa authentication login-lockout-time 300' (≥ 5 minutes).")
        elif pp.get("login_attempts", 99) > 5:
            self._issue("MEDIUM", "No Account Lockout Policy", "Password Policy",
                f"Login attempt limit is {pp['login_attempts']} (recommended ≤ 5).",
                "Reduce 'aaa authentication login-attempts' to 5 or fewer.",
                line=pp.get("line_lockout"))

        if not pp.get("min_length"):
            self._issue("HIGH", "Weak Password Minimum Length", "Password Policy",
                "No minimum password length is configured.",
                "Set 'password min-length 12' or higher.")
        elif pp.get("min_length", 0) < 12:
            self._issue("MEDIUM", "Weak Password Minimum Length", "Password Policy",
                f"Minimum password length is {pp['min_length']} characters (recommended ≥ 12).",
                "Increase 'password min-length' to 12 or more.",
                line=pp.get("line_length"))

        has_complexity = any(pp.get(k, 0) > 0 for k in
                             ("complexity_upper", "complexity_lower",
                              "complexity_numeric", "complexity_special"))
        if not has_complexity:
            self._issue("MEDIUM", "Password Complexity Not Enforced", "Password Policy",
                "No password complexity requirements are configured.",
                "Set 'password complexity min-uppercase 1', 'min-lowercase 1', "
                "'min-numeric 1', 'min-special-char 1'.")

        timeout = pp.get("session_timeout", 0)
        if timeout == 0:
            self._issue("HIGH", "Management Session Timeout Not Set", "Password Policy",
                "No session/idle timeout is configured — management sessions never expire.",
                "Set 'session-timeout 900' (15 minutes) or 'cli session-timeout 15'.")
        elif timeout > 1800:
            self._issue("MEDIUM", "Management Session Timeout Not Set", "Password Policy",
                f"Session timeout is {timeout} seconds ({timeout // 60} min) — too long.",
                "Reduce to 'session-timeout 900' (15 minutes) or less.",
                line=pp.get("line_session"))

        if self.ntp_servers and not pp.get("ntp_auth_enabled"):
            self._issue("LOW", "NTP Authentication Not Configured", "NTP",
                "NTP authentication is not enabled. The switch accepts time from any source, "
                "making it vulnerable to NTP spoofing attacks.",
                "Enable NTP authentication: 'ntp authentication', "
                "'ntp authentication-key <id> sha256 <key>', 'ntp trusted-key <id>'.")

        age = pp.get("password_age", 0)
        if age == 0:
            self._issue("MEDIUM", "Password Expiry Not Configured", "Password Policy",
                "No password age limit is configured. Passwords never expire.",
                "Set 'password age 90' (days) or less.")
        elif age > 90:
            self._issue("MEDIUM", "Password Expiry Not Configured", "Password Policy",
                f"Password expiry is {age} days; passwords should be rotated every 90 days or fewer.",
                "Reduce 'password age' to 90 days or fewer.",
                line=pp.get("line_age") or "")

    def _chk_ssh_version(self):
        """PCI DSS 4.2.1: Only SSH v2 should be permitted."""
        if self.system.get("ssh_vrfs") and not self.system.get("ssh_v1_disabled"):
            self._issue("HIGH", "SSH v1 Enabled", "SSH",
                "SSH is enabled but SSHv1 has not been explicitly disabled. "
                "SSHv1 has known cryptographic weaknesses and should not be permitted.",
                "Disable SSHv1: add 'no ssh server v1' or 'ssh server version 2'.",
                line=self.system.get("ssh_line", ""))

    def _chk_acl_default_deny(self):
        """PCI DSS 1.2.4: ACLs should end with an explicit deny-all."""
        for acl in self.acls:
            if not acl["entries"]:
                continue
            last = acl["entries"][-1]
            match_lower = last["match"].lower().strip()
            is_deny = last["action"] == "deny"
            is_any  = match_lower in ("any any any", "any", "ip any any")
            if not (is_deny and is_any):
                self._issue("MEDIUM", "ACL Missing Default Deny", f"ACL: {acl['name']}",
                    f"ACL '{acl['name']}' does not end with an explicit deny-all entry. "
                    "Traffic not matched by any rule will be permitted by the implicit permit-any.",
                    "Add a final entry: '<seq> deny any any any' to explicitly block "
                    "all unmatched traffic.",
                    f"Last entry: seq {last['seq']} {last['action']} {last['match']}",
                    line=last["line"])

    def _chk_syslog_tls(self):
        """PCI DSS 10.5.4: Syslog should use encrypted transport."""
        for srv in self.syslog_servers:
            if not srv.get("tls"):
                self._issue("MEDIUM", "Syslog Not Using TLS", f"Syslog: {srv['host']}",
                    f"Syslog server {srv['host']} does not use TLS transport. "
                    "Log data transmitted in cleartext can be intercepted or tampered with in transit.",
                    "Configure syslog with TLS: 'logging remote <ip> port 6514 tls'.",
                    line=srv["line"])

    def _chk_acl_insecure_protocols(self):
        INSECURE: dict[str, tuple[str, str, str]] = {
            "23":     ("HIGH",   "telnet", "Telnet — plaintext credential protocol"),
            "21":     ("MEDIUM", "ftp",    "FTP — plaintext credential/data protocol"),
            "69":     ("MEDIUM", "tftp",   "TFTP — unauthenticated file transfer"),
            "513":    ("HIGH",   "rlogin", "rlogin — unauthenticated remote login"),
            "514":    ("HIGH",   "rsh",    "RSH — Remote Shell, no encryption"),
            "telnet": ("HIGH",   "telnet", "Telnet — plaintext credential protocol"),
            "ftp":    ("MEDIUM", "ftp",    "FTP — plaintext credential/data protocol"),
            "tftp":   ("MEDIUM", "tftp",   "TFTP — unauthenticated file transfer"),
        }
        for acl in self.acls:
            for entry in acl["entries"]:
                if entry["action"] != "permit":
                    continue
                ml = entry["match"].lower()
                for token, (sev, app_name, desc) in INSECURE.items():
                    if re.search(rf"\beq\s+{re.escape(token)}\b", ml):
                        self._issue(sev, "ACL Permits Insecure Protocol",
                            f"ACL {acl['name']} seq {entry['seq']}",
                            f"ACL '{acl['name']}' entry {entry['seq']} permits {desc}.",
                            f"Block {app_name} (port {token}). "
                            "Replace with SSH/SFTP if remote access is needed.",
                            f"Match: {entry['match']}",
                            line=entry["line"])
                        break  # one finding per entry is enough


# ── Excel report writer ───────────────────────────────────────────────────────
class ExcelReporter:
    SEV_ORDER  = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"],   C["critical_l"]),
        "HIGH":     (C["high"],       C["high_l"]),
        "MEDIUM":   (C["medium"],     C["medium_l"]),
        "LOW":      (C["low"],        C["low_l"]),
        "INFO":     (C["info"],       C["info_l"]),
    }

    def __init__(self, parser: ArubaCXParser, output_file: str):
        self.p   = parser
        self.out = output_file
        self.wb  = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _hdr(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill      = _fill(C["hdr_bg"])
            c.font      = _font(bold=True, color=C["hdr_fg"])
            c.alignment = _align("center", wrap=False)
            c.border    = THIN
        ws.row_dimensions[row].height = 28

    def _set_widths(self, ws, widths: list[int | float]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _row_fill(self, row_idx: int) -> str | None:
        return C["alt_row"] if row_idx % 2 == 0 else None

    def _wr(self, ws, row: int, col: int, value, bold=False,
            bg: str | None = None, fg: str = "000000",
            h: str = "left", border: bool = True):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(bold=bold, color=fg)
        c.alignment = _align(h)
        if bg:
            c.fill = _fill(bg)
        if border:
            c.border = THIN
        return c

    # ── Summary ───────────────────────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Aruba CX Switch — Configuration Security Report"
        c.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        c.fill  = _fill(C["hdr_bg"])
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 42

        ws.merge_cells("A2:F2")
        c = ws["A2"]
        c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    "
                   f"Source: {os.path.basename(self.p.config_file)}    "
                   f"Hostname: {self.p.system.get('hostname', '(unknown)')}")
        c.font  = _font(italic=True, color=C["info"], size=9)
        c.fill  = _fill("F2F2F2")
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 18

        row = 4
        p = self.p

        def section(label):
            nonlocal row
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font  = _font(bold=True, color=C["hdr_bg"], size=12)
            row += 1

        def kv(label, value, good: bool | None = None):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font(); c1.alignment = _align()
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = _font(bold=True); c2.alignment = _align("center")
            if good is True:
                c2.fill = _fill(C["ok_l"]); c2.font = _font(bold=True, color=C["ok"])
            elif good is False:
                c2.fill = _fill(C["high_l"]); c2.font = _font(bold=True, color=C["high"])
            elif row % 2 == 0:
                c1.fill = c2.fill = _fill(C["alt_row"])
            row += 1

        section("DEVICE OVERVIEW")
        kv("Hostname",            p.system.get("hostname", "(not set)"))
        kv("Version (config)",    p.system.get("version", "(not found)"))
        kv("Location",            p.system.get("location") or "(not set)",
                                  bool(p.system.get("location")))
        kv("Contact",             p.system.get("contact") or "(not set)",
                                  bool(p.system.get("contact")))
        kv("SSH VRFs",            ", ".join(p.system.get("ssh_vrfs", [])) or "NONE",
                                  bool(p.system.get("ssh_vrfs")))
        kv("Telnet",              "Enabled" if p.system.get("telnet_enabled") else "Disabled",
                                  not p.system.get("telnet_enabled"))
        kv("NTP Servers",         ", ".join(p.ntp_servers) or "NONE",
                                  bool(p.ntp_servers))
        kv("Syslog Servers",      str(len(p.syslog_servers)),
                                  bool(p.syslog_servers))
        kv("Interfaces",          len(p.interfaces))
        kv("VLANs",               len(p.vlans))
        kv("ACLs",                len(p.acls))
        kv("RADIUS Servers",      len(p.radius_servers),
                                  bool(p.radius_servers))
        kv("TACACS+ Servers",     len(p.tacacs_servers))
        kv("SNMP Communities",    len(p.snmp.get("communities", [])))
        kv("SNMPv3 Users",        len(p.snmp.get("v3_users", [])))
        kv("DHCP Snooping",       "Enabled" if p.dhcp_snooping.get("enabled") else "Disabled",
                                  p.dhcp_snooping.get("enabled"))
        kv("ARP Inspection (DAI)","Enabled" if p.arp_inspection.get("enabled") else "Disabled",
                                  p.arp_inspection.get("enabled"))
        kv("Spanning Tree",       "Enabled" if p.spanning_tree.get("enabled") else "Not found",
                                  p.spanning_tree.get("enabled"))
        row += 1

        section("SECURITY FINDINGS BY SEVERITY")
        self._hdr(ws, ["Severity", "Count"], row=row)
        row += 1
        total = 0
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            cnt = sum(1 for i in p.issues if i["severity"] == sev)
            total += cnt
            fg, bg = self.SEV_COLORS[sev]
            for col, val in enumerate([sev, cnt], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = _fill(bg); c.border = THIN
                c.font = _font(bold=True, color=fg)
                c.alignment = _align("center")
            row += 1
        for col, val in enumerate(["TOTAL", total], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _font(bold=True); c.border = THIN
            c.alignment = _align("center")
        row += 2

        section("FINDINGS BY CATEGORY")
        self._hdr(ws, ["Category", "Count"], row=row)
        row += 1
        cat_counts: dict[str, int] = defaultdict(int)
        for iss in p.issues:
            cat_counts[iss["category"]] += 1
        for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
            rb = C["alt_row"] if row % 2 == 0 else None
            c1 = ws.cell(row=row, column=1, value=cat)
            c1.font = _font()
            if rb:
                c1.fill = _fill(rb)
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.font = _font(bold=True); c2.alignment = _align("center")
            if rb:
                c2.fill = _fill(rb)
            row += 1

        self._set_widths(ws, [38, 18, 18, 18, 18, 18])

    # ── Interfaces ────────────────────────────────────────────────────────────
    def _sheet_interfaces(self):
        ws = self.wb.create_sheet("Interfaces")
        ws.sheet_view.showGridLines = False
        headers = [
            "Interface", "Type", "Line #", "Admin",
            "VLAN Mode", "Access VLAN", "Native VLAN", "Trunk VLANs",
            "IP Address(es)", "Description",
            "BPDU Guard", "Root Guard", "Port Type", "Loop Guard",
            "DHCP Trust", "ARP Trust", "IP Src Guard",
            "802.1X", "MAC Auth", "Port Security", "Storm Control",
            "ACL In", "ACL Out", "IPv6 ACL In", "IPv6 ACL Out",
            "LLDP TX", "LLDP RX", "MTU",
        ]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        def yn(b: bool) -> str:
            return "yes" if b else ""

        for i, iface in enumerate(self.p.interfaces, 2):
            rb = self._row_fill(i)
            vals = [
                iface["name"], iface["type"], iface["line"],
                "up" if iface["admin_up"] else "down",
                iface["vlan_mode"], iface["access_vlan"],
                iface["native_vlan"], iface["trunk_vlans"],
                iface["ip_addresses"], iface["description"],
                yn(iface["bpduguard"]), yn(iface["root_guard"]),
                iface["port_type"], yn(iface["loop_guard"]),
                yn(iface["dhcp_trust"]), yn(iface["arp_trust"]),
                yn(iface["ip_src_guard"]),
                yn(iface["dot1x"]), yn(iface["mac_auth"]),
                yn(iface["port_security"]), yn(iface["storm_control"]),
                iface["acl_in"], iface["acl_out"],
                iface["acl6_in"], iface["acl6_out"],
                yn(iface["lldp_tx"]), yn(iface["lldp_rx"]),
                iface["mtu"],
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(color=C["disabled"] if not iface["admin_up"] else "000000")
                c.alignment = _align("center" if col in (1, 2, 3, 4, 5, 6) else "left")
                c.border = THIN
                if rb:
                    c.fill = _fill(rb)

        widths = [18, 10, 8, 8,  10, 12, 12, 30,  20, 30,
                  11, 11, 14, 11,  11, 11, 12,
                  8, 10, 13, 13,
                  15, 15, 15, 15,
                  10, 10, 8]
        self._set_widths(ws, widths)

    # ── VLANs ─────────────────────────────────────────────────────────────────
    def _sheet_vlans(self):
        ws = self.wb.create_sheet("VLANs")
        ws.sheet_view.showGridLines = False
        headers = ["VLAN ID", "Name", "Description", "Line #"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"

        for i, v in enumerate(self.p.vlans, 2):
            rb = self._row_fill(i)
            for col, val in enumerate([v["id"], v["name"], v["description"], v["line"]], 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)

        self._set_widths(ws, [10, 30, 45, 10])

    # ── ACLs ──────────────────────────────────────────────────────────────────
    def _sheet_acls(self):
        ws = self.wb.create_sheet("ACLs")
        ws.sheet_view.showGridLines = False
        headers = ["ACL Name", "Type", "Seq", "Action", "Match", "Line #"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        row = 2
        for acl in self.p.acls:
            if not acl["entries"]:
                rb = self._row_fill(row)
                for col, val in enumerate([acl["name"], acl["type"], "", "", "(empty)", acl["line"]], 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = _font(italic=True, color=C["info"]); c.border = THIN
                    if rb:
                        c.fill = _fill(rb)
                row += 1
            else:
                for ace in acl["entries"]:
                    rb = self._row_fill(row)
                    act = ace["action"].lower()
                    act_fg = C["ok"] if act == "permit" else C["critical"]
                    act_bg = C["ok_l"] if act == "permit" else C["critical_l"]
                    for col, val in enumerate(
                            [acl["name"], acl["type"], ace["seq"],
                             ace["action"].upper(), ace["match"], ace["line"]], 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.alignment = _align(); c.border = THIN
                        if col == 4:
                            c.fill = _fill(act_bg); c.font = _font(bold=True, color=act_fg)
                        else:
                            c.font = _font()
                            if rb:
                                c.fill = _fill(rb)
                    row += 1

        self._set_widths(ws, [28, 8, 8, 10, 60, 8])

    # ── AAA & Management ──────────────────────────────────────────────────────
    def _sheet_aaa(self):
        ws = self.wb.create_sheet("AAA & Management")
        ws.sheet_view.showGridLines = False

        def section(title: str, headers: list[str]) -> int:
            nonlocal _row
            ws.cell(row=_row, column=1).value = title
            ws.cell(row=_row, column=1).font  = _font(bold=True, color=C["hdr_bg"], size=11)
            ws.row_dimensions[_row].height = 20
            _row += 1
            self._hdr(ws, headers, row=_row)
            _row += 1
            return _row

        _row = 1
        p = self.p

        # AAA methods
        section("AAA AUTHENTICATION", ["Method List", "Methods"])
        for label, methods in [
            ("login default",  p.aaa.get("login_default", [])),
            ("login console",  p.aaa.get("login_console", [])),
            ("enable default", p.aaa.get("enable_default", [])),
        ]:
            rb = self._row_fill(_row)
            for col, val in enumerate([label, " → ".join(methods) or "(not configured)"], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            _row += 1
        _row += 1

        # RADIUS servers
        section("RADIUS SERVERS",
                ["Host", "Key Type", "VRF", "Line #"])
        for srv in p.radius_servers:
            rb = self._row_fill(_row)
            for col, val in enumerate([srv["host"], srv["key_type"], srv["vrf"], srv["line"]], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            if srv["key_type"] == "plaintext":
                ws.cell(row=_row, column=2).fill = _fill(C["critical_l"])
                ws.cell(row=_row, column=2).font = _font(bold=True, color=C["critical"])
            _row += 1
        if not p.radius_servers:
            ws.cell(row=_row, column=1).value = "(none configured)"
            ws.cell(row=_row, column=1).font = _font(italic=True, color=C["info"])
            _row += 1
        _row += 1

        # TACACS+ servers
        section("TACACS+ SERVERS", ["Host", "Key Type", "VRF", "Line #"])
        for srv in p.tacacs_servers:
            rb = self._row_fill(_row)
            for col, val in enumerate([srv["host"], srv["key_type"], srv["vrf"], srv["line"]], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            if srv["key_type"] == "plaintext":
                ws.cell(row=_row, column=2).fill = _fill(C["critical_l"])
                ws.cell(row=_row, column=2).font = _font(bold=True, color=C["critical"])
            _row += 1
        if not p.tacacs_servers:
            ws.cell(row=_row, column=1).value = "(none configured)"
            ws.cell(row=_row, column=1).font = _font(italic=True, color=C["info"])
            _row += 1
        _row += 1

        # Local users
        section("LOCAL USERS", ["Username", "Group", "Password Type", "Line #"])
        for u in p.users:
            rb = self._row_fill(_row)
            for col, val in enumerate([u["name"], u["group"], u["pwd_type"], u["line"]], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            if u["pwd_type"] == "plaintext":
                ws.cell(row=_row, column=3).fill = _fill(C["critical_l"])
                ws.cell(row=_row, column=3).font = _font(bold=True, color=C["critical"])
            elif u["pwd_type"] == "none":
                ws.cell(row=_row, column=3).fill = _fill(C["high_l"])
                ws.cell(row=_row, column=3).font = _font(bold=True, color=C["high"])
            _row += 1
        if not p.users:
            ws.cell(row=_row, column=1).value = "(none found)"
            ws.cell(row=_row, column=1).font = _font(italic=True, color=C["info"])
            _row += 1
        _row += 1

        # SNMP
        section("SNMP COMMUNITIES",
                ["Community Name", "Access", "ACL", "Line #"])
        for c_obj in p.snmp.get("communities", []):
            rb = self._row_fill(_row)
            for col, val in enumerate([c_obj["name"], c_obj["access"],
                                       c_obj["acl"] or "(none)", c_obj["line"]], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            if c_obj["access"] == "rw":
                ws.cell(row=_row, column=2).fill = _fill(C["high_l"])
                ws.cell(row=_row, column=2).font = _font(bold=True, color=C["high"])
            if not c_obj.get("acl"):
                ws.cell(row=_row, column=3).fill = _fill(C["medium_l"])
            _row += 1
        if not p.snmp.get("communities"):
            ws.cell(row=_row, column=1).value = "(none configured)"
            ws.cell(row=_row, column=1).font = _font(italic=True, color=C["info"])
            _row += 1
        _row += 1

        # NTP / Syslog
        section("NTP & SYSLOG", ["Type", "Host / Server", "Details"])
        for ntp in p.ntp_servers:
            rb = self._row_fill(_row)
            for col, val in enumerate(["NTP Server", ntp, ""], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            _row += 1
        for srv in p.syslog_servers:
            rb = self._row_fill(_row)
            details = f"severity: {srv['severity']}  vrf: {srv['vrf']}"
            for col, val in enumerate(["Syslog", srv["host"], details], 1):
                c = ws.cell(row=_row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            _row += 1
        if not p.ntp_servers and not p.syslog_servers:
            ws.cell(row=_row, column=1).value = "(none configured)"
            ws.cell(row=_row, column=1).font = _font(italic=True, color=C["info"])
            _row += 1

        self._set_widths(ws, [28, 22, 30, 10])

    # ── Security Issues ───────────────────────────────────────────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False
        headers = ["#", "Validated", "Severity", "Residual Risk", "Residual Risk Note",
                   "Category", "Rule / Object", "Config Line(s)", "CIS v8 Controls", "PCI DSS",
                   "Description", "Recommendation", "Details",
                   "Asset", "Target", "Vuln", "Output", "Source"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        hostname = self.p.system.get("hostname", "") or ""
        source = os.path.basename(self.p.config_file)

        sorted_issues = sorted(self.p.issues,
                               key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            rb = self._row_fill(row)

            obj = iss["object"]
            line = iss.get("line", "")
            target = f"{obj} ({line})" if line else obj
            details = iss.get("details", "")
            output = f"{iss['description']}\n{details}" if details else iss["description"]

            vals = [idx, "Y", sev, "", "",
                    iss["category"], obj, line,
                    iss.get("cis_controls", ""), iss.get("pci_dss", ""),
                    iss["description"], iss["recommendation"], details,
                    hostname, target, "", output, source]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 3:  # Severity
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                    c.alignment = _align("center")
                elif col in (1, 8):  # #, Config Line
                    c.font = _font(bold=(col == 1)); c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 9:  # CIS v8 Controls
                    c.font = _font(bold=True, color="17375E", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 10:  # PCI DSS
                    c.font = _font(bold=True, color="7B2D8B", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 2:  # Validated
                    c.font = _font(bold=True)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                else:
                    c.font = _font(); c.alignment = _align()
                    if rb:
                        c.fill = _fill(rb)
            ws.row_dimensions[row].height = 40

        self._set_widths(ws, [4, 12, 12, 18, 28, 32, 36, 14, 20, 18, 60, 60, 36, 24, 44, 16, 70, 30])

    # ── CIS v8 Mapping ────────────────────────────────────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS v8 Mapping")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v8 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each CIS safeguard lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        ctrl_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        def _sort_ctrl(k):
            parts = k.split(".")
            return (int(parts[0]), float("0." + parts[1]) if len(parts) > 1 else 0)

        row = 4
        for ctrl_id in sorted(CIS_CTRL_DESC.keys(), key=_sort_ctrl):
            desc   = CIS_CTRL_DESC[ctrl_id]
            issues = sorted(ctrl_issues.get(ctrl_id, []),
                            key=lambda x: self.SEV_ORDER.get(x["severity"], 9))

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1, value=f"CIS {ctrl_id} — {desc}")
            hc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill = _fill("17375E"); hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 22
            row += 1

            if not issues:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this control")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9"); nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object / Interface",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF")
                    c.fill = _fill("2E4057"); c.alignment = _align("center", wrap=False)
                    c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1

                for iss in issues:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = self._row_fill(row)
                    for col, val in enumerate(
                            [sev, iss["category"], iss["object"],
                             iss.get("line", ""), iss["description"],
                             iss["recommendation"]], 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb:
                                c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb:
                                c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1

            row += 1

        self._set_widths(ws, [12, 34, 30, 12, 60, 60])

    # ── PCI DSS v4.0 Mapping ──────────────────────────────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each PCI DSS v4.0 requirement lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        from collections import defaultdict as _dd
        req_issues: dict[str, list[dict]] = _dd(list)
        for iss in self.p.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        SEV_ORDER  = self.SEV_ORDER
        SEV_COLORS = self.SEV_COLORS
        row = 3
        for req_id in sorted(PCI_DSS_DESC.keys(),
                              key=lambda x: [int(p) for p in x.split(".")]):
            desc = PCI_DSS_DESC[req_id]
            issues_for_req = sorted(req_issues.get(req_id, []),
                                    key=lambda x: SEV_ORDER.get(x["severity"], 9))
            count = len(issues_for_req)
            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not issues_for_req:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this requirement")
                nc.font = _font(italic=True, color=C["info"]); nc.fill = _fill("F9F9F9")
                nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object / Interface",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in issues_for_req:
                    sev = iss["severity"]
                    fg, bg = SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 30, 12, 60, 60])

    # ── Save ──────────────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_interfaces()
        self._sheet_vlans()
        self._sheet_acls()
        self._sheet_aaa()
        self._sheet_issues()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Aruba CX Switch Config Analyzer — outputs Excel report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python aruba_cx_analyzer.py show-run.txt
  python aruba_cx_analyzer.py show-run.txt -o switch-audit.xlsx

Capture the config from the switch:
  show running-config > running-config.txt
""",
    )
    ap.add_argument("config", help="AOS-CX running-config text file")
    ap.add_argument("-o", "--output", default=None,
                    help="Output Excel file (default: <config-stem>_analysis.xlsx)")
    args = ap.parse_args()

    if not args.output:
        stem = os.path.splitext(os.path.basename(args.config))[0]
        args.output = f"{stem}_analysis.xlsx"

    print(f"[*] Parsing:  {args.config}")
    parser = ArubaCXParser(args.config)
    parser.parse()

    sev_counts: dict[str, int] = defaultdict(int)
    for iss in parser.issues:
        sev_counts[iss["severity"]] += 1

    print(f"[*] Parsed:")
    print(f"      Interfaces     : {len(parser.interfaces)}")
    print(f"      VLANs          : {len(parser.vlans)}")
    print(f"      ACLs           : {len(parser.acls)}")
    print(f"      RADIUS servers : {len(parser.radius_servers)}")
    print(f"      TACACS+ servers: {len(parser.tacacs_servers)}")
    print(f"      Local users    : {len(parser.users)}")
    print(f"      SNMP communities: {len(parser.snmp.get('communities', []))}")
    print(f"[*] Security findings: {len(parser.issues)}")
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        if sev_counts[sev]:
            print(f"      {sev:<10}: {sev_counts[sev]}")

    print(f"[*] Writing Excel report...")
    reporter = ExcelReporter(parser, args.output)
    reporter.save()
    print(f"[+] Saved: {args.output}")


if __name__ == "__main__":
    main()
