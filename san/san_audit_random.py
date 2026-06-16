#!/usr/bin/env python3
"""
SAN Environment Security Audit – Multi-Device Excel Reporter
Analyzes Brocade FOS PuTTY session logs, HP Alletra MP CLI logs, and VMware
ESXi storage CSVs to generate an Excel security audit report with CIS Controls
v8 and PCI DSS v4.0.1 mapping.

Output format matches the sample-aruba_analysis.xlsx report style exactly:
  Summary / Security Issues / CIS v8 Mapping / PCI DSS Mapping

Input files (resolved relative to this script's directory):
  Brocade01 2026-06-01.log              – SAN Fabric A switch PuTTY log
  Brocade02 2026-06-01.log              – SAN Fabric B switch PuTTY log
  Alletra 2026-06-09-01.log             – HP Alletra Storage MP CLI log
  storage-devices-export-data.csv       – VMware ESXi storage device export
  Volumes_Inventory_06092026_120519.csv – HP Alletra volume set inventory

Usage:
    python san_audit_random.py
    python san_audit_random.py -o report.xlsx
"""

import csv
import os
import re
import sys
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILES = {
    "brocade01": os.path.join(SCRIPT_DIR, "Brocade01 2026-06-01.log"),
    "brocade02": os.path.join(SCRIPT_DIR, "Brocade02 2026-06-01.log"),
    "alletra":   os.path.join(SCRIPT_DIR, "Alletra 2026-06-09-01.log"),
    "esxi_dev":  os.path.join(SCRIPT_DIR, "storage-devices-export-data.csv"),
    "esxi_vols": os.path.join(SCRIPT_DIR, "Volumes_Inventory_06092026_120519.csv"),
}

# ── CIS Controls v8 ───────────────────────────────────────────────────────────
CIS_CTRL_DESC = {
    "3.10":  "Encrypt Sensitive Data in Transit",
    "4.2":   "Secure Configuration Process for Network Infrastructure",
    "4.8":   "Disable Unnecessary Services on Enterprise Assets and Software",
    "5.2":   "Use Unique Passwords",
    "8.2":   "Collect Audit Logs",
    "8.9":   "Centralize Audit Logs",
    "11.1":  "Establish and Maintain a Data Recovery Practice",
    "11.2":  "Perform Automated Backups",
    "12.2":  "Establish and Maintain a Secure Network Architecture",
    "12.3":  "Securely Manage Network Infrastructure",
    "12.6":  "Use Secure Network Management and Communication Protocols",
    "13.4":  "Perform Traffic Filtering Between Network Segments",
}

CIS_CONTROL_MAP: dict[str, list[str]] = {
    "FC Authentication Policy Disabled":   ["12.2", "12.3"],
    "Default Zone Allows All Access":      ["12.2", "13.4"],
    "No CHAP on Storage Hosts":            ["3.10", "12.2"],
    "Security Baseline Not Verifiable":    ["4.2", "8.2"],
    "No Volume Replication Configured":    ["11.1", "11.2"],
    "No SSH Keys Configured":              ["3.10", "12.6"],
    "No Login Banner Configured":          ["4.2"],
    "Degraded FC Ports":                   ["12.2"],
    "REST API Exposed":                    ["4.8", "12.3"],
    "Stale Zone Definitions":              ["12.2"],
    "Default MAPS Monitoring Policy":      ["8.2"],
    "Perennially Reserved Not Configured": ["12.2"],
}

# ── PCI DSS v4.0.1 ────────────────────────────────────────────────────────────
PCI_DSS_DESC = {
    "1.2.4":  "All traffic between trusted/untrusted networks is explicitly controlled",
    "1.3.2":  "Outbound traffic from the CDE is restricted to only that which is necessary",
    "2.2.1":  "Configuration standards are defined for all system components",
    "2.2.4":  "Only necessary services, protocols, and functions are enabled",
    "2.2.7":  "All non-console administrative access is encrypted using strong cryptography",
    "4.2.1":  "Strong cryptography safeguards transmission of cardholder data",
    "10.5.4": "Audit log files are protected to prevent unauthorized access or modifications",
}

PCI_DSS_MAP: dict[str, list[str]] = {
    "FC Authentication Policy Disabled":   ["1.2.4", "4.2.1"],
    "Default Zone Allows All Access":      ["1.2.4", "1.3.2"],
    "No CHAP on Storage Hosts":            ["1.2.4", "4.2.1"],
    "Security Baseline Not Verifiable":    ["2.2.1", "10.5.4"],
    "No Volume Replication Configured":    [],
    "No SSH Keys Configured":              ["2.2.7", "4.2.1"],
    "No Login Banner Configured":          ["2.2.1"],
    "Degraded FC Ports":                   [],
    "REST API Exposed":                    ["2.2.4"],
    "Stale Zone Definitions":              [],
    "Default MAPS Monitoring Policy":      [],
    "Perennially Reserved Not Configured": [],
}


def _cis_label(ids: list[str]) -> str:
    return " · ".join(f"CIS {c}" for c in ids)


def _pci_label(ids: list[str]) -> str:
    return " · ".join(f"PCI {r}" for r in ids)


# ── Colour palette (matches aruba_cx_analyzer.py) ────────────────────────────
C = {
    "hdr_bg":    "1B3A5C", "hdr_fg":    "FFFFFF",
    "critical":  "C00000", "critical_l": "FFB3B3",
    "high":      "FF0000", "high_l":     "FFD9B3",
    "medium":    "FF8C00", "medium_l":   "FFF2CC",
    "low":       "0070C0", "low_l":      "BDD7EE",
    "info":      "595959", "info_l":     "F2F2F2",
    "ok":        "375623", "ok_l":       "E2EFDA",
    "disabled":  "A6A6A6",
    "alt_row":   "F5F5F5",
}
_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=size)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Issue store ───────────────────────────────────────────────────────────────
class IssueStore:
    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

    def __init__(self):
        self.issues: list[dict] = []

    def add(self, severity: str, category: str, device: str, obj: str,
            description: str, recommendation: str,
            details: str = "", line: int | str = 0, output: str = "",
            source_file: str = ""):
        cis_ids = CIS_CONTROL_MAP.get(category, [])
        pci_ids = PCI_DSS_MAP.get(category, [])
        self.issues.append({
            "severity":       severity,
            "category":       category,
            "device":         device,
            "source_file":    source_file,
            "object":         obj,
            "line":           str(line) if line else "",
            "description":    description,
            "recommendation": recommendation,
            "details":        details,
            "output":         output,
            "cis_controls":   _cis_label(cis_ids),
            "cis_ids":        cis_ids,
            "pci_dss":        _pci_label(pci_ids),
            "pci_ids":        pci_ids,
        })

    def sorted_issues(self) -> list[dict]:
        return sorted(self.issues,
                      key=lambda x: self.SEV_ORDER.get(x["severity"], 9))

    def counts(self) -> dict[str, int]:
        c: dict[str, int] = defaultdict(int)
        for iss in self.issues:
            c[iss["severity"]] += 1
        return c

    def by_category(self) -> dict[str, int]:
        c: dict[str, int] = defaultdict(int)
        for iss in self.issues:
            c[iss["category"]] += 1
        return c


# ── Brocade FOS PuTTY log parser ──────────────────────────────────────────────
class BrocadeParser:

    def __init__(self, path: str):
        self.path            = path
        self.lines:          list[str] = []
        self.switch_name:    str  = ""
        self.switch_domain:  str  = ""
        self.switch_wwn:     str  = ""
        self.switch_type:    str  = ""
        self.active_cfg:     str  = ""
        self.auth_policy:    int  = -1
        self.auth_policy_ln: int  = 0
        self.default_zone:   str  = "noaccess"
        self.defzone_ln:     int  = 0
        self.banner_empty:   bool = False
        self.banner_ln:      int  = 0
        self.maps_policy:    str  = ""
        self.maps_enabled:   bool = False
        self.zones:          list[dict] = []
        self.cfgs:           list[dict] = []
        self.fabric_label:   str  = ""

    def parse(self) -> bool:
        try:
            with open(self.path, encoding="utf-8", errors="replace") as fh:
                self.lines = fh.readlines()
        except FileNotFoundError:
            print(f"  [!] File not found: {self.path}", file=sys.stderr)
            return False
        self._parse_configshow()
        self._parse_switchshow()
        self._parse_cfgshow()
        base = os.path.basename(self.path)
        self.fabric_label = (f"{self.switch_name} (Domain {self.switch_domain})"
                             if self.switch_name else base)
        return True

    def _parse_configshow(self):
        in_banner = False
        banner_content_found = False
        for lineno, raw in enumerate(self.lines, 1):
            s = raw.strip()
            if re.match(r"\[Banner\]", s, re.I):
                in_banner = True
                self.banner_ln = lineno
                continue
            if in_banner:
                if re.match(r"\[", s):
                    in_banner = False
                    if not banner_content_found:
                        self.banner_empty = True
                elif s:
                    banner_content_found = True
                continue
            m = re.match(r"auth\.policy\s*:\s*(\d+)$", s)
            if m:
                self.auth_policy    = int(m.group(1))
                self.auth_policy_ln = lineno
            m = re.match(r"defzone\s*:\s*(\S+)", s, re.I)
            if m:
                self.default_zone = m.group(1).lower()
                self.defzone_ln   = lineno
            m = re.match(r"maps\.activePolicy\s*:\s*(\S+)", s)
            if m:
                self.maps_policy = m.group(1)
            m = re.match(r"maps\.enabled\s*:\s*(\d+)", s)
            if m:
                self.maps_enabled = (m.group(1) == "1")

    def _parse_switchshow(self):
        for raw in self.lines:
            s = raw.strip()
            m = re.match(r"switchName\s*:\s*(\S+)", s, re.I)
            if m: self.switch_name = m.group(1)
            m = re.match(r"switchDomain\s*:\s*(\d+)", s, re.I)
            if m: self.switch_domain = m.group(1)
            m = re.match(r"switchWwn\s*:\s*(\S+)", s, re.I)
            if m: self.switch_wwn = m.group(1)
            m = re.match(r"switchType\s*:\s*(\S+)", s, re.I)
            if m: self.switch_type = m.group(1)
            m = re.match(r"zoning\s*:\s*ON\s*\((.+?)\)", s, re.I)
            if m: self.active_cfg = m.group(1).strip()

    def _parse_cfgshow(self):
        in_defined   = False
        in_effective = False
        cur_zone: dict | None = None
        cur_cfg:  dict | None = None
        for lineno, raw in enumerate(self.lines, 1):
            s = raw.strip()
            if re.match(r"Defined configuration:", s, re.I):
                in_defined = True; in_effective = False; continue
            if re.match(r"Effective configuration:", s, re.I):
                in_effective = True; in_defined = False; continue
            if in_effective or not in_defined:
                continue
            m = re.match(r"cfg\s*:\s*(\S+)", s, re.I)
            if m:
                cur_cfg  = {"name": m.group(1).rstrip(";"), "zones": [], "line": lineno}
                cur_zone = None
                self.cfgs.append(cur_cfg)
                continue
            m = re.match(r"zone\s*:\s*(\S+)", s, re.I)
            if m:
                cur_zone = {"name": m.group(1).rstrip(";"), "members": [], "line": lineno}
                cur_cfg  = None
                self.zones.append(cur_zone)
                continue
            if re.match(r"alias\s*:\s*", s, re.I):
                cur_zone = None; cur_cfg = None; continue
            if cur_zone and s:
                cur_zone["members"].extend(
                    m2.strip().rstrip(";") for m2 in s.split(";") if m2.strip().rstrip(";"))
            elif cur_cfg and s:
                cur_cfg["zones"].extend(
                    z.strip().rstrip(";") for z in s.split(";") if z.strip().rstrip(";"))

    def orphaned_zones(self) -> list[dict]:
        all_cfg_zones: set[str] = set()
        for cfg in self.cfgs:
            all_cfg_zones.update(z.strip().rstrip(";") for z in cfg["zones"])
        return [z for z in self.zones if z["name"] not in all_cfg_zones]

    def run_checks(self, store: IssueStore):
        device = self.fabric_label
        src    = os.path.basename(self.path)

        if self.auth_policy == 0:
            store.add(
                "CRITICAL", "FC Authentication Policy Disabled", device,
                "auth.policy: 0 (off)",
                "FC Authentication Policy (DH-CHAP) is disabled. No authentication is "
                "required between fabric ports, allowing unauthorized switches or rogue "
                "HBAs to join the fabric without credentials and enabling fabric "
                "eavesdropping and spoofed SCSI commands.",
                "Enable FC authentication: 'authutil --set -policy active'. "
                "Configure DH-CHAP shared secrets on ISL partners: 'secauthsecret --set'. "
                "Minimum acceptable: passive (1) for mixed-firmware environments.",
                details=f"configShow line {self.auth_policy_ln}",
                line=self.auth_policy_ln,
                output=f"configShow: auth.policy=0 (off) at line {self.auth_policy_ln}. "
                       f"DH-CHAP authentication is not enforced on any fabric port. "
                       f"Active config: {self.active_cfg}.",
                source_file=src)

        if self.default_zone in ("allaccess", "all access", "all_access"):
            store.add(
                "CRITICAL", "Default Zone Allows All Access", device,
                f"defzone: {self.default_zone}",
                "The default zone policy is set to 'allaccess'. Devices not placed in a "
                "named zone can communicate with all other unzoned devices, creating an "
                "open fabric with no isolation between initiators and targets.",
                "Set default zone to 'noaccess': 'defzone --noaccess; cfgsave'. "
                "This is the Brocade security baseline and required for PCI DSS.",
                details=f"configShow line {self.defzone_ln}",
                line=self.defzone_ln,
                output=f"configShow: defzone={self.default_zone} at line {self.defzone_ln}. "
                       f"Active config: {self.active_cfg}. "
                       "All fabric ports without explicit zone membership have unrestricted access.",
                source_file=src)

        if self.banner_empty:
            store.add(
                "MEDIUM", "No Login Banner Configured", device,
                "[Banner] section empty",
                "No login warning banner (MOTD) is configured on the switch. A legal "
                "notice banner is required by most compliance frameworks to establish "
                "legal standing for access monitoring and deter unauthorized use.",
                "Configure a banner: 'bannercfg --set "
                "\"Authorized access only. All sessions are logged.\"'",
                details=f"configShow line {self.banner_ln}",
                line=self.banner_ln,
                output=f"configShow: [Banner] section is empty (line {self.banner_ln}). "
                       "No warning is presented to users at login.",
                source_file=src)

        orphans = self.orphaned_zones()
        if orphans:
            names = "; ".join(z["name"] for z in orphans)
            store.add(
                "MEDIUM", "Stale Zone Definitions", device,
                f"{len(orphans)} orphaned zone(s)",
                "Zone definitions exist in the fabric that are not included in any active "
                "zone configuration. Orphaned zones may represent decommissioned hosts or "
                "storage removed without cleaning up the zone database, and can be "
                "accidentally re-activated.",
                "Delete orphaned zones: 'zonedelete <zonename>; cfgsave'. "
                "Remove unused aliases: 'alicedelete <alias>'. "
                "Review all zone definitions against current fabric topology.",
                details=names,
                line=orphans[0]["line"],
                output=f"{len(orphans)} zone(s) defined but not in active cfg "
                       f"'{self.active_cfg}': {names}.",
                source_file=src)

        if self.maps_enabled and "dflt" in self.maps_policy.lower():
            store.add(
                "LOW", "Default MAPS Monitoring Policy", device,
                f"maps.activePolicy: {self.maps_policy}",
                "MAPS (Monitoring and Alerting Policy Suite) is using the default "
                "monitoring policy rather than a customized environment-specific policy. "
                "Default thresholds may miss or over-alert on issues specific to this fabric.",
                "Create a custom MAPS policy tuned to this fabric: "
                "'mapsconfig --addpolicy custom_policy --clonedfrom dflt_base_policy'. "
                "Adjust thresholds for port errors, frame drops, and link instability.",
                output=f"configShow: maps.activePolicy={self.maps_policy}. "
                       "Default policy thresholds not tuned to this environment's traffic patterns.",
                source_file=src)


# ── HP Alletra MP PuTTY log parser ─────────────────────────────────────────────
class AlletraParser:

    def __init__(self, path: str):
        self.path              = path
        self.lines:            list[str] = []
        self.system_name:      str       = ""
        self.system_model:     str       = ""
        self.system_serial:    str       = ""
        self.hosts_no_chap:    list[str] = []
        self.invalid_commands: list[str] = []
        self.ports_loss_sync:  list[str] = []
        self.wsapi_enabled:    bool      = False
        self.ssh_key_found:    bool      = True
        self.ntp_servers:      list[str] = []
        self.hosts:            list[dict] = []

    def parse(self) -> bool:
        try:
            with open(self.path, encoding="utf-8", errors="replace") as fh:
                self.lines = fh.readlines()
        except FileNotFoundError:
            print(f"  [!] File not found: {self.path}", file=sys.stderr)
            return False
        self._parse_system()
        self._parse_invalid_commands()
        self._parse_chap()
        self._parse_ports()
        self._parse_wsapi()
        self._parse_sshkey()
        self._parse_net()
        return True

    def _parse_system(self):
        for raw in self.lines:
            m = re.search(r"0x[\dA-Fa-f]+\s+(\S+)\s+(HPE\s+\S+(?:\s+\S+)*?)\s+(\S+)\s+\d",
                          raw)
            if m:
                self.system_name   = m.group(1)
                self.system_model  = m.group(2).strip()
                self.system_serial = m.group(3)
                break
            mp = re.match(r"(\S+)\s+cli%", raw)
            if mp and not self.system_name:
                self.system_name = mp.group(1)

    def _parse_invalid_commands(self):
        known = {"showpasswordpolicy", "showsnmp", "showsyslog", "showaudit", "showtime"}
        for raw in self.lines:
            s = raw.strip()
            if "invalid command name" in s.lower():
                m = re.search(r'"(\w+)"', s)
                if m and m.group(1).lower() in known:
                    cmd = m.group(1)
                    if cmd not in self.invalid_commands:
                        self.invalid_commands.append(cmd)

    def _parse_chap(self):
        in_chap = False
        for raw in self.lines:
            s = raw.strip()
            if re.search(r"Initiator_CHAP_Name", s):
                in_chap = True
                continue
            if in_chap:
                if re.match(r"[-]+$", s) or not s:
                    continue
                if re.match(r"\d+\s+total", s):
                    in_chap = False
                    continue
                m = re.match(r"\s*(\d+)\s+(\S+)\s+(--|\S+)\s+(--|\S+)", raw)
                if m:
                    host       = m.group(2)
                    init_chap  = m.group(3)
                    tgt_chap   = m.group(4)
                    chap_ok    = (init_chap != "--" and tgt_chap != "--")
                    self.hosts.append({"name": host, "chap": chap_ok,
                                       "init": init_chap, "target": tgt_chap})
                    if not chap_ok:
                        self.hosts_no_chap.append(host)

    def _parse_ports(self):
        for raw in self.lines:
            m = re.match(r"\s*([\d:]+)\s+target\s+loss_sync", raw, re.I)
            if m:
                self.ports_loss_sync.append(m.group(1))

    def _parse_wsapi(self):
        in_wsapi = False
        for raw in self.lines:
            s = raw.strip()
            if re.search(r"showwsapi", s, re.I):
                in_wsapi = True
                continue
            if in_wsapi and re.search(r"Enabled\s+Active", s, re.I):
                self.wsapi_enabled = True
                in_wsapi = False

    def _parse_sshkey(self):
        for raw in self.lines:
            if re.search(r"No SSH key found", raw, re.I):
                self.ssh_key_found = False
                break

    def _parse_net(self):
        for raw in self.lines:
            s = raw.strip()
            if re.search(r"^NTP server\s*:", s, re.I):
                parts = s.split(":", 1)[1].strip().split()
                for p in parts:
                    if p not in ("None", "none", ""):
                        self.ntp_servers.append(p)

    def run_checks(self, store: IssueStore):
        device = self.system_name or "HP Alletra MP"
        src    = os.path.basename(self.path)

        if self.hosts_no_chap:
            hosts = ", ".join(self.hosts_no_chap)
            store.add(
                "HIGH", "No CHAP on Storage Hosts", device,
                f"{len(self.hosts_no_chap)} host(s) without CHAP",
                "No CHAP (Challenge-Handshake Authentication Protocol) authentication is "
                "configured for storage hosts. Without CHAP, any host presenting the "
                "correct WWN can authenticate to the array, enabling unauthorized LUN "
                "access in a compromised or miscabled environment.",
                "Configure per-host CHAP credentials on the Alletra MP for all "
                "production hosts using 'createhost -chap <initiator_name> <secret>'. "
                "Enable target CHAP to prevent host-spoofing attacks.",
                details=f"showhost -chap: {hosts}",
                output=f"showhost -chap: {len(self.hosts_no_chap)} host(s) show '--' "
                       f"(not configured) for both Initiator CHAP Name and Target CHAP Name: "
                       f"{hosts}.",
                source_file=src)

        if self.invalid_commands:
            cmds = ", ".join(self.invalid_commands)
            store.add(
                "HIGH", "Security Baseline Not Verifiable", device,
                "commands: " + cmds,
                "Security baseline commands are not available on this firmware version. "
                "Password policy, SNMP configuration, syslog, audit logging, and time "
                "synchronization settings cannot be verified from the CLI, leaving key "
                "security controls in an unknown compliance state.",
                "Verify password policy, SNMP, syslog, and NTP via the GreenLake Cloud "
                "Console or HPE SSMC. Update firmware to restore CLI security commands, "
                "or document equivalent GUI controls for the compliance record.",
                details=f"Invalid on this firmware: {cmds}",
                output=f"Commands returning 'invalid command name' on this firmware: {cmds}. "
                       "These settings cannot be audited via CLI.",
                source_file=src)

        if not self.ssh_key_found:
            store.add(
                "HIGH", "No SSH Keys Configured", device,
                "showsshkey: No SSH key found",
                "No SSH public keys are configured on the storage array. Administrative "
                "access relies exclusively on password authentication, which does not "
                "meet requirements for strong cryptographic access controls.",
                "Install SSH public keys for admin accounts: "
                "'setsshkey -add <keyfile>' or via GreenLake Console SSH key page. "
                "Prefer Ed25519 or RSA-4096 keys.",
                details="showsshkey: No SSH key found",
                output="showsshkey: 'No SSH key found'. "
                       "All administrative authentication uses passwords only; "
                       "no public-key authentication is configured.",
                source_file=src)

        if self.ports_loss_sync:
            ports = ", ".join(self.ports_loss_sync)
            store.add(
                "MEDIUM", "Degraded FC Ports", device,
                f"loss_sync: {ports}",
                "One or more FC target ports are in 'loss_sync' state, indicating lost "
                "synchronization with the fabric. Degraded ports are not passing traffic, "
                "reducing path redundancy for connected hosts and indicating a potential "
                "hardware or cabling fault.",
                "Investigate each port: check SFP health ('showport -sfp'), verify "
                "cabling to the Brocade switch, and check Brocade switchShow for the "
                "matching F-Port state. Replace faulty SFPs or cables.",
                details=f"Ports in loss_sync: {ports}",
                output=f"showport: {len(self.ports_loss_sync)} port(s) in loss_sync state: "
                       f"{ports}. These ports have lost link-level synchronization with "
                       "the connected Brocade switch and are not carrying I/O.",
                source_file=src)

        if self.wsapi_enabled:
            store.add(
                "MEDIUM", "REST API Exposed", device,
                "WSAPI: Enabled / Active on HTTPS 443",
                "The array REST API (WSAPI) is enabled and active. The API exposes full "
                "array management capabilities including volume, host, and VLUN management "
                "without requiring FC fabric access to execute changes.",
                "Restrict WSAPI access by source IP via management network ACLs. "
                "Use dedicated service accounts with least-privilege for automation. "
                "Disable WSAPI if REST API management is not actively used: 'stopwsapi'.",
                details="WSAPI: Enabled / Active, HTTPS 443",
                output="showwsapi: Service=Enabled, State=Active on HTTPS port 443. "
                       "REST API is accessible and accepting connections.",
                source_file=src)


# ── HP Alletra Volume CSV parser ──────────────────────────────────────────────
class AlletraVolumesParser:

    def __init__(self, path: str):
        self.path    = path
        self.volumes: list[dict] = []

    def parse(self) -> bool:
        try:
            with open(self.path, encoding="utf-8-sig", errors="replace", newline="") as fh:
                content = fh.readlines()
        except FileNotFoundError:
            print(f"  [!] File not found: {self.path}", file=sys.stderr)
            return False
        data_start = 0
        for i, line in enumerate(content):
            if "Volume Count" in line and "Protection Level" in line:
                data_start = i
                break
        if not data_start:
            return False
        reader = csv.DictReader(content[data_start:])
        for row in reader:
            name  = row.get("Name", "").strip().strip('"')
            if not name:
                continue
            self.volumes.append({
                "name":             name,
                "volume_count":     row.get("Volume Count", "").strip().strip('"'),
                "system":           row.get("System", "").strip().strip('"'),
                "export_status":    row.get("Export Status", "").strip().strip('"'),
                "protection_level": row.get("Protection Level", "").strip().strip('"'),
                "replication_type": row.get("Replication Type", "").strip().strip('"'),
            })
        return True

    def run_checks(self, store: IssueStore):
        src      = os.path.basename(self.path)
        exported = [v for v in self.volumes if v["export_status"] == "Exported"]
        for vol in exported:
            if vol["replication_type"].lower() in ("none", "--", ""):
                device = vol["system"] or "HP Alletra MP"
                store.add(
                    "HIGH", "No Volume Replication Configured", device,
                    f"Volume Set: {vol['name']}",
                    "No remote replication is configured for this exported volume set. "
                    "Local snapshots alone cannot protect against site-level failure, "
                    "storage hardware failure, or ransomware that corrupts all local "
                    "snapshots before a recovery point is reached.",
                    "Configure remote replication to a secondary Alletra MP or cloud target "
                    "via GreenLake Data Services. Define RPO/RTO targets for each volume "
                    "set, test failover regularly, and document the DR procedure.",
                    details=f"Protection: {vol['protection_level']} | "
                            f"Replication: {vol['replication_type']}",
                    output=f"Volumes CSV: Volume Set '{vol['name']}' — "
                           f"Protection Level='{vol['protection_level']}', "
                           f"Replication Type='{vol['replication_type']}', "
                           f"Export Status=Exported. No DR replication target configured.",
                    source_file=src)


# ── VMware ESXi storage device CSV parser ─────────────────────────────────────
class ESXiStorageParser:

    def __init__(self, path: str):
        self.path    = path
        self.devices: list[dict] = []

    def parse(self) -> bool:
        try:
            with open(self.path, encoding="utf-8-sig", errors="replace", newline="") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    self.devices.append(
                        {k.strip().strip('"'): v.strip().strip('"') for k, v in row.items()})
        except FileNotFoundError:
            print(f"  [!] File not found: {self.path}", file=sys.stderr)
            return False
        return True

    def run_checks(self, store: IssueStore):
        src      = os.path.basename(self.path)
        fc_disks = [d for d in self.devices
                    if d.get("Transport", "").lower() == "fibre channel"
                    and d.get("Type", "").lower() == "disk"]
        not_reserved = [d for d in fc_disks
                        if d.get("Perennially Reserved", "").lower() != "yes"]
        if not_reserved:
            ids = "; ".join(d.get("Identifier", d.get("Name", "?"))[:40]
                            for d in not_reserved[:5])
            store.add(
                "LOW", "Perennially Reserved Not Configured",
                "VMware ESXi",
                f"{len(not_reserved)} FC disk(s)",
                "FC storage devices presented to ESXi hosts do not have 'Perennially "
                "Reserved' enabled. For FC LUNs not actively consumed as VMFS datastores, "
                "this causes unnecessary SCSI reservation probes and potential path "
                "thrashing at ESXi host boot.",
                "Set Perennially Reserved for non-datastore FC LUNs: "
                "esxcli storage core device setconfig -d <naa.id> --perennially-reserved=true. "
                "Review whether all presented LUNs are required on each ESXi host.",
                details=ids,
                output=f"storage-devices-export-data.csv: {len(not_reserved)} FC disk(s) "
                       f"with Perennially Reserved=No: {ids}. "
                       "ESXi will probe these LUNs for SCSI reservations on every boot.",
                source_file=src)


# ── Excel Report Builder ───────────────────────────────────────────────────────
class ExcelReporter:

    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"], C["critical_l"]),
        "HIGH":     (C["high"],     C["high_l"]),
        "MEDIUM":   (C["medium"],   C["medium_l"]),
        "LOW":      (C["low"],      C["low_l"]),
        "INFO":     (C["info"],     C["info_l"]),
    }

    def __init__(self, store: IssueStore, out_path: str,
                 brocades: list["BrocadeParser"],
                 alletra:  "AlletraParser | None",
                 volumes:  "AlletraVolumesParser | None",
                 esxi:     "ESXiStorageParser | None"):
        self.store    = store
        self.out      = out_path
        self.brocades = brocades
        self.alletra  = alletra
        self.volumes  = volumes
        self.esxi     = esxi
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _hdr(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = _font(bold=True, color=C["hdr_fg"])
            c.fill      = _fill(C["hdr_bg"])
            c.alignment = _align("center", wrap=False)
            c.border    = THIN
        ws.row_dimensions[row].height = 20

    @staticmethod
    def _set_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    @staticmethod
    def _row_fill(row: int) -> str | None:
        return C["alt_row"] if row % 2 == 0 else None

    # ── Sheet 1: Summary (matches Aruba layout exactly) ───────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "SAN / Storage Environment — Configuration Security Report"
        c.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        c.fill  = _fill(C["hdr_bg"])
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 42

        brocade_names = "; ".join(
            f"{b.switch_name} Domain {b.switch_domain}" for b in self.brocades)
        alletra_id = ((self.alletra.system_name + " " + self.alletra.system_model).strip()
                      if self.alletra else "(not provided)")

        ws.merge_cells("A2:F2")
        c = ws["A2"]
        c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    "
                   f"Switches: {brocade_names}    "
                   f"Array: {alletra_id}")
        c.font  = _font(italic=True, color=C["info"], size=9)
        c.fill  = _fill("F2F2F2")
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 18

        row = 4

        def section(label: str):
            nonlocal row
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font  = _font(bold=True, color=C["hdr_bg"], size=12)
            row += 1

        def kv(label: str, value, good: bool | None = None):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font(); c1.alignment = _align()
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = _font(bold=True); c2.alignment = _align("center")
            if good is True:
                c2.fill = _fill(C["ok_l"]); c2.font = _font(bold=True, color=C["ok"])
            elif good is False:
                c2.fill = _fill(C["high_l"]); c2.font = _font(bold=True, color=C["high"])
            elif row % 2 == 0:
                c1.fill = c2.fill = _fill(C["alt_row"])
            row += 1

        section("DEVICE OVERVIEW")
        for b in self.brocades:
            kv(f"Switch ({os.path.basename(b.path)})",
               f"{b.switch_name}  Domain {b.switch_domain}  WWN: {b.switch_wwn}")
            kv(f"  FC Auth Policy",
               f"auth.policy={b.auth_policy} ({'DISABLED' if b.auth_policy == 0 else 'OK'})",
               good=(b.auth_policy != 0))
            kv(f"  Default Zone",
               b.default_zone.upper(),
               good=(b.default_zone not in ("allaccess", "all_access")))
            kv(f"  Active Config", b.active_cfg or "(none)")
            kv(f"  Zones Defined", len(b.zones))
            kv(f"  Orphaned Zones", len(b.orphaned_zones()),
               good=(len(b.orphaned_zones()) == 0))

        if self.alletra:
            a = self.alletra
            kv("Storage Array", f"{a.system_name}  {a.system_model}  S/N: {a.system_serial}")
            kv("  WSAPI (REST API)",
               "Enabled" if a.wsapi_enabled else "Disabled",
               good=not a.wsapi_enabled)
            kv("  SSH Keys",
               "Configured" if a.ssh_key_found else "NONE",
               good=a.ssh_key_found)
            kv("  Hosts Without CHAP", len(a.hosts_no_chap),
               good=(len(a.hosts_no_chap) == 0))
            kv("  FC Ports in loss_sync", len(a.ports_loss_sync),
               good=(len(a.ports_loss_sync) == 0))
            kv("  NTP Servers", ", ".join(a.ntp_servers) or "(unknown)")
            kv("  Security CLI Commands", "UNAVAILABLE" if a.invalid_commands else "OK",
               good=(not a.invalid_commands))

        if self.volumes:
            exported  = [v for v in self.volumes.volumes if v["export_status"] == "Exported"]
            no_repl   = [v for v in exported
                         if v["replication_type"].lower() in ("none", "--", "")]
            kv("Volume Sets Exported", len(exported))
            kv("  Without Replication", len(no_repl), good=(len(no_repl) == 0))

        if self.esxi:
            fc_disks = [d for d in self.esxi.devices
                        if d.get("Transport", "").lower() == "fibre channel"
                        and d.get("Type", "").lower() == "disk"]
            kv("FC Storage Devices (ESXi)", len(fc_disks))

        row += 1
        section("SECURITY FINDINGS BY SEVERITY")
        self._hdr(ws, ["Severity", "Count"], row=row)
        row += 1
        total = 0
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            cnt = sum(1 for i in self.store.issues if i["severity"] == sev)
            total += cnt
            fg, bg = self.SEV_COLORS[sev]
            for col, val in enumerate([sev, cnt], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = _fill(bg); c.border = THIN
                c.font = _font(bold=True, color=fg)
                c.alignment = _align("center")
            row += 1
        for col, val in enumerate(["TOTAL", total], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _font(bold=True); c.border = THIN
            c.alignment = _align("center")
        row += 2

        section("FINDINGS BY CATEGORY")
        self._hdr(ws, ["Category", "Count"], row=row)
        row += 1
        for cat, cnt in sorted(self.store.by_category().items(), key=lambda x: -x[1]):
            rb = C["alt_row"] if row % 2 == 0 else None
            c1 = ws.cell(row=row, column=1, value=cat)
            c1.font = _font()
            if rb: c1.fill = _fill(rb)
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.font = _font(bold=True); c2.alignment = _align("center")
            if rb: c2.fill = _fill(rb)
            row += 1

        self._set_widths(ws, [38, 18, 18, 18, 18, 18])

    # ── Sheet 2: Zones (SAN-specific detail) ──────────────────────────────────
    def _sheet_zones(self):
        ws = self.wb.create_sheet("Zones")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "SAN Fabric — Zone Configuration"
        t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 30

        row = 2
        for b in self.brocades:
            ws.merge_cells(f"A{row}:G{row}")
            fh = ws.cell(row=row, column=1,
                         value=f"Fabric: {b.fabric_label}  |  Active Config: "
                               f"{b.active_cfg or '(none)'}  |  "
                               f"Default Zone: {b.default_zone.upper()}")
            fh.font = _font(bold=True, color="FFFFFF")
            dz_bad  = b.default_zone in ("allaccess", "all_access")
            fh.fill = _fill(C["critical"] if dz_bad else "2E4E7A")
            fh.alignment = _align("left", wrap=False)
            ws.row_dimensions[row].height = 22
            row += 1

            orphan_names = {z["name"] for z in b.orphaned_zones()}
            self._hdr(ws, ["Zone Name", "Members", "In Active Config?", "Member Count",
                           "Notes"], row=row)
            row += 1
            for z in b.zones:
                alt      = self._row_fill(row)
                orphaned = z["name"] in orphan_names
                members  = "; ".join(z["members"][:6])
                if len(z["members"]) > 6:
                    members += f" … (+{len(z['members'])-6} more)"
                vals = [z["name"], members,
                        "NO — orphaned" if orphaned else "Yes",
                        len(z["members"]),
                        "Not in any cfg — review/delete" if orphaned else ""]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = _font(bold=orphaned,
                                   color=C["medium"] if orphaned else "000000")
                    c.alignment = _align(); c.border = THIN
                    if alt: c.fill = _fill(alt)
                row += 1
            row += 1

        self._set_widths(ws, [44, 70, 18, 14, 32])

    # ── Sheet 3: Hosts & Volumes (SAN-specific detail) ────────────────────────
    def _sheet_hosts_volumes(self):
        ws = self.wb.create_sheet("Hosts & Volumes")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "HP Alletra MP — Hosts and Volume Sets"
        t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 30

        row = 3
        if self.alletra and self.alletra.hosts:
            ws.cell(row=row, column=1, value="FC HOSTS (showhost -chap)").font = _font(bold=True, size=11)
            row += 1
            self._hdr(ws, ["Host Name", "Initiator CHAP", "Target CHAP", "CHAP Status"],
                      row=row)
            row += 1
            for h in self.alletra.hosts:
                alt    = self._row_fill(row)
                bad    = not h["chap"]
                status = "OK" if h["chap"] else "MISSING — HIGH risk"
                vals   = [h["name"], h["init"], h["target"], status]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = _font(bold=(col == 4 and bad),
                                   color=C["high"] if (col == 4 and bad) else "000000")
                    c.alignment = _align(); c.border = THIN
                    if alt: c.fill = _fill(alt)
                row += 1
            row += 1

        if self.volumes and self.volumes.volumes:
            ws.cell(row=row, column=1, value="VOLUME SETS").font = _font(bold=True, size=11)
            row += 1
            self._hdr(ws, ["Volume Set", "Volumes", "System", "Export Status",
                           "Protection Level", "Replication Type"], row=row)
            row += 1
            for v in self.volumes.volumes:
                alt     = self._row_fill(row)
                no_repl = (v["replication_type"].lower() in ("none", "--", "")
                           and v["export_status"] == "Exported")
                vals = [v["name"], v["volume_count"], v["system"],
                        v["export_status"], v["protection_level"], v["replication_type"]]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = _font(bold=(col == 6 and no_repl),
                                   color=C["high"] if (col == 6 and no_repl) else "000000")
                    c.alignment = _align(); c.border = THIN
                    if alt: c.fill = _fill(alt)
                row += 1
            row += 1

        if self.esxi and self.esxi.devices:
            ws.cell(row=row, column=1, value="ESXI FC STORAGE DEVICES").font = _font(bold=True, size=11)
            row += 1
            self._hdr(ws, ["Name", "LUN", "Capacity", "Transport",
                           "Driver", "Perennially Reserved"], row=row)
            row += 1
            for d in self.esxi.devices:
                alt = self._row_fill(row)
                pr  = d.get("Perennially Reserved", "No")
                bad = (pr != "Yes" and d.get("Transport", "").lower() == "fibre channel"
                       and d.get("Type", "").lower() == "disk")
                vals = [d.get("Name", "")[:60], d.get("LUN", ""),
                        d.get("Capacity", ""), d.get("Transport", ""),
                        d.get("Owner", ""), pr]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = _font(bold=(col == 6 and bad),
                                   color=C["low"] if (col == 6 and bad) else "000000")
                    c.alignment = _align(); c.border = THIN
                    if alt: c.fill = _fill(alt)
                row += 1

        self._set_widths(ws, [50, 10, 14, 16, 14, 20])

    # ── Sheet 4: Security Issues (exact Aruba column layout) ──────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False

        headers = ["#", "Severity", "Category", "Object / Interface",
                   "Config Line", "CIS v8 Controls", "PCI DSS",
                   "Description", "Recommendation", "Details", "Output",
                   "Verified", "Asset", "Target", "Vuln"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for idx, iss in enumerate(self.store.sorted_issues(), 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            rb = self._row_fill(row)

            line   = iss.get("line", "")
            target = iss["object"] + (f" ({line})" if line else "")
            vuln   = "SAN - " + iss["category"]

            vals = [idx, sev, iss["category"], iss["object"],
                    line,
                    iss.get("cis_controls", ""),
                    iss.get("pci_dss", ""),
                    iss["description"], iss["recommendation"],
                    iss.get("details", ""), iss.get("output", ""),
                    "Y",
                    f"{iss.get('source_file', '')}: {iss['device']}" if iss.get('source_file') else iss["device"],
                    target, vuln]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 2:
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                    c.alignment = _align("center")
                elif col in (1, 5):
                    c.font = _font(bold=(col == 1)); c.alignment = _align("center")
                    if rb: c.fill = _fill(rb)
                elif col == 6:   # CIS v8 Controls
                    c.font = _font(bold=True, color="17375E", size=9)
                    c.alignment = _align("center")
                    if rb: c.fill = _fill(rb)
                elif col == 7:   # PCI DSS
                    c.font = _font(bold=True, color="7B2D8B", size=9)
                    c.alignment = _align("center")
                    if rb: c.fill = _fill(rb)
                elif col == 12:  # Verified
                    c.font = _font(bold=True, color=C["ok"]); c.alignment = _align("center")
                    c.fill = _fill(C["ok_l"])
                else:
                    c.font = _font(); c.alignment = _align()
                    if rb: c.fill = _fill(rb)
            ws.row_dimensions[row].height = 40

        self._set_widths(ws, [4, 12, 32, 30, 12, 22, 18, 60, 60, 30, 60, 10, 24, 36, 34])

    # ── Sheet 5: CIS v8 Mapping (exact Aruba format) ──────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS v8 Mapping")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v8 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each CIS safeguard lists all findings from this environment that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        ctrl_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.store.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        def _sort_ctrl(k: str):
            parts = k.split(".")
            return (int(parts[0]), float("0." + parts[1]) if len(parts) > 1 else 0)

        row = 4
        for ctrl_id in sorted(CIS_CTRL_DESC.keys(), key=_sort_ctrl):
            desc   = CIS_CTRL_DESC[ctrl_id]
            issues = sorted(ctrl_issues.get(ctrl_id, []),
                            key=lambda x: self.SEV_ORDER.get(x["severity"], 9))

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1, value=f"CIS {ctrl_id} — {desc}")
            hc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill = _fill("17375E"); hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 22
            row += 1

            if not issues:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this control")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9"); nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object / Interface",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF")
                    c.fill = _fill("2E4057"); c.alignment = _align("center", wrap=False)
                    c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in issues:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = self._row_fill(row)
                    for col, val in enumerate(
                            [sev, iss["category"], iss["object"],
                             iss.get("line", ""), iss["description"],
                             iss["recommendation"]], 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 30, 12, 60, 60])

    # ── Sheet 6: PCI DSS Mapping (exact Aruba format) ─────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each PCI DSS v4.0 requirement lists all findings from this environment that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        req_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.store.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        row = 3
        for req_id in sorted(PCI_DSS_DESC.keys(),
                              key=lambda x: [int(p) for p in x.split(".")]):
            desc   = PCI_DSS_DESC[req_id]
            issues = sorted(req_issues.get(req_id, []),
                            key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count  = len(issues)

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not issues:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this requirement")
                nc.font = _font(italic=True, color=C["info"]); nc.fill = _fill("F9F9F9")
                nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object / Interface",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in issues:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = self._row_fill(row)
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 30, 12, 60, 60])

    # ── Save workbook ─────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_zones()
        self._sheet_hosts_volumes()
        self._sheet_issues()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)
        print(f"Report saved: {os.path.abspath(self.out)}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="SAN / Storage Security Audit Report Generator — outputs Excel report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Input files are resolved relative to the script's own directory:
  Brocade01 2026-06-01.log
  Brocade02 2026-06-01.log
  Alletra 2026-06-09-01.log
  storage-devices-export-data.csv
  Volumes_Inventory_06092026_120519.csv
""")
    ap.add_argument("-o", "--output",
                    default=os.path.join(SCRIPT_DIR, "sample-san_analysis.xlsx"),
                    help="Output Excel file (default: sample-san_analysis.xlsx)")
    args = ap.parse_args()

    store    = IssueStore()
    brocades: list[BrocadeParser] = []

    for key in ("brocade01", "brocade02"):
        path = INPUT_FILES[key]
        print(f"[*] Parsing: {os.path.basename(path)}")
        b = BrocadeParser(path)
        if b.parse():
            b.run_checks(store)
            brocades.append(b)

    print(f"[*] Parsing: {os.path.basename(INPUT_FILES['alletra'])}")
    alletra = AlletraParser(INPUT_FILES["alletra"])
    if not alletra.parse():
        alletra = None

    print(f"[*] Parsing: {os.path.basename(INPUT_FILES['esxi_vols'])}")
    volumes = AlletraVolumesParser(INPUT_FILES["esxi_vols"])
    volumes.parse()

    print(f"[*] Parsing: {os.path.basename(INPUT_FILES['esxi_dev'])}")
    esxi = ESXiStorageParser(INPUT_FILES["esxi_dev"])
    esxi.parse()

    if alletra:
        alletra.run_checks(store)
    volumes.run_checks(store)
    esxi.run_checks(store)

    counts = store.counts()
    print(f"\n[*] Findings: "
          f"{counts.get('CRITICAL',0)} CRITICAL  "
          f"{counts.get('HIGH',0)} HIGH  "
          f"{counts.get('MEDIUM',0)} MEDIUM  "
          f"{counts.get('LOW',0)} LOW  "
          f"{counts.get('INFO',0)} INFO")

    reporter = ExcelReporter(store, args.output, brocades, alletra, volumes, esxi)
    reporter.save()


if __name__ == "__main__":
    main()
