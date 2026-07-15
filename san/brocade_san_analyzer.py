#!/usr/bin/env python3
"""
Brocade FOS SAN Switch Configuration Analyzer
Parses Brocade Fabric OS (FOS) configuration export and produces an Excel
security audit report with CIS Controls v8 and PCI DSS v4.0 mapping.

Designed for Brocade FOS 9.x on 6xxx / G-series / DCX SAN switches.
Includes VMware ESXi / vSphere SAN fabric-specific checks.

Usage:
    python brocade_san_analyzer.py configexport.txt
    python brocade_san_analyzer.py configexport.txt -o audit.xlsx

Input: Concatenated output from the following commands (pipe to a file):
    switchshow ; configshow ; userconfig --showlist ;
    passwdcfg --show ; snmpconfig --show snmpv1 ;
    snmpconfig --show snmpv3 ; syslogadmin --show ;
    tsclockserver ; ipfilter --show ; sshutil --show ;
    cfgshow ; defzone --show ; version
"""

import re
import os
import sys
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── CIS Controls v8 mapping ───────────────────────────────────────────────────
CIS_CTRL_DESC = {
    "3.10":  "Encrypt Sensitive Data in Transit",
    "4.2":   "Secure Configuration Process for Network Infrastructure",
    "4.7":   "Manage Default Accounts on Enterprise Assets and Software",
    "4.8":   "Disable Unnecessary Services on Enterprise Assets and Software",
    "5.2":   "Use Unique Passwords",
    "5.4":   "Restrict Administrator Privileges to Dedicated Accounts",
    "6.3":   "Require MFA for Externally-Exposed Applications",
    "6.5":   "Require MFA for Administrative Access",
    "6.7":   "Centralize Access Control",
    "6.8":   "Define and Maintain Role-Based Access Control",
    "8.2":   "Collect Audit Logs",
    "8.4":   "Standardize Time Synchronization",
    "8.9":   "Centralize Audit Logs",
    "12.2":  "Establish and Maintain a Secure Network Architecture",
    "12.3":  "Securely Manage Network Infrastructure",
    "12.6":  "Use Secure Network Management and Communication Protocols",
    "13.4":  "Perform Traffic Filtering Between Network Segments",
}

CIS_CONTROL_MAP: dict[str, list[str]] = {
    # ── Credentials / accounts ────────────────────────────────────────────────
    "Default Account Not Disabled":          ["4.7"],
    "Default Admin Password Not Changed":    ["4.7", "5.2"],
    "Account Without Password":              ["5.2"],
    "Excessive Admin Accounts":              ["5.4", "6.8"],
    # ── Authentication / access ───────────────────────────────────────────────
    "No RADIUS or LDAP Configured":          ["6.5", "6.7"],
    "No Account Lockout Policy":             ["5.2", "6.5"],
    "Telnet Enabled":                        ["4.2", "12.6"],
    "HTTP Management Enabled":               ["4.2", "12.6"],
    "SSH Disabled":                          ["4.2", "12.6"],
    "No IP Filter Configured":               ["12.3", "6.7"],
    "rsh/rlogin Enabled":                    ["4.2", "12.6"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1 Enabled":                        ["4.2", "12.3", "12.6"],
    "Default SNMP Community String":         ["4.7", "12.3"],
    "SNMP Write Community Configured":       ["12.3"],
    "No SNMPv3 AuthPriv":                    ["4.2", "12.3", "12.6"],
    # ── FC fabric security ────────────────────────────────────────────────────
    "FC Authentication Policy Disabled":     ["12.2", "12.3"],
    # ── Fabric / zone security ────────────────────────────────────────────────
    "Default Zone Allows All Access":        ["12.2", "13.4"],
    "No Active Zone Configuration":          ["12.2", "13.4"],
    "Orphaned Zone Members":                 ["12.2"],
    "Port-Based Zoning Used":               ["12.2", "13.4"],
    "Zone With Single Member":               ["12.2"],
    "Orphaned Zone (Not in Active Config)":  ["12.2"],
    "Oversized Zone":                        ["12.2", "13.4"],
    "No Fabric-Wide Consistency Policy":     ["12.2", "12.3"],
    "NPIV Not Enabled":                      ["12.2"],
    # ── Logging / time ────────────────────────────────────────────────────────
    "No Syslog Server Configured":           ["8.2", "8.9"],
    "Audit Logging Disabled":                ["8.2"],
    "NTP Not Configured":                    ["8.4"],
    "Multiple NTP Servers Not Configured":   ["8.4"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Weak Minimum Password Length":          ["5.2"],
    "No Password Complexity Requirements":   ["5.2"],
    "Password Never Expires":                ["5.2"],
    "No Password History":                   ["5.2"],
    # ── Session / management ──────────────────────────────────────────────────
    "No Login Banner Configured":            ["4.2"],
    "Weak SSH Ciphers Permitted":            ["3.10", "12.6"],
}


def _cis_label(ctrl_ids: list[str]) -> str:
    return " · ".join(f"CIS {c}" for c in ctrl_ids)


CIS_BENCHMARK_MAP: dict[str, list[str]] = {}


def _cis_benchmark_label(check_ids: list[str]) -> str:
    return " · ".join(check_ids)


# ── PCI DSS v4.0 mapping ─────────────────────────────────────────────────────
PCI_DSS_DESC = {
    "1.2.4":  "All traffic between trusted/untrusted networks is explicitly controlled",
    "1.2.7":  "Unused network access points are disabled",
    "2.2.1":  "Configuration standards are defined for all system components",
    "2.2.4":  "Only necessary services, protocols, and functions are enabled",
    "2.2.7":  "All non-console administrative access is encrypted",
    "4.2.1":  "Strong cryptography is used to safeguard data during transmission",
    "7.2.1":  "All user access is appropriate and assigned by business need",
    "8.2.1":  "All user IDs and authentication credentials are managed securely",
    "8.2.8":  "Session idle time-out is 15 minutes or less",
    "8.3.4":  "Invalid authentication attempts are limited",
    "8.3.6":  "Passwords meet minimum length and complexity requirements",
    "8.3.9":  "Passwords are changed at least once every 90 days",
    "8.4.1":  "MFA is implemented for all non-console administrative access",
    "10.5.4": "Audit log files are protected (written to external log servers)",
    "10.6.1": "System clocks are synchronized using time-synchronization technology",
}

PCI_DSS_MAP: dict[str, list[str]] = {
    "Default Account Not Disabled":         ["8.2.1"],
    "Default Admin Password Not Changed":   ["8.2.1"],
    "Account Without Password":             ["8.2.1"],
    "Excessive Admin Accounts":             ["7.2.1"],
    "No RADIUS or LDAP Configured":         ["8.4.1"],
    "No Account Lockout Policy":            ["8.3.4"],
    "Telnet Enabled":                       ["2.2.4", "2.2.7"],
    "HTTP Management Enabled":              ["2.2.4", "2.2.7"],
    "SSH Disabled":                         ["2.2.7"],
    "rsh/rlogin Enabled":                   ["2.2.4", "2.2.7"],
    "No IP Filter Configured":              ["1.2.4"],
    "SNMPv1 Enabled":                       ["2.2.4"],
    "Default SNMP Community String":        ["2.2.4"],
    "SNMP Write Community Configured":      ["2.2.4"],
    "No SNMPv3 AuthPriv":                   ["2.2.4", "4.2.1"],
    "FC Authentication Policy Disabled":    ["1.2.4"],
    "Default Zone Allows All Access":       ["1.2.4"],
    "No Active Zone Configuration":         ["1.2.4"],
    "Port-Based Zoning Used":               ["1.2.4"],
    "No Syslog Server Configured":          ["10.5.4"],
    "Audit Logging Disabled":               ["10.5.4"],
    "NTP Not Configured":                   ["10.6.1"],
    "Weak Minimum Password Length":         ["8.3.6"],
    "No Password Complexity Requirements":  ["8.3.6"],
    "Password Never Expires":               ["8.3.9"],
    "No Login Banner Configured":           ["2.2.1"],
    "Weak SSH Ciphers Permitted":           ["4.2.1"],
}


def _pci_label(req_ids: list[str]) -> str:
    return " · ".join(f"PCI {r}" for r in req_ids)


# ── Secure Controls Framework (SCF) mapping ──────────────────────────────────
SCF_MAP: dict[str, list[str]] = {
    # ── Credentials / accounts ────────────────────────────────────────────────
    "Default Account Not Disabled":          ["IAC-21"],
    "Default Admin Password Not Changed":    ["IAC-06"],
    "Account Without Password":              ["IAC-06"],
    "Excessive Admin Accounts":              ["IAC-07"],
    # ── Authentication / access ───────────────────────────────────────────────
    "No RADIUS or LDAP Configured":          ["IAC-01"],
    "No Account Lockout Policy":             ["IAC-06"],
    "Telnet Enabled":                        ["CRY-03"],
    "HTTP Management Enabled":               ["CRY-03", "NET-06"],
    "SSH Disabled":                          ["CRY-03"],
    "No IP Filter Configured":               ["NET-04", "IAC-10"],
    "rsh/rlogin Enabled":                    ["CRY-03"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1 Enabled":                        ["NET-06", "CRY-03"],
    "Default SNMP Community String":         ["IAC-06"],
    "SNMP Write Community Configured":       ["IAC-07"],
    "No SNMPv3 AuthPriv":                    ["CRY-03", "NET-06"],
    # ── FC fabric security ────────────────────────────────────────────────────
    "FC Authentication Policy Disabled":     ["IAC-01"],
    # ── Fabric / zone security ────────────────────────────────────────────────
    "Default Zone Allows All Access":        ["NET-04"],
    "No Active Zone Configuration":          ["NET-04"],
    "Orphaned Zone Members":                 ["NET-04"],
    "Port-Based Zoning Used":               ["NET-04"],
    "Zone With Single Member":               ["NET-04"],
    "Orphaned Zone (Not in Active Config)":  ["NET-04"],
    "Oversized Zone":                        ["NET-04"],
    "No Fabric-Wide Consistency Policy":     ["NET-06"],
    "NPIV Not Enabled":                      ["NET-04"],
    # ── Logging / time ────────────────────────────────────────────────────────
    "No Syslog Server Configured":           ["MON-06"],
    "Audit Logging Disabled":                ["MON-06"],
    "NTP Not Configured":                    ["OPS-01"],
    "Multiple NTP Servers Not Configured":   ["OPS-01"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Weak Minimum Password Length":          ["IAC-06"],
    "No Password Complexity Requirements":   ["IAC-06"],
    "Password Never Expires":                ["IAC-06"],
    "No Password History":                   ["IAC-06"],
    # ── Session / management ──────────────────────────────────────────────────
    "No Login Banner Configured":            ["IAC-09"],
    "Weak SSH Ciphers Permitted":            ["CRY-03"],
}


def _scf_label(ctrl_ids: list[str]) -> str:
    return " · ".join(ctrl_ids)


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "hdr_bg":   "1B3A5C", "hdr_fg":   "FFFFFF",
    "critical": "C00000", "critical_l": "FFB3B3",
    "high":     "FF0000", "high_l":     "FFD9B3",
    "medium":   "FF8C00", "medium_l":   "FFF2CC",
    "low":      "0070C0", "low_l":      "BDD7EE",
    "info":     "595959", "info_l":     "F2F2F2",
    "ok":       "375623", "ok_l":       "E2EFDA",
    "disabled": "A6A6A6",
    "alt_row":  "F5F5F5",
}

_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=size)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── FOS Config Parser ─────────────────────────────────────────────────────────
class BrocadeFOSParser:
    """
    Parses concatenated Brocade FOS command output collected for a security
    audit. Sections are identified by command prompts and keyword headers.
    """

    # Known default community strings for SNMP
    _DEFAULT_COMMUNITIES = {"public", "private", "common", "fibrechannel",
                            "snmpv1_public", "snmpv1_private"}

    def __init__(self, config_file: str):
        self.config_file = config_file
        self.lines: list[str] = []

        # Parsed data
        self.switch_info: dict        = {}
        self.fos_version: str         = ""
        self.users: list[dict]        = []
        self.password_policy: dict    = {}
        self.snmp: dict               = {}
        self.syslog_servers: list[str] = []
        self.audit_enabled: bool      = False
        self.ntp_servers: list[str]   = []
        self.ip_filters: list[dict]   = []
        self.ssh_info: dict           = {}
        self.telnet_enabled: bool     = False
        self.http_enabled: bool       = False
        self.rsh_enabled: bool        = False
        self.login_banner: str        = ""
        self.aaa_config: dict         = {}
        self.zones: list[dict]        = []
        self.zone_aliases: list[dict] = []
        self.cfgs: list[dict]         = []
        self.active_cfg: str          = ""
        self.default_zone: str        = "noaccess"
        self.defzone_line: int        = 0
        self.npiv_enabled: bool       = False
        self.auth_policy: int         = -1   # -1 = not found; 0=off,1=passive,2=active,3=strict
        self.auth_policy_line: int    = 0
        self.maps_policy: str         = ""
        self.maps_enabled: bool       = False
        self.banner_empty: bool       = False
        self.banner_line: int         = 0
        self.fabric_consistency: str  = ""
        self.issues: list[dict]       = []

    # ── Issue helper ──────────────────────────────────────────────────────────
    def _issue(self, severity: str, category: str, obj: str,
               description: str, recommendation: str, line: int = 0):
        cis_ids   = CIS_CONTROL_MAP.get(category, [])
        pci_ids   = PCI_DSS_MAP.get(category, [])
        scf_ids   = SCF_MAP.get(category, [])
        bench_ids = CIS_BENCHMARK_MAP.get(category, [])
        self.issues.append({
            "severity":        severity,
            "category":        category,
            "object":          obj,
            "description":     description,
            "recommendation":  recommendation,
            "line":            line if line else "",
            "cis_ids":         cis_ids,
            "pci_ids":         pci_ids,
            "cis_label":       _cis_label(cis_ids),
            "cis_benchmark":   _cis_benchmark_label(bench_ids),
            "pci_label":       _pci_label(pci_ids),
            "scf":             _scf_label(scf_ids),
        })

    # ── Entry point ───────────────────────────────────────────────────────────
    def parse(self):
        try:
            with open(self.config_file, encoding="utf-8", errors="replace") as fh:
                self.lines = fh.readlines()
        except FileNotFoundError:
            sys.exit(f"File not found: {self.config_file}")

        self._parse_switch_info()
        self._parse_version()
        self._parse_users()
        self._parse_password_policy()
        self._parse_snmp()
        self._parse_syslog()
        self._parse_audit()
        self._parse_ntp()
        self._parse_ip_filter()
        self._parse_ssh()
        self._parse_protocols()
        self._parse_aaa()
        self._parse_zones()
        self._parse_defzone()
        self._parse_fabric()
        self._run_checks()

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _find_section(self, *patterns: str) -> list[tuple[int, str]]:
        """Return lines in the section(s) matching any of the header patterns."""
        in_section = False
        result: list[tuple[int, str]] = []
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()
            if not stripped:
                continue
            for p in patterns:
                if re.search(p, line, re.I):
                    in_section = True
                    break
            else:
                if in_section and (
                    re.match(r"^[A-Z][A-Za-z\s/]+:?\s*$", line) or
                    re.match(r"^={4,}", line) or
                    re.match(r"^-{4,}", line)
                ):
                    in_section = False
            if in_section:
                result.append((lineno, line))
        return result

    def _section_lines(self, start_pattern: str,
                       end_patterns: list[str] | None = None) -> list[tuple[int, str]]:
        """Collect all lines between start_pattern and the next end_pattern."""
        collecting = False
        result: list[tuple[int, str]] = []
        end_re = [re.compile(e, re.I) for e in (end_patterns or [])]

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()

            if re.search(start_pattern, line, re.I):
                collecting = True
                result.append((lineno, line))
                continue

            if collecting:
                if end_re and any(e.search(line) for e in end_re):
                    break
                result.append((lineno, line))
        return result

    # ── Section parsers ───────────────────────────────────────────────────────
    def _parse_switch_info(self):
        """Extract switchname, model, WWN, domain ID from switchshow output."""
        for lineno, line in enumerate(self.lines, 1):
            m = re.match(r"switchName:\s*(.+)", line, re.I)
            if m:
                self.switch_info["name"] = m.group(1).strip()
                self.switch_info["name_line"] = lineno
            m = re.match(r"switchType:\s*(.+)", line, re.I)
            if m:
                self.switch_info["model"] = m.group(1).strip()
            m = re.match(r"switchWwn:\s*(.+)", line, re.I)
            if m:
                self.switch_info["wwn"] = m.group(1).strip()
            m = re.match(r"domainId:\s*(\d+)", line, re.I)
            if m:
                self.switch_info["domain_id"] = m.group(1).strip()
            m = re.match(r"fabricName:\s*(.+)", line, re.I)
            if m:
                self.switch_info["fabric_name"] = m.group(1).strip()
            m = re.match(r"FC Router:\s*(ON|OFF)", line, re.I)
            if m:
                self.switch_info["fc_router"] = m.group(1).upper() == "ON"

    def _parse_version(self):
        """Extract FOS firmware version."""
        for line in self.lines:
            m = re.match(r"Fabric OS:\s*v?(\S+)", line, re.I)
            if m:
                self.fos_version = m.group(1)
                break
            m = re.match(r"FabricOS:\s*v?(\S+)", line, re.I)
            if m:
                self.fos_version = m.group(1)
                break

    def _parse_users(self):
        """
        Parse userconfig --showlist output.
        Format:
            Account name: admin
            Roles: admin
            Account enabled: yes
            Password locked: no
            ...
        """
        i = 0
        lines = self.lines
        while i < len(lines):
            m = re.match(r"\s*Account name:\s*(\S+)", lines[i], re.I)
            if m:
                user: dict = {
                    "name":     m.group(1).strip(),
                    "roles":    [],
                    "enabled":  True,
                    "locked":   False,
                    "line":     i + 1,
                }
                j = i + 1
                while j < len(lines):
                    l = lines[j].strip()
                    if not l:
                        break
                    rm = re.match(r"Roles?:\s*(.+)", l, re.I)
                    if rm:
                        user["roles"] = [r.strip() for r in rm.group(1).split(",")]
                    em = re.match(r"Account enabled:\s*(yes|no)", l, re.I)
                    if em:
                        user["enabled"] = em.group(1).lower() == "yes"
                    lm = re.match(r"Password locked:\s*(yes|no)", l, re.I)
                    if lm:
                        user["locked"] = lm.group(1).lower() == "yes"
                    if re.match(r"Account name:", l, re.I):
                        break
                    j += 1
                self.users.append(user)
                i = j
            else:
                i += 1

    def _parse_password_policy(self):
        """
        Parse passwdcfg --show output.
        Format:
            Minimum password length: 8
            Maximum password length: 40
            Minimum uppercase character: 0
            ...
        """
        pp: dict = {
            "min_length":       None,
            "max_age":          None,
            "min_upper":        0,
            "min_lower":        0,
            "min_numeric":      0,
            "min_special":      0,
            "history":          0,
            "lockout_threshold": 0,
            "lockout_duration": 0,
        }
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"Minimum password length:\s*(\d+)", line, re.I)
            if m:
                pp["min_length"] = int(m.group(1))
                pp["min_length_line"] = lineno
            m = re.match(r"Maximum age.*?:\s*(\d+)", line, re.I)
            if m:
                pp["max_age"] = int(m.group(1))
            m = re.match(r"Minimum uppercase.*?:\s*(\d+)", line, re.I)
            if m:
                pp["min_upper"] = int(m.group(1))
            m = re.match(r"Minimum lowercase.*?:\s*(\d+)", line, re.I)
            if m:
                pp["min_lower"] = int(m.group(1))
            m = re.match(r"Minimum numeric.*?:\s*(\d+)", line, re.I)
            if m:
                pp["min_numeric"] = int(m.group(1))
            m = re.match(r"Minimum special.*?:\s*(\d+)", line, re.I)
            if m:
                pp["min_special"] = int(m.group(1))
            m = re.match(r"(?:Number of )?[Pp]assword history.*?:\s*(\d+)", line, re.I)
            if m:
                pp["history"] = int(m.group(1))
            m = re.match(r"Lockout threshold.*?:\s*(\d+)", line, re.I)
            if m:
                pp["lockout_threshold"] = int(m.group(1))
            m = re.match(r"Lockout duration.*?:\s*(\d+)", line, re.I)
            if m:
                pp["lockout_duration"] = int(m.group(1))
        self.password_policy = pp

    def _parse_snmp(self):
        """
        Parse snmpconfig --show snmpv1 and snmpv3 output.
        """
        snmp: dict = {
            "v1_enabled":    None,
            "v3_enabled":    None,
            "communities":   [],
            "trap_targets":  [],
            "v3_users":      [],
            "contact":       "",
            "location":      "",
        }

        in_snmpv3_user = False
        cur_v3_user: dict = {}

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()

            # v1 enable/disable
            m = re.match(r"SNMPv1:\s*(yes|no|enabled|disabled)", line, re.I)
            if m:
                snmp["v1_enabled"] = m.group(1).lower() in ("yes", "enabled")

            # v3 enable/disable
            m = re.match(r"SNMPv3:\s*(yes|no|enabled|disabled)", line, re.I)
            if m:
                snmp["v3_enabled"] = m.group(1).lower() in ("yes", "enabled")

            # Community strings (v1/v2c)
            m = re.match(r"Community\s+\d+\s+\((?:read|write)\):\s*(\S+)", line, re.I)
            if m:
                comm = m.group(1).strip()
                rw = "write" if "write" in line.lower() else "read"
                snmp["communities"].append({
                    "name": comm, "access": rw, "line": lineno
                })

            # Contact/Location
            m = re.match(r"(?:System\s+)?Contact:\s*(.+)", line, re.I)
            if m and not snmp["contact"]:
                snmp["contact"] = m.group(1).strip()
            m = re.match(r"(?:System\s+)?Location:\s*(.+)", line, re.I)
            if m and not snmp["location"]:
                snmp["location"] = m.group(1).strip()

            # Trap recipients
            m = re.match(r"Trap Recipient\s*\d*[:\s]+(\d+\.\d+\.\d+\.\d+)", line, re.I)
            if m and m.group(1) not in ("0.0.0.0", ""):
                snmp["trap_targets"].append({"ip": m.group(1), "line": lineno})

            # v3 users
            m = re.match(r"SNMPv3 User\s*\d*[:\s]+(.+)", line, re.I)
            if m:
                if cur_v3_user:
                    snmp["v3_users"].append(cur_v3_user)
                cur_v3_user = {"name": m.group(1).strip(), "auth": "", "priv": "", "line": lineno}
            m = re.match(r"Authentication:\s*(\S+)", line, re.I)
            if m and cur_v3_user:
                cur_v3_user["auth"] = m.group(1).strip()
            m = re.match(r"Privacy:\s*(\S+)", line, re.I)
            if m and cur_v3_user:
                cur_v3_user["priv"] = m.group(1).strip()

        if cur_v3_user:
            snmp["v3_users"].append(cur_v3_user)
        self.snmp = snmp

    def _parse_syslog(self):
        """Parse syslogadmin --show output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            # syslogadmin --show format: <IP>:<port> or just IP
            m = re.match(r"(\d+\.\d+\.\d+\.\d+)(?::(\d+))?", line)
            if m and "syslog" in "".join(
                    self.lines[max(0, lineno - 10):lineno]).lower():
                self.syslog_servers.append(m.group(1))
            # Alternative format: Syslog server: <ip>
            m = re.match(r"[Ss]yslog\s+server[:\s]+(\d+\.\d+\.\d+\.\d+)", line)
            if m:
                ip = m.group(1)
                if ip not in ("0.0.0.0", "") and ip not in self.syslog_servers:
                    self.syslog_servers.append(ip)

    def _parse_audit(self):
        """Parse audit log configuration."""
        for line in self.lines:
            if re.search(r"audit\s+log.*?:\s*(enabled|yes)", line, re.I):
                self.audit_enabled = True
            if re.search(r"auditcfg.*?--enable", line, re.I):
                self.audit_enabled = True

    def _parse_ntp(self):
        """Parse tsclockserver output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            # tsclockserver output: LOCL or IP(s) separated by semicolons
            m = re.match(r"TSCLOCKSERVER\s*:\s*(.+)", line, re.I)
            if m:
                servers = [s.strip() for s in m.group(1).split(";")]
                for s in servers:
                    if s and s.upper() not in ("LOCL", "LOCAL", ""):
                        self.ntp_servers.append(s)
            # Alternative format
            m = re.match(r"NTP server[:\s]+(\S+)", line, re.I)
            if m and m.group(1) not in ("0.0.0.0", "LOCL", ""):
                if m.group(1) not in self.ntp_servers:
                    self.ntp_servers.append(m.group(1))

    def _parse_ip_filter(self):
        """
        Parse ipfilter --show output.
        Format:
            Name: default_ipv4
            Type: IPv4
            Rule    Action  Src IP/Mask   Proto  Dst Port
            ----    ------  -----------   -----  --------
            1       PERMIT  any           tcp    22
            ...
        """
        in_filter = False
        cur_filter: dict | None = None

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()

            m = re.match(r"Name:\s*(\S+)", line, re.I)
            if m:
                if cur_filter:
                    self.ip_filters.append(cur_filter)
                cur_filter = {"name": m.group(1), "type": "", "rules": [], "line": lineno}
                in_filter = True
                continue

            if in_filter and cur_filter:
                tm = re.match(r"Type:\s*(\S+)", line, re.I)
                if tm:
                    cur_filter["type"] = tm.group(1)
                    continue
                # Rule lines: number action src proto dstport
                rm = re.match(
                    r"(\d+)\s+(PERMIT|DENY)\s+(\S+)\s+(\S+)\s+(\S+)",
                    line, re.I)
                if rm:
                    cur_filter["rules"].append({
                        "seq":    int(rm.group(1)),
                        "action": rm.group(2).upper(),
                        "src":    rm.group(3),
                        "proto":  rm.group(4),
                        "dst_port": rm.group(5),
                        "line":   lineno,
                    })

        if cur_filter:
            self.ip_filters.append(cur_filter)

    def _parse_ssh(self):
        """Parse sshutil --show output."""
        ssh: dict = {
            "enabled":  None,
            "ciphers":  [],
            "macs":     [],
            "kex":      [],
            "auth_methods": [],
            "v2_only":  False,
        }
        for line in self.lines:
            m = re.match(r"SSH\s+(?:Server\s+)?[Ee]nabled:\s*(yes|no)", line, re.I)
            if m:
                ssh["enabled"] = m.group(1).lower() == "yes"
            if re.search(r"protocol\s+2", line, re.I):
                ssh["v2_only"] = True
            m = re.match(r"[Cc]iphers?:\s*(.+)", line, re.I)
            if m:
                ssh["ciphers"] = [c.strip() for c in m.group(1).split(",")]
            m = re.match(r"MACs?:\s*(.+)", line, re.I)
            if m:
                ssh["macs"] = [c.strip() for c in m.group(1).split(",")]
        self.ssh_info = ssh

    def _parse_protocols(self):
        """
        Detect enabled/disabled protocols and key security settings from both
        the configShow key:value format and the traditional --show command output.
        """
        in_banner_section = False
        banner_content_found = False

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()

            # ── configShow [Banner] section ────────────────────────────────────
            if re.match(r"\[Banner\]", stripped, re.I):
                in_banner_section = True
                self.banner_line = lineno
                continue
            if in_banner_section:
                if re.match(r"\[", stripped):          # next section header
                    in_banner_section = False
                    if not banner_content_found:
                        self.banner_empty = True
                elif stripped:
                    banner_content_found = True
                    self.login_banner = stripped
                continue

            # ── configShow key:value format ────────────────────────────────────
            # auth.policy:0  (0=off, 1=passive, 2=active, 3=strict)
            m = re.match(r"auth\.policy\s*:\s*(\d+)$", stripped)
            if m and not re.search(r"auth\.policy\.dev", stripped):
                self.auth_policy = int(m.group(1))
                self.auth_policy_line = lineno

            # MAPS
            m = re.match(r"maps\.activePolicy\s*:\s*(\S+)", stripped)
            if m:
                self.maps_policy = m.group(1).strip()
            m = re.match(r"maps\.enabled\s*:\s*(\d+)", stripped)
            if m:
                self.maps_enabled = m.group(1) == "1"

            # Fabric remote FOS exec
            if re.match(r"fabric\.remoteFosexec\s*:\s*1", stripped):
                pass   # could be a finding in future

            # ── Traditional protocol lines ─────────────────────────────────────
            if re.search(r"telnetd?\s*[=:]\s*(yes|enabled|on)\b", line, re.I):
                self.telnet_enabled = True
            if re.search(r"httpd?\s*[=:]\s*(yes|enabled|on)\b", line, re.I):
                self.http_enabled = True
            if re.search(r"rsh\s*[=:]\s*(yes|enabled|on)\b", line, re.I):
                self.rsh_enabled = True

            # NPIV detection from switchShow port listing
            if re.search(r"NPIV\s+public", line):
                self.npiv_enabled = True

    def _parse_aaa(self):
        """Parse AAA / RADIUS / LDAP configuration."""
        aaa: dict = {
            "radius_servers":  [],
            "ldap_servers":    [],
            "primary_method":  "local",
        }
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"Primary\s+auth\s+method:\s*(\S+)", line, re.I)
            if m:
                aaa["primary_method"] = m.group(1).lower()
            m = re.match(r"RADIUS\s+server[:\s]+(\S+)", line, re.I)
            if m:
                aaa["radius_servers"].append({"host": m.group(1), "line": lineno})
            m = re.match(r"LDAP\s+server[:\s]+(\S+)", line, re.I)
            if m:
                aaa["ldap_servers"].append({"host": m.group(1), "line": lineno})
            # aaaconfig --showlist format
            m = re.match(r"Authentication\s+server\s*\d*[:\s]+(\S+)", line, re.I)
            if m and m.group(1) not in ("local", "none"):
                aaa["radius_servers"].append({"host": m.group(1), "line": lineno})
        self.aaa_config = aaa

    def _parse_zones(self):
        """
        Parse cfgshow output.
        Format:
            Defined configuration:
             cfg:   PROD_CFG
                     zone1; zone2
             zone:  zone_ESX01_array01
                     10:00:00:00:c9:xx:xx:xx; 50:00:14:40:xx:xx:xx
             alias: esx01_hba0
                     10:00:00:00:c9:xx:xx:xx
            Effective configuration:
             cfg:   PROD_CFG
             zone:  zone_ESX01_array01
        """
        zones: list[dict] = []
        aliases: list[dict] = []
        cfgs: list[dict] = []
        active_cfg = ""

        i = 0
        lines = self.lines
        in_effective = False

        while i < len(lines):
            line = lines[i].rstrip("\n")
            stripped = line.strip()

            if re.search(r"Effective configuration", line, re.I):
                in_effective = True

            # cfg definition — handles both " cfg: NAME" and " cfg:NAME"
            m = re.match(r"\s+cfg:\s*(\S+)", line)
            if m:
                cfg_name = m.group(1).strip().rstrip(";")
                members_str = ""
                j = i + 1
                while j < len(lines):
                    cont = lines[j].rstrip("\n").strip()
                    if not cont or re.match(r"\s+(?:cfg|zone|alias):\s*\S", lines[j]):
                        break
                    members_str += cont
                    j += 1
                members = [x.strip().rstrip(";") for x in re.split(r"[;,]", members_str) if x.strip()]
                if in_effective:
                    active_cfg = cfg_name
                else:
                    cfgs.append({"name": cfg_name, "zones": members, "line": i + 1})
                i = j
                continue

            # zone definition — handles both " zone: NAME" and " zone:NAME"
            m = re.match(r"\s+zone:\s*(\S+)", line)
            if m and not in_effective:
                zone_name = m.group(1).strip().rstrip(";")
                members_str = ""
                j = i + 1
                while j < len(lines):
                    cont = lines[j].rstrip("\n").strip()
                    if not cont or re.match(r"\s+(?:cfg|zone|alias):\s*\S", lines[j]):
                        break
                    members_str += cont
                    j += 1
                raw_members = [x.strip().rstrip(";") for x in re.split(r"[;\s,]+", members_str) if x.strip()]
                # Determine member types
                port_based = any(re.match(r"^\d+,\d+$", m) for m in raw_members)
                wwn_based = any(re.match(r"^[\da-fA-F]{2}:[\da-fA-F]{2}:", m) for m in raw_members)
                zones.append({
                    "name":       zone_name,
                    "members":    raw_members,
                    "port_based": port_based,
                    "wwn_based":  wwn_based,
                    "line":       i + 1,
                })
                i = j
                continue

            # alias definition — handles both " alias: NAME" and " alias:NAME"
            m = re.match(r"\s+alias:\s*(\S+)", line)
            if m and not in_effective:
                alias_name = m.group(1).strip().rstrip(";")
                members_str = ""
                j = i + 1
                while j < len(lines):
                    cont = lines[j].rstrip("\n").strip()
                    if not cont or re.match(r"\s+(?:cfg|zone|alias):\s*\S", lines[j]):
                        break
                    members_str += cont
                    j += 1
                raw_members = [x.strip().rstrip(";") for x in re.split(r"[;\s,]+", members_str) if x.strip()]
                aliases.append({"name": alias_name, "members": raw_members, "line": i + 1})
                i = j
                continue

            i += 1

        self.zones = zones
        self.zone_aliases = aliases
        self.cfgs = cfgs
        self.active_cfg = active_cfg

    def _parse_defzone(self):
        """Parse defzone setting from both configShow k:v and defzone --show output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()
            m = re.match(r"(?:Default Zone Access|defzone)\s*[:\-=]\s*(\S+)", stripped, re.I)
            if m:
                self.default_zone = m.group(1).lower().strip()
                self.defzone_line = lineno
                break
            if re.search(r"all.?access", stripped, re.I) and re.search(r"def.*zone|defzone", stripped, re.I):
                self.default_zone = "allaccess"
                self.defzone_line = lineno
                break
            if re.search(r"no.?access", stripped, re.I) and re.search(r"def.*zone|defzone", stripped, re.I):
                self.default_zone = "noaccess"
                self.defzone_line = lineno
                break

    def _parse_fabric(self):
        """Parse fabric-wide consistency policy."""
        for line in self.lines:
            m = re.match(r"(?:Fabric|fabricwide).*?consistency.*?:\s*(\S+)", line, re.I)
            if m:
                self.fabric_consistency = m.group(1)
                break

    # ── Security checks ───────────────────────────────────────────────────────
    def _run_checks(self):
        self._chk_users()
        self._chk_password_policy()
        self._chk_aaa()
        self._chk_protocols()
        self._chk_ssh()
        self._chk_snmp()
        self._chk_syslog()
        self._chk_audit()
        self._chk_ntp()
        self._chk_ip_filter()
        self._chk_auth_policy()
        self._chk_zones()
        self._chk_orphaned_zones()
        self._chk_fabric()

    def _chk_users(self):
        admin_count = 0
        default_names = {"admin", "user", "root", "factory"}
        for u in self.users:
            roles_lower = [r.lower() for r in u["roles"]]
            if any(r in ("admin", "root", "factory") for r in roles_lower):
                admin_count += 1
            if u["name"].lower() in default_names and u["enabled"]:
                self._issue(
                    "HIGH", "Default Account Not Disabled",
                    f"account: {u['name']}",
                    f"Default account '{u['name']}' is still enabled. "
                    "Attackers commonly target well-known default FOS accounts.",
                    "Disable unused default accounts with 'userconfig --delete' or lock "
                    "with 'userconfig --change <name> -e no'.",
                    line=u["line"])

        if admin_count > 3:
            self._issue(
                "MEDIUM", "Excessive Admin Accounts", f"{admin_count} admin-role accounts",
                f"{admin_count} accounts have admin/root-level roles. Excessive privileged "
                "accounts increase the attack surface.",
                "Limit admin-role accounts to named individuals. Use role-based accounts "
                "with least-privilege roles (operator, securityadmin) where possible.")

    def _chk_password_policy(self):
        pp = self.password_policy
        if not pp:
            return

        min_len = pp.get("min_length")
        if min_len is not None:
            if min_len < 8:
                self._issue(
                    "HIGH", "Weak Minimum Password Length",
                    f"min length: {min_len}",
                    f"Minimum password length is only {min_len} characters (recommended: ≥14).",
                    "Run 'passwdcfg --set -minpasswordlength 14' to enforce stronger passwords.",
                    line=pp.get("min_length_line", 0))
            elif min_len < 12:
                self._issue(
                    "MEDIUM", "Weak Minimum Password Length",
                    f"min length: {min_len}",
                    f"Minimum password length is {min_len} characters (recommended: ≥14).",
                    "Run 'passwdcfg --set -minpasswordlength 14'.")

        has_complexity = (pp.get("min_upper", 0) >= 1 or
                          pp.get("min_lower", 0) >= 1 or
                          pp.get("min_numeric", 0) >= 1 or
                          pp.get("min_special", 0) >= 1)
        if not has_complexity:
            self._issue(
                "MEDIUM", "No Password Complexity Requirements",
                "passwdcfg",
                "No minimum uppercase, lowercase, numeric, or special character requirements "
                "are enforced for passwords.",
                "Use 'passwdcfg --set -minuppercase 1 -minlowercase 1 -minnumeric 1 "
                "-minspecialchar 1' to enforce complexity.")

        max_age = pp.get("max_age")
        if max_age is not None and max_age == 0:
            self._issue(
                "MEDIUM", "Password Never Expires",
                "passwdcfg max age: 0",
                "Maximum password age is set to 0 (passwords never expire). "
                "Stale credentials increase risk of long-term credential compromise.",
                "Set maximum password age: 'passwdcfg --set -maxpasswdage 90'.")

        history = pp.get("history", 0)
        if history < 5:
            self._issue(
                "LOW", "No Password History",
                f"history: {history}",
                f"Password history is set to {history} (recommended: ≥5). "
                "Users can immediately reuse old passwords.",
                "Set password history: 'passwdcfg --set -passwdhistorysize 5'.")

        threshold = pp.get("lockout_threshold", 0)
        if threshold == 0:
            self._issue(
                "HIGH", "No Account Lockout Policy",
                "lockout threshold: 0",
                "Account lockout is disabled (threshold=0). Brute-force attacks "
                "can proceed indefinitely without triggering a lockout.",
                "Enable lockout: 'passwdcfg --set -lockoutthreshold 5 "
                "-lockoutduration 30'.")

    def _chk_aaa(self):
        aaa = self.aaa_config
        has_external = bool(aaa.get("radius_servers")) or bool(aaa.get("ldap_servers"))
        if not has_external:
            self._issue(
                "HIGH", "No RADIUS or LDAP Configured",
                "AAA",
                "No external authentication (RADIUS/LDAP) is configured. "
                "All authentication uses local accounts with no centralized control.",
                "Configure RADIUS or LDAP for centralized authentication. "
                "Use 'aaaconfig --add <server> -conf radius' or 'ldapcfg --set'.")

    def _chk_protocols(self):
        if self.telnet_enabled:
            self._issue(
                "CRITICAL", "Telnet Enabled",
                "telnet: enabled",
                "Telnet transmits credentials and all data in cleartext. "
                "Any attacker with network access can capture switch credentials.",
                "Disable Telnet: 'ipfilter --create custom_deny_telnet' blocking TCP 23, "
                "or 'configure switch' and disable telnetd.")

        if self.http_enabled:
            self._issue(
                "HIGH", "HTTP Management Enabled",
                "httpd: enabled",
                "HTTP (port 80) management exposes credentials and session tokens in cleartext.",
                "Disable HTTP management and require HTTPS only. "
                "Use 'httpd --disable' or configure IP filter to block TCP 80.")

        if self.rsh_enabled:
            self._issue(
                "CRITICAL", "rsh/rlogin Enabled",
                "rsh: enabled",
                "rsh/rlogin are legacy protocols with no encryption and no modern "
                "authentication. Their use violates virtually all security frameworks.",
                "Disable rsh: 'configure' then disable rshd, or block with IP filter.")

    def _chk_ssh(self):
        ssh = self.ssh_info
        if ssh.get("enabled") is False:
            self._issue(
                "CRITICAL", "SSH Disabled",
                "SSH: disabled",
                "SSH is the only secure remote management protocol on FOS. "
                "If it is disabled, administrators are forced to use Telnet or console.",
                "Enable SSH: 'sshutil enable'. Ensure public-key auth is configured "
                "for privileged accounts.")

        weak_ciphers = {"des", "3des", "rc4", "blowfish", "arcfour"}
        found_weak = [c for c in ssh.get("ciphers", []) if c.lower() in weak_ciphers]
        if found_weak:
            self._issue(
                "HIGH", "Weak SSH Ciphers Permitted",
                f"ciphers: {', '.join(found_weak)}",
                f"Weak or deprecated SSH ciphers are permitted: {', '.join(found_weak)}. "
                "These are susceptible to cryptographic attacks.",
                "Remove weak ciphers with 'sshutil --set -cipher <list>'. "
                "Permit only AES-128-CTR, AES-256-CTR, AES-128-GCM, AES-256-GCM.")

    def _chk_snmp(self):
        snmp = self.snmp

        if snmp.get("v1_enabled"):
            self._issue(
                "HIGH", "SNMPv1 Enabled",
                "SNMPv1: enabled",
                "SNMPv1 uses cleartext community strings with no authentication or "
                "encryption. It allows unauthenticated read (or write) access to MIB data.",
                "Disable SNMPv1: 'snmpconfig --set snmpv1 -name 1 -enable no'. "
                "Migrate to SNMPv3 with authPriv security level.")

        default_comms = [c for c in snmp.get("communities", [])
                         if c["name"].lower() in self._DEFAULT_COMMUNITIES]
        for comm in default_comms:
            self._issue(
                "CRITICAL", "Default SNMP Community String",
                f"community: {comm['name']} ({comm['access']})",
                f"Default SNMP community string '{comm['name']}' is in use. "
                "This is universally known and gives attackers read (or write) fabric access.",
                "Change all community strings to unique, complex values: "
                "'snmpconfig --set snmpv1 -name 1 -community <new-string>'.",
                line=comm["line"])

        write_comms = [c for c in snmp.get("communities", []) if c["access"] == "write"]
        for comm in write_comms:
            self._issue(
                "HIGH", "SNMP Write Community Configured",
                f"community: {comm['name']} (write)",
                f"SNMP write community string '{comm['name']}' is configured. "
                "Write access allows an attacker to modify switch configuration via SNMP.",
                "Remove SNMP write community strings unless specifically required. "
                "Prefer SNMPv3 with read-only access for monitoring.",
                line=comm["line"])

        has_authpriv = any(
            u.get("auth") not in ("", "none", None) and u.get("priv") not in ("", "none", None)
            for u in snmp.get("v3_users", [])
        )
        if not has_authpriv and not snmp.get("v3_enabled"):
            self._issue(
                "MEDIUM", "No SNMPv3 AuthPriv",
                "SNMPv3 authPriv: not configured",
                "No SNMPv3 users with authPriv security level are configured. "
                "AuthPriv provides both authentication and encryption for SNMP.",
                "Configure SNMPv3 users with authPriv: "
                "'snmpconfig --set snmpv3 -name 1 -secname <user> "
                "-authtype SHA -privtype AES128'.")

    def _chk_syslog(self):
        if not self.syslog_servers:
            self._issue(
                "HIGH", "No Syslog Server Configured",
                "syslog: none",
                "No remote syslog server is configured. Security events, login attempts, "
                "and zone changes are only stored locally and may be overwritten.",
                "Configure syslog: 'syslogadmin --set -ip <syslog-server>'. "
                "Forward to a SIEM for correlation and long-term retention.")

    def _chk_audit(self):
        if not self.audit_enabled:
            self._issue(
                "HIGH", "Audit Logging Disabled",
                "audit logging: disabled",
                "FOS audit logging is not enabled. Administrative changes, config exports, "
                "and zone modifications will not be tracked in the audit trail.",
                "Enable audit logging: 'auditcfg --enable'. "
                "Audit records are critical for change accountability in SAN environments.")

    def _chk_ntp(self):
        if not self.ntp_servers:
            self._issue(
                "HIGH", "NTP Not Configured",
                "NTP: none",
                "No NTP server is configured (switch uses LOCL/internal clock). "
                "Clock drift invalidates log timestamps and breaks cross-device correlation.",
                "Configure NTP: 'tsclockserver <ntp-server>'. "
                "Use the same NTP source as your vSphere hosts for log correlation.")

        elif len(self.ntp_servers) < 2:
            self._issue(
                "LOW", "Multiple NTP Servers Not Configured",
                f"NTP servers: {len(self.ntp_servers)}",
                "Only one NTP server is configured. A single NTP server is a "
                "single point of failure for time synchronization.",
                "Add a secondary NTP server: 'tsclockserver <primary>;<secondary>'.")

    def _chk_ip_filter(self):
        if not self.ip_filters:
            self._issue(
                "HIGH", "No IP Filter Configured",
                "ipfilter: none",
                "No IP filter policy is applied to the management interfaces. "
                "Any host with IP connectivity can attempt to authenticate to the switch.",
                "Create an IP filter restricting SSH (TCP 22) and HTTPS (TCP 443) "
                "to management subnet: 'ipfilter --create mgmt_filter -type ipv4' "
                "then 'ipfilter --save; ipfilter --activate mgmt_filter'.")

    def _chk_auth_policy(self):
        """
        Check Brocade FC Authentication Policy (DH-CHAP / SCC).
        auth.policy:0  = off  — no FC-level authentication between switches/devices
        auth.policy:1  = passive (authenticate only if peer initiates)
        auth.policy:2  = active
        auth.policy:3  = strict (required for all E-ports)
        """
        if self.auth_policy == -1:
            return   # command not in input — can't assess

        if self.auth_policy == 0:
            self._issue(
                "HIGH", "FC Authentication Policy Disabled",
                "auth.policy: 0 (off)",
                "FC Authentication Policy is disabled (auth.policy=0 found in configShow). "
                "No DH-CHAP authentication is required between E-ports or N-ports. "
                "An unauthorized switch or rogue device can join the fabric without credentials, "
                "enabling eavesdropping or spoofed SCSI commands on VMware datastores.",
                "Enable FC authentication: 'authutil --set -policy active'. "
                "Configure DH-CHAP shared secrets on all ISL partners: 'secauthsecret --set'. "
                "Minimum acceptable: passive mode (1) for legacy device compatibility.",
                line=self.auth_policy_line)

    def _chk_orphaned_zones(self):
        """
        Detect zones that are defined in cfgshow but not included in any
        zone configuration (cfg). These may be stale, forgotten, or accidentally
        omitted — and in some cases indicate hosts that lost storage access.
        """
        if not self.zones:
            return

        # Collect all zone names referenced by any cfg
        all_cfg_zones: set[str] = set()
        for cfg in self.cfgs:
            for z in cfg.get("zones", []):
                all_cfg_zones.add(z.strip().rstrip(";"))

        orphaned = [z for z in self.zones if z["name"] not in all_cfg_zones]
        for z in orphaned:
            self._issue(
                "MEDIUM", "Orphaned Zone (Not in Active Config)",
                f"zone: {z['name']}",
                f"Zone '{z['name']}' (cfgShow line {z['line']}) is defined but not included "
                "in any zone configuration (cfg). It is inactive. "
                "Orphaned zones may represent stale access paths, hosts that lost "
                "connectivity, or zones accidentally omitted from the active cfg.",
                "Review and either delete orphaned zones ('zonedelete') or add them "
                "to the appropriate cfg. Run 'cfgsave' after changes.",
                line=z["line"])

    def _chk_zones(self):
        if self.default_zone.lower() in ("allaccess", "all access", "all_access"):
            self._issue(
                "CRITICAL", "Default Zone Allows All Access",
                f"defzone: {self.default_zone}",
                "The default zone is set to 'allaccess'. Any device not explicitly "
                "placed in a zone can communicate with ALL other unzoned devices — "
                "an open fabric. This violates SAN isolation requirements for VMware.",
                "Set default zone to 'noaccess': 'defzone --noaccess; cfgsave'.",
                line=self.defzone_line)

        if not self.cfgs and not self.zones:
            self._issue(
                "CRITICAL", "No Active Zone Configuration",
                "cfgshow: empty",
                "No zone configuration exists in this fabric. All ports can communicate "
                "with all other ports regardless of defzone setting on some FOS versions.",
                "Create dedicated per-initiator zones: one initiator WWN and one "
                "target port WWN per zone (single-initiator single-target). "
                "Define a cfg and activate it: 'cfgenable <cfgname>'.")

        elif self.active_cfg == "" and self.cfgs:
            self._issue(
                "HIGH", "No Active Zone Configuration",
                "effective cfg: none",
                f"{len(self.cfgs)} zone configuration(s) are defined but none is active. "
                "An inactive cfg means defzone policy governs all fabric traffic.",
                "Activate the appropriate zone configuration: 'cfgenable <cfgname>'.")

        for zone in self.zones:
            if len(zone["members"]) <= 1:
                self._issue(
                    "MEDIUM", "Zone With Single Member",
                    f"zone: {zone['name']}",
                    f"Zone '{zone['name']}' has only {len(zone['members'])} member(s). "
                    "A single-member zone provides no connectivity and may be an orphaned entry.",
                    "Review and clean up single-member zones. Each zone should have "
                    "at least one initiator and one target.",
                    line=zone["line"])

            if len(zone["members"]) > 20:
                self._issue(
                    "MEDIUM", "Oversized Zone",
                    f"zone: {zone['name']} ({len(zone['members'])} members)",
                    f"Zone '{zone['name']}' has {len(zone['members'])} members. "
                    "Oversized zones reduce isolation and increase fabric-wide RSCN storm risk "
                    "in VMware environments.",
                    "Restructure to single-initiator zones: one ESXi HBA per zone with "
                    "only its target ports. This minimises RSCN impact per host.",
                    line=zone["line"])

            if zone["port_based"]:
                self._issue(
                    "MEDIUM", "Port-Based Zoning Used",
                    f"zone: {zone['name']}",
                    f"Zone '{zone['name']}' uses port-based (D,P) zoning. "
                    "Port-based zoning ties security to physical port location — "
                    "replacing a cable bypasses zone isolation.",
                    "Migrate to WWN-based zoning for VMware environments. "
                    "WWN zoning persists across cabling changes and is recommended by VMware.",
                    line=zone["line"])

            # Multi-member zone: >2 aliases = likely multiple initiators or a
            # many-to-many zone, both of which violate single-initiator best practice
            if len(zone["members"]) > 2:
                member_list = "; ".join(zone["members"][:5])
                self._issue(
                    "LOW", "Oversized Zone",
                    f"zone: {zone['name']} ({len(zone['members'])} members)",
                    f"Zone '{zone['name']}' contains {len(zone['members'])} members "
                    f"({member_list}{'...' if len(zone['members'])>5 else ''}). "
                    "Best practice for VMware FC SAN is single-initiator / single-target-alias "
                    "zoning to minimise RSCN blast radius and isolate host-to-storage paths.",
                    "Restructure as one initiator alias per zone paired with the appropriate "
                    "target alias. Use 'zonecreate' for each initiator-target pair, then "
                    "update the cfg with 'cfgadd' and activate with 'cfgenable'.",
                    line=zone["line"])

    def _chk_fabric(self):
        if not self.fabric_consistency:
            self._issue(
                "INFO", "No Fabric-Wide Consistency Policy",
                "fabricConsistencyPolicy: not set",
                "No fabric-wide consistency policy is configured. Without this policy, "
                "switches in the fabric may have divergent security settings.",
                "Configure a fabric-wide consistency policy to enforce uniform "
                "settings: 'fwconsistencypolicy --set tolerant' or 'strict'.")

        if self.banner_empty or not self.login_banner:
            self._issue(
                "LOW", "No Login Banner Configured",
                "[Banner] section is empty" if self.banner_empty else "motd: none",
                "No login banner (MOTD) is configured on the switch. "
                "The [Banner] section in configShow is empty. "
                "A legal notice banner is required by most compliance frameworks "
                "and establishes legal standing for access monitoring.",
                "Configure a login banner: 'motd' command or 'bannercfg --set "
                "\"Unauthorized access prohibited. All sessions are logged.\"'.",
                line=self.banner_line)


# ── Excel Reporter ────────────────────────────────────────────────────────────
class ExcelReporter:

    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"], C["critical_l"]),
        "HIGH":     (C["high"],     C["high_l"]),
        "MEDIUM":   (C["medium"],   C["medium_l"]),
        "LOW":      (C["low"],      C["low_l"]),
        "INFO":     (C["info"],     C["info_l"]),
    }

    def __init__(self, parser: BrocadeFOSParser, out_path: str):
        self.p   = parser
        self.out = out_path
        self.wb  = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    # ── Shared helpers ────────────────────────────────────────────────────────
    def _hdr(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = _font(bold=True, color=C["hdr_fg"])
            c.fill      = _fill(C["hdr_bg"])
            c.alignment = _align("center", wrap=False)
            c.border    = THIN
        ws.row_dimensions[row].height = 20

    @staticmethod
    def _set_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    @staticmethod
    def _row_fill(row: int) -> str | None:
        return C["alt_row"] if row % 2 == 0 else None

    def _sev_counts(self) -> dict[str, int]:
        counts: dict[str, int] = defaultdict(int)
        for iss in self.p.issues:
            counts[iss["severity"]] += 1
        return counts

    # ── Summary sheet ─────────────────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "Brocade FOS SAN Switch — Security Audit Report"
        t.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 40

        # Metadata
        sw = self.p.switch_info
        meta = [
            ("Switch Name",      sw.get("name", "Unknown")),
            ("Switch WWN",       sw.get("wwn", "Unknown")),
            ("Model",            sw.get("model", "Unknown")),
            ("Fabric Name",      sw.get("fabric_name", "Unknown")),
            ("Domain ID",        sw.get("domain_id", "Unknown")),
            ("FOS Version",      self.p.fos_version or "Unknown"),
            ("Report Date",      datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Input File",       os.path.basename(self.p.config_file)),
        ]
        for row, (label, value) in enumerate(meta, 2):
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font(bold=True); c1.alignment = _align(); c1.border = THIN
            c1.fill = _fill("EAF0FB")
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = _font(); c2.alignment = _align(); c2.border = THIN
            ws.merge_cells(f"B{row}:F{row}")

        # Finding summary
        row = len(meta) + 3
        ws.cell(row=row, column=1, value="FINDING SUMMARY").font = _font(bold=True, size=12)
        row += 1

        sev_counts = self._sev_counts()
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            count = sev_counts.get(sev, 0)
            fg, bg = self.SEV_COLORS[sev]
            c1 = ws.cell(row=row, column=1, value=sev)
            c1.fill = _fill(bg); c1.font = _font(bold=True, color=fg)
            c1.alignment = _align("center"); c1.border = THIN
            c2 = ws.cell(row=row, column=2, value=count)
            c2.font = _font(bold=(count > 0)); c2.alignment = _align("center")
            c2.border = THIN
            ws.merge_cells(f"B{row}:F{row}")
            row += 1

        row += 1
        total = sum(sev_counts.values())
        ws.cell(row=row, column=1, value="TOTAL").font = _font(bold=True)
        ws.cell(row=row, column=1).border = THIN
        ws.cell(row=row, column=2, value=total).font = _font(bold=True)
        ws.cell(row=row, column=2).border = THIN

        self._set_widths(ws, [28, 35, 20, 20, 20, 20])

    # ── Zones sheet ───────────────────────────────────────────────────────────
    def _sheet_zones(self):
        ws = self.wb.create_sheet("Zone Configuration")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "Zone Configuration — Active & Defined"
        t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 32

        row = 2
        # Active config banner
        ws.merge_cells(f"A{row}:G{row}")
        ac = ws.cell(row=row, column=1,
                     value=f"Active Config: {self.p.active_cfg or '(none — fabric is OPEN)'}")
        ac.font = _font(bold=True, color="FFFFFF")
        ac.fill = _fill("C00000" if not self.p.active_cfg else "375623")
        ac.alignment = _align("center", wrap=False); ac.border = THIN
        ws.row_dimensions[row].height = 24
        row += 2

        # Default zone
        dz = self.p.default_zone.lower()
        ws.merge_cells(f"A{row}:G{row}")
        dzc = ws.cell(row=row, column=1,
                      value=f"Default Zone: {dz.upper()}")
        dzc.font = _font(bold=True, color="FFFFFF")
        dzc.fill = _fill("C00000" if "allaccess" in dz else "375623")
        dzc.alignment = _align("center", wrap=False); dzc.border = THIN
        ws.row_dimensions[row].height = 20
        row += 2

        # Zones table
        headers = ["Zone Name", "Member Count", "Member Type",
                   "Port-Based", "Members (truncated)", "Config Line", "Notes"]
        self._hdr(ws, headers, row=row)
        row += 1

        for zone in self.p.zones:
            rb = self._row_fill(row)
            mtype = "WWN" if zone["wwn_based"] else ("Port" if zone["port_based"] else "Alias/Mixed")
            notes = []
            if len(zone["members"]) <= 1:
                notes.append("ORPHANED")
            if len(zone["members"]) > 20:
                notes.append("OVERSIZED")
            if zone["port_based"]:
                notes.append("PORT-BASED")
            members_preview = "; ".join(zone["members"][:4])
            if len(zone["members"]) > 4:
                members_preview += f" ... (+{len(zone['members'])-4} more)"
            vals = [zone["name"], len(zone["members"]), mtype,
                    "yes" if zone["port_based"] else "",
                    members_preview, zone["line"], ", ".join(notes)]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
                if col == 7 and notes:
                    c.font = _font(bold=True, color=C["high"])
            row += 1

        row += 1
        # Aliases table
        ws.cell(row=row, column=1, value="ZONE ALIASES").font = _font(bold=True, size=11)
        row += 1
        alias_hdrs = ["Alias Name", "Member Count", "Members", "Config Line"]
        self._hdr(ws, alias_hdrs, row=row)
        row += 1
        for alias in self.p.zone_aliases:
            rb = self._row_fill(row)
            members_preview = "; ".join(alias["members"][:3])
            if len(alias["members"]) > 3:
                members_preview += f" ... (+{len(alias['members'])-3} more)"
            vals = [alias["name"], len(alias["members"]), members_preview, alias["line"]]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            row += 1

        self._set_widths(ws, [40, 14, 14, 12, 60, 12, 30])

    # ── Users sheet ───────────────────────────────────────────────────────────
    def _sheet_users(self):
        ws = self.wb.create_sheet("Users & AAA")
        ws.sheet_view.showGridLines = False

        headers = ["Username", "Roles", "Enabled", "Locked", "Config Line", "Notes"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"

        default_names = {"admin", "user", "root", "factory"}
        for i, u in enumerate(self.p.users, 2):
            rb = self._row_fill(i)
            notes = []
            if u["name"].lower() in default_names:
                notes.append("DEFAULT ACCOUNT")
            if not u["enabled"]:
                notes.append("disabled")
            if u["locked"]:
                notes.append("LOCKED")
            vals = [u["name"], ", ".join(u["roles"]),
                    "yes" if u["enabled"] else "no",
                    "yes" if u["locked"] else "no",
                    u["line"], ", ".join(notes)]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(color=C["disabled"] if not u["enabled"] else "000000")
                c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
                if col == 6 and "DEFAULT ACCOUNT" in notes:
                    c.font = _font(bold=True, color=C["high"])

        row = len(self.p.users) + 3

        # AAA summary
        ws.cell(row=row, column=1, value="AAA CONFIGURATION").font = _font(bold=True, size=11)
        row += 1
        aaa = self.p.aaa_config
        for label, val in [
            ("Primary auth method", aaa.get("primary_method", "local")),
            ("RADIUS servers",      len(aaa.get("radius_servers", []))),
            ("LDAP servers",        len(aaa.get("ldap_servers", []))),
        ]:
            ws.cell(row=row, column=1, value=label).font = _font(bold=True)
            ws.cell(row=row, column=1).border = THIN
            ws.cell(row=row, column=2, value=str(val)).border = THIN
            row += 1

        self._set_widths(ws, [20, 25, 10, 10, 12, 40])

    # ── SNMP sheet ────────────────────────────────────────────────────────────
    def _sheet_snmp(self):
        ws = self.wb.create_sheet("SNMP & Protocols")
        ws.sheet_view.showGridLines = False

        row = 1
        # Protocol summary
        ws.cell(row=row, column=1, value="MANAGEMENT PROTOCOLS").font = _font(bold=True, size=11)
        row += 1
        self._hdr(ws, ["Protocol", "Status", "Risk"], row=row)
        row += 1
        proto_data = [
            ("SSH",     "enabled" if self.p.ssh_info.get("enabled") is not False else "DISABLED",
             "Low" if self.p.ssh_info.get("enabled") is not False else "CRITICAL"),
            ("Telnet",  "ENABLED" if self.p.telnet_enabled else "disabled",
             "CRITICAL" if self.p.telnet_enabled else "Low"),
            ("HTTP",    "ENABLED" if self.p.http_enabled else "disabled",
             "HIGH" if self.p.http_enabled else "Low"),
            ("rsh",     "ENABLED" if self.p.rsh_enabled else "disabled",
             "CRITICAL" if self.p.rsh_enabled else "Low"),
            ("NPIV",    "enabled" if self.p.npiv_enabled else "disabled",
             "Info"),
        ]
        for proto, status, risk in proto_data:
            rb = self._row_fill(row)
            bad = risk in ("CRITICAL", "HIGH")
            for col, val in enumerate([proto, status, risk], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(bold=bad, color=C["critical"] if bad else "000000")
                c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="SNMP COMMUNITIES").font = _font(bold=True, size=11)
        row += 1
        self._hdr(ws, ["Community String", "Access", "Is Default?", "Config Line"], row=row)
        row += 1
        snmp = self.p.snmp
        for comm in snmp.get("communities", []):
            rb = self._row_fill(row)
            is_def = comm["name"].lower() in BrocadeFOSParser._DEFAULT_COMMUNITIES
            for col, val in enumerate(
                    [comm["name"], comm["access"],
                     "YES - change immediately!" if is_def else "no",
                     comm["line"]], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(bold=(col == 3 and is_def),
                               color=C["critical"] if (col == 3 and is_def) else "000000")
                c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            row += 1

        if not snmp.get("communities"):
            ws.cell(row=row, column=1, value="(no communities configured)").font = _font(italic=True)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="SNMPv3 USERS").font = _font(bold=True, size=11)
        row += 1
        self._hdr(ws, ["Username", "Auth Protocol", "Privacy Protocol", "Config Line"], row=row)
        row += 1
        for u in snmp.get("v3_users", []):
            rb = self._row_fill(row)
            for col, val in enumerate(
                    [u["name"], u.get("auth", "none"), u.get("priv", "none"), u["line"]], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            row += 1

        if not snmp.get("v3_users"):
            ws.cell(row=row, column=1, value="(no SNMPv3 users configured)").font = _font(italic=True)

        self._set_widths(ws, [30, 18, 28, 14])

    # ── Issues sheet ─────────────────────────────────────────────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False

        headers = ["#", "Validated", "Severity", "Residual Risk", "Residual Risk Note",
                   "Category", "Rule / Object", "Config Line(s)", "CIS v8", "CIS Benchmark",
                   "PCI DSS", "SCF",
                   "Description", "Recommendation", "Details",
                   "Asset", "Target", "Vuln", "Output", "Source"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        hostname = self.p.switch_info.get("name", "") or ""
        source = os.path.basename(self.p.config_file)

        sorted_issues = sorted(
            self.p.issues,
            key=lambda x: self.SEV_ORDER.get(x["severity"], 9))

        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            rb = self._row_fill(row)
            obj = iss["object"]
            line = iss.get("line", "")
            target = f"{obj} ({line})" if line else obj
            output = iss["description"]
            vals = [idx, "Y", sev, "", "",
                    iss["category"], obj, line,
                    iss.get("cis_label", ""), iss.get("cis_benchmark", ""),
                    iss.get("pci_label", ""), iss.get("scf", ""),
                    iss["description"], iss["recommendation"], "",
                    hostname, target, "", output, source]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 3:  # Severity
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                    c.alignment = _align("center")
                elif col in (1, 8):  # #, Config Line
                    c.font = _font(bold=(col == 1)); c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 9:  # CIS v8
                    c.font = _font(bold=True, color="17375E", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 10:  # CIS Benchmark
                    c.font = _font(bold=True, color="1F618D", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 11:  # PCI DSS
                    c.font = _font(bold=True, color="7B2D8B", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 12:  # SCF
                    c.font = _font(bold=True, color="1A5C3A", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 2:  # Validated
                    c.font = _font(bold=True)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                else:
                    c.font = _font(); c.alignment = _align()
                    if rb:
                        c.fill = _fill(rb)
            ws.row_dimensions[row].height = 40

        self._set_widths(ws, [4, 12, 12, 18, 28, 32, 36, 14, 20, 20, 18, 18, 60, 60, 36, 24, 44, 16, 70, 30])

    # ── CIS Controls mapping ──────────────────────────────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS Controls Mapping")
        ws.sheet_view.showGridLines = False
        CIS_HDR = "1B3A5C"

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v8 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(CIS_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each CIS Control lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        ctrl_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        row = 3
        for ctrl_id in sorted(CIS_CTRL_DESC.keys(),
                               key=lambda x: [float(p) for p in x.split(".")]):
            desc  = CIS_CTRL_DESC[ctrl_id]
            items = sorted(ctrl_issues.get(ctrl_id, []),
                           key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count = len(items)

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"CIS {ctrl_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(CIS_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not items:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this control")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9"); nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in items:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 28, 10, 60, 60])

    # ── PCI DSS mapping ───────────────────────────────────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each PCI DSS v4.0 requirement lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        req_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        row = 3
        for req_id in sorted(PCI_DSS_DESC.keys(),
                              key=lambda x: [int(p) for p in x.split(".")]):
            desc  = PCI_DSS_DESC[req_id]
            items = sorted(req_issues.get(req_id, []),
                           key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count = len(items)

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not items:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this requirement")
                nc.font = _font(italic=True, color=C["info"]); nc.fill = _fill("F9F9F9")
                nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in items:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 28, 10, 60, 60])

    # ── Save ──────────────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_zones()
        self._sheet_users()
        self._sheet_snmp()
        self._sheet_issues()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Brocade FOS SAN Switch Config Analyzer — outputs Excel security report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python brocade_san_analyzer.py configexport.txt
  python brocade_san_analyzer.py configexport.txt -o san-audit.xlsx

Collect input on the switch (run as admin):
  switchshow > configexport.txt
  configshow >> configexport.txt
  userconfig --showlist >> configexport.txt
  passwdcfg --show >> configexport.txt
  snmpconfig --show snmpv1 >> configexport.txt
  snmpconfig --show snmpv3 >> configexport.txt
  syslogadmin --show >> configexport.txt
  tsclockserver >> configexport.txt
  ipfilter --show >> configexport.txt
  sshutil --show >> configexport.txt
  cfgshow >> configexport.txt
  defzone --show >> configexport.txt
  version >> configexport.txt
""",
    )
    ap.add_argument("config", help="Concatenated Brocade FOS command output file")
    ap.add_argument("-o", "--output", default=None,
                    help="Output Excel file (default: <config-stem>_brocade_analysis.xlsx)")
    args = ap.parse_args()

    if not args.output:
        stem = os.path.splitext(os.path.basename(args.config))[0]
        args.output = f"{stem}_brocade_analysis.xlsx"

    print(f"[*] Parsing:  {args.config}")
    parser = BrocadeFOSParser(args.config)
    parser.parse()

    sev_counts: dict[str, int] = defaultdict(int)
    for iss in parser.issues:
        sev_counts[iss["severity"]] += 1

    sw = parser.switch_info
    print(f"[*] Parsed:")
    print(f"      Switch name    : {sw.get('name', 'Unknown')}")
    print(f"      FOS version    : {parser.fos_version or 'Unknown'}")
    print(f"      Users          : {len(parser.users)}")
    print(f"      Zones          : {len(parser.zones)}")
    print(f"      Zone aliases   : {len(parser.zone_aliases)}")
    print(f"      Zone configs   : {len(parser.cfgs)}")
    print(f"      Active cfg     : {parser.active_cfg or '(none)'}")
    print(f"      Default zone   : {parser.default_zone}")
    print(f"      SNMP communities: {len(parser.snmp.get('communities', []))}")
    print(f"      NTP servers    : {len(parser.ntp_servers)}")
    print(f"      Syslog servers : {len(parser.syslog_servers)}")
    print(f"[*] Security findings: {len(parser.issues)}")
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        if sev_counts[sev]:
            print(f"      {sev:<10}: {sev_counts[sev]}")

    print(f"[*] Writing Excel report...")
    reporter = ExcelReporter(parser, args.output)
    reporter.save()
    print(f"[+] Saved: {args.output}")


if __name__ == "__main__":
    main()
