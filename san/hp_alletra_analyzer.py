#!/usr/bin/env python3
"""
HP Alletra Storage MP Security Analyzer
Parses HPE Alletra MP / Primera / 3PAR CLI command output and produces an
Excel security audit report with CIS Controls v8 and PCI DSS v4.0 mapping.

Based on HPE Alletra MP Security Guide (a00138815enw),
HPE Alletra MP Administrator Hardening Guide (a00146015enw),
and HPE DISA STIG guidance.

Targeted for VMware vSphere datastores (Fibre Channel, iSCSI, NFS).

Usage:
    python hp_alletra_analyzer.py alletra-export.txt
    python hp_alletra_analyzer.py alletra-export.txt -o audit.xlsx

Input: Concatenated output from the following CLI commands (pipe to a file):
    showsys
    showuser -d
    showpasswordpolicy
    showsnmp
    showsyslog
    showaudit
    showtime -zone
    shownet
    showhost
    showhost -chap
    showhostset
    showvv -col Name,VSize,UsrCPG,SnpCPG,Prov,State,Host
    showvlun
    showport
    showwsapi
    showsshkey
"""

import re
import os
import sys
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── CIS Controls v8 mapping ───────────────────────────────────────────────────
CIS_CTRL_DESC = {
    "3.10":  "Encrypt Sensitive Data in Transit",
    "3.11":  "Encrypt Sensitive Data at Rest",
    "4.2":   "Secure Configuration Process for Network Infrastructure",
    "4.7":   "Manage Default Accounts on Enterprise Assets and Software",
    "4.8":   "Disable Unnecessary Services on Enterprise Assets and Software",
    "5.2":   "Use Unique Passwords",
    "5.4":   "Restrict Administrator Privileges to Dedicated Accounts",
    "6.3":   "Require MFA for Externally-Exposed Applications",
    "6.5":   "Require MFA for Administrative Access",
    "6.7":   "Centralize Access Control",
    "6.8":   "Define and Maintain Role-Based Access Control",
    "8.2":   "Collect Audit Logs",
    "8.4":   "Standardize Time Synchronization",
    "8.9":   "Centralize Audit Logs",
    "12.2":  "Establish and Maintain a Secure Network Architecture",
    "12.3":  "Securely Manage Network Infrastructure",
    "12.6":  "Use Secure Network Management and Communication Protocols",
    "13.4":  "Perform Traffic Filtering Between Network Segments",
}

CIS_CONTROL_MAP: dict[str, list[str]] = {
    # ── Credentials / accounts ────────────────────────────────────────────────
    "Default Service Account Active":        ["4.7"],
    "Excessive Super-Role Accounts":         ["5.4", "6.8"],
    "Account Without Password":              ["5.2"],
    "No LDAP / AD Integration":              ["6.5", "6.7"],
    "No MFA for Administrative Access":      ["6.3", "6.5"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Weak Minimum Password Length":          ["5.2"],
    "No Password Complexity Requirements":   ["5.2"],
    "Password Never Expires":                ["5.2"],
    "No Password History":                   ["5.2"],
    "No Account Lockout Policy":             ["5.2", "6.5"],
    # ── Management access ─────────────────────────────────────────────────────
    "Telnet / HTTP Management Enabled":      ["4.2", "12.6"],
    "No Management IP Restriction":          ["12.3", "6.7"],
    "SSH Password Auth Enabled":             ["4.2", "12.6"],
    "No SSH Keys Configured":                ["4.2", "12.6"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1/v2 Enabled":                     ["4.2", "12.3", "12.6"],
    "Default SNMP Community String":         ["4.7", "12.3"],
    "SNMP Write Community Configured":       ["12.3"],
    "No SNMPv3 Configured":                  ["4.2", "12.3", "12.6"],
    # ── Logging / audit ───────────────────────────────────────────────────────
    "Audit Logging Not Enabled":             ["8.2"],
    "No Syslog Server Configured":           ["8.2", "8.9"],
    "NTP Not Configured":                    ["8.4"],
    "Single NTP Server":                     ["8.4"],
    # ── Storage access control ────────────────────────────────────────────────
    "iSCSI Without CHAP Authentication":     ["12.2", "13.4"],
    "Host Without Host Set Membership":      ["12.2", "13.4"],
    "Wildcard Volume Mapping":               ["12.2", "13.4"],
    "No VMware Persona Configured":          ["4.2", "12.2"],
    # ── Encryption ────────────────────────────────────────────────────────────
    "Encryption at Rest Not Enabled":        ["3.11"],
    "WSAPI Using HTTP (Not HTTPS)":          ["3.10", "12.6"],
    "Self-Signed Certificate in Use":        ["3.10", "12.6"],
    # ── Configuration hygiene ─────────────────────────────────────────────────
    "No Snapshot Schedule Configured":       ["4.2"],
    "No Management VLAN Separation":         ["12.2", "13.4"],
}


def _cis_label(ctrl_ids: list[str]) -> str:
    return " · ".join(f"CIS {c}" for c in ctrl_ids)


CIS_BENCHMARK_MAP: dict[str, list[str]] = {}


def _cis_benchmark_label(check_ids: list[str]) -> str:
    return " · ".join(check_ids)


# ── PCI DSS v4.0 mapping ─────────────────────────────────────────────────────
PCI_DSS_DESC = {
    "1.2.4":  "All traffic between trusted/untrusted networks is explicitly controlled",
    "1.2.7":  "Unused network access points are disabled",
    "2.2.1":  "Configuration standards are defined for all system components",
    "2.2.4":  "Only necessary services, protocols, and functions are enabled",
    "2.2.7":  "All non-console administrative access is encrypted",
    "3.5.1":  "Primary account number is secured with strong cryptography",
    "4.2.1":  "Strong cryptography is used to safeguard data during transmission",
    "7.2.1":  "All user access is appropriate and assigned by business need",
    "8.2.1":  "All user IDs and authentication credentials are managed securely",
    "8.3.4":  "Invalid authentication attempts are limited",
    "8.3.6":  "Passwords meet minimum length and complexity requirements",
    "8.3.9":  "Passwords are changed at least once every 90 days",
    "8.4.1":  "MFA is implemented for all non-console administrative access",
    "10.5.4": "Audit log files are protected (written to external log servers)",
    "10.6.1": "System clocks are synchronized using time-synchronization technology",
}

PCI_DSS_MAP: dict[str, list[str]] = {
    "Default Service Account Active":       ["8.2.1"],
    "Excessive Super-Role Accounts":        ["7.2.1"],
    "Account Without Password":             ["8.2.1"],
    "No LDAP / AD Integration":             ["8.4.1"],
    "No MFA for Administrative Access":     ["8.4.1"],
    "Weak Minimum Password Length":         ["8.3.6"],
    "No Password Complexity Requirements":  ["8.3.6"],
    "Password Never Expires":               ["8.3.9"],
    "No Account Lockout Policy":            ["8.3.4"],
    "Telnet / HTTP Management Enabled":     ["2.2.4", "2.2.7"],
    "No Management IP Restriction":         ["1.2.4"],
    "SSH Password Auth Enabled":            ["2.2.7"],
    "SNMPv1/v2 Enabled":                    ["2.2.4"],
    "Default SNMP Community String":        ["2.2.4"],
    "SNMP Write Community Configured":      ["2.2.4"],
    "No SNMPv3 Configured":                 ["2.2.4", "4.2.1"],
    "Audit Logging Not Enabled":            ["10.5.4"],
    "No Syslog Server Configured":          ["10.5.4"],
    "NTP Not Configured":                   ["10.6.1"],
    "iSCSI Without CHAP Authentication":    ["1.2.4"],
    "Host Without Host Set Membership":     ["1.2.4"],
    "Wildcard Volume Mapping":              ["1.2.4"],
    "Encryption at Rest Not Enabled":       ["3.5.1"],
    "WSAPI Using HTTP (Not HTTPS)":         ["4.2.1"],
    "Self-Signed Certificate in Use":       ["4.2.1"],
}


def _pci_label(req_ids: list[str]) -> str:
    return " · ".join(f"PCI {r}" for r in req_ids)


# ── Secure Controls Framework (SCF) mapping ──────────────────────────────────
SCF_MAP: dict[str, list[str]] = {
    # ── Credentials / accounts ────────────────────────────────────────────────
    "Default Service Account Active":      ["IAC-21"],
    "Excessive Super-Role Accounts":       ["IAC-07"],
    "Account Without Password":            ["IAC-06"],
    "No LDAP / AD Integration":            ["IAC-01"],
    "No MFA for Administrative Access":    ["IAC-01"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Weak Minimum Password Length":        ["IAC-06"],
    "No Password Complexity Requirements": ["IAC-06"],
    "Password Never Expires":              ["IAC-06"],
    "No Password History":                 ["IAC-06"],
    "No Account Lockout Policy":           ["IAC-06"],
    # ── Management access ─────────────────────────────────────────────────────
    "Telnet / HTTP Management Enabled":    ["CRY-03", "NET-06"],
    "No Management IP Restriction":        ["NET-04", "IAC-10"],
    "SSH Password Auth Enabled":           ["CRY-03"],
    "No SSH Keys Configured":              ["CRY-03"],
    # ── SNMP ──────────────────────────────────────────────────────────────────
    "SNMPv1/v2 Enabled":                   ["NET-06", "CRY-03"],
    "Default SNMP Community String":       ["IAC-06"],
    "SNMP Write Community Configured":     ["IAC-07"],
    "No SNMPv3 Configured":                ["CRY-03", "NET-06"],
    # ── Logging / audit ───────────────────────────────────────────────────────
    "Audit Logging Not Enabled":           ["MON-06"],
    "No Syslog Server Configured":         ["MON-06"],
    "NTP Not Configured":                  ["OPS-01"],
    "Single NTP Server":                   ["OPS-01"],
    # ── Storage access control ────────────────────────────────────────────────
    "iSCSI Without CHAP Authentication":   ["IAC-01", "NET-04"],
    "Host Without Host Set Membership":    ["NET-04"],
    "Wildcard Volume Mapping":             ["NET-04"],
    "No VMware Persona Configured":        ["NET-04"],
    # ── Encryption ────────────────────────────────────────────────────────────
    "Encryption at Rest Not Enabled":      ["CRY-03"],
    "WSAPI Using HTTP (Not HTTPS)":        ["CRY-03", "NET-06"],
    "Self-Signed Certificate in Use":      ["CRY-03"],
    # ── Configuration hygiene ─────────────────────────────────────────────────
    "No Snapshot Schedule Configured":     ["OPS-01"],
    "No Management VLAN Separation":       ["NET-04"],
}


def _scf_label(ctrl_ids: list[str]) -> str:
    return " · ".join(ctrl_ids)


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "hdr_bg":   "0E5484", "hdr_fg":   "FFFFFF",   # HPE blue
    "critical": "C00000", "critical_l": "FFB3B3",
    "high":     "FF0000", "high_l":     "FFD9B3",
    "medium":   "FF8C00", "medium_l":   "FFF2CC",
    "low":      "0070C0", "low_l":      "BDD7EE",
    "info":     "595959", "info_l":     "F2F2F2",
    "ok":       "375623", "ok_l":       "E2EFDA",
    "disabled": "A6A6A6",
    "alt_row":  "F5F5F5",
}

_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=size)

def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── HP Alletra MP / Primera / 3PAR CLI Parser ────────────────────────────────
class HPAlletraParser:
    """
    Parses concatenated HPE Alletra MP CLI command output collected for a
    security audit. Supports Alletra MP B10000, Primera, and 3PAR OS.
    """

    _DEFAULT_COMMUNITIES = {"public", "private", "common", "alletra", "hpe",
                            "3parsnmp", "inserv"}
    _DEFAULT_ACCOUNTS    = {"3parsvc", "browse", "guest", "support", "factory",
                            "installer"}
    _SUPER_ROLES         = {"super", "superuser", "administrator", "admin"}
    _VMWARE_PERSONAS     = {"vmware", "vmware_4", "vmware_5", "vmware_6",
                            "vmware_7", "vmware_8", "4", "5", "6", "7", "8"}

    def __init__(self, config_file: str):
        self.config_file = config_file
        self.lines: list[str] = []

        # Parsed data
        self.sys_info: dict           = {}
        self.users: list[dict]        = []
        self.password_policy: dict    = {}
        self.snmp: dict               = {}
        self.syslog_servers: list[dict] = []
        self.audit_enabled: bool      = False
        self.audit_line: int          = 0
        self.ntp_servers: list[str]   = []
        self.ntp_line: int            = 0
        self.ldap_configured: bool    = False
        self.network: dict            = {}
        self.hosts: list[dict]        = []
        self.host_sets: list[dict]    = []
        self.vluns: list[dict]        = []
        self.volumes: list[dict]      = []
        self.ports: list[dict]        = []
        self.wsapi: dict              = {}
        self.ssh_keys: list[str]      = []
        self.encryption_enabled: bool = False
        self.issues: list[dict]       = []

    # ── Issue helper ──────────────────────────────────────────────────────────
    def _issue(self, severity: str, category: str, obj: str,
               description: str, recommendation: str, line: int = 0):
        cis_ids   = CIS_CONTROL_MAP.get(category, [])
        pci_ids   = PCI_DSS_MAP.get(category, [])
        scf_ids   = SCF_MAP.get(category, [])
        bench_ids = CIS_BENCHMARK_MAP.get(category, [])
        self.issues.append({
            "severity":        severity,
            "category":        category,
            "object":          obj,
            "description":     description,
            "recommendation":  recommendation,
            "line":            line if line else "",
            "cis_ids":         cis_ids,
            "pci_ids":         pci_ids,
            "cis_label":       _cis_label(cis_ids),
            "cis_benchmark":   _cis_benchmark_label(bench_ids),
            "pci_label":       _pci_label(pci_ids),
            "scf":             _scf_label(scf_ids),
        })

    # ── Entry point ───────────────────────────────────────────────────────────
    def parse(self):
        try:
            with open(self.config_file, encoding="utf-8", errors="replace") as fh:
                self.lines = fh.readlines()
        except FileNotFoundError:
            sys.exit(f"File not found: {self.config_file}")

        self._parse_sys_info()
        self._parse_users()
        self._parse_password_policy()
        self._parse_snmp()
        self._parse_syslog()
        self._parse_audit()
        self._parse_ntp()
        self._parse_ldap()
        self._parse_network()
        self._parse_hosts()
        self._parse_host_sets()
        self._parse_vluns()
        self._parse_volumes()
        self._parse_ports()
        self._parse_wsapi()
        self._parse_ssh_keys()
        self._parse_encryption()
        self._run_checks()

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _col_table(self, start_lineno: int, start_pat: str,
                   end_pat: str | None = None) -> list[dict]:
        """
        Parse an HPE 3PAR/Primera/Alletra column-aligned table.
        Returns a list of dicts keyed by the header column names.
        """
        in_table = False
        headers: list[str] = []
        col_starts: list[int] = []
        results: list[dict] = []

        for lineno, raw in enumerate(self.lines, 1):
            if lineno < start_lineno:
                continue
            line = raw.rstrip("\n")

            if not in_table:
                if re.search(start_pat, line, re.I):
                    in_table = True
                continue

            if end_pat and re.search(end_pat, line, re.I):
                break

            stripped = line.strip()
            if not stripped:
                if results:   # blank line after table data = end
                    break
                continue

            if not headers:
                # First non-blank line is the header
                if re.match(r"^-+", stripped):
                    continue   # skip separator
                headers = re.findall(r"\S+", line)
                col_starts = [line.index(h) for h in headers]
                continue

            if re.match(r"^[-=]+", stripped):
                continue   # skip separator line

            row: dict = {}
            for idx, (key, start) in enumerate(zip(headers, col_starts)):
                end = col_starts[idx + 1] if idx + 1 < len(col_starts) else None
                val = (line[start:end] if end else line[start:]).strip()
                row[key] = val
            row["_line"] = lineno
            results.append(row)

        return results

    def _find_value(self, patterns: list[str]) -> tuple[str, int]:
        """Return (value, lineno) for the first matching pattern in the file."""
        for lineno, raw in enumerate(self.lines, 1):
            for pat in patterns:
                m = re.search(pat, raw, re.I)
                if m:
                    return m.group(1).strip(), lineno
        return "", 0

    # ── Section parsers ───────────────────────────────────────────────────────
    def _parse_sys_info(self):
        """Parse showsys output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            for key, pats in [
                ("name",    [r"System [Nn]ame\s*:\s*(.+)", r"SystemName\s*:\s*(.+)"]),
                ("model",   [r"System [Mm]odel\s*:\s*(.+)", r"(?:Alletra|Primera|3PAR)\s+(\S+)"]),
                ("serial",  [r"Serial [Nn]umber\s*:\s*(\S+)"]),
                ("fw",      [r"[Ss]ystem [Ff]irmware\s*:\s*(\S+)",
                              r"InformOS[/ ][Vv]ersion\s*:\s*(\S+)",
                              r"(?:OS|Firmware)\s+[Vv]ersion\s*:\s*(\S+)"]),
                ("nodes",   [r"Number of [Nn]odes\s*:\s*(\d+)"]),
            ]:
                if key in self.sys_info:
                    continue
                for pat in pats:
                    m = re.search(pat, line, re.I)
                    if m:
                        self.sys_info[key] = m.group(1).strip()
                        self.sys_info[f"{key}_line"] = lineno
                        break

    def _parse_users(self):
        """
        Parse showuser -d output.
        Format (column-aligned):
            Name        Role    Enabled  PwdExp   Last_Login
            ----------  ------  -------  -------  ----------
            3parsvc     super   yes      never    -
            admin       super   yes      never    2024-01-15
        """
        in_section = False
        headers: list[str] = []
        col_starts: list[int] = []
        past_header = False

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()

            if not in_section:
                if re.search(r"showuser|^Name\s+Role\s+Enabled", line, re.I):
                    in_section = True
                continue

            if not stripped:
                if past_header:
                    break
                continue

            if not headers:
                # Detect header row
                if re.match(r"(?:Name|User)\s+(?:Role|Permission)", stripped, re.I):
                    headers = re.findall(r"\S+", line)
                    col_starts = [line.index(h) for h in headers]
                continue

            if re.match(r"^-{3,}", stripped):
                past_header = True
                continue

            if not past_header:
                continue

            # Parse data row
            if col_starts:
                row: dict = {}
                for idx, (key, start) in enumerate(zip(headers, col_starts)):
                    end = col_starts[idx + 1] if idx + 1 < len(col_starts) else None
                    val = (line[start:end] if end else line[start:]).strip()
                    row[key] = val
                row["_line"] = lineno

                name = row.get("Name", row.get("name", ""))
                role = row.get("Role", row.get("role", row.get("Permission", "")))
                enabled = row.get("Enabled", row.get("enabled", "yes")).lower()
                pwd_exp = row.get("PwdExp", row.get("Expires", "never"))

                if name:
                    self.users.append({
                        "name":    name,
                        "role":    role,
                        "enabled": enabled in ("yes", "true", "1"),
                        "pwd_exp": pwd_exp,
                        "line":    lineno,
                    })
            else:
                # Fallback: parse key: value format
                m = re.match(r"(\S+)\s+(\S+)\s+(yes|no)\s+(\S+)", stripped, re.I)
                if m:
                    self.users.append({
                        "name":    m.group(1),
                        "role":    m.group(2),
                        "enabled": m.group(3).lower() == "yes",
                        "pwd_exp": m.group(4),
                        "line":    lineno,
                    })

    def _parse_password_policy(self):
        """Parse showpasswordpolicy output."""
        pp: dict = {
            "min_length":        None,
            "max_age":           None,
            "min_upper":         0,
            "min_lower":         0,
            "min_numeric":       0,
            "min_special":       0,
            "history":           0,
            "lockout_threshold": 0,
            "lockout_duration":  0,
        }

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"Minimum password length\s*:?\s*(\d+)", line, re.I)
            if m:
                pp["min_length"] = int(m.group(1))
                pp["min_length_line"] = lineno
            m = re.match(r"Maximum password age.*?:\s*(\d+)", line, re.I)
            if m:
                pp["max_age"] = int(m.group(1))
                pp["max_age_line"] = lineno
            m = re.match(r"Require uppercase\s*:?\s*(yes|no|1|0)", line, re.I)
            if m:
                pp["min_upper"] = 1 if m.group(1).lower() in ("yes", "1") else 0
            m = re.match(r"Require lowercase\s*:?\s*(yes|no|1|0)", line, re.I)
            if m:
                pp["min_lower"] = 1 if m.group(1).lower() in ("yes", "1") else 0
            m = re.match(r"Require numeric\s*:?\s*(yes|no|1|0)", line, re.I)
            if m:
                pp["min_numeric"] = 1 if m.group(1).lower() in ("yes", "1") else 0
            m = re.match(r"Require special\s*:?\s*(yes|no|1|0)", line, re.I)
            if m:
                pp["min_special"] = 1 if m.group(1).lower() in ("yes", "1") else 0
            m = re.match(r"Password history\s*:?\s*(\d+)", line, re.I)
            if m:
                pp["history"] = int(m.group(1))
            m = re.match(r"(?:Failed login|Lockout) threshold\s*:?\s*(\d+)", line, re.I)
            if m:
                pp["lockout_threshold"] = int(m.group(1))
            m = re.match(r"(?:Account lockout|Lockout) duration\s*:?\s*(\d+)", line, re.I)
            if m:
                pp["lockout_duration"] = int(m.group(1))

        self.password_policy = pp

    def _parse_snmp(self):
        """Parse showsnmp output."""
        snmp: dict = {
            "v1_v2_enabled": None,
            "v3_enabled":    False,
            "communities":   [],
            "trap_targets":  [],
            "v3_users":      [],
            "contact":       "",
            "location":      "",
        }

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"Read community string\s*:\s*(.+)", line, re.I)
            if m:
                comm = m.group(1).strip()
                snmp["communities"].append({"name": comm, "access": "read", "line": lineno})
                snmp["v1_v2_enabled"] = True
            m = re.match(r"Write community string\s*:\s*(.+)", line, re.I)
            if m:
                comm = m.group(1).strip()
                if comm.lower() not in ("none", "", "-"):
                    snmp["communities"].append({"name": comm, "access": "write", "line": lineno})
            m = re.match(r"MIB-II sysContact\s*:\s*(.+)", line, re.I)
            if m:
                snmp["contact"] = m.group(1).strip()
            m = re.match(r"MIB-II sysLocation\s*:\s*(.+)", line, re.I)
            if m:
                snmp["location"] = m.group(1).strip()
            m = re.match(r"MIB-II sysName\s*:\s*(.+)", line, re.I)
            if m and not self.sys_info.get("name"):
                self.sys_info["name"] = m.group(1).strip()
            # SNMPv3
            if re.search(r"SNMPv3\s*:\s*(yes|enabled|true)", line, re.I):
                snmp["v3_enabled"] = True
            # Trap recipients
            m = re.match(r"(?:SNMP\s+[Tt]rap|[Tt]rap)\s+[Rr]ecipient\s*(?:\d+)?\s*:\s*(.+)", line, re.I)
            if m:
                addr = m.group(1).strip()
                if addr not in ("0.0.0.0", "none", ""):
                    snmp["trap_targets"].append({"addr": addr, "line": lineno})

        self.snmp = snmp

    def _parse_syslog(self):
        """Parse showsyslog output (column-aligned table)."""
        in_section = False
        past_header = False
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if not in_section:
                if re.search(r"showsyslog|^IP_Address\s+Port|^Syslog", line, re.I):
                    in_section = True
                continue
            if re.match(r"^-{3,}", line):
                past_header = True
                continue
            if not line and past_header:
                break
            if past_header and line:
                parts = line.split()
                if parts and re.match(r"\d+\.\d+\.\d+\.\d+", parts[0]):
                    self.syslog_servers.append({"ip": parts[0], "port": parts[1] if len(parts) > 1 else "514", "line": lineno})

    def _parse_audit(self):
        """Parse showaudit output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if re.search(r"[Aa]udit [Ll]ogging\s*:\s*(enabled|yes|on)", line, re.I):
                self.audit_enabled = True
                self.audit_line = lineno
            if re.search(r"[Aa]udit [Ll]ogging\s*:\s*(disabled|no|off)", line, re.I):
                self.audit_enabled = False
                self.audit_line = lineno

    def _parse_ntp(self):
        """Parse showtime -zone output."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"NTP\s+[Ss]erver\s*:\s*(.+)", line, re.I)
            if m:
                srv = m.group(1).strip()
                if srv not in ("none", "0.0.0.0", "", "-"):
                    self.ntp_servers.append(srv)
                    if not self.ntp_line:
                        self.ntp_line = lineno
            m = re.match(r"Time\s+[Ss]ource\s*:\s*(.+)", line, re.I)
            if m and "ntp" in m.group(1).lower():
                pass   # just confirms NTP is in use

    def _parse_ldap(self):
        """Detect LDAP/AD configuration."""
        for line in self.lines:
            if re.search(r"ldap|active.?directory|LDAP\s+server", line, re.I):
                if not re.search(r"not\s+configured|disabled|none", line, re.I):
                    self.ldap_configured = True
                    break

    def _parse_network(self):
        """Parse shownet output."""
        net: dict = {"interfaces": [], "vlan_separated": False}
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            m = re.match(r"(\S+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)", line)
            if m:
                net["interfaces"].append({
                    "name": m.group(1), "ip": m.group(2),
                    "mask": m.group(3), "gw": m.group(4),
                    "line": lineno,
                })
        self.network = net

    def _parse_hosts(self):
        """
        Parse showhost and showhost -chap output.
        Format (column-aligned):
            Id Name       Persona  -WWN/iSCSI_Name-          Port  IP_addr
            -- ------     -------  -------------------------  ----  -------
             1 esxi01     VMware6  10:00:08:f1:ea:b9:2e:c4   1:1   -
        """
        in_section = False
        past_header = False
        headers: list[str] = []
        col_starts: list[int] = []
        chap_data: dict[str, dict] = {}

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()

            if not in_section:
                if re.search(r"showhost|^Id\s+Name\s+Persona", line, re.I):
                    in_section = True
                    past_header = False
                    headers = []
                continue

            if not stripped:
                if past_header:
                    in_section = False
                continue

            if not headers:
                if re.match(r"Id\s+Name", stripped, re.I):
                    headers = re.findall(r"\S+", line)
                    col_starts = [line.index(h) for h in headers]
                continue

            if re.match(r"^-{3,}", stripped):
                past_header = True
                continue

            if not past_header:
                continue

            if col_starts:
                row: dict = {}
                for idx, (key, start) in enumerate(zip(headers, col_starts)):
                    end = col_starts[idx + 1] if idx + 1 < len(col_starts) else None
                    val = (line[start:end] if end else line[start:]).strip()
                    row[key] = val

                name = row.get("Name", "")
                persona = row.get("Persona", row.get("persona", ""))
                proto = "FC"
                wwn_iqn = ""
                for k in row:
                    if "WWN" in k.upper() or "iSCSI" in k.upper() or "ISCSI" in k.upper():
                        wwn_iqn = row[k]
                        if "iqn" in wwn_iqn.lower() or ":" not in wwn_iqn:
                            proto = "iSCSI"
                        break

                if name:
                    self.hosts.append({
                        "name":    name,
                        "persona": persona,
                        "proto":   proto,
                        "wwn_iqn": wwn_iqn,
                        "chap":    False,   # updated by CHAP parse
                        "line":    lineno,
                    })

            # CHAP: showhost -chap
            mc = re.match(r"(\S+)\s+(yes|no)\s+", stripped, re.I)
            if mc:
                chap_data[mc.group(1)] = {"chap": mc.group(2).lower() == "yes"}

        # Merge CHAP data
        for h in self.hosts:
            if h["name"] in chap_data:
                h["chap"] = chap_data[h["name"]]["chap"]

    def _parse_host_sets(self):
        """Parse showhostset output."""
        in_section = False
        past_header = False
        cur_set: dict | None = None

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if not in_section:
                if re.search(r"showhostset|^Id\s+Name\s+Members|^Host\s+Set", line, re.I):
                    in_section = True
                continue
            if not line:
                if past_header and cur_set:
                    self.host_sets.append(cur_set)
                    cur_set = None
                continue
            if re.match(r"^-{3,}", line):
                past_header = True
                continue
            if past_header:
                parts = line.split()
                if parts and parts[0].isdigit():
                    if cur_set:
                        self.host_sets.append(cur_set)
                    cur_set = {
                        "id": parts[0], "name": parts[1] if len(parts) > 1 else "",
                        "members": parts[2:] if len(parts) > 2 else [],
                        "line": lineno
                    }
                elif cur_set and parts:
                    cur_set["members"].extend(parts)
        if cur_set:
            self.host_sets.append(cur_set)

    def _parse_vluns(self):
        """
        Parse showvlun output.
        Format:
            Lun  VVname     Hostname    Port  Type
            ---  -------    --------    ----  ----
              0  ds01-rw    esxi01      -     host
              1  ds02-rw    *           -     matched
        """
        in_section = False
        past_header = False
        headers: list[str] = []
        col_starts: list[int] = []

        for lineno, raw in enumerate(self.lines, 1):
            line = raw.rstrip("\n")
            stripped = line.strip()
            if not in_section:
                if re.search(r"showvlun|^Lun\s+VVname|^LUN\s+", line, re.I):
                    in_section = True
                continue
            if not stripped:
                if past_header:
                    in_section = False
                continue
            if not headers:
                if re.match(r"Lun\s+VVname|LUN\s+Volume", stripped, re.I):
                    headers = re.findall(r"\S+", line)
                    col_starts = [line.index(h) for h in headers]
                continue
            if re.match(r"^-{3,}", stripped):
                past_header = True
                continue
            if not past_header:
                continue
            if col_starts:
                row: dict = {}
                for idx, (key, start) in enumerate(zip(headers, col_starts)):
                    end = col_starts[idx + 1] if idx + 1 < len(col_starts) else None
                    val = (line[start:end] if end else line[start:]).strip()
                    row[key] = val
                row["_line"] = lineno
                self.vluns.append(row)

    def _parse_volumes(self):
        """Parse showvv output for encryption and provisioning info."""
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if re.search(r"[Ee]ncryption.*?:\s*(enabled|yes|on)", line, re.I):
                self.encryption_enabled = True

    def _parse_ports(self):
        """Parse showport output for protocol and state info."""
        in_section = False
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if not in_section:
                if re.search(r"showport|^N:S:P\s+Mode", line, re.I):
                    in_section = True
                continue
            if not line:
                in_section = False
                continue
            m = re.match(r"(\d+:\d+:\d+)\s+(\S+)\s+(\S+)\s+(\S+)", line)
            if m:
                self.ports.append({
                    "port":  m.group(1),
                    "mode":  m.group(2),
                    "state": m.group(3),
                    "proto": m.group(4),
                    "line":  lineno,
                })

    def _parse_wsapi(self):
        """Parse showwsapi output."""
        wsapi: dict = {"enabled": False, "http": False, "https": False,
                       "http_port": 8008, "https_port": 8080, "version": ""}
        for lineno, raw in enumerate(self.lines, 1):
            line = raw.strip()
            if re.search(r"[Ss]tate\s*:\s*(Active|enabled|running)", line, re.I):
                wsapi["enabled"] = True
            if re.search(r"HTTP\s*:\s*(enabled|yes)", line, re.I):
                wsapi["http"] = True
            if re.search(r"HTTPS\s*:\s*(enabled|yes)", line, re.I):
                wsapi["https"] = True
            m = re.match(r"WSAPI\s+[Vv]ersion\s*:\s*(\S+)", line, re.I)
            if m:
                wsapi["version"] = m.group(1)
        self.wsapi = wsapi

    def _parse_ssh_keys(self):
        """Parse showsshkey output."""
        for line in self.lines:
            if re.search(r"ssh-rsa|ssh-ed25519|ecdsa-sha2|ssh-dss", line, re.I):
                self.ssh_keys.append(line.strip())

    def _parse_encryption(self):
        """Detect encryption at rest from showsys or volume info."""
        for line in self.lines:
            if re.search(r"[Ee]ncryption\s*(at\s*[Rr]est)?\s*:\s*(enabled|yes|on)", line, re.I):
                self.encryption_enabled = True
            if re.search(r"SED\s*:\s*(enabled|yes|active)", line, re.I):
                self.encryption_enabled = True

    # ── Security checks ───────────────────────────────────────────────────────
    def _run_checks(self):
        self._chk_users()
        self._chk_password_policy()
        self._chk_ldap_mfa()
        self._chk_snmp()
        self._chk_syslog()
        self._chk_audit()
        self._chk_ntp()
        self._chk_hosts()
        self._chk_vluns()
        self._chk_wsapi()
        self._chk_encryption()

    def _chk_users(self):
        super_count = 0
        hosts_in_sets = {h for hs in self.host_sets for h in hs.get("members", [])}

        for u in self.users:
            role_lower = u["role"].lower()

            # Default service accounts
            if u["name"].lower() in self._DEFAULT_ACCOUNTS and u["enabled"]:
                self._issue(
                    "HIGH", "Default Service Account Active",
                    f"account: {u['name']}",
                    f"Default HPE service/support account '{u['name']}' is enabled. "
                    "These accounts are known to attackers and should be disabled "
                    "unless actively needed for an HPE support session.",
                    f"Disable with: 'removeuser {u['name']}' or "
                    f"'setuser -e false {u['name']}'. "
                    "Re-enable only when HPE support is actively engaged.",
                    line=u["line"])

            if role_lower in self._SUPER_ROLES:
                super_count += 1

            # Passwords that never expire for privileged accounts
            if role_lower in self._SUPER_ROLES and u["pwd_exp"].lower() == "never":
                self._issue(
                    "MEDIUM", "Password Never Expires",
                    f"account: {u['name']} (super role)",
                    f"Super-role account '{u['name']}' has a password set to never expire. "
                    "Stale admin credentials are a top ransomware vector for storage platforms.",
                    "Set password expiry: 'setpasswd -maxage 90'. "
                    "Enforce a 90-day rotation policy for all privileged accounts.",
                    line=u["line"])

        if super_count > 2:
            self._issue(
                "MEDIUM", "Excessive Super-Role Accounts",
                f"{super_count} super-role accounts",
                f"{super_count} accounts have the super role. Excessive privileged "
                "accounts expand the blast radius of a credential compromise and "
                "complicate audit trails.",
                "Limit super-role accounts to 2 named individuals at most. "
                "Create 'audit' role accounts for read-only security monitoring "
                "and 'operator' role for day-to-day tasks.")

    def _chk_password_policy(self):
        pp = self.password_policy
        if not pp:
            return

        min_len = pp.get("min_length")
        if min_len is not None:
            if min_len < 8:
                self._issue(
                    "HIGH", "Weak Minimum Password Length",
                    f"min length: {min_len}",
                    f"Minimum password length is only {min_len} characters. "
                    "Short passwords are easily brute-forced. Recommended: ≥14 characters.",
                    "Increase minimum password length via the Alletra management console "
                    "or CLI: 'setpasswd -minlength 14'.",
                    line=pp.get("min_length_line", 0))
            elif min_len < 12:
                self._issue(
                    "MEDIUM", "Weak Minimum Password Length",
                    f"min length: {min_len}",
                    f"Minimum password length is {min_len} (recommended: ≥14).",
                    "Increase minimum password length: 'setpasswd -minlength 14'.")

        has_complexity = (pp.get("min_upper", 0) >= 1 or
                          pp.get("min_lower", 0) >= 1 or
                          pp.get("min_numeric", 0) >= 1 or
                          pp.get("min_special", 0) >= 1)
        if not has_complexity:
            self._issue(
                "MEDIUM", "No Password Complexity Requirements",
                "showpasswordpolicy",
                "No uppercase, lowercase, numeric, or special character requirements "
                "are enforced for passwords. Simple passwords are easily guessed.",
                "Enable complexity requirements via the management console or CLI.")

        max_age = pp.get("max_age")
        if max_age is not None and max_age == 0:
            self._issue(
                "MEDIUM", "Password Never Expires",
                "max age: 0 (never)",
                "Maximum password age is 0 — passwords never expire. "
                "Long-lived credentials increase exposure after a breach.",
                "Set a maximum password age of 90 days: 'setpasswd -maxage 90'.",
                line=pp.get("max_age_line", 0))

        if pp.get("history", 0) < 5:
            self._issue(
                "LOW", "No Password History",
                f"history: {pp.get('history', 0)}",
                f"Password history is {pp.get('history', 0)} (recommended: ≥5). "
                "Users can immediately reuse old passwords.",
                "Increase password history: 'setpasswd -historysize 5'.")

        if pp.get("lockout_threshold", 0) == 0:
            self._issue(
                "HIGH", "No Account Lockout Policy",
                "lockout threshold: 0",
                "Account lockout is disabled. Brute-force or credential-stuffing attacks "
                "can proceed indefinitely. This is especially dangerous for internet-facing "
                "WSAPI or GreenLake management endpoints.",
                "Enable lockout: set a threshold of 5 attempts. "
                "Configure via the Alletra management console Security settings.")

    def _chk_ldap_mfa(self):
        if not self.ldap_configured:
            self._issue(
                "HIGH", "No LDAP / AD Integration",
                "LDAP/AD: not configured",
                "No LDAP or Active Directory integration is configured. "
                "All authentication uses local accounts only, providing no centralized "
                "control, no MFA enforcement, and no de-provisioning on employee departure.",
                "Integrate with Active Directory or LDAP. "
                "For GreenLake-connected deployments, enable SSO/SAML and MFA. "
                "Reference: HPE Alletra MP Administrator Hardening Guide sec. 4.")

        if not self.ldap_configured:
            self._issue(
                "HIGH", "No MFA for Administrative Access",
                "MFA: not verified",
                "Multi-factor authentication for administrative access cannot be confirmed. "
                "Without LDAP/SSO integration, MFA cannot be enforced at the storage layer. "
                "Storage systems are Tier-0 infrastructure — a compromised admin credential "
                "can result in total data loss or ransomware encryption of all datastores.",
                "Enable MFA via GreenLake IAM or through an LDAP provider that enforces MFA. "
                "Restrict management access to a jump host / bastion that requires MFA.")

    def _chk_snmp(self):
        snmp = self.snmp

        if snmp.get("v1_v2_enabled"):
            self._issue(
                "HIGH", "SNMPv1/v2 Enabled",
                "SNMP: v1/v2 community strings present",
                "SNMPv1/v2c uses cleartext community strings with no authentication "
                "or encryption. An attacker on the management VLAN can read the full "
                "storage MIB (volumes, hosts, network config) without credentials.",
                "Disable SNMPv1/v2 community strings via the management console. "
                "Migrate monitoring to SNMPv3 with authPriv mode.")

        for comm in snmp.get("communities", []):
            if comm["name"].lower() in self._DEFAULT_COMMUNITIES:
                self._issue(
                    "CRITICAL", "Default SNMP Community String",
                    f"community: {comm['name']} ({comm['access']})",
                    f"Default SNMP community string '{comm['name']}' is configured. "
                    "This string is universally known. Any host on the management network "
                    "can read (or write) storage MIB data without authentication.",
                    "Change all SNMP community strings to unique, complex values "
                    "immediately. Rotate quarterly. Consider migrating to SNMPv3.",
                    line=comm["line"])

        write_comms = [c for c in snmp.get("communities", []) if c["access"] == "write"]
        for comm in write_comms:
            self._issue(
                "HIGH", "SNMP Write Community Configured",
                f"community: {comm['name']} (write)",
                f"SNMP write community '{comm['name']}' is configured. "
                "Write access allows an attacker to modify storage configuration via SNMP.",
                "Remove SNMP write community strings. Use SNMPv3 read-only for monitoring.",
                line=comm["line"])

        if not snmp.get("v3_enabled") and not snmp.get("v3_users"):
            self._issue(
                "MEDIUM", "No SNMPv3 Configured",
                "SNMPv3: not configured",
                "No SNMPv3 is configured. SNMPv3 authPriv provides both authentication "
                "and encryption for SNMP management traffic.",
                "Configure SNMPv3 with authPriv via the management console. "
                "Use SHA-256 for authentication and AES-128+ for privacy.")

    def _chk_syslog(self):
        if not self.syslog_servers:
            self._issue(
                "HIGH", "No Syslog Server Configured",
                "syslog: none",
                "No remote syslog server is configured. Authentication events, "
                "volume mapping changes, and admin activity are only retained locally "
                "and may be lost or tampered with during a ransomware event.",
                "Configure a remote syslog target via the management console. "
                "Forward to a SIEM for correlation. Logs should be treated as immutable.")

    def _chk_audit(self):
        if not self.audit_enabled:
            self._issue(
                "HIGH", "Audit Logging Not Enabled",
                "showaudit: disabled",
                "Audit logging is not enabled. Without audit logs, there is no record "
                "of who changed volume mappings, user permissions, or snapshot schedules. "
                "This is a critical gap for detecting insider threats or post-breach activity.",
                "Enable audit logging via the management console or CLI. "
                "Forward audit logs to a SIEM immediately. "
                "Per HPE guidance (Hardening Guide sec. 8): enable centralized logging.",
                line=self.audit_line if self.audit_line else 0)

    def _chk_ntp(self):
        if not self.ntp_servers:
            self._issue(
                "HIGH", "NTP Not Configured",
                "NTP: not configured",
                "No NTP server is configured. Timestamp drift invalidates log correlation "
                "with VMware vSphere events and makes forensic investigation unreliable.",
                "Configure NTP via the management console. Use the same NTP source as "
                "your vSphere hosts for consistent log timestamps across the environment.")

        elif len(self.ntp_servers) < 2:
            self._issue(
                "LOW", "Single NTP Server",
                f"NTP servers: {len(self.ntp_servers)}",
                "Only one NTP server is configured. A single server is a single point "
                "of failure for time synchronization.",
                "Add a secondary NTP server for redundancy.")

    def _chk_hosts(self):
        """Check host definitions for security issues."""
        hosts_in_sets: set[str] = set()
        for hs in self.host_sets:
            hosts_in_sets.update(hs.get("members", []))

        for h in self.hosts:
            # iSCSI hosts without CHAP
            if h["proto"] == "iSCSI" and not h["chap"]:
                self._issue(
                    "HIGH", "iSCSI Without CHAP Authentication",
                    f"host: {h['name']}",
                    f"iSCSI host '{h['name']}' does not have CHAP authentication configured. "
                    "Without CHAP, any iSCSI initiator that knows the target IQN can "
                    "connect to storage, bypassing host-based access control in VMware.",
                    "Enable mutual CHAP for all iSCSI hosts: "
                    "configure CHAP credentials on both the Alletra and the ESXi host's "
                    "software iSCSI initiator. Reference: HPE Hardening Guide sec. 6.",
                    line=h["line"])

            # Hosts without VMware persona
            if h["proto"] == "FC" and h["persona"].lower() not in self._VMWARE_PERSONAS:
                if any(x in h["name"].lower() for x in ("esx", "vmware", "vsphere", "vcsa")):
                    self._issue(
                        "MEDIUM", "No VMware Persona Configured",
                        f"host: {h['name']} (persona: {h['persona'] or 'none'})",
                        f"Host '{h['name']}' appears to be a VMware ESXi host but does not "
                        "have a VMware persona configured. Without the correct persona, "
                        "SCSI reservation handling and multi-pathing may behave incorrectly, "
                        "causing data corruption or host connectivity issues.",
                        "Set the VMware persona: update host definition to use "
                        "'VMware' persona. Reference: HPE Alletra MP Host Configuration Guide.",
                        line=h["line"])

            # Hosts not in a host set
            if h["name"] not in hosts_in_sets:
                self._issue(
                    "LOW", "Host Without Host Set Membership",
                    f"host: {h['name']}",
                    f"Host '{h['name']}' is not a member of any host set. "
                    "Host sets provide a structured way to manage volume mappings "
                    "and apply access policies consistently across hosts.",
                    "Create or assign a host set: 'createhostset <setname>' and "
                    "add the host: 'addhosttoset <setname> <hostname>'. "
                    "Use host sets for all VMware cluster nodes.",
                    line=h["line"])

    def _chk_vluns(self):
        """Check volume-to-host mappings for wildcard (any-host) exports."""
        for vlun in self.vluns:
            hostname = vlun.get("Hostname", vlun.get("hostname", "")).strip()
            vvname   = vlun.get("VVname", vlun.get("vvname", "")).strip()
            if hostname == "*":
                self._issue(
                    "CRITICAL", "Wildcard Volume Mapping",
                    f"volume: {vvname} → host: * (any host)",
                    f"Volume '{vvname}' is mapped with a wildcard hostname ('*'). "
                    "This means ANY host that presents the correct LUN number can "
                    "access this volume — there is no host-level access restriction. "
                    "In a VMware environment this can lead to data corruption if "
                    "multiple ESXi hosts mount a VMFS volume simultaneously without "
                    "SCSI reservations.",
                    "Remove wildcard mappings: 'removevlun -f'. "
                    "Re-create specific host or host-set mappings. "
                    "Run: 'createvlun <vvname> <lun> <hostname-or-hostset>'.",
                    line=vlun.get("_line", 0))

    def _chk_wsapi(self):
        """Check WSAPI (REST API) configuration."""
        wsapi = self.wsapi
        if not wsapi.get("enabled"):
            return

        if wsapi.get("http"):
            self._issue(
                "HIGH", "WSAPI Using HTTP (Not HTTPS)",
                "WSAPI HTTP: enabled",
                "The HPE Alletra storage REST API (WSAPI) is accessible over HTTP. "
                "HTTP transmits API credentials and storage data in cleartext, "
                "exposing authentication tokens and configuration data on the wire.",
                "Disable HTTP WSAPI: 'setwsapi -http disabled'. "
                "Enable HTTPS only: 'setwsapi -https enabled'. "
                "Reference: HPE Hardening Guide sec. 5 — Encryption in Transit.")

        if not wsapi.get("https"):
            self._issue(
                "MEDIUM", "WSAPI Using HTTP (Not HTTPS)",
                "WSAPI HTTPS: not confirmed",
                "WSAPI HTTPS mode is not confirmed. If WSAPI is in use, ensure "
                "it is only accessible over HTTPS with a valid certificate.",
                "Enable WSAPI HTTPS: 'setwsapi -https enabled'. "
                "Replace the default self-signed certificate immediately.")

    def _chk_encryption(self):
        """Check encryption at rest status."""
        if not self.encryption_enabled:
            self._issue(
                "MEDIUM", "Encryption at Rest Not Enabled",
                "SED/encryption: not detected",
                "Encryption at rest is not detected on this Alletra MP system. "
                "Without encryption, physical disk theft or unauthorized drive removal "
                "provides full access to stored data, including VMware VMDK files. "
                "Per HPE Hardening Guide sec. 5: enable SED encryption where available.",
                "Enable Self-Encrypting Drive (SED) encryption. "
                "Store recovery keys offline in an access-controlled, encrypted vault — "
                "never on the storage system itself. Validate the license requirement.")


# ── Excel Reporter ────────────────────────────────────────────────────────────
class ExcelReporter:

    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"], C["critical_l"]),
        "HIGH":     (C["high"],     C["high_l"]),
        "MEDIUM":   (C["medium"],   C["medium_l"]),
        "LOW":      (C["low"],      C["low_l"]),
        "INFO":     (C["info"],     C["info_l"]),
    }

    def __init__(self, parser: HPAlletraParser, out_path: str):
        self.p   = parser
        self.out = out_path
        self.wb  = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    # ── Shared helpers ────────────────────────────────────────────────────────
    def _hdr(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = _font(bold=True, color=C["hdr_fg"])
            c.fill      = _fill(C["hdr_bg"])
            c.alignment = _align("center", wrap=False)
            c.border    = THIN
        ws.row_dimensions[row].height = 20

    @staticmethod
    def _set_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    @staticmethod
    def _row_fill(row: int) -> str | None:
        return C["alt_row"] if row % 2 == 0 else None

    def _sev_counts(self) -> dict[str, int]:
        counts: dict[str, int] = defaultdict(int)
        for iss in self.p.issues:
            counts[iss["severity"]] += 1
        return counts

    # ── Summary sheet ─────────────────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "HPE Alletra Storage MP — Security Audit Report"
        t.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 40

        si = self.p.sys_info
        meta = [
            ("System Name",      si.get("name", "Unknown")),
            ("System Model",     si.get("model", "Unknown")),
            ("Serial Number",    si.get("serial", "Unknown")),
            ("Firmware Version", si.get("fw", "Unknown")),
            ("Nodes",            si.get("nodes", "Unknown")),
            ("Report Date",      datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Input File",       os.path.basename(self.p.config_file)),
            ("Hosts Defined",    len(self.p.hosts)),
            ("Host Sets",        len(self.p.host_sets)),
            ("VLUNs",            len(self.p.vluns)),
        ]
        for row, (label, value) in enumerate(meta, 2):
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font(bold=True); c1.alignment = _align(); c1.border = THIN
            c1.fill = _fill("E8F4FC")
            c2 = ws.cell(row=row, column=2, value=str(value))
            c2.font = _font(); c2.alignment = _align(); c2.border = THIN
            ws.merge_cells(f"B{row}:F{row}")

        row = len(meta) + 3
        ws.cell(row=row, column=1, value="FINDING SUMMARY").font = _font(bold=True, size=12)
        row += 1

        sev_counts = self._sev_counts()
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            count = sev_counts.get(sev, 0)
            fg, bg = self.SEV_COLORS[sev]
            c1 = ws.cell(row=row, column=1, value=sev)
            c1.fill = _fill(bg); c1.font = _font(bold=True, color=fg)
            c1.alignment = _align("center"); c1.border = THIN
            c2 = ws.cell(row=row, column=2, value=count)
            c2.font = _font(bold=(count > 0)); c2.alignment = _align("center")
            c2.border = THIN
            ws.merge_cells(f"B{row}:F{row}")
            row += 1

        row += 1
        total = sum(sev_counts.values())
        ws.cell(row=row, column=1, value="TOTAL").font = _font(bold=True)
        ws.cell(row=row, column=1).border = THIN
        ws.cell(row=row, column=2, value=total).font = _font(bold=True)
        ws.cell(row=row, column=2).border = THIN

        # Hardening guide reference
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        ref = ws.cell(row=row, column=1,
                      value="Reference: HPE Alletra MP Security Guide (a00138815enw) | "
                            "HPE Alletra MP Hardening Guide (a00146015enw)")
        ref.font = _font(italic=True, color=C["info"], size=9)
        ref.alignment = _align("center", wrap=False)

        self._set_widths(ws, [28, 35, 20, 20, 20, 20])

    # ── Hosts sheet ───────────────────────────────────────────────────────────
    def _sheet_hosts(self):
        ws = self.wb.create_sheet("Hosts & Host Sets")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "Defined Hosts"
        t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 28
        row = 2

        hosts_in_sets: set[str] = set()
        for hs in self.p.host_sets:
            hosts_in_sets.update(hs.get("members", []))

        headers = ["Host Name", "Persona", "Protocol", "WWN / iSCSI Name",
                   "CHAP", "In Host Set?", "Config Line"]
        self._hdr(ws, headers, row=row)
        row += 1

        for h in self.p.hosts:
            rb = self._row_fill(row)
            in_set = h["name"] in hosts_in_sets
            chap_risk = h["proto"] == "iSCSI" and not h["chap"]
            vals = [h["name"], h["persona"], h["proto"], h["wwn_iqn"],
                    "yes" if h["chap"] else ("N/A" if h["proto"] != "iSCSI" else "NO"),
                    "yes" if in_set else "NO",
                    h["line"]]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
                if col == 5 and chap_risk:
                    c.font = _font(bold=True, color=C["high"])
                if col == 6 and not in_set:
                    c.font = _font(bold=True, color=C["medium"])
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="HOST SETS").font = _font(bold=True, size=11)
        row += 1
        self._hdr(ws, ["Set ID", "Set Name", "Members", "Config Line"], row=row)
        row += 1
        for hs in self.p.host_sets:
            rb = self._row_fill(row)
            for col, val in enumerate(
                    [hs["id"], hs["name"], ", ".join(hs["members"]), hs["line"]], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
            row += 1

        self._set_widths(ws, [28, 14, 10, 40, 8, 14, 10])

    # ── VLUNs sheet ───────────────────────────────────────────────────────────
    def _sheet_vluns(self):
        ws = self.wb.create_sheet("Volume Mappings (VLUNs)")
        ws.sheet_view.showGridLines = False

        headers = list({k for v in self.p.vluns for k in v if not k.startswith("_")})
        headers = sorted(headers)
        if not headers:
            headers = ["Lun", "VVname", "Hostname", "Port", "Type"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"

        for i, vlun in enumerate(self.p.vluns, 2):
            rb = self._row_fill(i)
            is_wildcard = vlun.get("Hostname", vlun.get("hostname", "")).strip() == "*"
            for col, key in enumerate(headers, 1):
                val = vlun.get(key, "")
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(bold=is_wildcard and key in ("Hostname", "hostname"),
                               color=C["critical"] if is_wildcard and key in ("Hostname", "hostname") else "000000")
                c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)

        self._set_widths(ws, [10, 30, 28, 10, 15, 15])

    # ── Users sheet ───────────────────────────────────────────────────────────
    def _sheet_users(self):
        ws = self.wb.create_sheet("Users & Auth")
        ws.sheet_view.showGridLines = False

        headers = ["Username", "Role", "Enabled", "Pwd Expires", "Config Line", "Notes"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"

        for i, u in enumerate(self.p.users, 2):
            rb = self._row_fill(i)
            notes = []
            if u["name"].lower() in HPAlletraParser._DEFAULT_ACCOUNTS:
                notes.append("DEFAULT ACCOUNT")
            if u["role"].lower() in HPAlletraParser._SUPER_ROLES:
                notes.append("SUPER ROLE")
            if u["pwd_exp"].lower() == "never":
                notes.append("pwd never expires")
            vals = [u["name"], u["role"],
                    "yes" if u["enabled"] else "no",
                    u["pwd_exp"], u["line"], ", ".join(notes)]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(color=C["disabled"] if not u["enabled"] else "000000")
                c.alignment = _align(); c.border = THIN
                if rb:
                    c.fill = _fill(rb)
                if col == 6 and "DEFAULT ACCOUNT" in notes:
                    c.font = _font(bold=True, color=C["high"])

        row = len(self.p.users) + 3
        ws.cell(row=row, column=1, value="AUTHENTICATION CONFIGURATION").font = _font(bold=True, size=11)
        row += 1
        for label, val in [
            ("LDAP / AD configured", "yes" if self.p.ldap_configured else "NO"),
            ("Audit logging",        "yes" if self.p.audit_enabled else "NO"),
            ("NTP servers",          str(len(self.p.ntp_servers))),
            ("SSH keys configured",  str(len(self.p.ssh_keys))),
        ]:
            bad = val == "NO"
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font(bold=True); c1.border = THIN
            c2 = ws.cell(row=row, column=2, value=val)
            c2.font = _font(bold=bad, color=C["high"] if bad else "000000")
            c2.border = THIN
            row += 1

        self._set_widths(ws, [22, 16, 10, 14, 10, 40])

    # ── Issues sheet ─────────────────────────────────────────────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False

        headers = ["#", "Validated", "Severity", "Residual Risk", "Residual Risk Note",
                   "Category", "Rule / Object", "Config Line(s)", "CIS v8", "CIS Benchmark",
                   "PCI DSS", "SCF",
                   "Description", "Recommendation", "Details",
                   "Asset", "Target", "Vuln", "Output", "Source"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        hostname = self.p.sys_info.get("name", "") or ""
        source = os.path.basename(self.p.config_file)

        sorted_issues = sorted(
            self.p.issues,
            key=lambda x: self.SEV_ORDER.get(x["severity"], 9))

        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            rb = self._row_fill(row)
            obj = iss["object"]
            line = iss.get("line", "")
            target = f"{obj} ({line})" if line else obj
            output = iss["description"]
            vals = [idx, "Y", sev, "", "",
                    iss["category"], obj, line,
                    iss.get("cis_label", ""), iss.get("cis_benchmark", ""),
                    iss.get("pci_label", ""), iss.get("scf", ""),
                    iss["description"], iss["recommendation"], "",
                    hostname, target, "", output, source]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 3:  # Severity
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                    c.alignment = _align("center")
                elif col in (1, 8):  # #, Config Line
                    c.font = _font(bold=(col == 1)); c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 9:  # CIS v8
                    c.font = _font(bold=True, color="17375E", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 10:  # CIS Benchmark
                    c.font = _font(bold=True, color="1F618D", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 11:  # PCI DSS
                    c.font = _font(bold=True, color="7B2D8B", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 12:  # SCF
                    c.font = _font(bold=True, color="1A5C3A", size=9)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                elif col == 2:  # Validated
                    c.font = _font(bold=True)
                    c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                else:
                    c.font = _font(); c.alignment = _align()
                    if rb:
                        c.fill = _fill(rb)
            ws.row_dimensions[row].height = 40

        self._set_widths(ws, [4, 12, 12, 18, 28, 32, 36, 14, 20, 20, 18, 18, 60, 60, 36, 24, 44, 16, 70, 30])

    # ── CIS Controls mapping ──────────────────────────────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS Controls Mapping")
        ws.sheet_view.showGridLines = False
        CIS_HDR = C["hdr_bg"]

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v8 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(CIS_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each CIS Control lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        ctrl_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        row = 3
        for ctrl_id in sorted(CIS_CTRL_DESC.keys(),
                               key=lambda x: [float(p) for p in x.split(".")]):
            desc  = CIS_CTRL_DESC[ctrl_id]
            items = sorted(ctrl_issues.get(ctrl_id, []),
                           key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count = len(items)

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"CIS {ctrl_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(CIS_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not items:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this control")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9"); nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in items:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 28, 10, 60, 60])

    # ── PCI DSS mapping ───────────────────────────────────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each PCI DSS v4.0 requirement lists all findings from this config that map to it."
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        req_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        row = 3
        for req_id in sorted(PCI_DSS_DESC.keys(),
                              key=lambda x: [int(p) for p in x.split(".")]):
            desc  = PCI_DSS_DESC[req_id]
            items = sorted(req_issues.get(req_id, []),
                           key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count = len(items)

            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not items:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this requirement")
                nc.font = _font(italic=True, color=C["info"]); nc.fill = _fill("F9F9F9")
                nc.alignment = _align(); nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Object",
                            "Config Line", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("2E4057")
                    c.alignment = _align("center", wrap=False); c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in items:
                    sev = iss["severity"]
                    fg, bg = self.SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["object"],
                            iss.get("line", ""), iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill = _fill(bg); c.font = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb: c.fill = _fill(rb)
                        else:
                            c.font = _font(); c.alignment = _align()
                            if rb: c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 28, 10, 60, 60])

    # ── Save ──────────────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_hosts()
        self._sheet_vluns()
        self._sheet_users()
        self._sheet_issues()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)


# ── CLI ───────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="HPE Alletra Storage MP Security Analyzer — outputs Excel report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python hp_alletra_analyzer.py alletra-export.txt
  python hp_alletra_analyzer.py alletra-export.txt -o alletra-audit.xlsx

Collect input on the array (run as super/admin):
  showsys > alletra-export.txt
  showuser -d >> alletra-export.txt
  showpasswordpolicy >> alletra-export.txt
  showsnmp >> alletra-export.txt
  showsyslog >> alletra-export.txt
  showaudit >> alletra-export.txt
  showtime -zone >> alletra-export.txt
  shownet >> alletra-export.txt
  showhost >> alletra-export.txt
  showhost -chap >> alletra-export.txt
  showhostset >> alletra-export.txt
  showvlun >> alletra-export.txt
  showvv -showcols Name,VSize,Prov,State >> alletra-export.txt
  showport >> alletra-export.txt
  showwsapi >> alletra-export.txt
  showsshkey >> alletra-export.txt

References:
  HPE Alletra MP Security Guide          a00138815enw
  HPE Alletra MP Administrator Hardening a00146015enw
  HPE DISA STIG Guidance                 community.hpe.com
""",
    )
    ap.add_argument("config", help="Concatenated HPE Alletra MP CLI output file")
    ap.add_argument("-o", "--output", default=None,
                    help="Output Excel file (default: <config-stem>_alletra_analysis.xlsx)")
    args = ap.parse_args()

    if not args.output:
        stem = os.path.splitext(os.path.basename(args.config))[0]
        args.output = f"{stem}_alletra_analysis.xlsx"

    print(f"[*] Parsing:  {args.config}")
    parser = HPAlletraParser(args.config)
    parser.parse()

    sev_counts: dict[str, int] = defaultdict(int)
    for iss in parser.issues:
        sev_counts[iss["severity"]] += 1

    si = parser.sys_info
    print(f"[*] Parsed:")
    print(f"      System name      : {si.get('name', 'Unknown')}")
    print(f"      Model            : {si.get('model', 'Unknown')}")
    print(f"      Firmware         : {si.get('fw', 'Unknown')}")
    print(f"      Users            : {len(parser.users)}")
    print(f"      Hosts            : {len(parser.hosts)}")
    print(f"      Host sets        : {len(parser.host_sets)}")
    print(f"      VLUNs            : {len(parser.vluns)}")
    print(f"      SNMP communities : {len(parser.snmp.get('communities', []))}")
    print(f"      NTP servers      : {len(parser.ntp_servers)}")
    print(f"      Syslog servers   : {len(parser.syslog_servers)}")
    print(f"      Audit enabled    : {parser.audit_enabled}")
    print(f"      Encryption       : {'yes' if parser.encryption_enabled else 'not detected'}")
    print(f"[*] Security findings: {len(parser.issues)}")
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        if sev_counts[sev]:
            print(f"      {sev:<10}: {sev_counts[sev]}")

    print(f"[*] Writing Excel report...")
    reporter = ExcelReporter(parser, args.output)
    reporter.save()
    print(f"[+] Saved: {args.output}")


if __name__ == "__main__":
    main()
