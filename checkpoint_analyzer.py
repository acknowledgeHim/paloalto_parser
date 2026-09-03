#!/usr/bin/env python3
"""
Check Point Firewall Configuration Analyzer
Parses a Check Point Gaia OS configuration — a 'show configuration' (clish)
CLI text capture from the gateway or management server — checks it against
the CIS Check Point Firewall Benchmark L1/L2 using the actual Tenable Nessus
.audit files for that benchmark (bundled in audits.tar.gz, same as the other
analyzers in this repo), and exports findings + rulebase inventory to Excel.
A SmartConsole Rule Base "Export to CSV" can supply the Security Rulebase,
since Gaia's own 'show configuration' only covers gateway/OS-level settings
(password policy, SNMP, NTP, DNS, session timeouts, ...) — the access
rulebase itself lives on the Management Server, not in the Gaia config.

Usage:
    python checkpoint_analyzer.py "show configuration.txt"
    python checkpoint_analyzer.py "show configuration.txt" --rules-csv rulebase.csv
    python checkpoint_analyzer.py "show configuration.txt" -o audit.xlsx
    python checkpoint_analyzer.py rulebase.csv                    # rules-only mode
"""

import re
import os
import sys
import csv as _csv
import tarfile
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── Colour palette (matches the other analyzers in this repo) ───────────────
C = {
    "hdr_bg":   "1F3864", "hdr_fg":   "FFFFFF",
    "critical": "C00000", "critical_l": "FFB3B3",
    "high":     "FF0000", "high_l":     "FFD9B3",
    "medium":   "FF8C00", "medium_l":   "FFF2CC",
    "low":      "0070C0", "low_l":      "BDD7EE",
    "info":     "595959", "info_l":     "F2F2F2",
    "allow":    "375623", "allow_l":    "E2EFDA",
    "deny":     "C00000", "deny_l":     "FFB3B3",
    "disabled": "A6A6A6",
    "alt_row":  "F5F5F5",
}
_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── vulns.txt (shared vendor-agnostic vulnerability taxonomy) ───────────────
def _load_vulns(path: str) -> dict[int, str]:
    vulns: dict[int, str] = {}
    try:
        with open(path, encoding="utf-8") as fh:
            for vuln_id, line in enumerate(fh, 1):
                text = line.strip()
                if text:
                    vulns[vuln_id] = text
        return vulns
    except FileNotFoundError:
        return vulns


def _find_vulns_file() -> dict[int, str]:
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "vulns.txt"),
        os.path.join(os.getcwd(), "vulns.txt"),
    ]
    for path in candidates:
        vulns = _load_vulns(path)
        if vulns:
            return vulns
    print("[!] Warning: vulns.txt not found — Vuln column will be empty. "
          "Place vulns.txt alongside checkpoint_analyzer.py.")
    return {}


# Maps native (non-benchmark-catalog) finding categories to a vulns.txt line ID.
# Benchmark-catalog-driven findings (the CIS Check Point items) don't use this —
# their evidence/mapping comes straight from the .audit file's own reference data.
CATEGORY_VULN_ID: dict[str, int] = {
    "Disabled Rulebase Rule":              2,
    "Missing Rule Comment":                5,
    "Rule Allows Any Source":              9,
    "Rule Allows Any Destination":         9,
    "Rule Allows Any Service":             9,
    "Allow Rule Not Logging":              28,
    "No Hit Count Data":                   25,
    "No Default Drop/Cleanup Rule":        9,
}

# Maps each CIS Check Point Firewall Benchmark ID to a vulns.txt line ID —
# the fallback used for every benchmark-catalog-driven finding (the "CIS
# x.x — ..." categories), since those category strings are built dynamically
# per check and can't be enumerated in CATEGORY_VULN_ID above. IDs with no
# reasonable fit in vulns.txt's fixed taxonomy are left unmapped (blank Vuln
# column) rather than forced into a category that doesn't really describe them.
BENCHMARK_ID_VULN_ID: dict[str, int] = {
    "1.1": 27, "1.2": 27, "1.3": 27, "1.4": 27, "1.5": 27, "1.6": 27,
    "1.7": 3, "1.8": 3, "1.9": 3, "1.10": 27, "1.11": 3, "1.12": 3, "1.13": 3,
    "2.1.1": 21, "2.1.2": 8, "2.1.5": 34, "2.1.6": 17, "2.1.8": 24, "2.1.9": 4,
    "2.2.1": 15, "2.2.2": 15, "2.2.3": 28, "2.2.4": 28,
    "2.3.1": 18, "2.3.2": 19,
    "2.5.1": 27, "2.5.2": 27, "2.5.3": 4, "2.5.4": 27, "2.5.5": 14,
    "2.6.1": 28, "2.6.2": 28, "2.6.3": 28,
    "3.1": 14, "3.2": 9, "3.4": 25, "3.5": 9, "3.6": 9, "3.7": 9, "3.8": 28, "3.9": 28,
    "3.15": 17, "3.16": 17, "3.20": 28,
}


# ── audits.tar.gz locate (same search strategy as the other analyzers) ──────
_TAR_SEARCH_SKIP_DIRS = {".git", "__pycache__", "node_modules", ".venv", "venv", ".tox", "dist", "build"}


def _find_audits_tar() -> "str | None":
    roots = list(dict.fromkeys([os.path.dirname(os.path.abspath(__file__)), os.getcwd()]))
    for base in roots:
        direct = os.path.join(base, "audits.tar.gz")
        if os.path.isfile(direct):
            return direct
    for base in roots:
        for root, dirs, files in os.walk(base):
            dirs[:] = [d for d in dirs if d not in _TAR_SEARCH_SKIP_DIRS and not d.startswith(".")]
            if "audits.tar.gz" in files:
                return os.path.join(root, "audits.tar.gz")
    return None


# ── CIS Controls v7 descriptions ─────────────────────────────────────────────
# The CIS Check Point Firewall Benchmark's .audit reference data cites CIS
# Controls v7 (CSCv7), not v8 — this table only covers the sub-controls that
# actually appear there.
CIS_CTRL_DESC: dict[str, str] = {
    "5.1":  "Establish Secure Configurations",
    "5.2":  "Maintain Secure Images",
    "5.3":  "Securely Store Master Images",
    "5.5":  "Implement Automated Configuration Monitoring Systems",
    "6.2":  "Activate Audit Logging",
    "6.4":  "Ensure Adequate Storage for Logs",
    "6.7":  "Regularly Review Logs",
    "9.2":  "Ensure Only Approved Ports, Protocols and Services Are Running",
    "11.1": "Maintain Standard Security Configurations for Network Devices",
    "11.2": "Document Traffic Configuration Rules",
    "11.3": "Use Automated Tools to Verify Standard Device Configurations and Detect Changes",
    "11.7": "Manage Network Infrastructure Through a Dedicated Network",
    "12.1": "Maintain an Inventory of Network Boundaries",
    "12.2": "Scan for Unauthorized Connections Across Trusted Network Boundaries",
    "12.3": "Deny Communications with Known Malicious or Unused Internet IP Addresses",
    "12.4": "Deny Communication over Unauthorized Ports",
}

# ── PCI DSS v4.0 descriptions ────────────────────────────────────────────────
# Same approach: only the requirement IDs that actually appear in the
# benchmark's reference data.
PCI_DSS_DESC: dict[str, str] = {
    "1.2.1":  "Configuration standards for NSC rulesets are defined, implemented, and maintained",
    "1.3.1":  "Inbound traffic to the CDE is restricted",
    "1.4.1":  "Network security controls are implemented between trusted and untrusted networks",
    "1.4.2":  "Inbound traffic from untrusted networks to trusted networks is restricted",
    "1.4.3":  "Anti-spoofing measures are implemented to detect and block forged source IP addresses",
    "2.2.3":  "Primary functions requiring different security levels are managed to prevent security issues",
    "2.2.4":  "Only necessary services, protocols, daemons, and functions are enabled",
    "8.2.8":  "If a session is idle for more than 15 minutes, the user is required to re-authenticate",
    "10.2.2": "Audit logs capture all actions taken by individuals with administrative access",
    "10.6":   "Time-synchronization mechanisms support consistent time settings across all systems",
    "10.6.1": "System clocks and time are synchronized using time-synchronization technology",
    "10.6.2": "Systems are configured to correct time settings from a designated authoritative source",
    "10.6.3": "Time synchronization settings and data are protected from unauthorized changes",
}


def _cis_label(ids: list[str]) -> str:
    return " · ".join(f"CIS {c}" for c in ids)


def _cis_benchmark_label(ids: list[str]) -> str:
    return " · ".join(ids)


def _pci_label(ids: list[str]) -> str:
    return " · ".join(f"PCI DSS {c}" for c in ids)


def _refs_from_reference(reference: str) -> tuple[list[str], list[str]]:
    """Pull CIS Controls v7 (CSCv7|x.x) and PCI DSS v4.0 (PCI-DSSv4.0|x.x) IDs
    out of a .audit 'reference' field."""
    cis_ids = sorted(set(re.findall(r'CSCv7\|([\d.]+)', reference)),
                      key=lambda s: [int(p) for p in s.split(".")])
    pci_ids = sorted(set(re.findall(r'PCI-DSSv4\.0\|([\d.]+)', reference)),
                      key=lambda s: [int(p) for p in s.split(".")])
    return cis_ids, pci_ids


# ── Gaia 'show configuration' (clish) text loading ───────────────────────────
_ANSI_CSI_RE  = re.compile(r'\x1b\[[0-9;?]*[A-Za-z]')
_ANSI_MISC_RE = re.compile(r'\x1b[=>]')


def _load_gaia_config(filename: str) -> list[tuple[int, str]]:
    """Return [(line_no, text), ...] for every 'set'/'add'/'delete' clish
    statement in a captured Gaia 'show configuration' (or 'show configuration
    all') session. Strips ANSI escapes/PuTTY artifacts and comment/blank/
    prompt lines."""
    try:
        with open(filename, "r", encoding="utf-8", errors="replace") as fh:
            raw = fh.read()
    except FileNotFoundError:
        sys.exit(f"File not found: {filename}")

    raw = _ANSI_CSI_RE.sub("", raw)
    raw = _ANSI_MISC_RE.sub("", raw)

    out: list[tuple[int, str]] = []
    for lineno, line in enumerate(raw.splitlines(), 1):
        text = line.rstrip()
        stripped = text.strip()
        if not stripped or stripped.startswith("#"):
            continue
        # Gaia clish statements of interest all start with one of these verbs.
        if re.match(r'^(set|add|delete)\s+\S', stripped):
            out.append((lineno, stripped))
    return out


# ── Check Point CIS Benchmark .audit parser ──────────────────────────────────
def _extract_quoted(field: str, block: str) -> "str | None":
    """Extract a 'field : "..."' value from a .audit block, honoring \\" escapes."""
    m = re.search(rf'(?m)^\s*{field}\s*:\s*"((?:[^"\\]|\\.)*)"', block)
    return m.group(1).replace('\\"', '"') if m else None


def _first_para(text: str) -> str:
    for marker in ("\nRationale:", "\n\nRationale:", "\nImpact:", "\n\nImpact:"):
        if marker in text:
            text = text[:text.index(marker)]
    paras = [p.strip() for p in text.split("\n\n") if p.strip()]
    return re.sub(r'\s+', ' ', paras[0]) if paras else re.sub(r'\s+', ' ', text.strip())


_ID_RE = re.compile(r'^(\d+(?:\.\d+)*)\s+(.*)$')


def _split_id(desc: str) -> tuple[str, str]:
    """'2.2.3 Ensure SNMP traps is enabled - lowDiskSpace' ->
    ('2.2.3', 'Ensure SNMP traps is enabled - lowDiskSpace')"""
    m = _ID_RE.match(desc.strip())
    return (m.group(1), m.group(2)) if m else ("", desc.strip())


def _parse_checkpoint_audit(content: str, level: str) -> list[dict]:
    """Parse a CIS_Check_Point_Firewall_*.audit file into a flat list of
    catalog items — both automated CONFIG_CHECK entries (have regex+expect)
    and non-automated <report> entries (manual review / conditionally
    evaluated by Nessus). Each item carries the raw benchmark ID it belongs
    to so callers can group multi-part checks (e.g. '1.4 ... - history-length'
    and '1.4 ... - history-checking') under one finding."""
    items: list[dict] = []

    for block in re.findall(r'<custom_item>(.*?)</custom_item>', content, re.DOTALL):
        if 'CONFIG_CHECK' not in block:
            continue
        desc = _extract_quoted("description", block) or ""
        base_id, title = _split_id(desc)
        if not base_id:
            continue  # internal AND-condition helper item, not a real benchmark control
        items.append({
            "level": level, "id": base_id, "title": title, "kind": "custom_item",
            "info": _first_para(_extract_quoted("info", block) or ""),
            "solution": (_extract_quoted("solution", block) or "").strip(),
            "reference": _extract_quoted("reference", block) or "",
            "regex": _extract_quoted("regex", block),
            "expect": _extract_quoted("expect", block),
            "not_expect": _extract_quoted("not_expect", block),
        })

    for block in re.findall(r'<report[^>]*>(.*?)</report>', content, re.DOTALL):
        desc = _extract_quoted("description", block) or ""
        base_id, title = _split_id(desc)
        if not base_id:
            continue
        items.append({
            "level": level, "id": base_id, "title": title, "kind": "report",
            "info": _first_para(_extract_quoted("info", block) or ""),
            "solution": (_extract_quoted("solution", block) or "").strip(),
            "reference": _extract_quoted("reference", block) or "",
            "regex": None, "expect": None, "not_expect": None,
        })

    return items


# Manual, best-judgment severities per benchmark ID — the .audit file itself
# specifies a 'severity' field on only 2 of the 68 catalog items, so every
# other severity here is assigned the same way the other analyzers in this
# repo assign theirs: by what the control actually protects against.
SEVERITY_OVERRIDE: dict[str, str] = {
    "1.1": "HIGH", "1.2": "MEDIUM", "1.3": "HIGH", "1.4": "MEDIUM", "1.5": "MEDIUM",
    "1.6": "LOW", "1.7": "LOW", "1.8": "MEDIUM", "1.9": "LOW", "1.10": "MEDIUM",
    "1.11": "HIGH", "1.12": "HIGH", "1.13": "MEDIUM",
    "2.1.1": "LOW", "2.1.2": "LOW", "2.1.3": "LOW", "2.1.4": "LOW",
    "2.1.5": "MEDIUM", "2.1.6": "MEDIUM", "2.1.7": "LOW", "2.1.8": "LOW",
    "2.1.9": "HIGH", "2.1.10": "LOW",
    "2.2.1": "HIGH", "2.2.2": "MEDIUM", "2.2.3": "LOW", "2.2.4": "LOW",
    "2.3.1": "MEDIUM", "2.3.2": "LOW",
    "2.4.1": "LOW", "2.4.2": "LOW", "2.4.3": "LOW",
    "2.5.1": "MEDIUM", "2.5.2": "MEDIUM", "2.5.3": "MEDIUM", "2.5.4": "MEDIUM", "2.5.5": "HIGH",
    "2.6.1": "MEDIUM", "2.6.2": "MEDIUM", "2.6.3": "MEDIUM",
    "3.1": "HIGH", "3.2": "HIGH", "3.3": "INFO", "3.4": "LOW",
    "3.5": "HIGH", "3.6": "HIGH", "3.7": "HIGH", "3.8": "MEDIUM", "3.9": "LOW",
    "3.10": "MEDIUM", "3.11": "MEDIUM", "3.12": "HIGH", "3.13": "LOW",
    "3.14": "MEDIUM", "3.15": "MEDIUM", "3.16": "LOW", "3.17": "LOW",
    "3.18": "LOW", "3.19": "LOW", "3.20": "MEDIUM",
}


def _severity_for(base_id: str) -> str:
    return SEVERITY_OVERRIDE.get(base_id, "MEDIUM")


_VAR_RE = re.compile(r'@[A-Z_]+@')


def _compile_pattern(pattern: str) -> tuple[re.Pattern, bool]:
    """A .audit expect/regex value may embed a Nessus policy variable like
    @DNS_PRIMARY@ or @TIMEZONE@ — a site-specific value this script has no
    way to know. Those are relaxed to 'any non-space token' so the check
    still verifies the *feature* is configured, and the finding text says
    to confirm the value by hand."""
    had_vars = bool(_VAR_RE.search(pattern))
    relaxed = _VAR_RE.sub(r'\\S+', pattern)
    return re.compile(relaxed), had_vars


def _eval_config_check(check: dict, gaia_lines: list[tuple[int, str]]) -> dict:
    """Evaluate one automated CONFIG_CHECK item against the parsed Gaia
    config lines. Returns {'status': 'PASS'|'FAIL', 'line': int, 'text': str,
    'had_vars': bool}."""
    regex_pat = re.compile(check["regex"])
    matches = [(ln, txt) for ln, txt in gaia_lines if regex_pat.search(txt)]

    if check["expect"] is None and check["not_expect"] is not None:
        # e.g. DHCP: absent entirely, or present-but-not-matching-not_expect, both PASS.
        nexp_pat, had_vars = _compile_pattern(check["not_expect"])
        bad = next(((ln, txt) for ln, txt in matches if nexp_pat.search(txt)), None)
        if bad:
            return {"status": "FAIL", "line": bad[0], "text": bad[1], "had_vars": had_vars}
        last = matches[-1] if matches else (0, "")
        return {"status": "PASS", "line": last[0], "text": last[1], "had_vars": had_vars}

    exp_pat, had_vars = _compile_pattern(check["expect"] or "")
    good = next(((ln, txt) for ln, txt in matches if exp_pat.search(txt)), None)
    if good:
        return {"status": "PASS", "line": good[0], "text": good[1], "had_vars": had_vars}
    if matches:
        ln, txt = matches[0]
        return {"status": "FAIL", "line": ln, "text": txt, "had_vars": had_vars}
    return {"status": "FAIL", "line": 0, "text": "", "had_vars": had_vars}


# ── SmartConsole Rule Base CSV export ────────────────────────────────────────
_RULE_HDR_ALIASES: dict[str, list[str]] = {
    "num":         ["no.", "no", "rule number", "#"],
    "name":        ["name", "rule name"],
    "source":      ["source"],
    "destination": ["destination"],
    "vpn":         ["vpn"],
    "services":    ["services & applications", "service", "services", "application"],
    "action":      ["action"],
    "track":       ["track", "log"],
    "install_on":  ["install on"],
    "time":        ["time"],
    "comments":    ["comments", "comment"],
    "enabled":     ["enabled", "rule enabled", "hits enabled"],
    "hits":        ["hits", "hit count"],
}


def _csv_cell(v: str) -> str:
    v = (v or "").strip()
    return "" if v in ("-", "--", "n/a", "N/A") else v


def _parse_rulebase_csv(path: str) -> list[dict]:
    try:
        fh = open(path, "r", encoding="utf-8-sig", newline="")
    except FileNotFoundError:
        sys.exit(f"Rules CSV not found: {path}")

    with fh:
        reader = _csv.DictReader(fh)
        if reader.fieldnames is None:
            sys.exit(f"Rules CSV has no header row: {path}")
        lower_hdr = {(h or "").strip().lower(): h for h in reader.fieldnames}

        def col(key: str) -> "str | None":
            for alias in _RULE_HDR_ALIASES[key]:
                if alias in lower_hdr:
                    return lower_hdr[alias]
            return None

        cols = {k: col(k) for k in _RULE_HDR_ALIASES}
        rules: list[dict] = []
        for rule_num, raw_row in enumerate(reader, 1):
            row = {k: _csv_cell(v) for k, v in raw_row.items()}
            name = row.get(cols["name"], "") if cols["name"] else ""
            action = row.get(cols["action"], "") if cols["action"] else ""
            if not name and not action:
                continue  # section-header / blank row in the export
            enabled_raw = (row.get(cols["enabled"], "") if cols["enabled"] else "").strip().lower()
            rules.append({
                "num": rule_num,
                "name": name or f"Rule {rule_num}",
                "source":      row.get(cols["source"], "") if cols["source"] else "",
                "destination": row.get(cols["destination"], "") if cols["destination"] else "",
                "vpn":         row.get(cols["vpn"], "") if cols["vpn"] else "",
                "services":    row.get(cols["services"], "") if cols["services"] else "",
                "action":      action,
                "track":       row.get(cols["track"], "") if cols["track"] else "",
                "install_on":  row.get(cols["install_on"], "") if cols["install_on"] else "",
                "time":        row.get(cols["time"], "") if cols["time"] else "",
                "comments":    row.get(cols["comments"], "") if cols["comments"] else "",
                "enabled":     enabled_raw not in ("no", "false", "disabled", "0"),
                "hits":        row.get(cols["hits"], "") if cols["hits"] else "",
            })
    print(f"[+] Loaded {len(rules)} rulebase row(s) from CSV: {path}")
    return rules


def _has_any(field: str) -> bool:
    return any(tok.strip().lower() == "any" for tok in re.split(r'[;,]', field or ""))


def _is_allow(action: str) -> bool:
    return (action or "").strip().lower() in ("accept", "allow")


def _is_drop(action: str) -> bool:
    return (action or "").strip().lower() in ("drop", "reject", "deny")


# ── Dialect detection: enterprise Gaia vs. Quantum Spark (Gaia Embedded) ────
# The CIS Check Point Firewall Benchmark's .audit checks (the generic
# CONFIG_CHECK engine above) are written against enterprise Gaia's clish
# 'show configuration' syntax (e.g. 'set password-controls ...',
# 'set net-access telnet off'). Quantum Spark / SMB-series appliances (the
# 1500/1600/1800/1900 line, running Gaia Embedded) export a materially
# different command grammar for the same settings — e.g. password policy
# lives in one 'set administrator session-settings ...' line with quoted
# values, not 'set password-controls ...'. Running the enterprise regex
# checks against a Spark capture produces false "not configured" findings
# for settings that ARE configured, just under a different command name.
_ENTERPRISE_SIGNATURE_RE = re.compile(
    r'(?m)^(?:set password-controls|set net-access telnet|set message banner|'
    r'set message motd|set hostname\s|set syslog (?:mgmtauditlogs|auditlog|cplogs))\b')
_SPARK_SIGNATURE_RE = re.compile(
    r'(?m)^(?:set device-details|set administrator session-settings|'
    r'set message type "(?:banner|motd)"|set administrators (?:radius|tacacs)-auth)\b')


def _detect_dialect(gaia_lines: list[tuple[int, str]]) -> str:
    """Return 'spark' or 'enterprise'. Defaults to 'enterprise' (the original,
    already-validated behavior) when the signal is weak or absent, so this
    never changes behavior for a config that isn't clearly Quantum Spark."""
    text = "\n".join(txt for _ln, txt in gaia_lines)
    spark_hits = len(_SPARK_SIGNATURE_RE.findall(text))
    enterprise_hits = len(_ENTERPRISE_SIGNATURE_RE.findall(text))
    return "spark" if spark_hits > enterprise_hits else "enterprise"


_SPARK_ATTR_RE = re.compile(r'([A-Za-z][A-Za-z0-9_.-]*)\s+"((?:[^"\\]|\\.)*)"')


def _spark_attrs(line: str) -> dict[str, str]:
    """Extract 'key "value"' pairs from a Quantum Spark config line, e.g.
    'set administrator session-settings lockout-enable "on" ...' ->
    {'lockout-enable': 'on', ...}. Positional (unlabeled) quoted values —
    e.g. 'set message type "banner" "on" msgvalue "..."' — aren't captured
    this way; those are pulled out with their own dedicated regex below."""
    return {m.group(1): m.group(2).replace('\\"', '"') for m in _SPARK_ATTR_RE.finditer(line)}


def _spark_has_msgvalue_text(line: str) -> bool:
    """Whether a banner/MOTD line's msgvalue has real content. The message
    text itself commonly embeds literal newlines in the source file (Spark
    lets banners span many lines) — _load_gaia_config keeps only the first
    physical line of such a statement, so the quote is often left
    unterminated on it. Checking for *any* non-blank text after the opening
    quote (closed or not) is what actually reflects whether a banner is set,
    rather than requiring a closing quote that may be several lines away."""
    m = re.search(r'msgvalue\s+"(.*)', line)
    return bool(m and m.group(1).strip())


def _parse_spark_facts(gaia_lines: list[tuple[int, str]]) -> dict[str, tuple]:
    """Walk a Quantum Spark config once and pull out every setting the Spark
    CIS-check evaluator below needs, as {fact_name: (value, line_no)}."""
    facts: dict[str, tuple] = {}
    admin_access: list[tuple[str, int]] = []

    for ln, txt in gaia_lines:
        if txt.startswith('set administrator session-settings'):
            a = _spark_attrs(txt)
            for key in ("lockout-enable", "max-lockout-attempts", "lock-period",
                        "inactivity-timeout", "password-complexity-level",
                        "password-history-mechanism", "password-expiration-timeout"):
                if key in a:
                    facts[key] = (a[key], ln)
        elif txt.startswith('set dns primary'):
            for part in ("primary", "secondary", "tertiary"):
                m = re.search(rf'{part}\s+ipv4-address\s+"([^"]*)"', txt)
                if m:
                    facts[f"dns-{part}"] = (m.group(1), ln)
        elif txt.startswith('set message type "banner"'):
            m = re.search(r'"banner"\s+"(on|off)"', txt)
            if m:
                facts["banner-state"] = (m.group(1), ln)
            facts["banner-has-text"] = (_spark_has_msgvalue_text(txt), ln)
        elif txt.startswith('set message type "motd"'):
            m = re.search(r'"motd"\s+"(on|off)"', txt)
            if m:
                facts["motd-state"] = (m.group(1), ln)
            facts["motd-has-text"] = (_spark_has_msgvalue_text(txt), ln)
        elif txt.startswith('set device-details'):
            a = _spark_attrs(txt)
            if "hostname" in a:
                facts["hostname"] = (a["hostname"], ln)
        elif txt.startswith('set snmp agent'):
            a = _spark_attrs(txt)
            if "agent" in a:
                facts["snmp-agent"] = (a["agent"], ln)
            if "agent-version" in a:
                facts["snmp-agent-version"] = (a["agent-version"], ln)
            facts["snmp-community"] = (a.get("community", ""), ln)
        elif txt.startswith('set ntp local-time-zone'):
            m = re.search(r'local-time-zone\s+"([^"]*)"', txt)
            if m:
                facts["timezone"] = (m.group(1), ln)
        elif txt.startswith('set ntp active'):
            m = re.search(r'active\s+"(on|off)"', txt)
            if m:
                facts["ntp-active"] = (m.group(1), ln)
        elif txt.startswith('set ntp server'):
            a = _spark_attrs(txt)
            if "primary" in a:
                facts["ntp-primary"] = (a["primary"], ln)
            if "secondary" in a:
                facts["ntp-secondary"] = (a["secondary"], ln)
        elif txt.startswith('set administrators radius-auth'):
            m = re.search(r'radius-auth\s+"(enable|disable)"', txt)
            if m:
                facts["radius-auth"] = (m.group(1), ln)
        elif txt.startswith('set administrators tacacs-auth'):
            m = re.search(r'tacacs-auth\s+"(enable|disable)"', txt)
            if m:
                facts["tacacs-auth"] = (m.group(1), ln)
        elif txt.startswith('set logs-config'):
            a = _spark_attrs(txt)
            if "send-audit-on-db-change" in a:
                facts["audit-on-change"] = (a["send-audit-on-db-change"], ln)
            if "display-audit-logs" in a:
                facts["display-audit-logs"] = (a["display-audit-logs"], ln)
        elif txt == 'delete admin-access-ipv4-address-all' or txt.startswith('add admin-access-ipv4-address'):
            admin_access.append((txt, ln))

    if admin_access:
        facts["admin-access-entries"] = admin_access
    return facts


# Every benchmark ID _eval_spark_check() below knows how to evaluate on
# Quantum Spark — including 2.5.5, which the enterprise .audit itself marks
# "Manual Review Required" (Spark's admin-access-ipv4-address list makes it
# automatable there, unlike enterprise Gaia's 'allowed-client').
SPARK_IMPLEMENTED_IDS: frozenset[str] = frozenset({
    "1.3", "1.4", "1.5", "1.11", "1.12", "1.13",
    "2.1.1", "2.1.2", "2.1.6", "2.1.8",
    "2.2.1", "2.2.2",
    "2.3.1", "2.3.2",
    "2.5.1", "2.5.2", "2.5.4", "2.5.5",
    "2.6.1", "2.6.3",
})


def _eval_spark_check(base_id: str, facts: dict) -> "dict | None":
    """Evaluate one CIS Check Point Firewall Benchmark ID against Quantum
    Spark facts. Returns {'status': 'PASS'|'FAIL', 'line': int, 'evidence': str},
    or None when this platform has no equivalent setting to evaluate (the
    caller reports that as 'Not Applicable — Quantum Spark', not a failure)."""

    def get(key):
        return facts.get(key, (None, 0))

    if base_id == "1.3":
        val, ln = get("password-complexity-level")
        if val is None:
            return None
        ok = val.lower() in ("high", "strict")
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'password-complexity-level = "{val}"'}

    if base_id == "1.4":
        val, ln = get("password-history-mechanism")
        if val is None:
            return None
        ok = val.lower() == "true"
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'password-history-mechanism = "{val}"'}

    if base_id == "1.5":
        val, ln = get("password-expiration-timeout")
        if val is None:
            return None
        days = int(val) if val.isdigit() else None
        ok = days is not None and 0 < days <= 90
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'password-expiration-timeout = "{val}" days'}

    if base_id == "1.11":
        val, ln = get("lockout-enable")
        if val is None:
            return None
        ok = val.lower() == "on"
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'lockout-enable = "{val}"'}

    if base_id == "1.12":
        val, ln = get("max-lockout-attempts")
        if val is None:
            return None
        n = int(val) if val.isdigit() else None
        ok = n is not None and 1 <= n <= 5
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'max-lockout-attempts = "{val}"'}

    if base_id == "1.13":
        val, ln = get("lock-period")
        if val is None:
            return None
        minutes = int(val) if val.isdigit() else None
        ok = minutes is not None and minutes * 60 >= 300
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'lock-period = "{val}" (assumed minutes — the WebUI field is '
                            f'"Block user for ... minutes"; verify units if this looks wrong)'}

    if base_id == "2.1.1":
        state, ln1 = get("banner-state")
        has_text, ln2 = get("banner-has-text")
        if state is None:
            return None
        ok = state.lower() == "on" and bool(has_text)
        return {"status": "PASS" if ok else "FAIL", "line": ln1 or ln2,
                "evidence": f'message type "banner" = "{state}", text configured = {bool(has_text)}'}

    if base_id == "2.1.2":
        state, ln1 = get("motd-state")
        has_text, ln2 = get("motd-has-text")
        if state is None:
            return None
        ok = state.lower() == "on" and bool(has_text)
        return {"status": "PASS" if ok else "FAIL", "line": ln1 or ln2,
                "evidence": f'message type "motd" = "{state}", text configured = {bool(has_text)}'}

    if base_id == "2.1.6":
        p, lnp = get("dns-primary")
        s, lns = get("dns-secondary")
        t, lnt = get("dns-tertiary")
        if p is None and s is None and t is None:
            return None
        missing = [name for name, v in (("primary", p), ("secondary", s), ("tertiary", t)) if not v]
        return {"status": "FAIL" if missing else "PASS", "line": lnp or lns or lnt,
                "evidence": f'primary={p or "(none)"}, secondary={s or "(none)"}, tertiary={t or "(none)"}'}

    if base_id == "2.1.8":
        val, ln = get("hostname")
        if val is None:
            return None
        ok = bool(val.strip())
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": f'hostname = "{val}"'}

    if base_id == "2.2.1":
        val, ln = get("snmp-agent")
        if val is None:
            return None
        ok = val.lower() == "off"
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": f'snmp agent = "{val}"'}

    if base_id == "2.2.2":
        ver, ln = get("snmp-agent-version")
        comm, _ = get("snmp-community")
        if ver is None:
            return None
        ok = ver.lower() == "v3-only" and not (comm or "").strip()
        evidence = f'agent-version = "{ver}"'
        if comm:
            evidence += f', but a community string ("{comm}") is also configured — undermines v3-only enforcement'
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": evidence}

    if base_id == "2.3.1":
        active, ln1 = get("ntp-active")
        p, ln2 = get("ntp-primary")
        s, ln3 = get("ntp-secondary")
        if active is None and p is None:
            return None
        ok = (active or "").lower() == "on" and bool(p) and bool(s)
        return {"status": "PASS" if ok else "FAIL", "line": ln1 or ln2,
                "evidence": f'active={active}, primary={p or "(none)"}, secondary={s or "(none)"}'}

    if base_id == "2.3.2":
        val, ln = get("timezone")
        if val is None:
            return None
        ok = bool(val.strip()) and "NOT_SET" not in val.upper()
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": f'local-time-zone = "{val}"'}

    if base_id in ("2.5.1", "2.5.2"):
        val, ln = get("inactivity-timeout")
        if val is None:
            return None
        minutes = int(val) if val.isdigit() else None
        ok = minutes is not None and 0 < minutes <= 10
        which = "CLI" if base_id == "2.5.1" else "WebUI"
        return {"status": "PASS" if ok else "FAIL", "line": ln,
                "evidence": f'administrator inactivity-timeout = "{val}" min (Quantum Spark uses one '
                            f'shared timeout for CLI and WebUI — evaluated here for {which})'}

    if base_id == "2.5.4":
        radius, ln1 = get("radius-auth")
        tacacs, ln2 = get("tacacs-auth")
        if radius is None and tacacs is None:
            return None
        ok = (radius or "").lower() == "enable" or (tacacs or "").lower() == "enable"
        return {"status": "PASS" if ok else "FAIL", "line": ln2 or ln1,
                "evidence": f'radius-auth={radius or "(unset)"}, tacacs-auth={tacacs or "(unset)"}'}

    if base_id == "2.5.5":
        entries = facts.get("admin-access-entries")
        if not entries:
            return None
        add_entries = [e for e in entries if e[0].startswith("add ")]
        any_open = any(re.search(r'"0\.0\.0\.0"', txt) for txt, _ln in add_entries)
        ok = bool(add_entries) and not any_open
        ln = add_entries[0][1] if add_entries else entries[0][1]
        evidence = f'{len(add_entries)} admin-access-ipv4-address entr{"y" if len(add_entries) == 1 else "ies"} configured'
        if any_open:
            evidence += " (includes 0.0.0.0 — effectively unrestricted)"
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": evidence}

    if base_id == "2.6.1":
        val, ln = get("audit-on-change")
        if val is None:
            return None
        ok = val.lower() == "true"
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": f'send-audit-on-db-change = "{val}"'}

    if base_id == "2.6.3":
        val, ln = get("display-audit-logs")
        if val is None:
            return None
        ok = val.lower() == "true"
        return {"status": "PASS" if ok else "FAIL", "line": ln, "evidence": f'display-audit-logs = "{val}"'}

    return None


# ══════════════════════════════════════════════════════════════════════════
class CheckpointParser:
    def __init__(self, config_file: "str | None", rules_csv: "str | None" = None):
        self.config_file = config_file
        self.rules_csv = rules_csv
        self.source_label = os.path.basename(config_file) if config_file else \
            (os.path.basename(rules_csv) if rules_csv else "(none)")

        self.gaia_lines: list[tuple[int, str]] = []
        self.security_rules: list[dict] = []
        self.system: dict[str, tuple[str, int]] = {}   # label -> (value, line#)
        self.dialect = "enterprise"     # or "spark" — see _detect_dialect()
        self.spark_facts: dict = {}
        self.issues: list[dict] = []
        self._seen_issues: set[tuple[str, str]] = set()
        self._native_ids: set[str] = set()   # benchmark IDs a native rulebase check already covered

        self.audits_tar_path = ""
        self.cis_l1_audit_used = ""
        self.cis_l2_audit_used = ""
        self.catalog: list[dict] = []   # every parsed CIS Check Point Firewall Benchmark item

    # ── Parse entry point ────────────────────────────────────────────────────
    def parse(self):
        if self.config_file is not None:
            self.gaia_lines = _load_gaia_config(self.config_file)
            self.dialect = _detect_dialect(self.gaia_lines)
            print(f"[+] Loaded {len(self.gaia_lines)} configuration statement(s) from "
                  f"{os.path.basename(self.config_file)} (dialect: {self.dialect})")
            if self.dialect == "spark":
                self.spark_facts = _parse_spark_facts(self.gaia_lines)
            self._parse_system_settings()
        if self.rules_csv:
            self.security_rules = _parse_rulebase_csv(self.rules_csv)
        if not self.config_file and not self.rules_csv:
            sys.exit("No Gaia config and no --rules-csv given — nothing to analyze.")

        if self.security_rules:
            self._run_rulebase_checks()
        if self.gaia_lines:
            self._run_cis_checks()

    # ── System settings (for the Gaia Configuration sheet) ──────────────────
    def _sys_set(self, label: str, regex: str, group: int = 0):
        pat = re.compile(regex)
        for ln, txt in self.gaia_lines:
            m = pat.search(txt)
            if m:
                self.system[label] = (m.group(group) if group else txt, ln)

    def _parse_system_settings(self):
        if self.dialect == "spark":
            self._parse_system_settings_spark()
            return
        self._sys_set("Hostname",              r'^set hostname (\S+)', 1)
        self._sys_set("Timezone",               r'^set timezone (.+)$', 1)
        self._sys_set("IPv6 State",             r'^set ipv6-state (\S+)', 1)
        self._sys_set("DNS Primary",            r'^set dns primary (\S+)', 1)
        self._sys_set("DNS Secondary",          r'^set dns secondary (\S+)', 1)
        self._sys_set("NTP Active",             r'^set ntp active (\S+)', 1)
        self._sys_set("Telnet",                 r'^set net-access telnet (\S+)', 1)
        self._sys_set("DHCP Server",            r'^set dhcp server')
        self._sys_set("SNMP Agent",             r'^set snmp agent (\S+)$', 1)
        self._sys_set("SNMP Version",           r'^set snmp agent-version (\S+)', 1)
        self._sys_set("Core Dump",              r'^set core-dump enable')
        self._sys_set("CLI Inactivity Timeout", r'^set inactivity-timeout (\S+)', 1)
        self._sys_set("Web Session Timeout",    r'^set web session-timeout (\S+)', 1)
        self._sys_set("Login Banner",           r'^set message banner on')
        self._sys_set("MOTD Banner",            r'^set message motd on')
        self._sys_set("Password Min Length",    r'^set password-controls min-password-length (\S+)', 1)
        self._sys_set("Password Complexity",    r'^set password-controls complexity (\S+)', 1)
        self._sys_set("Mgmt Audit Logs",        r'^set syslog mgmtauditlogs (\S+)', 1)
        self._sys_set("Audit Log Retention",    r'^set syslog auditlog (\S+)', 1)
        self._sys_set("RADIUS/TACACS+ State",   r'^set aaa tacacs-servers state (\S+)', 1)

    def _parse_system_settings_spark(self):
        """Same 'Gaia Configuration' sheet, populated from Quantum Spark's
        own facts instead of the enterprise-Gaia regexes above (which don't
        match Spark's command grammar at all)."""
        f = self.spark_facts

        def put(label, key):
            if key in f:
                val, ln = f[key]
                self.system[label] = (str(val), ln)

        put("Hostname", "hostname")
        put("Timezone", "timezone")
        put("DNS Primary", "dns-primary")
        put("DNS Secondary", "dns-secondary")
        put("NTP Active", "ntp-active")
        put("SNMP Agent", "snmp-agent")
        put("SNMP Version", "snmp-agent-version")
        put("CLI Inactivity Timeout", "inactivity-timeout")
        put("Web Session Timeout", "inactivity-timeout")
        put("Login Banner", "banner-state")
        put("MOTD Banner", "motd-state")
        put("Password Complexity", "password-complexity-level")
        put("Mgmt Audit Logs", "audit-on-change")
        if "radius-auth" in f or "tacacs-auth" in f:
            radius = f.get("radius-auth", ("", 0))
            tacacs = f.get("tacacs-auth", ("", 0))
            self.system["RADIUS/TACACS+ State"] = (
                f"radius={radius[0] or '(unset)'}, tacacs={tacacs[0] or '(unset)'}",
                tacacs[1] or radius[1])

    # ── Findings ──────────────────────────────────────────────────────────
    def _issue(self, severity, category, item_id, name, description, recommendation,
               details="", line="", cis_ids=None, pci_ids=None):
        key = (category, name)
        if key in self._seen_issues:
            return
        self._seen_issues.add(key)
        cis_ids = cis_ids or []
        pci_ids = pci_ids or []
        self.issues.append({
            "severity": severity, "category": category, "item_id": item_id, "rule_name": name,
            "line": str(line) if line else "", "description": description,
            "recommendation": recommendation, "details": details,
            "cis_controls": _cis_label(cis_ids), "cis_ids": cis_ids,
            "cis_benchmark": item_id, "pci_dss": _pci_label(pci_ids), "pci_ids": pci_ids,
        })

    # ── Native rulebase checks (need --rules-csv) ────────────────────────────
    def _catalog_text(self, item_id: str) -> tuple[str, str]:
        """(description, solution) for a benchmark ID, if the catalog has been
        loaded yet — falls back to empty strings when checked before _run_cis_checks."""
        for it in self.catalog:
            if it["id"] == item_id:
                return it["info"], it["solution"]
        return "", ""

    def _run_rulebase_checks(self):
        active = [r for r in self.security_rules if r["enabled"]]
        allow_active = [r for r in active if _is_allow(r["action"])]

        for r in self.security_rules:
            if not r["enabled"]:
                self._issue("LOW", "Disabled Rulebase Rule", "", r["name"],
                             f"Rule '{r['name']}' is disabled but still present in the rulebase.",
                             "Remove rules that are no longer needed instead of leaving them "
                             "disabled indefinitely — disabled rules accumulate and obscure intent.",
                             line=r["num"])
        for r in active:
            if not r["comments"].strip():
                self._issue("LOW", "Missing Rule Comment", "", r["name"],
                             f"Rule '{r['name']}' has no comment describing its business purpose.",
                             "Add a comment to every rule explaining why it exists and who owns it.",
                             line=r["num"])

        for r in allow_active:
            if _has_any(r["source"]):
                self._issue("HIGH", "Rule Allows Any Source", "3.6", r["name"],
                             f"Allow rule '{r['name']}' permits traffic from Any source.",
                             "Restrict the source to the specific networks/hosts that require access.",
                             details=f"Source: {r['source']}", line=r["num"])
            if _has_any(r["destination"]):
                self._issue("HIGH", "Rule Allows Any Destination", "3.5", r["name"],
                             f"Allow rule '{r['name']}' permits traffic to Any destination.",
                             "Restrict the destination to the specific networks/hosts/services required.",
                             details=f"Destination: {r['destination']}", line=r["num"])
            if _has_any(r["services"]):
                self._issue("HIGH", "Rule Allows Any Service", "3.7", r["name"],
                             f"Allow rule '{r['name']}' permits Any service/application.",
                             "Restrict the rule to the specific services/applications required.",
                             details=f"Services: {r['services']}", line=r["num"])
            track = r["track"].strip().lower()
            if track in ("", "none", "log disabled", "no log"):
                self._issue("MEDIUM", "Allow Rule Not Logging", "3.8", r["name"],
                             f"Allow rule '{r['name']}' does not have logging (Track) enabled.",
                             "Set Track to Log (or Detailed/Extended Log) on every allow rule.",
                             details=f"Track: {r['track'] or '(blank)'}", line=r["num"])
        self._native_ids |= {"3.5", "3.6", "3.7", "3.8"}

        # 3.2 — a default drop/cleanup rule should be the last rule in the base.
        if active:
            last = active[-1]
            is_cleanup = (_is_drop(last["action"]) and _has_any(last["source"])
                          and _has_any(last["destination"]) and _has_any(last["services"]))
            if not is_cleanup:
                desc, sol = self._catalog_text("3.2")
                self._issue("HIGH", "No Default Drop/Cleanup Rule", "3.2", "(rulebase)",
                             desc or "The last rule in the rulebase does not explicitly drop all "
                             "traffic not matched by an earlier rule.",
                             sol or "Add an explicit Any/Any/Any Drop rule (with logging) as the "
                             "final rule in the rulebase.",
                             details=f"Last rule: '{last['name']}' — action={last['action']}, "
                             f"source={last['source']}, destination={last['destination']}, "
                             f"services={last['services']}")
        self._native_ids.add("3.2")

        # 3.4 — Hit Count column present and populated for at least one rule.
        has_hits = any(r["hits"].strip() for r in active)
        if not has_hits:
            desc, sol = self._catalog_text("3.4")
            self._issue("LOW", "No Hit Count Data", "3.4", "(rulebase)",
                         desc or "No rule in the exported rulebase shows Hit Count data.",
                         sol or "Enable Hit Count under SmartConsole > Global Properties, and "
                         "review it periodically to find and remove unused rules.")
        self._native_ids.add("3.4")

    # ── CIS Check Point Firewall Benchmark checks (need the Gaia config) ────
    def _run_cis_checks(self):
        tar_path = _find_audits_tar()
        if not tar_path:
            print("[!] audits.tar.gz not found — skipping CIS Check Point Firewall Benchmark checks.")
            return
        self.audits_tar_path = tar_path

        members = {
            "L1": "portal_audits/CheckPoint_Compliance/CIS_Check_Point_Firewall_Level_1_v1.1.0.audit",
            "L2": "portal_audits/CheckPoint_Compliance/CIS_Check_Point_Firewall_Level_2_v1.1.0.audit",
        }
        raw_items: list[dict] = []
        with tarfile.open(tar_path, "r:*") as tf:
            for level, member in members.items():
                try:
                    f = tf.extractfile(member)
                except KeyError:
                    f = None
                if f is None:
                    print(f"[!] {member} not found inside {tar_path} — skipping {level} checks.")
                    continue
                content = f.read().decode("utf-8", errors="replace")
                raw_items.extend(_parse_checkpoint_audit(content, level))
                if level == "L1":
                    self.cis_l1_audit_used = os.path.basename(member)
                else:
                    self.cis_l2_audit_used = os.path.basename(member)

        self.catalog = raw_items

        groups: dict[str, list[dict]] = defaultdict(list)
        for it in raw_items:
            groups[it["id"]].append(it)

        def _sort_key(k):
            return [int(p) for p in k.split(".")]

        for base_id in sorted(groups.keys(), key=_sort_key):
            if base_id in self._native_ids:
                continue  # already given a definitive, evidence-based verdict from the rulebase CSV
            subitems = groups[base_id]
            title = subitems[0]["title"]
            # Multi-part checks share one title with a ' - <sub-label>' suffix
            # that differs per sub-item (e.g. '... - history-checking' /
            # '... - history-length') — the umbrella title is the shared prefix.
            umbrella_title = title.split(" - ", 1)[0] if len(subitems) > 1 else title
            level = subitems[0]["level"]
            info = subitems[0]["info"]
            solution = subitems[0]["solution"]
            cis_ids, pci_ids = _refs_from_reference(subitems[0]["reference"])
            severity = _severity_for(base_id)

            automated = [s for s in subitems
                         if s["kind"] == "custom_item" and s["regex"]
                         and s["expect"] != "Manual Review Required"]

            # Quantum Spark has its own evaluator for some IDs (SPARK_IMPLEMENTED_IDS)
            # that's independent of whether the *enterprise* audit could automate
            # them (2.5.5 is enterprise-manual but Spark-automatable, via
            # admin-access-ipv4-address) — so this has to run before the
            # "not automated -> manual review" short-circuit below, and it takes
            # over entirely for those IDs regardless of the enterprise classification.
            if self.dialect == "spark" and base_id in SPARK_IMPLEMENTED_IDS:
                self._emit_spark_check(base_id, umbrella_title, level, info, solution,
                                        severity, cis_ids, pci_ids)
                continue

            if not automated:
                # Manual-review / conditionally-evaluated item — the benchmark
                # (or Nessus itself) can't determine this from a static config
                # capture alone, on either platform.
                self._issue(severity,
                             f"CIS {base_id} — Manual Review Required", base_id,
                             f"[{level}] {umbrella_title}",
                             info or "This control requires manual review — it isn't derivable "
                             "from a static Gaia configuration capture (it lives in the SmartConsole "
                             "database / Global Properties, or needs judgment calls Nessus itself "
                             "doesn't automate either).",
                             solution or "Review the Check Point CIS Benchmark control by hand in "
                             "SmartConsole.",
                             cis_ids=cis_ids, pci_ids=pci_ids)
                continue

            if self.dialect == "spark":
                # Enterprise-automatable, but no Spark equivalent implemented —
                # the enterprise regex would just false-FAIL on Spark's syntax.
                self._issue("INFO", f"CIS {base_id} — Not Applicable (Quantum Spark)", base_id,
                             f"[{level}] {umbrella_title}",
                             "This Check Point CIS Benchmark control has no equivalent setting "
                             "implemented for Quantum Spark / SMB-series appliances (Gaia Embedded) "
                             "in this analyzer — the benchmark is written for enterprise Gaia. "
                             + (info or ""),
                             solution or "Not applicable to this platform.",
                             cis_ids=cis_ids, pci_ids=pci_ids)
                continue

            failures = []
            fail_lines, pass_lines = [], []
            var_note = False
            for s in automated:
                result = _eval_config_check(s, self.gaia_lines)
                if result["had_vars"]:
                    var_note = True
                if result["status"] == "FAIL":
                    sub_label = s["title"].split(" - ", 1)[-1] if " - " in s["title"] else ""
                    ev = f"'{result['text']}' (line {result['line']})" if result["text"] else "not configured"
                    failures.append(f"{sub_label + ': ' if sub_label else ''}{ev}")
                    if result["line"]:
                        fail_lines.append(result["line"])
                elif result["line"]:
                    pass_lines.append(result["line"])

            if failures:
                desc = info or f"{umbrella_title} — one or more required settings are not configured correctly."
                details = "; ".join(failures)
                if var_note:
                    details += "  [expected value depends on your environment — verify against your DNS/NTP/AAA/SNMP/timezone/banner text]"
                self._issue(severity, f"CIS {base_id} — {umbrella_title}", base_id,
                             f"[{level}] {umbrella_title}", desc,
                             solution or "See the CIS Check Point Firewall Benchmark for remediation steps.",
                             details=details, line=", ".join(str(l) for l in fail_lines),
                             cis_ids=cis_ids, pci_ids=pci_ids)
            elif var_note:
                # Passed, but the expected value was a site-specific variable —
                # surface an informational note so it gets a human look.
                self._issue("INFO", f"CIS {base_id} — Verify Environment-Specific Value", base_id,
                             f"[{level}] {umbrella_title}",
                             f"{umbrella_title} is configured — confirm the configured value "
                             "(DNS/NTP/AAA server, timezone, or banner text) is actually correct "
                             "for your environment.",
                             "No action needed if the value shown is correct.",
                             line=", ".join(str(l) for l in pass_lines),
                             cis_ids=cis_ids, pci_ids=pci_ids)

    def _emit_spark_check(self, base_id, umbrella_title, level, info, solution,
                           severity, cis_ids, pci_ids):
        """Evaluate one benchmark ID against Quantum Spark facts instead of
        the enterprise-Gaia regex engine, and emit the matching finding."""
        result = _eval_spark_check(base_id, self.spark_facts)
        if result is None:
            self._issue("INFO", f"CIS {base_id} — Not Applicable (Quantum Spark)", base_id,
                         f"[{level}] {umbrella_title}",
                         "This Check Point CIS Benchmark control has no equivalent setting on a "
                         "Quantum Spark / SMB-series appliance (Gaia Embedded) — the benchmark is "
                         "written for enterprise Gaia. " + (info or ""),
                         solution or "Not applicable to this platform.",
                         cis_ids=cis_ids, pci_ids=pci_ids)
            return
        if result["status"] == "FAIL":
            self._issue(severity, f"CIS {base_id} — {umbrella_title}", base_id,
                         f"[{level}] {umbrella_title}",
                         info or f"{umbrella_title} — not configured correctly on this Quantum Spark appliance.",
                         solution or "See the CIS Check Point Firewall Benchmark for remediation steps "
                         "(adapted here to this platform's own settings — see Details).",
                         details=result["evidence"], line=str(result["line"]) if result["line"] else "",
                         cis_ids=cis_ids, pci_ids=pci_ids)


# ══════════════════════════════════════════════════════════════════════════
class ExcelReporter:
    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"], C["critical_l"]),
        "HIGH":     (C["high"],     C["high_l"]),
        "MEDIUM":   (C["medium"],   C["medium_l"]),
        "LOW":      (C["low"],      C["low_l"]),
        "INFO":     (C["info"],     C["info_l"]),
    }

    def __init__(self, parser: CheckpointParser, output_file: str):
        self.p = parser
        self.out = output_file
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self._vulns = _find_vulns_file()

    # ── Sheet helpers ────────────────────────────────────────────────────────
    def _hdr(self, ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = _fill(C["hdr_bg"])
            c.font = _font(bold=True, color=C["hdr_fg"])
            c.alignment = _align("center", wrap=False)
            c.border = THIN
        ws.row_dimensions[row].height = 28

    def _row_fill(self, row_idx, disabled=False):
        if disabled:
            return C["info_l"]
        return C["alt_row"] if row_idx % 2 == 0 else None

    def _set_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Summary ───────────────────────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Check Point Firewall — Configuration Security Report"
        c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        c.fill = _fill(C["hdr_bg"])
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 42

        ws.merge_cells("A2:F2")
        c = ws["A2"]
        c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    "
                   f"Source: {self.p.source_label}")
        c.font = _font(italic=True, color="595959", size=9)
        c.fill = _fill("F2F2F2")
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 18

        row = 4
        p = self.p

        def section_header(label):
            nonlocal row
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _font(bold=True, color=C["hdr_bg"], size=12)
            row += 1

        def kv(label, value):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font()
            c1.alignment = _align()
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = _font(bold=True)
            c2.alignment = _align("center")
            if row % 2 == 0:
                c1.fill = c2.fill = _fill(C["alt_row"])
            row += 1

        section_header("CONFIGURATION OVERVIEW")
        kv("Gaia config supplied", "yes" if p.config_file else "no")
        if p.config_file:
            kv("Gaia config statements parsed", len(p.gaia_lines))
            dialect_label = {"enterprise": "Enterprise Gaia",
                              "spark": "Quantum Spark / SMB (Gaia Embedded)"}[p.dialect]
            kv("Detected config dialect", dialect_label)
            kv("audits.tar.gz used", p.audits_tar_path or "not found")
            kv("CIS Benchmark — L1 .audit", p.cis_l1_audit_used or "n/a")
            kv("CIS Benchmark — L2 .audit", p.cis_l2_audit_used or "n/a")
        else:
            kv("CIS Benchmark checks", "skipped — no Gaia config supplied")
        kv("Rules CSV supplied", "yes" if p.rules_csv else "no")
        if p.rules_csv:
            kv("Rulebase rows", len(p.security_rules))
            kv("  Active", sum(1 for r in p.security_rules if r["enabled"]))
            kv("  Disabled", sum(1 for r in p.security_rules if not r["enabled"]))
            kv("  Allow", sum(1 for r in p.security_rules if _is_allow(r["action"])))
        row += 1

        section_header("SECURITY FINDINGS BY SEVERITY")
        self._hdr(ws, ["Severity", "Count"], row=row)
        row += 1
        total = 0
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            cnt = sum(1 for i in p.issues if i["severity"] == sev)
            total += cnt
            fg, bg = self.SEV_COLORS[sev]
            c1 = ws.cell(row=row, column=1, value=sev)
            c1.fill = _fill(bg); c1.font = _font(bold=True, color=fg)
            c1.alignment = _align("center"); c1.border = THIN
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.fill = _fill(bg); c2.font = _font(bold=True, color=fg)
            c2.alignment = _align("center"); c2.border = THIN
            row += 1
        c1 = ws.cell(row=row, column=1, value="TOTAL")
        c1.font = _font(bold=True); c1.border = THIN
        c2 = ws.cell(row=row, column=2, value=total)
        c2.font = _font(bold=True); c2.border = THIN
        c2.alignment = _align("center")
        row += 2

        section_header("FINDINGS BY CATEGORY")
        self._hdr(ws, ["Category", "Count"], row=row)
        row += 1
        cat_counts: dict[str, int] = defaultdict(int)
        for iss in p.issues:
            cat_counts[iss["category"]] += 1
        for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat).font = _font()
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.font = _font(bold=True)
            c2.alignment = _align("center")
            if row % 2 == 0:
                ws.cell(row=row, column=1).fill = _fill(C["alt_row"])
                c2.fill = _fill(C["alt_row"])
            row += 1

        self._set_widths(ws, [46, 15, 15, 15, 15, 15])

    # ── Gaia Configuration ───────────────────────────────────────────────────
    def _sheet_gaia_config(self):
        ws = self.wb.create_sheet("Gaia Configuration")
        ws.sheet_view.showGridLines = False
        if not self.p.config_file:
            ws["A1"] = "No Gaia 'show configuration' capture was supplied."
            ws["A1"].font = _font(italic=True, color=C["info"])
            return
        headers = ["Setting", "Value / Config Line", "Line #"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        row = 2
        for label, (value, ln) in self.p.system.items():
            row_bg = self._row_fill(row)
            for col, val in enumerate([label, value, ln or ""], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                c.font = _font(bold=(col == 1))
                c.alignment = _align("center" if col == 3 else "left")
                if row_bg:
                    c.fill = _fill(row_bg)
            row += 1
        self._set_widths(ws, [28, 60, 10])

    # ── Security rules ────────────────────────────────────────────────────────
    def _sheet_security_rules(self):
        ws = self.wb.create_sheet("Security Rules")
        ws.sheet_view.showGridLines = False
        if not self.p.rules_csv:
            ws["A1"] = ("No rulebase CSV was supplied — export the Rule Base from SmartConsole "
                        "(right-click the rule base > Export to CSV) and pass it via --rules-csv "
                        "to populate this sheet and the rulebase-shaped CIS Benchmark checks "
                        "(3.2, 3.4–3.8).")
            ws["A1"].font = _font(italic=True, color=C["info"])
            ws.column_dimensions["A"].width = 110
            return
        headers = ["#", "Name", "Status", "Source", "Destination", "VPN", "Services & Applications",
                   "Action", "Track", "Install On", "Time", "Hits", "Comments"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for i, r in enumerate(self.p.security_rules, 1):
            row = i + 1
            disabled = not r["enabled"]
            row_bg = self._row_fill(row, disabled)
            action = (r["action"] or "").strip().lower()
            action_bg = C["allow_l"] if _is_allow(action) else (C["deny_l"] if _is_drop(action) else row_bg)
            action_fg = C["allow"] if _is_allow(action) else (C["deny"] if _is_drop(action) else "000000")

            values = [r["num"], r["name"], "Disabled" if disabled else "Active",
                      r["source"], r["destination"], r["vpn"], r["services"],
                      r["action"].upper(), r["track"], r["install_on"], r["time"], r["hits"], r["comments"]]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 8:
                    c.fill = _fill(action_bg)
                    c.font = _font(bold=True, color=action_fg)
                    c.alignment = _align("center")
                elif col in (1, 3):
                    c.font = _font(bold=(col == 1))
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                else:
                    c.font = _font()
                    c.alignment = _align()
                    if row_bg:
                        c.fill = _fill(row_bg)
            ws.row_dimensions[row].height = 30
        self._set_widths(ws, [5, 24, 10, 26, 26, 12, 30, 10, 12, 16, 12, 10, 34])

    # ── Security Issues ──────────────────────────────────────────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False
        headers = ["#", "Validated", "Severity", "Residual Risk", "Residual Risk Note",
                   "Category", "Rule / Object", "Config Line(s)",
                   "CIS Controls v7", "CIS Benchmark ID", "PCI DSS",
                   "Description", "Recommendation", "Details",
                   "Asset", "Target", "Vuln", "Source"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        source = self.p.source_label
        asset = self.p.source_label                                # the config/CSV filename
        hostname = self.p.system.get("Hostname", ("", 0))[0]       # the firewall's hostname
        sorted_issues = sorted(self.p.issues, key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            row_bg = self._row_fill(row)
            vuln_id = CATEGORY_VULN_ID.get(iss["category"]) or BENCHMARK_ID_VULN_ID.get(iss.get("item_id", ""))
            vuln = self._vulns.get(vuln_id, "") if vuln_id else ""
            details = iss.get("details", "")
            validated = "N" if "Manual Review Required" in iss["category"] else "Y"
            line = iss.get("line", "")
            if hostname:
                target = f"{hostname} ({line})" if line else hostname
            else:
                # No Gaia config to read a hostname from (e.g. an orphaned
                # rules-only CSV run) — fall back to the finding's own
                # rule/object and config line, same shape as pa_analyzer.py's
                # Target column.
                target = f"{iss['rule_name']} ({line})" if line else iss["rule_name"]

            values = [idx, validated, sev, "", iss["category"],
                      iss["category"], iss["rule_name"], iss.get("line", ""),
                      iss.get("cis_controls", ""), iss.get("cis_benchmark", ""), iss.get("pci_dss", ""),
                      iss["description"], iss["recommendation"], details,
                      asset, target, vuln, source]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 3:
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg); c.alignment = _align("center")
                elif col in (1, 8):
                    c.font = _font(bold=(col == 1)); c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 9:
                    c.font = _font(bold=True, color="17375E", size=9); c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 11:
                    c.font = _font(bold=True, color="7B2D8B", size=9); c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 2:
                    c.font = _font(bold=True); c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                else:
                    c.font = _font(); c.alignment = _align()
                    if row_bg:
                        c.fill = _fill(row_bg)
            ws.row_dimensions[row].height = 40
        self._set_widths(ws, [4, 10, 12, 14, 26, 40, 24, 12, 18, 16, 16, 55, 55, 50, 22, 18, 44, 26])

    # ── Ticketing Export ──────────────────────────────────────────────────────
    def _sheet_export(self):
        ws = self.wb.create_sheet("Export")
        ws.sheet_view.showGridLines = False
        headers = ["Hostname", "line#", "protocol", "port", "output"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        hostname = self.p.system.get("Hostname", ("", 0))[0]
        sorted_issues = sorted(self.p.issues, key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            output_text = f"{iss['rule_name']}\n{iss['description']}\n{iss['recommendation']}"
            values = [hostname, iss.get("line", ""), "tcp", 0, output_text]
            row_bg = self._row_fill(row)
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                c.font = _font()
                c.alignment = _align("left" if col == 5 else "center", wrap=(col == 5))
                if row_bg:
                    c.fill = _fill(row_bg)
            ws.row_dimensions[row].height = 60
        self._set_widths(ws, [24, 10, 10, 8, 80])

    # ── CIS Controls v7 Mapping ──────────────────────────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS Controls Mapping")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v7 — Finding Cross-Reference"
        t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = ("Each CIS Control lists all findings from this config that map to it "
                   "(per the CIS Check Point Firewall Benchmark's own reference data).")
        s.font = _font(italic=True, color=C["info"], size=9)
        s.fill = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        ctrl_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        def _sort_key(k):
            parts = k.split(".")
            return (int(parts[0]), float("0." + parts[1]) if len(parts) > 1 else 0)

        row = 4
        for ctrl_id in sorted(CIS_CTRL_DESC.keys(), key=_sort_key):
            ctrl_desc = CIS_CTRL_DESC[ctrl_id]
            issues_for_ctrl = sorted(ctrl_issues.get(ctrl_id, []),
                                      key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1, value=f"CIS {ctrl_id} — {ctrl_desc}")
            hc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill = _fill("17375E")
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 22
            row += 1
            row = self._mapping_rows(ws, row, issues_for_ctrl)
            row += 1
        self._set_widths(ws, [12, 40, 24, 14, 55, 55])

    # ── PCI DSS v4.0 Mapping ─────────────────────────────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = "Each PCI DSS v4.0 requirement lists all findings from this config that map to it."
        s.font = _font(italic=True, color=C["info"], size=9)
        s.fill = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        req_issues: dict[str, list[dict]] = defaultdict(list)
        for iss in self.p.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        row = 4
        for req_id in sorted(PCI_DSS_DESC.keys(), key=lambda x: [int(p) for p in x.split(".")]):
            desc = PCI_DSS_DESC[req_id]
            issues_for_req = sorted(req_issues.get(req_id, []),
                                     key=lambda x: self.SEV_ORDER.get(x["severity"], 9))
            count = len(issues_for_req)
            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1
            row = self._mapping_rows(ws, row, issues_for_req)
            row += 1
        self._set_widths(ws, [12, 40, 24, 14, 55, 55])

    def _mapping_rows(self, ws, row, issues_for):
        if not issues_for:
            ws.merge_cells(f"A{row}:F{row}")
            nc = ws.cell(row=row, column=1, value="No findings for this control/requirement")
            nc.font = _font(italic=True, color=C["info"])
            nc.fill = _fill("F9F9F9")
            nc.alignment = _align()
            nc.border = THIN
            return row + 1

        sub_hdrs = ["Severity", "Category", "Rule / Object", "Config Line(s)", "Description", "Recommendation"]
        for col, h in enumerate(sub_hdrs, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = _font(bold=True, color="FFFFFF")
            c.fill = _fill("2E4057")
            c.alignment = _align("center", wrap=False)
            c.border = THIN
        ws.row_dimensions[row].height = 20
        row += 1
        for iss in issues_for:
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            rb = C["alt_row"] if row % 2 == 0 else None
            vals = [sev, iss["category"], iss["rule_name"], iss.get("line", ""),
                    iss["description"], iss["recommendation"]]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 1:
                    c.fill = _fill(bg); c.font = _font(bold=True, color=fg); c.alignment = _align("center")
                elif col == 4:
                    c.font = _font(color=C["info"], size=9); c.alignment = _align("center")
                    if rb:
                        c.fill = _fill(rb)
                else:
                    c.font = _font(); c.alignment = _align()
                    if rb:
                        c.fill = _fill(rb)
            ws.row_dimensions[row].height = 36
            row += 1
        return row

    # ── Save ──────────────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_gaia_config()
        self._sheet_security_rules()
        self._sheet_issues()
        self._sheet_export()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)


# ── Directory (batch) mode ───────────────────────────────────────────────────
_BATCH_SKIP_EXT = {".xlsx", ".xls", ".pyc", ".zip", ".tar", ".gz"}


def _looks_like_gaia_config(path: str) -> bool:
    """Sniff a file's content (not just its extension) for Gaia clish
    statements, so a directory scan doesn't have to guess from filenames
    alone. Requires a few matches, not just one, to avoid false positives
    on unrelated text files that happen to contain a stray 'set ...' line."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            sample = fh.read(200_000)
    except (OSError, UnicodeDecodeError):
        return False
    sample = _ANSI_CSI_RE.sub("", sample)
    return len(re.findall(r'(?m)^\s*(?:set|add|delete)\s+\S', sample)) >= 3


# Descriptive tokens that commonly differ between a gateway's config capture
# and its rulebase export even when they're clearly "the same gateway" —
# e.g. 'gw01_show_configuration.txt' vs 'gw01_rules.csv'. Stripped out before
# comparing filenames so what's left is (ideally) just the gateway identifier.
_PAIR_NOISE_TOKENS = {
    "show", "configuration", "config", "cfg", "running", "merged", "all",
    "rules", "rulebase", "ruleset", "policy", "export", "csv", "smartconsole",
    "gaia", "clish", "gw", "gateway", "fw", "firewall",
}


def _pair_sig(path: str) -> str:
    """A filename 'signature' for pairing a config to its rulebase CSV:
    lowercased, split on non-alnum, common descriptive tokens dropped."""
    stem = os.path.splitext(os.path.basename(path))[0].lower()
    tokens = [t for t in re.split(r'[^a-z0-9]+', stem) if t and t not in _PAIR_NOISE_TOKENS]
    return "".join(tokens)


def _scan_directory(directory: str, recursive: bool) -> tuple[list[str], list[str]]:
    """Return (config_paths, csv_paths) found under directory."""
    configs, csvs = [], []
    if recursive:
        walk = os.walk(directory)
    else:
        walk = [(directory, [], sorted(os.listdir(directory)))]
    for root, _dirs, files in walk:
        for fname in sorted(files):
            if fname.startswith("."):
                continue
            path = os.path.join(root, fname)
            if not os.path.isfile(path):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext in _BATCH_SKIP_EXT:
                continue
            if ext == ".csv":
                csvs.append(path)
            elif _looks_like_gaia_config(path):
                configs.append(path)
    return configs, csvs


def _pair_rules_csv(config_path: str, csvs: list[str]) -> "str | None":
    """Best-effort match of a config file to its rulebase CSV by filename
    signature (e.g. 'gw01_show_configuration.txt' <-> 'gw01_rules.csv' both
    reduce to 'gw01' once descriptive words are stripped — see _pair_sig).
    Returns None when nothing lines up unambiguously — missing or ambiguous
    pairings are left for the user to run by hand rather than guessed at."""
    csig = _pair_sig(config_path)
    if not csig:
        return None
    sig_pairs = [(c, _pair_sig(c)) for c in csvs]
    sig_pairs = [(c, s) for c, s in sig_pairs if s]  # an empty signature (e.g. a bare
    #                                                   'rules.csv') can't be tied to any one config

    exact = [c for c, s in sig_pairs if s == csig]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return None  # more than one CSV reduces to the same signature — ambiguous

    prefix = [c for c, s in sig_pairs if s.startswith(csig) or csig.startswith(s)]
    if len(prefix) == 1:
        return prefix[0]
    return None


def _run_one(config_file: "str | None", rules_csv: "str | None", output_path: str) -> "CheckpointParser | None":
    """Parse one config (+ optional rulebase CSV) and write its Excel report.
    Returns the parser on success, None on a handled per-file failure (so
    batch mode can keep going instead of aborting the whole run)."""
    label = config_file or rules_csv
    print(f"[*] Parsing: {label}")
    try:
        parser = CheckpointParser(config_file, rules_csv=rules_csv)
        parser.parse()
    except SystemExit as exc:
        print(f"[!] Skipped {label}: {exc}")
        return None

    sev_counts: dict[str, int] = defaultdict(int)
    for iss in parser.issues:
        sev_counts[iss["severity"]] += 1

    print("[*] Parsed:")
    print(f"      Gaia config statements: {len(parser.gaia_lines)}")
    print(f"      Rulebase rows         : {len(parser.security_rules)}")
    print(f"[*] Security findings: {len(parser.issues)}")
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        if sev_counts[sev]:
            print(f"      {sev:<10}: {sev_counts[sev]}")

    reporter = ExcelReporter(parser, output_path)
    reporter.save()
    print(f"[+] Saved: {output_path}")
    return parser


def _run_batch(directory: str, output_dir: "str | None", recursive: bool,
                shared_rules_csv: "str | None"):
    configs, csvs = _scan_directory(directory, recursive)
    if not configs and not csvs:
        sys.exit(f"No Gaia 'show configuration' captures or rulebase CSVs found under {directory}")

    print(f"[*] Found {len(configs)} config file(s) and {len(csvs)} CSV file(s) under {directory}"
          f"{' (recursive)' if recursive else ''}")

    results = []
    used_csvs: set[str] = set()
    for config_file in configs:
        if shared_rules_csv:
            rules_csv = shared_rules_csv
        else:
            rules_csv = _pair_rules_csv(config_file, csvs)
            if rules_csv:
                print(f"[+] Paired {os.path.basename(config_file)}  <->  {os.path.basename(rules_csv)}")
                used_csvs.add(rules_csv)

        out_dir = output_dir or os.path.dirname(config_file) or "."
        os.makedirs(out_dir, exist_ok=True)
        stem = os.path.splitext(os.path.basename(config_file))[0]
        output_path = os.path.join(out_dir, f"{stem}_analysis.xlsx")

        parser = _run_one(config_file, rules_csv, output_path)
        results.append((config_file, parser))
        print()

    # CSVs with no matching (or shared) config — e.g. a directory that's
    # entirely rulebase exports, or one CSV nothing else paired with — run
    # each standalone in rules-only mode, same as a bare .csv positional in
    # single-file mode, rather than silently skipping them.
    if not shared_rules_csv:
        for csv_file in csvs:
            if csv_file in used_csvs:
                continue
            out_dir = output_dir or os.path.dirname(csv_file) or "."
            os.makedirs(out_dir, exist_ok=True)
            stem = os.path.splitext(os.path.basename(csv_file))[0]
            output_path = os.path.join(out_dir, f"{stem}_analysis.xlsx")

            parser = _run_one(None, csv_file, output_path)
            results.append((csv_file, parser))
            print()

    ok = sum(1 for _, p in results if p is not None)
    print(f"[*] Batch complete: {ok}/{len(results)} report(s) written.")
    failed = [c for c, p in results if p is None]
    if failed:
        print("[!] Failed:")
        for c in failed:
            print(f"      {c}")


# ── CLI entry point ───────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Check Point Firewall Config Analyzer — outputs Excel report(s)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python checkpoint_analyzer.py "show configuration.txt"
  python checkpoint_analyzer.py "show configuration.txt" --rules-csv rulebase.csv
  python checkpoint_analyzer.py "show configuration.txt" -o audit-$(date +%%Y%%m%%d).xlsx
  python checkpoint_analyzer.py rulebase.csv
  python checkpoint_analyzer.py ./configs/                       # batch mode
  python checkpoint_analyzer.py ./configs/ --output-dir ./audits/
  python checkpoint_analyzer.py ./configs/ --recursive
""",
    )
    ap.add_argument("config",
                     help="A Gaia 'show configuration' (or 'show configuration all') clish CLI "
                          "text capture — e.g. a PuTTY/SecureCRT session log from the gateway or "
                          "management server — OR a directory containing multiple such captures "
                          "and/or rulebase CSVs, which runs every one of them and writes one Excel "
                          "report per config or per unpaired CSV ('batch mode'; a directory with "
                          "only CSVs runs each in rules-only mode). On its own, with no rulebase "
                          "CSV, only the CIS Check Point Firewall Benchmark (Gaia OS-level) checks "
                          "run — the access rulebase lives on the Management Server, not in Gaia's "
                          "config. A single '.csv' file passed here (not a directory) is routed to "
                          "--rules-csv automatically and runs rulebase checks only.")
    ap.add_argument("-o", "--output", default=None,
                     help="Output Excel file for a single config (default: "
                          "<config-stem>_analysis.xlsx). Ignored in batch mode — use "
                          "--output-dir instead.")
    ap.add_argument("--rules-csv", default=None,
                     help="SmartConsole Rule Base 'Export to CSV' (No., Name, Source, "
                          "Destination, VPN, Services & Applications, Action, Track, Install On, "
                          "Time, Comments, ...). Populates the Security Rules sheet and the "
                          "rulebase-shaped CIS Benchmark checks (3.2, 3.4–3.8). In batch mode, "
                          "if given, the SAME CSV is applied to every config found — omit it and "
                          "each config is auto-paired with a same-named CSV in the directory "
                          "instead (e.g. 'gw01.txt' <-> 'gw01.csv' / 'gw01_rules.csv'), when one "
                          "exists unambiguously.")
    ap.add_argument("--output-dir", default=None,
                     help="Batch mode only: directory to write the *_analysis.xlsx reports into "
                          "(default: next to each config file). Created if it doesn't exist.")
    ap.add_argument("--recursive", action="store_true",
                     help="Batch mode only: also scan subdirectories for config/CSV files.")
    args = ap.parse_args()

    if os.path.isdir(args.config):
        if args.output:
            sys.exit("--output isn't used in batch mode (one directory -> many reports) — "
                      "use --output-dir instead.")
        _run_batch(args.config, args.output_dir, args.recursive, args.rules_csv)
        return

    config_file = args.config
    rules_csv = args.rules_csv
    if config_file.lower().endswith(".csv"):
        if rules_csv and os.path.abspath(rules_csv) != os.path.abspath(config_file):
            sys.exit("Both a .csv positional and --rules-csv were given — pass the Gaia config "
                      "as the positional and the CSV via --rules-csv.")
        rules_csv, config_file = config_file, None

    output_path = args.output
    if not output_path:
        stem = os.path.splitext(os.path.basename(args.config))[0]
        output_path = f"{stem}_analysis.xlsx"

    _run_one(config_file, rules_csv, output_path)


if __name__ == "__main__":
    main()
