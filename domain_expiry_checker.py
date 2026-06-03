#!/usr/bin/env python3
"""
Domain Expiry Checker with VirusTotal Reputation Lookup

Usage:
    python domain_expiry_checker.py -d domains.txt -k YOUR_VT_API_KEY
    python domain_expiry_checker.py -d domains.txt -k YOUR_VT_API_KEY --days 180 --output report.json
    echo "example.com" | python domain_expiry_checker.py -k YOUR_VT_API_KEY

Requires:
    pip install python-whois requests
"""

import argparse
import csv
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import whois
except ImportError:
    whois = None

try:
    import requests
except ImportError:
    print("Missing dependency: pip install requests", file=sys.stderr)
    sys.exit(1)


VT_API_BASE = "https://www.virustotal.com/api/v3"
RDAP_BASE = "https://rdap.org/domain"
# Pause between VirusTotal requests
VT_REQUEST_DELAY = 30


def _parse_expiry_date(date_str: str) -> Optional[datetime]:
    """Parse ISO 8601 or common WHOIS date strings into a UTC-aware datetime."""
    if not date_str:
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(date_str[:26], fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _expiry_from_rdap(domain: str) -> tuple[Optional[datetime], Optional[str]]:
    """Query RDAP over HTTPS and return (expiry_datetime, error_string). Retries up to 3 times."""
    last_err = ""
    for attempt in range(3):
        try:
            resp = requests.get(
                f"{RDAP_BASE}/{domain}",
                timeout=20,
                headers={"Accept": "application/rdap+json"},
                allow_redirects=True,
            )
            if resp.status_code == 404:
                return None, "Domain not found in RDAP (may be unregistered or unsupported TLD)"
            resp.raise_for_status()
            data = resp.json()
            for event in data.get("events", []):
                if event.get("eventAction") == "expiration":
                    dt = _parse_expiry_date(event.get("eventDate", ""))
                    if dt:
                        return dt, None
            return None, "No expiration event in RDAP response"
        except requests.Timeout:
            last_err = "timed out"
            time.sleep(2 ** attempt)
        except requests.RequestException as exc:
            last_err = str(exc)
            time.sleep(2 ** attempt)
    return None, f"RDAP failed after 3 attempts: {last_err}"


def _expiry_from_whois(domain: str) -> tuple[Optional[datetime], Optional[str]]:
    """Fall back to python-whois and return (expiry_datetime, error_string)."""
    if whois is None:
        return None, "python-whois not installed (pip install python-whois)"
    try:
        w = whois.whois(domain)
        exp = w.expiration_date
        if exp is None:
            return None, "No expiration date in WHOIS response"
        if isinstance(exp, list):
            exp = exp[0]
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        return exp, None
    except Exception as exc:
        return None, f"WHOIS error: {exc}"


def get_expiry_info(domain: str) -> dict:
    """Return expiry status for a domain, trying RDAP first then WHOIS."""
    result = {
        "domain": domain,
        "expiry_date": None,
        "expired": False,
        "days_since_expiry": None,
        "whois_error": None,
        "lookup_method": None,
    }

    exp, err = _expiry_from_rdap(domain)
    if exp:
        result["lookup_method"] = "RDAP"
    else:
        rdap_err = err
        exp, err = _expiry_from_whois(domain)
        if exp:
            result["lookup_method"] = "WHOIS"
        else:
            result["whois_error"] = f"RDAP: {rdap_err} | WHOIS: {err}"
            return result

    now = datetime.now(timezone.utc)
    result["expiry_date"] = exp.isoformat()
    if exp < now:
        result["expired"] = True
        result["days_since_expiry"] = (now - exp).days

    return result


def query_virustotal(domain: str, api_key: str) -> dict:
    """Query VirusTotal domain report endpoint."""
    headers = {"x-apikey": api_key}
    url = f"{VT_API_BASE}/domains/{domain}"

    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 404:
            return {"vt_error": "Domain not found in VirusTotal"}
        if resp.status_code == 401:
            return {"vt_error": "Invalid VirusTotal API key"}
        if resp.status_code == 429:
            return {"vt_error": "VirusTotal rate limit exceeded — slow down requests"}
        resp.raise_for_status()
        data = resp.json().get("data", {}).get("attributes", {})

        stats = data.get("last_analysis_stats", {})
        total_engines = sum(stats.values()) if stats else 0
        malicious = stats.get("malicious", 0)
        suspicious = stats.get("suspicious", 0)

        # Build a flat summary of per-engine verdicts (non-clean only)
        flagged_engines = {}
        for engine, result in data.get("last_analysis_results", {}).items():
            if result.get("category") not in ("harmless", "undetected", "clean"):
                flagged_engines[engine] = {
                    "category": result.get("category"),
                    "result": result.get("result"),
                }

        return {
            "reputation": data.get("reputation"),
            "categories": data.get("categories", {}),
            "tags": data.get("tags", []),
            "last_analysis_date": (
                datetime.fromtimestamp(
                    data["last_analysis_date"], tz=timezone.utc
                ).isoformat()
                if data.get("last_analysis_date")
                else None
            ),
            "analysis_stats": stats,
            "total_engines": total_engines,
            "malicious_count": malicious,
            "suspicious_count": suspicious,
            "flagged_engines": flagged_engines,
            "creation_date": data.get("creation_date"),
            "whois": data.get("whois", "")[:500] if data.get("whois") else None,
            "registrar": data.get("registrar"),
            "country": data.get("country"),
            "popularity_ranks": data.get("popularity_ranks", {}),
        }

    except requests.RequestException as exc:
        return {"vt_error": str(exc)}


def risk_label(vt: dict) -> str:
    """Derive a simple risk label from VirusTotal results."""
    if "vt_error" in vt:
        return "UNKNOWN"
    malicious = vt.get("malicious_count", 0)
    suspicious = vt.get("suspicious_count", 0)
    reputation = vt.get("reputation", 0) or 0
    if malicious >= 3 or reputation < -10:
        return "HIGH RISK"
    if malicious >= 1 or suspicious >= 2 or reputation < 0:
        return "SUSPICIOUS"
    return "CLEAN"


def print_report(results: list[dict], verbose: bool = False) -> None:
    """Print a human-readable report to stdout."""
    sep = "=" * 70
    print(f"\n{sep}")
    print(f"  DOMAIN EXPIRY & REPUTATION REPORT  —  {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")
    print(sep)

    expired_count = sum(1 for r in results if r.get("expired"))
    checked_vt = sum(1 for r in results if r.get("virustotal"))
    print(f"  Domains checked : {len(results)}")
    print(f"  Expired         : {expired_count}")
    print(f"  VirusTotal hits : {checked_vt}")
    print(sep)

    for r in results:
        domain = r["domain"]
        print(f"\n  Domain  : {domain}")

        if r.get("whois_error"):
            print(f"  Lookup  : ERROR — {r['whois_error']}")
        elif r.get("expired"):
            method = r.get("lookup_method", "?")
            print(f"  Status  : EXPIRED  ({r['days_since_expiry']} days ago — {r['expiry_date']})  [{method}]")
        else:
            method = r.get("lookup_method", "?")
            print(f"  Status  : Active  (expires {r['expiry_date']})  [{method}]")

        vt = r.get("virustotal")
        if vt:
            if "vt_error" in vt:
                print(f"  VT      : ERROR — {vt['vt_error']}")
            else:
                label = risk_label(vt)
                print(f"  Risk    : {label}")
                print(f"  Reputation score : {vt['reputation']}")
                stats = vt.get("analysis_stats", {})
                print(
                    f"  Engines : {vt['total_engines']} total | "
                    f"{vt['malicious_count']} malicious | "
                    f"{vt['suspicious_count']} suspicious | "
                    f"{stats.get('harmless', 0)} harmless"
                )
                if vt.get("categories"):
                    cats = ", ".join(f"{k}: {v}" for k, v in vt["categories"].items())
                    print(f"  Categories : {cats}")
                if vt.get("tags"):
                    print(f"  Tags       : {', '.join(vt['tags'])}")
                if vt.get("registrar"):
                    print(f"  Registrar  : {vt['registrar']}")
                if vt.get("last_analysis_date"):
                    print(f"  Last scan  : {vt['last_analysis_date']}")
                if verbose and vt.get("flagged_engines"):
                    print("  Flagged engines:")
                    for eng, det in vt["flagged_engines"].items():
                        print(f"    [{det['category']}] {eng}: {det['result']}")

        print(f"  {'-' * 66}")


CSV_FIELDS = [
    "domain",
    "expired",
    "days_since_expiry",
    "expiry_date",
    "lookup_method",
    "risk",
    "vt_reputation",
    "vt_categories",
    "vt_tags",
    "vt_malicious",
    "vt_suspicious",
    "vt_harmless",
    "vt_total_engines",
    "vt_last_analysis",
    "registrar",
    "country",
    "vt_error",
    "whois_error",
]


def _flatten(r: dict) -> dict:
    """Flatten a result dict into a single-level row for CSV/XLSX."""
    vt = r.get("virustotal") or {}
    stats = vt.get("analysis_stats") or {}
    cats = vt.get("categories") or {}
    return {
        "domain": r["domain"],
        "expired": "Yes" if r.get("expired") else "No",
        "days_since_expiry": r.get("days_since_expiry") or "",
        "expiry_date": r.get("expiry_date") or "",
        "lookup_method": r.get("lookup_method") or "",
        "risk": r.get("risk") or "",
        "vt_reputation": vt.get("reputation") if vt and "vt_error" not in vt else "",
        "vt_categories": "; ".join(f"{k}: {v}" for k, v in cats.items()),
        "vt_tags": "; ".join(vt.get("tags") or []),
        "vt_malicious": vt.get("malicious_count") if vt and "vt_error" not in vt else "",
        "vt_suspicious": vt.get("suspicious_count") if vt and "vt_error" not in vt else "",
        "vt_harmless": stats.get("harmless", "") if vt and "vt_error" not in vt else "",
        "vt_total_engines": vt.get("total_engines") if vt and "vt_error" not in vt else "",
        "vt_last_analysis": vt.get("last_analysis_date") or "",
        "registrar": vt.get("registrar") or "",
        "country": vt.get("country") or "",
        "vt_error": vt.get("vt_error") or "",
        "whois_error": r.get("whois_error") or "",
    }


def write_csv(results: list[dict], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in results:
            writer.writerow(_flatten(r))


def write_xlsx(results: list[dict], path: str) -> None:
    if not HAS_OPENPYXL:
        print("Missing dependency for xlsx: pip install openpyxl", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Domain Report"

    HEADERS = {
        "domain": "Domain",
        "expired": "Expired",
        "days_since_expiry": "Days Since Expiry",
        "expiry_date": "Expiry Date",
        "lookup_method": "Lookup Method",
        "risk": "Risk Level",
        "vt_reputation": "VT Reputation",
        "vt_categories": "VT Categories",
        "vt_tags": "VT Tags",
        "vt_malicious": "Malicious Engines",
        "vt_suspicious": "Suspicious Engines",
        "vt_harmless": "Harmless Engines",
        "vt_total_engines": "Total Engines",
        "vt_last_analysis": "VT Last Scan",
        "registrar": "Registrar",
        "country": "Country",
        "vt_error": "VT Error",
        "whois_error": "Lookup Error",
    }

    COL_WIDTHS = {
        "domain": 30, "expired": 10, "days_since_expiry": 18,
        "expiry_date": 28, "lookup_method": 14, "risk": 14,
        "vt_reputation": 16, "vt_categories": 40, "vt_tags": 30,
        "vt_malicious": 18, "vt_suspicious": 18, "vt_harmless": 16,
        "vt_total_engines": 15, "vt_last_analysis": 28, "registrar": 28,
        "country": 10, "vt_error": 35, "whois_error": 35,
    }

    RISK_FILLS = {
        "HIGH RISK":  PatternFill("solid", fgColor="FF4444"),
        "SUSPICIOUS": PatternFill("solid", fgColor="FFAA00"),
        "CLEAN":      PatternFill("solid", fgColor="44BB44"),
    }
    EXPIRED_FILL  = PatternFill("solid", fgColor="FFE0E0")
    HEADER_FILL   = PatternFill("solid", fgColor="2D4A8A")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    BOLD          = Font(bold=True)
    thin          = Side(style="thin", color="CCCCCC")
    CELL_BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fields = list(HEADERS.keys())

    # Header row
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS[field])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[field]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, r in enumerate(results, 2):
        flat = _flatten(r)
        is_expired = r.get("expired", False)
        risk = flat["risk"]
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=flat[field])
            cell.border = CELL_BORDER
            cell.alignment = CENTER if field in (
                "expired", "days_since_expiry", "lookup_method", "risk",
                "vt_reputation", "vt_malicious", "vt_suspicious",
                "vt_harmless", "vt_total_engines", "country",
            ) else LEFT

            if field == "risk" and risk in RISK_FILLS:
                cell.fill = RISK_FILLS[risk]
                cell.font = Font(bold=True, color="FFFFFF")
            elif field == "expired" and is_expired:
                cell.fill = EXPIRED_FILL
                cell.font = BOLD
            elif field == "domain":
                cell.font = BOLD
        ws.row_dimensions[row_idx].height = 18

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total = len(results)
    expired = sum(1 for r in results if r.get("expired"))
    high_risk = sum(1 for r in results if r.get("risk") == "HIGH RISK")
    suspicious = sum(1 for r in results if r.get("risk") == "SUSPICIOUS")
    vt_checked = sum(1 for r in results if r.get("virustotal"))

    summary_rows = [
        ("Report generated", now_str),
        ("Total domains", total),
        ("Expired", expired),
        ("Active", total - expired),
        ("VirusTotal checked", vt_checked),
        ("High risk", high_risk),
        ("Suspicious", suspicious),
        ("Clean", vt_checked - high_risk - suspicious),
    ]
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 28
    for i, (label, value) in enumerate(summary_rows, 1):
        a = ws2.cell(row=i, column=1, value=label)
        b = ws2.cell(row=i, column=2, value=value)
        a.font = BOLD
        a.border = CELL_BORDER
        b.border = CELL_BORDER

    wb.save(path)


def load_domains(args) -> list[str]:
    """Load domains from file, CLI args, or stdin."""
    domains: list[str] = []
    if args.domains:
        for entry in args.domains:
            domains.append(entry.strip().lower())
    if args.file:
        with open(args.file) as fh:
            for line in fh:
                line = line.strip().lower()
                if line and not line.startswith("#"):
                    domains.append(line)
    if not domains and not sys.stdin.isatty():
        for line in sys.stdin:
            line = line.strip().lower()
            if line and not line.startswith("#"):
                domains.append(line)
    return list(dict.fromkeys(domains))  # deduplicate, preserve order


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Check domain expiry and query VirusTotal for recently expired domains."
    )
    parser.add_argument(
        "-d", "--domain", dest="domains", metavar="DOMAIN", nargs="+",
        help="One or more domains to check",
    )
    parser.add_argument(
        "-f", "--file", metavar="FILE",
        help="File containing one domain per line",
    )
    parser.add_argument(
        "-k", "--api-key", metavar="KEY",
        help="VirusTotal API key (or set VT_API_KEY env var)",
    )
    parser.add_argument(
        "--days", type=int, default=90,
        help="Query VirusTotal for domains expired within this many days (default: 90). "
             "Use 0 to check ALL expired domains.",
    )
    parser.add_argument(
        "--check-all", action="store_true",
        help="Query VirusTotal for every domain, not just recently expired ones",
    )
    parser.add_argument(
        "--output", metavar="FILE",
        help="Save full JSON results to this file",
    )
    parser.add_argument(
        "--csv", metavar="FILE",
        help="Save results to a CSV file",
    )
    parser.add_argument(
        "--xlsx", metavar="FILE",
        help="Save results to a formatted Excel (.xlsx) file",
    )
    parser.add_argument(
        "--workers", type=int, default=10, metavar="N",
        help="Parallel workers for WHOIS/RDAP lookups (default: 10)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Show per-engine flagging details",
    )
    args = parser.parse_args()

    # Resolve API key
    import os
    api_key: Optional[str] = args.api_key or os.environ.get("VT_API_KEY")

    domains = load_domains(args)
    if not domains:
        parser.error("No domains provided. Use -d, -f, or pipe a list via stdin.")

    if not api_key:
        print(
            "Warning: No VirusTotal API key provided (-k / VT_API_KEY). "
            "WHOIS checks only.\n",
            file=sys.stderr,
        )

    # Phase 1 — parallel RDAP/WHOIS lookups
    print(f"Looking up {len(domains)} domain(s) with {args.workers} workers ...", file=sys.stderr)
    expiry_map: dict[str, dict] = {}
    completed = 0
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(get_expiry_info, d): d for d in domains}
        for fut in as_completed(futures):
            completed += 1
            info = fut.result()
            expiry_map[info["domain"]] = info
            print(f"  [{completed}/{len(domains)}] {info['domain']} — "
                  f"{'EXPIRED ' + str(info.get('days_since_expiry','?')) + 'd ago' if info.get('expired') else info.get('expiry_date', info.get('whois_error', '?'))[:28]}",
                  file=sys.stderr)

    # Preserve original input order
    ordered = [expiry_map[d] for d in domains]

    # Phase 2 — sequential VirusTotal (rate-limited)
    vt_queue = []
    for info in ordered:
        if api_key:
            if args.check_all:
                vt_queue.append(info)
            elif info.get("expired"):
                days = info.get("days_since_expiry") or 0
                if args.days == 0 or days <= args.days:
                    vt_queue.append(info)

    if vt_queue:
        print(f"\nQuerying VirusTotal for {len(vt_queue)} domain(s) "
              f"({VT_REQUEST_DELAY}s between requests) ...", file=sys.stderr)

    for i, info in enumerate(vt_queue, 1):
        print(f"  [VT {i}/{len(vt_queue)}] {info['domain']} ...", file=sys.stderr)
        info["virustotal"] = query_virustotal(info["domain"], api_key)
        info["risk"] = risk_label(info["virustotal"])
        if i < len(vt_queue):
            time.sleep(VT_REQUEST_DELAY)

    for info in ordered:
        if "virustotal" not in info:
            info["virustotal"] = None
            info["risk"] = None

    results = ordered

    print_report(results, verbose=args.verbose)

    if args.output:
        with open(args.output, "w") as fh:
            json.dump(results, fh, indent=2, default=str)
        print(f"JSON saved to: {args.output}", file=sys.stderr)

    if args.csv:
        write_csv(results, args.csv)
        print(f"CSV saved to:  {args.csv}", file=sys.stderr)

    xlsx_path = args.xlsx or f"domain_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    write_xlsx(results, xlsx_path)
    print(f"XLSX saved to: {xlsx_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
