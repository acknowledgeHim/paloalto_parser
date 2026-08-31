#!/usr/bin/env python3
"""
Palo Alto Firewall Configuration Analyzer
Parses a PAN-OS configuration — either an XML export (device or Panorama) or a
'show config merged'/'show config running' CLI text capture — checks for
security issues/misconfigurations (including CIS Benchmark L1/L2 checks run
against Tenable Nessus .audit files), and exports findings + rule inventory
to Excel. A Policy Optimizer / rulebase CSV export can supply the Security
Rulebase (with rule-usage data) when a CLI capture's own rulebase is empty.

Usage:
    python pa_analyzer.py running-config.xml
    python pa_analyzer.py running-config.xml -o audit.xlsx
    python pa_analyzer.py "show merged combined.txt" --rules-csv rules.csv
"""

import re
import tarfile
import xml.etree.ElementTree as ET
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run:  pip install openpyxl")
    sys.exit(1)


# ── Line-number-aware XML parser ─────────────────────────────────────────────
def _parse_xml_with_linenos(filename: str):
    """Parse an XML file and return (root_element, {id(elem): lineno}) dict.

    Uses the expat C parser directly so we can read CurrentLineNumber at
    the moment each start-tag is processed — before the Python-level
    XMLParser C-accelerator drops that information.
    """
    from xml.parsers import expat

    builder = ET.TreeBuilder()
    linemap: dict[int, int] = {}
    ep = expat.ParserCreate()
    ep.ordered_attributes = 1   # attrs arrive as flat [name, val, ...] list

    def _start(tag, attr_flat):
        attrs = {}
        for i in range(0, len(attr_flat), 2):
            attrs[attr_flat[i]] = attr_flat[i + 1]
        elem = builder.start(tag, attrs)
        linemap[id(elem)] = ep.CurrentLineNumber

    ep.StartElementHandler = _start
    ep.EndElementHandler = lambda tag: builder.end(tag)
    ep.CharacterDataHandler = lambda data: builder.data(data)

    with open(filename, "rb") as fh:
        ep.ParseFile(fh)

    return builder.close(), linemap


# ── "show config merged" (curly-brace CLI) parser ────────────────────────────
# PAN-OS also renders its config as an indented curly-brace tree — the format
# produced by `show config merged` / `show config running` on the CLI, typically
# captured via a PuTTY/SecureCRT session log. This section converts that text
# into the same ElementTree shape the XML API export produces, so every
# downstream parser/check/CIS-audit routine below runs unmodified regardless of
# which format was supplied.
#
# Grammar (per block): a statement is one of
#   key { ... }        -> nested block
#   key [ v1 v2 ... ];  -> member-list (always unambiguous)
#   key value;          -> scalar (or a single-value member-list, see below)
#   key;                -> "flag" — either a bare list value (permitted-ip-style
#                          identity, zone/tag member) or a structural/boolean
#                          leaf (e.g. "none;", "rules;" for an empty rulebase).
#                          Disambiguated heuristically in _emit_curly_flag().
#
# The one genuine ambiguity the curly syntax can't resolve on its own is
# whether a block's *named* children (`key { childname { ... } }`) are
# individually-named objects (PAN-OS XML: <key><entry name="childname">) or a
# fixed structural field reused verbatim across configs (PAN-OS XML:
# <key><childname>). CURLY_ENTRY_LIST_TAGS is the curated (non-exhaustive)
# answer for the tags this script actually inspects; anything not listed
# defaults to the structural (direct-nest) interpretation, which matches the
# majority of the PAN-OS schema.

_ANSI_CSI_RE       = re.compile(r'\x1b\[[0-9;?]*[A-Za-z]')
_ANSI_MISC_RE      = re.compile(r'\x1b[=>]')
_PAGER_ARTIFACT_RE = re.compile(r'(?m)^[ \t]*lines\s+\d+-\d+[ \t]*')
_VALID_TAG_RE      = re.compile(r'^[A-Za-z_][A-Za-z0-9_.\-]*$')
_CONFIG_BLOCK_RE   = re.compile(r'(?m)^\s*config\s*\{')

# Tags whose children are individually-named objects -> <entry name="child">
CURLY_ENTRY_LIST_TAGS: set[str] = {
    "address", "address-group", "region", "external-list",
    "application", "application-group", "application-filter",
    "service", "service-group", "tag",
    "rules",
    "zone", "virtual-router", "virtual-wire", "vlan",
    "ethernet", "aggregate-ethernet", "loopback", "tunnel", "units",
    "static-route", "static-route-ipv6", "redistribution-profile",
    "profile-group",
    "virus", "spyware", "vulnerability", "url-filtering", "file-blocking",
    "wildfire-analysis", "data-filtering", "dos-protection",
    "zone-protection-profile", "decryption-profile", "certificate-profile",
    "certificate", "ssl-tls-service-profile",
    "ike-crypto-profiles", "ipsec-crypto-profiles",
    "ike-gateways", "gateway", "ipsec",
    "administrator", "users",
    "server", "syslog", "email", "http", "snmptrap",
    "permitted-ip",
    "device-group", "vsys", "devices", "template", "template-stack",
    "schedule", "match-list",
    "custom-url-category", "log-forwarding", "authentication-profile",
    "local-user-database", "user", "user-group",
}

# Tags whose bare children are plain values -> <member>value</member>
CURLY_MEMBER_TAGS: set[str] = {
    "from", "to", "source", "destination", "application", "service", "category",
    "hip-profiles", "tag", "member", "os", "source-user",
    "access-route", "exclude-access-route", "trusted-root-CA",
    "send-syslog", "send-email", "send-http", "send-snmptrap",
    "encryption", "authentication", "dh-group", "hash",
    "interface", "static", "group",
    "virus", "vulnerability", "spyware", "url-filtering", "file-blocking",
    "wildfire-analysis", "data-filtering", "dos-protection",
    "severity", "file-types", "applications",
}

# Tags whose bare children are identities, not plain values -> <entry name="value"/>
CURLY_ENTRY_VALUE_TAGS: set[str] = {"permitted-ip"}

# Bare "key;" flags that are boolean/choice leaves (become <key/>), not
# container/identity data — checked only once CURLY_MEMBER_TAGS /
# CURLY_ENTRY_VALUE_TAGS / tag-validity have ruled out the data cases.
CURLY_BOOLEAN_FLAGS: set[str] = {
    "none", "any", "enable", "disable", "yes", "no", "default", "all",
    "drop", "reset-both", "reset-client", "reset-server", "block-ip",
    "alert", "allow", "alarm", "block", "continue", "override", "deny", "reset",
}

# (enclosing tag, curly child key) -> renamed XML tag, for known CLI/API
# naming divergences (PAN-OS's local-admin CLI tree is "mgt-config/users/*"
# but the XML API schema calls the same list "administrator").
CURLY_TAG_RENAME: dict[tuple[str, str], str] = {
    ("mgt-config", "users"): "administrator",
}


def _looks_like_curly_config(text: str) -> bool:
    """True if text looks like a 'show config merged/running' CLI capture."""
    return _CONFIG_BLOCK_RE.search(text) is not None


def _valid_tag_name(s: str) -> bool:
    return bool(_VALID_TAG_RE.match(s))


def _safe_tag(name: str) -> str:
    """Coerce an arbitrary token into a valid XML tag name.

    Almost every curly-format key is already schema-safe, but a handful of
    config fields can hold free-form text (HTML templates, banners) with
    unbalanced quotes that desync the tokenizer; a stray fragment of that text
    can otherwise end up used as a tag (e.g. a trailing-colon CSS property
    like "font-family:", which lxml rejects as an invalid QName and aborts
    CIS/.audit XSLT execution for the whole config). Sanitizing every
    dynamically-derived tag name keeps that kind of corruption local instead
    of taking down the benchmark run.
    """
    if _valid_tag_name(name):
        return name
    s = re.sub(r"[^A-Za-z0-9_.\-]", "_", name) or "_"
    if not re.match(r"^[A-Za-z_]", s):
        s = "_" + s
    return s


def _tokenize_curly(text: str, start: int) -> list[tuple[str, str, int]]:
    """Tokenize text[start:] into (kind, value, lineno) tuples.

    kind is one of IDENT, LBRACE, RBRACE, LBRACKET, RBRACKET, SEMI, EOF.
    Quoted strings (with backslash escapes, possibly spanning lines) become a
    single IDENT token holding the unescaped contents.
    """
    tokens: list[tuple[str, str, int]] = []
    i, n = start, len(text)
    line = text.count("\n", 0, start) + 1
    single = {"{": "LBRACE", "}": "RBRACE", "[": "LBRACKET", "]": "RBRACKET", ";": "SEMI"}
    while i < n:
        c = text[i]
        if c == "\n":
            line += 1
            i += 1
            continue
        if c in " \t\r":
            i += 1
            continue
        if c in single:
            tokens.append((single[c], c, line))
            i += 1
            continue
        if c == '"':
            tok_line = line
            j = i + 1
            buf: list[str] = []
            while j < n and text[j] != '"':
                if text[j] == "\\" and j + 1 < n:
                    buf.append(text[j + 1])
                    if text[j + 1] == "\n":
                        line += 1
                    j += 2
                    continue
                if text[j] == "\n":
                    line += 1
                buf.append(text[j])
                j += 1
            tokens.append(("IDENT", "".join(buf), tok_line))
            i = j + 1
            continue
        j = i
        while j < n and text[j] not in ' \t\r\n{}[];"':
            j += 1
        if j == i:
            i += 1  # stray unrecognized char — skip defensively
            continue
        tokens.append(("IDENT", text[i:j], line))
        i = j
    tokens.append(("EOF", "", line))
    return tokens


def _parse_curly_statements(tokens: list, pos: int, until_rbrace: bool):
    """Parse statements from tokens[pos:] until a matching RBRACE or EOF.

    Returns (list of (kind, key, payload, lineno) statements, new_pos).
    kind is one of: 'block' (payload=child statements), 'list' (payload=[str]),
    'scalar' (payload=str), 'flag' (payload=None).
    """
    stmts = []
    while True:
        kind, val, line = tokens[pos]
        if kind == "EOF":
            break
        if until_rbrace and kind == "RBRACE":
            break
        if kind != "IDENT":
            pos += 1  # unexpected token — skip defensively
            continue
        key, key_line = val, line
        pos += 1
        nkind, nval, _nline = tokens[pos]
        if nkind == "LBRACE":
            children, pos = _parse_curly_statements(tokens, pos + 1, True)
            if tokens[pos][0] == "RBRACE":
                pos += 1
            stmts.append(("block", key, children, key_line))
        elif nkind == "LBRACKET":
            pos += 1
            values = []
            while tokens[pos][0] not in ("RBRACKET", "EOF"):
                if tokens[pos][0] == "IDENT":
                    values.append(tokens[pos][1])
                pos += 1
            if tokens[pos][0] == "RBRACKET":
                pos += 1
            if tokens[pos][0] == "SEMI":
                pos += 1
            stmts.append(("list", key, values, key_line))
        elif nkind == "IDENT":
            pos += 1
            if tokens[pos][0] == "SEMI":
                pos += 1
            stmts.append(("scalar", key, nval, key_line))
        elif nkind == "SEMI":
            pos += 1
            stmts.append(("flag", key, None, key_line))
        else:
            stmts.append(("flag", key, None, key_line))
    return stmts, pos


def _emit_curly_flag(builder: ET.TreeBuilder, container_tag: str, flag_val: str,
                      line: int, linemap: dict):
    """Emit one bare 'key;' statement found inside `container_tag`'s block."""
    if container_tag in CURLY_MEMBER_TAGS:
        m = builder.start("member", {}); linemap[id(m)] = line
        builder.data(flag_val)
        builder.end("member")
        return
    if container_tag in CURLY_ENTRY_VALUE_TAGS:
        e = builder.start("entry", {"name": flag_val}); linemap[id(e)] = line
        builder.end("entry")
        return
    if not _valid_tag_name(flag_val):
        # Can't be a tag name, so it must be identity/value data (e.g. a bare
        # CIDR under an unrecognized container) — emit both shapes so either
        # XPath convention (<member> or <entry name=...>) matches downstream.
        m = builder.start("member", {}); linemap[id(m)] = line
        builder.data(flag_val)
        builder.end("member")
        e = builder.start("entry", {"name": flag_val}); linemap[id(e)] = line
        builder.end("entry")
        return
    # Structural/boolean leaf, e.g. "none;", "rules;" (empty rulebase), "v1;".
    e = builder.start(flag_val, {}); linemap[id(e)] = line
    builder.end(flag_val)


def _emit_curly_block(builder: ET.TreeBuilder, container_tag: str, stmts: list, linemap: dict):
    """Populate the currently-open `container_tag` element with `stmts`."""
    for kind, key, payload, line in stmts:
        xml_key = CURLY_TAG_RENAME.get((container_tag, key), key)
        safe_key = _safe_tag(xml_key)
        if kind == "flag":
            _emit_curly_flag(builder, container_tag, key, line, linemap)
        elif kind == "scalar":
            e = builder.start(safe_key, {}); linemap[id(e)] = line
            if xml_key in CURLY_MEMBER_TAGS or key in CURLY_MEMBER_TAGS:
                m = builder.start("member", {}); linemap[id(m)] = line
                builder.data(payload)
                builder.end("member")
            else:
                builder.data(payload)
            builder.end(safe_key)
        elif kind == "list":
            e = builder.start(safe_key, {}); linemap[id(e)] = line
            for v in payload:
                m = builder.start("member", {}); linemap[id(m)] = line
                builder.data(v)
                builder.end("member")
            builder.end(safe_key)
        elif kind == "block":
            if container_tag in CURLY_ENTRY_LIST_TAGS:
                e = builder.start("entry", {"name": xml_key}); linemap[id(e)] = line
                _emit_curly_block(builder, xml_key, payload, linemap)
                builder.end("entry")
            else:
                e = builder.start(safe_key, {}); linemap[id(e)] = line
                _emit_curly_block(builder, xml_key, payload, linemap)
                builder.end(safe_key)


def _load_curly_config(filename: str):
    """Parse a 'show config merged/running' CLI capture into (root, linemap)."""
    with open(filename, "r", encoding="utf-8", errors="replace") as fh:
        raw = fh.read()

    text = raw.replace("\r\n", "\n").replace("\r", "\n")
    text = _ANSI_CSI_RE.sub("", text)
    text = _ANSI_MISC_RE.sub("", text)
    text = _PAGER_ARTIFACT_RE.sub("", text)  # PuTTY/pager "lines N-M" redraw artifacts
    text = text.replace(" \x08", "")         # terminal soft-wrap space+backspace artifact
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)  # remaining control chars (invalid in XML)

    m = _CONFIG_BLOCK_RE.search(text)
    if not m:
        raise ValueError("No 'config {' block found — not a recognized CLI config capture")

    tokens = _tokenize_curly(text, m.start())
    pos = 0
    if tokens[pos][0] != "IDENT" or tokens[pos][1] != "config":
        raise ValueError("Expected 'config' at start of matched block")
    root_line = tokens[pos][2]
    pos += 1
    if tokens[pos][0] != "LBRACE":
        raise ValueError("Expected '{' after 'config'")
    children, _pos = _parse_curly_statements(tokens, pos + 1, True)

    builder = ET.TreeBuilder()
    linemap: dict[int, int] = {}
    root = builder.start("config", {})
    linemap[id(root)] = root_line
    _emit_curly_block(builder, "config", children, linemap)
    builder.end("config")
    return builder.close(), linemap


# ── CIS Controls v8 mapping ──────────────────────────────────────────────────
# Short descriptions used in the CIS reference tab
CIS_CTRL_DESC = {
    "3.10":  "Encrypt Sensitive Data in Transit",
    "4.2":   "Secure Configuration Process for Network Infrastructure",
    "4.7":   "Manage Default Accounts on Enterprise Assets and Software",
    "4.8":   "Disable Unnecessary Services on Enterprise Assets and Software",
    "4.9":   "Configure Trusted DNS Servers on Enterprise Assets",
    "5.2":   "Use Unique Passwords",
    "5.4":   "Restrict Administrator Privileges to Dedicated Accounts",
    "6.5":   "Require MFA for Administrative Access",
    "6.7":   "Centralize Access Control",
    "6.8":   "Define and Maintain Role-Based Access Control",
    "8.2":   "Collect Audit Logs",
    "8.4":   "Standardize Time Synchronization",
    "8.5":   "Collect Detailed Audit Logs",
    "8.9":   "Centralize Audit Logs",
    "10.1":  "Deploy and Maintain Anti-Malware Software",
    "10.7":  "Use Behavior-Based Anti-Malware Software",
    "12.2":  "Establish and Maintain a Secure Network Architecture",
    "12.3":  "Securely Manage Network Infrastructure",
    "12.6":  "Use Secure Network Management and Communication Protocols",
    "13.3":  "Deploy a Network Intrusion Detection Solution",
    "13.4":  "Perform Traffic Filtering Between Network Segments",
    "13.8":  "Deploy a Network Intrusion Prevention Solution",
    "13.10": "Perform Application Layer Filtering",
}

# Maps every check category string → list of CIS v8 safeguard IDs
CIS_CONTROL_MAP: dict[str, list[str]] = {
    # ── Firewall rule checks ───────────────────────────────────────────────────
    "Any/Any/Any Allow Rule":                    ["12.2", "13.4"],
    "Missing Security Profiles":                 ["10.1", "13.8", "13.10"],
    "No Logging Configured":                     ["8.2",  "8.5"],
    "Allow Rule Not Logging Session End":        ["8.2",  "8.5"],
    "Unrestricted Source Address":               ["12.2", "13.4"],
    "Unrestricted Destination Address":          ["12.2", "13.4"],
    "Exposed RDP from Any Source":               ["12.2", "12.6"],
    "Cleartext Telnet Allowed":                  ["4.8",  "12.6"],
    "SSH Exposed from Any Source":               ["12.2", "12.3"],
    "SMB Exposed from Any Source":               ["12.2", "13.4"],
    "VNC Exposed from Any Source":               ["12.2", "12.6"],
    "Risky Application Allowed Without User-ID Restriction": ["12.2", "4.2"],
    "Disabled Rule":                             ["4.2"],
    "Missing Rule Description":                  ["4.2"],
    "Negated Source Address":                    ["4.2",  "12.2"],
    "Negated Destination Address":               ["4.2",  "12.2"],
    "Zone Missing Protection Profile":           ["13.3", "13.4"],
    "Potential Shadow Rule":                     ["4.2",  "12.2"],
    "Application+Service Both Any":              ["13.4", "13.10"],
    "Inbound Allow Without Inspection":          ["10.1", "13.4", "13.8"],
    "Service=Any with Specific Application":     ["12.2", "13.4"],
    # ── Crypto checks ─────────────────────────────────────────────────────────
    "Weak IKE Encryption":                       ["3.10", "12.6"],
    "Weak IKE Hash/PRF":                         ["3.10", "12.6"],
    "Weak IKE DH Group":                         ["3.10", "12.6"],
    "Weak IPSec Encryption":                     ["3.10", "12.6"],
    "Weak IPSec Authentication":                 ["3.10", "12.6"],
    "Weak IPSec DH Group (PFS)":                 ["3.10", "12.6"],
    "IPSec PFS Disabled":                        ["3.10"],
    "Weak Minimum TLS Version":                  ["3.10", "12.6"],
    "IKEv1 in Use":                              ["12.6"],
    "IKE Pre-Shared Key Authentication":         ["12.6"],
    # ── Management / system checks ────────────────────────────────────────────
    "HTTP Management Enabled":                   ["4.2",  "12.3", "12.6"],
    "Telnet Management Enabled":                 ["4.2",  "12.3", "12.6"],
    "No Management IP Restrictions":             ["12.3", "6.7"],
    "NTP Not Configured":                        ["8.4"],
    "No Login Banner":                           ["4.2"],
    "DNS Not Configured":                        ["4.9"],
    "Admin Without Authentication Profile":      ["6.5"],
    "Admin Account Has No Password":             ["5.2",  "6.5"],
    "Excessive Superuser Accounts":              ["5.4",  "6.8"],
    "SNMPv1 Enabled":                            ["4.2",  "12.3", "12.6"],
    "SNMPv2c Enabled":                           ["4.2",  "12.3"],
    "Default/Weak SNMP Community String":        ["4.7",  "12.3"],
    "SNMP Enabled Without Source Restrictions":  ["12.3", "6.7"],
    "No Syslog Servers Configured":              ["8.2",  "8.9"],
    "Syslog Transmitted Over UDP":               ["8.9"],
    # ── Password & session policy ─────────────────────────────────────────────
    "Password Complexity Not Enforced":          ["5.2"],
    "Weak Password Minimum Length":              ["5.2"],
    "No Account Lockout Policy":                 ["5.2", "6.5"],
    "Long or No Management Session Timeout":     ["12.3"],
    # ── Content updates ───────────────────────────────────────────────────────
    "AV/Threat Content Updates Not Automatic":   ["10.1"],
    "WildFire Updates Not Automatic":            ["10.7"],
    # ── Security profile quality ──────────────────────────────────────────────
    "Vulnerability Profile Allows Critical/High Threats": ["13.3", "13.8"],
    "WildFire Profile Missing Rules":            ["10.7"],
    "WildFire Profile Incomplete Coverage":      ["10.7"],
    # ── Zone / User-ID ────────────────────────────────────────────────────────
    "User-ID Enabled on Untrusted Zone":         ["12.2", "4.2"],
    # ── NTP ───────────────────────────────────────────────────────────────────
    "Only One NTP Server":                       ["8.4"],
    "NTP Authentication Not Configured":         ["8.4", "12.6"],
    # ── Insecure protocols / certificates ────────────────────────────────────
    "Insecure Protocol Allowed in Rule":         ["4.8", "12.6"],
    "TLS Profile Using Default Certificate":     ["4.2", "12.6"],
    "Password Expiry Not Configured":            ["5.2"],
    "Insufficient Password History":             ["5.2"],
    "No Default Deny Rule":                      ["12.2", "13.4"],
    "File Blocking Not Applied":                 ["10.1"],
    # ── CIS L2 benchmark checks ───────────────────────────────────────────────
    "Admin Interface Default Certificate":       ["3.10", "12.6"],
    "WMI Probing Enabled":                       ["4.8", "12.3"],
    "Zone Flood Protection Disabled":            ["13.3", "13.4"],
    "Decryption Certificate Untrusted":          ["3.10", "12.6"],
    "SNMPv3 Trap Not Configured":                ["8.2", "8.9"],
    # ── CIS L1 benchmark checks ───────────────────────────────────────────────
    "High DP Load Logging Disabled":             ["8.2"],
    "SNMPv3 Polling Not Configured":             ["8.2", "8.9"],
    "Update Server Verification Disabled":       ["4.6", "12.6"],
    "Password Profile Bypass Configured":        ["5.2", "6.5"],
    "User-ID Not Configured":                    ["6.1"],
    "User-ID Network Exclusions Not Configured": ["4.2", "6.1"],
    "User-ID Agent Traffic Unrestricted":        ["4.2", "12.3"],
    "High Availability Not Configured":          ["12.2"],
    "WildFire File Size Limits Not Maximized":   ["10.7"],
    "WildFire Profile Not Applied":              ["10.7"],
    "WildFire Decrypted Content Not Forwarded":  ["10.7"],
    "WildFire Session Information Incomplete":   ["10.7"],
    "WildFire Malicious File Alerts Disabled":   ["10.7"],
    "WildFire Inline ML Not Enabled":            ["10.7"],
    "Antivirus Profile Not Blocking":            ["10.1", "13.4"],
    "Antivirus Profile Not Applied":             ["10.1", "13.4"],
    "Anti-Spyware Profile Not Blocking":         ["10.1", "13.4"],
    "Anti-Spyware Profile Not Applied":          ["10.1", "13.4"],
    "Vulnerability Profile Not Applied":         ["13.3", "13.8"],
    "URL Filtering Not Configured":              ["9.2", "13.4"],
    "URL Category Action Not Blocking":          ["9.2", "13.4"],
    "URL Filtering Not Logging":                 ["8.2", "9.2"],
    "URL Filtering Profile Not Applied":         ["9.2", "13.4"],
    "Data Filtering Not Applied":                ["3.13", "13.4"],
    "Reconnaissance Protection Disabled":        ["13.3", "13.4"],
    "Packet-Based Attack Protection Disabled":   ["13.3", "13.4"],
    "User Credential Submission Uncontrolled":   ["13.4", "6.4"],
    "Advanced Threat Prevention Not Enabled":    ["10.1", "13.4"],
    "Threat Intelligence Blocking Not Configured": ["10.1", "13.4"],
    "Default Policy Logging Disabled":           ["8.2"],
    "Management Interface Not Restricted":       ["12.3", "6.7"],
    "SSL Decryption Not Configured":             ["3.10", "13.4"],
}


def _cis_label(ctrl_ids: list[str]) -> str:
    """Return a compact string like 'CIS 12.2 · CIS 13.4'."""
    return " · ".join(f"CIS {c}" for c in ctrl_ids)


# ── CIS PAN-OS Benchmark check number mapping (L1 + L2) ─────────────────────
# Maps category → list of CIS PAN-OS Firewall Benchmark check IDs
CIS_BENCHMARK_MAP: dict[str, list[str]] = {
    # L1 checks
    "No Login Banner":                         ["1.1.2"],
    "High DP Load Logging Disabled":           ["1.1.3"],
    "Management Interface Not Restricted":     ["1.2.1", "1.2.2"],
    "HTTP Management Enabled":                 ["1.2.3"],
    "Telnet Management Enabled":               ["1.2.3"],
    "Password Complexity Not Enforced":        ["1.3.1", "1.3.3", "1.3.4", "1.3.5", "1.3.6"],
    "Weak Password Minimum Length":            ["1.3.2"],
    "Password Expiry Not Configured":          ["1.3.7"],
    "Insufficient Password History":           ["1.3.8", "1.3.9"],
    "Password Profile Bypass Configured":      ["1.3.10"],
    "Long or No Management Session Timeout":   ["1.4.1"],
    "No Account Lockout Policy":               ["1.4.1"],
    "SNMPv3 Polling Not Configured":           ["1.5.1"],
    "Update Server Verification Disabled":     ["1.6.1"],
    "User-ID Not Configured":                  ["2.7"],
    "User-ID Network Exclusions Not Configured": ["2.4"],
    "User-ID Agent Traffic Unrestricted":      ["2.8"],
    "High Availability Not Configured":        ["3.1"],
    "AV/Threat Content Updates Not Automatic": ["4.1", "4.2"],
    "WildFire File Size Limits Not Maximized": ["5.1"],
    "WildFire Profile Not Applied":            ["5.2"],
    "WildFire Decrypted Content Not Forwarded": ["5.3"],
    "WildFire Session Information Incomplete": ["5.4"],
    "WildFire Malicious File Alerts Disabled": ["5.5"],
    "WildFire Updates Not Automatic":          ["5.6"],
    "WildFire Inline ML Not Enabled":          ["5.8"],
    "Antivirus Profile Not Blocking":          ["6.1", "6.20", "6.21"],
    "Antivirus Profile Not Applied":           ["6.2"],
    "Anti-Spyware Profile Not Blocking":       ["6.3", "6.24"],
    "Anti-Spyware Profile Not Applied":        ["6.5"],
    "Vulnerability Profile Allows Critical/High Threats": ["6.6"],
    "Vulnerability Profile Not Applied":       ["6.7"],
    "URL Filtering Not Configured":            ["6.8"],
    "URL Category Action Not Blocking":        ["6.9"],
    "URL Filtering Not Logging":               ["6.10"],
    "URL Filtering Profile Not Applied":       ["6.12"],
    "Data Filtering Not Applied":              ["6.14"],
    "Zone Flood Protection Disabled":          ["6.15", "6.16"],
    "Reconnaissance Protection Disabled":      ["6.17"],
    "Packet-Based Attack Protection Disabled": ["6.18"],
    "User Credential Submission Uncontrolled": ["6.19"],
    "Advanced Threat Prevention Not Enabled":  ["6.22", "6.23"],
    "Service=Any with Specific Application":   ["7.2"],
    "Threat Intelligence Blocking Not Configured": ["7.3"],
    "Default Policy Logging Disabled":         ["7.4"],
    "SSL Decryption Not Configured":           ["7.4"],
    # L2 checks
    "SNMPv3 Trap Not Configured":              ["1.1.1.2"],
    "Admin Interface Default Certificate":     ["1.2.5"],
    "WMI Probing Enabled":                     ["2.2"],
    "Decryption Certificate Untrusted":        ["8.3"],
}


def _cis_benchmark_label(check_ids: list[str]) -> str:
    return " · ".join(check_ids)


# ── PCI DSS v4.0 mapping ─────────────────────────────────────────────────────
PCI_DSS_DESC = {
    "1.2.4":  "All traffic between trusted/untrusted networks is explicitly controlled",
    "1.3.1":  "Inbound traffic to the CDE is restricted to what is necessary",
    "1.3.2":  "Outbound traffic from the CDE is restricted to what is necessary",
    "2.2.1":  "Configuration standards are defined for all system components",
    "2.2.4":  "Only necessary services, protocols, and functions are enabled",
    "2.2.7":  "All non-console administrative access is encrypted",
    "4.2.1":  "Strong cryptography is used to safeguard PAN during transmission",
    "5.3.1":  "Anti-malware solution is deployed on all applicable system components",
    "5.3.2":  "Anti-malware mechanisms are kept current and actively running",
    "6.3.1":  "Security vulnerabilities are identified and managed",
    "7.2.1":  "All user access is appropriate and access is assigned by business need",
    "8.2.1":  "All user IDs and authentication credentials are managed securely",
    "8.2.8":  "If idle for more than 15 minutes, the session re-authenticates the user",
    "8.3.4":  "Invalid authentication attempts are limited to no more than 10",
    "8.3.6":  "Passwords/passphrases meet minimum complexity and length requirements",
    "8.3.7":  "Passwords/passphrases cannot be the same as any of the last four used",
    "8.3.9":  "Passwords/passphrases for user accounts are changed at least every 90 days",
    "8.4.1":  "MFA is implemented for all non-console administrative access",
    "10.2.1": "Audit logs are enabled and active for all system components",
    "10.2.2": "Audit logs capture all activities by individuals with root or admin privileges",
    "10.5.4": "Audit log files are protected against destruction and unauthorized modifications",
    "10.6.1": "System clocks are synchronized using time-synchronization technology",
}

PCI_DSS_MAP: dict[str, list[str]] = {
    # ── Firewall rule checks ───────────────────────────────────────────────────
    "Any/Any/Any Allow Rule":              ["1.2.4", "1.3.1"],
    "Missing Security Profiles":           ["5.3.1", "6.3.1"],
    "No Logging Configured":               ["10.2.1", "10.2.2"],
    "Allow Rule Not Logging Session End":  ["10.2.1"],
    "Unrestricted Source Address":         ["1.3.1"],
    "Unrestricted Destination Address":    ["1.3.2"],
    "Exposed RDP from Any Source":         ["1.3.1"],
    "Cleartext Telnet Allowed":            ["2.2.4", "2.2.7"],
    "SSH Exposed from Any Source":         ["1.3.1"],
    "SMB Exposed from Any Source":         ["1.3.1"],
    "VNC Exposed from Any Source":         ["1.3.1"],
    "Risky Application Allowed Without User-ID Restriction": ["7.2.1", "8.2.1"],
    "Disabled Rule":                       ["1.2.4"],
    "Missing Rule Description":            ["1.2.4"],
    "Negated Source Address":              ["1.2.4"],
    "Negated Destination Address":         ["1.2.4"],
    "Zone Missing Protection Profile":     ["1.2.4", "1.3.1"],
    "Potential Shadow Rule":               ["1.2.4"],
    "Application+Service Both Any":        ["1.3.1"],
    "Inbound Allow Without Inspection":    ["5.3.1", "6.3.1"],
    "Service=Any with Specific Application": ["1.2.4"],
    "No Default Deny Rule":                ["1.2.4"],
    "File Blocking Not Applied":           ["5.3.1"],
    # ── Crypto checks ─────────────────────────────────────────────────────────
    "Weak IKE Encryption":                 ["4.2.1"],
    "Weak IKE Hash/PRF":                   ["4.2.1"],
    "Weak IKE DH Group":                   ["4.2.1"],
    "Weak IPSec Encryption":               ["4.2.1"],
    "Weak IPSec Authentication":           ["4.2.1"],
    "Weak IPSec DH Group (PFS)":           ["4.2.1"],
    "IPSec PFS Disabled":                  ["4.2.1"],
    "Weak Minimum TLS Version":            ["4.2.1"],
    "IKEv1 in Use":                        ["4.2.1"],
    "IKE Pre-Shared Key Authentication":   ["4.2.1"],
    "TLS Profile Using Default Certificate": ["4.2.1"],
    # ── Management / system ───────────────────────────────────────────────────
    "HTTP Management Enabled":             ["2.2.4", "2.2.7"],
    "Telnet Management Enabled":           ["2.2.4", "2.2.7"],
    "No Management IP Restrictions":       ["1.2.4"],
    "NTP Not Configured":                  ["10.6.1"],
    "Only One NTP Server":                 ["10.6.1"],
    "NTP Authentication Not Configured":   ["10.6.1"],
    "No Login Banner":                     ["2.2.1"],
    "Admin Without Authentication Profile": ["8.4.1"],
    "Admin Account Has No Password":       ["8.2.1", "8.3.6"],
    "Excessive Superuser Accounts":        ["7.2.1"],
    "SNMPv1 Enabled":                      ["2.2.4"],
    "SNMPv2c Enabled":                     ["2.2.4"],
    "Default/Weak SNMP Community String":  ["2.2.4"],
    "SNMP Enabled Without Source Restrictions": ["1.2.4"],
    "No Syslog Servers Configured":        ["10.5.4"],
    "Syslog Transmitted Over UDP":         ["10.5.4"],
    # ── Password policy ───────────────────────────────────────────────────────
    "Password Complexity Not Enforced":    ["8.3.6"],
    "Weak Password Minimum Length":        ["8.3.6"],
    "No Account Lockout Policy":           ["8.3.4"],
    "Long or No Management Session Timeout": ["8.2.8"],
    "Password Expiry Not Configured":      ["8.3.9"],
    "Insufficient Password History":       ["8.3.7"],
    # ── Content updates / profile quality ────────────────────────────────────
    "AV/Threat Content Updates Not Automatic": ["5.3.2"],
    "WildFire Updates Not Automatic":      ["5.3.2"],
    "Vulnerability Profile Allows Critical/High Threats": ["6.3.1"],
    "WildFire Profile Missing Rules":      ["5.3.2"],
    "WildFire Profile Incomplete Coverage": ["5.3.2"],
    # ── Zone / protocol ───────────────────────────────────────────────────────
    "User-ID Enabled on Untrusted Zone":   ["1.3.1"],
    "Insecure Protocol Allowed in Rule":   ["2.2.4", "4.2.1"],
    # ── CIS L2 benchmark checks ───────────────────────────────────────────────
    "Admin Interface Default Certificate": ["2.2.7", "4.2.1"],
    "WMI Probing Enabled":                 ["2.2.4"],
    "Zone Flood Protection Disabled":      ["1.2.4", "1.3.1"],
    "Decryption Certificate Untrusted":    ["4.2.1"],
    "SNMPv3 Trap Not Configured":          ["10.2.1", "10.5.4"],
    # ── CIS L1 benchmark checks ───────────────────────────────────────────────
    "High DP Load Logging Disabled":             ["10.2.1"],
    "SNMPv3 Polling Not Configured":             ["2.2.4"],
    "Update Server Verification Disabled":       ["6.3.1"],
    "Password Profile Bypass Configured":        ["8.3.6"],
    "User-ID Not Configured":                    [],
    "User-ID Network Exclusions Not Configured": [],
    "User-ID Agent Traffic Unrestricted":        ["1.2.4"],
    "High Availability Not Configured":          [],
    "WildFire File Size Limits Not Maximized":   ["5.3.1"],
    "WildFire Profile Not Applied":              ["5.3.1"],
    "WildFire Decrypted Content Not Forwarded":  ["5.3.1"],
    "WildFire Session Information Incomplete":   ["5.3.1"],
    "WildFire Malicious File Alerts Disabled":   ["5.3.1"],
    "WildFire Inline ML Not Enabled":            ["5.3.1"],
    "Antivirus Profile Not Blocking":            ["5.3.1", "5.3.2"],
    "Antivirus Profile Not Applied":             ["5.3.1"],
    "Anti-Spyware Profile Not Blocking":         ["5.3.1", "6.3.1"],
    "Anti-Spyware Profile Not Applied":          ["5.3.1"],
    "Vulnerability Profile Not Applied":         ["6.3.1"],
    "URL Filtering Not Configured":              ["1.2.4", "1.3.1"],
    "URL Category Action Not Blocking":          ["1.2.4", "1.3.1"],
    "URL Filtering Not Logging":                 ["10.2.1"],
    "URL Filtering Profile Not Applied":         ["1.2.4", "1.3.1"],
    "Data Filtering Not Applied":                ["1.3.2"],
    "Reconnaissance Protection Disabled":        ["1.2.4", "1.3.1"],
    "Packet-Based Attack Protection Disabled":   ["1.2.4", "1.3.1"],
    "User Credential Submission Uncontrolled":   ["1.2.4"],
    "Advanced Threat Prevention Not Enabled":    ["5.3.1", "6.3.1"],
    "Threat Intelligence Blocking Not Configured": ["1.3.1"],
    "Default Policy Logging Disabled":           ["10.2.1"],
    "Management Interface Not Restricted":       ["1.2.4"],
    "SSL Decryption Not Configured":             ["4.2.1"],
}


def _pci_label(req_ids: list[str]) -> str:
    return " · ".join(f"PCI {r}" for r in req_ids)


# ── Secure Controls Framework (SCF) mapping ──────────────────────────────────
SCF_MAP: dict[str, list[str]] = {
    # ── Firewall rule checks ───────────────────────────────────────────────────
    "Any/Any/Any Allow Rule":                        ["NET-04"],
    "Missing Security Profiles":                     ["NET-14"],
    "No Logging Configured":                         ["MON-06"],
    "Allow Rule Not Logging Session End":            ["MON-06"],
    "Unrestricted Source Address":                   ["NET-04"],
    "Unrestricted Destination Address":              ["NET-04"],
    "Exposed RDP from Any Source":                   ["NET-04", "IAC-10"],
    "Cleartext Telnet Allowed":                      ["CRY-03"],
    "SSH Exposed from Any Source":                   ["NET-04"],
    "SMB Exposed from Any Source":                   ["NET-04"],
    "VNC Exposed from Any Source":                   ["NET-04"],
    "Risky Application Allowed Without User-ID Restriction": ["IAC-01", "IAC-10"],
    "Disabled Rule":                                 ["NET-04"],
    "Missing Rule Description":                      ["OPS-01"],
    "Negated Source Address":                        ["NET-04"],
    "Negated Destination Address":                   ["NET-04"],
    "Zone Missing Protection Profile":               ["NET-04"],
    "Potential Shadow Rule":                         ["NET-04"],
    "Application+Service Both Any":                  ["NET-04"],
    "Inbound Allow Without Inspection":              ["NET-14"],
    "Service=Any with Specific Application":         ["NET-04"],
    # ── Crypto checks ─────────────────────────────────────────────────────────
    "Weak IKE Encryption":                           ["CRY-03"],
    "Weak IKE Hash/PRF":                             ["CRY-03"],
    "Weak IKE DH Group":                             ["CRY-03"],
    "Weak IPSec Encryption":                         ["CRY-03"],
    "Weak IPSec Authentication":                     ["CRY-03"],
    "Weak IPSec DH Group (PFS)":                     ["CRY-03"],
    "IPSec PFS Disabled":                            ["CRY-03"],
    "Weak Minimum TLS Version":                      ["CRY-03"],
    "IKEv1 in Use":                                  ["CRY-03"],
    "IKE Pre-Shared Key Authentication":             ["CRY-01"],
    # ── Management / system checks ────────────────────────────────────────────
    "HTTP Management Enabled":                       ["CRY-03", "NET-06"],
    "Telnet Management Enabled":                     ["CRY-03", "NET-06"],
    "No Management IP Restrictions":                 ["NET-04", "IAC-10"],
    "NTP Not Configured":                            ["OPS-01"],
    "No Login Banner":                               ["IAC-09"],
    "DNS Not Configured":                            ["OPS-01"],
    "Admin Without Authentication Profile":          ["IAC-01"],
    "Admin Account Has No Password":                 ["IAC-06"],
    "Excessive Superuser Accounts":                  ["IAC-07"],
    "SNMPv1 Enabled":                                ["NET-06", "CRY-03"],
    "SNMPv2c Enabled":                               ["NET-06", "CRY-03"],
    "Default/Weak SNMP Community String":            ["IAC-06"],
    "SNMP Enabled Without Source Restrictions":      ["NET-04"],
    "No Syslog Servers Configured":                  ["MON-06"],
    "Syslog Transmitted Over UDP":                   ["MON-06", "CRY-03"],
    # ── Password & session policy ─────────────────────────────────────────────
    "Password Complexity Not Enforced":              ["IAC-06"],
    "Weak Password Minimum Length":                  ["IAC-06"],
    "No Account Lockout Policy":                     ["IAC-06"],
    "Long or No Management Session Timeout":         ["IAC-09"],
    "Password Expiry Not Configured":                ["IAC-06"],
    "Insufficient Password History":                 ["IAC-06"],
    # ── Content updates ───────────────────────────────────────────────────────
    "AV/Threat Content Updates Not Automatic":       ["VPM-10", "TDA-02"],
    "WildFire Updates Not Automatic":                ["VPM-10", "TDA-02"],
    # ── Security profile quality ──────────────────────────────────────────────
    "Vulnerability Profile Allows Critical/High Threats": ["VPM-01"],
    "WildFire Profile Missing Rules":                ["TDA-02"],
    "WildFire Profile Incomplete Coverage":          ["TDA-02"],
    "File Blocking Not Applied":                     ["TDA-02"],
    # ── Zone / User-ID ────────────────────────────────────────────────────────
    "User-ID Enabled on Untrusted Zone":             ["IAC-01", "NET-04"],
    # ── NTP ───────────────────────────────────────────────────────────────────
    "Only One NTP Server":                           ["OPS-01"],
    "NTP Authentication Not Configured":             ["OPS-01", "CRY-03"],
    # ── Insecure protocols / certificates ────────────────────────────────────
    "Insecure Protocol Allowed in Rule":             ["CRY-03"],
    "TLS Profile Using Default Certificate":         ["CRY-03"],
    # ── Rule completeness ─────────────────────────────────────────────────────
    "No Default Deny Rule":                          ["NET-04"],
    # ── CIS L2 benchmark checks ───────────────────────────────────────────────
    "Admin Interface Default Certificate":           ["CRY-03", "IAC-10"],
    "WMI Probing Enabled":                           ["IAC-01", "NET-04"],
    "Zone Flood Protection Disabled":                ["NET-04"],
    "Decryption Certificate Untrusted":              ["CRY-03"],
    "SNMPv3 Trap Not Configured":                    ["MON-06"],
    # ── CIS L1 benchmark checks ───────────────────────────────────────────────
    "High DP Load Logging Disabled":                 ["MON-06"],
    "SNMPv3 Polling Not Configured":                 ["MON-06", "CRY-03"],
    "Update Server Verification Disabled":           ["VPM-10"],
    "Password Profile Bypass Configured":            ["IAC-06"],
    "User-ID Not Configured":                        ["IAC-01"],
    "User-ID Network Exclusions Not Configured":     ["IAC-01", "NET-04"],
    "User-ID Agent Traffic Unrestricted":            ["IAC-01", "NET-04"],
    "High Availability Not Configured":              ["OPS-04"],
    "WildFire File Size Limits Not Maximized":       ["TDA-02"],
    "WildFire Profile Not Applied":                  ["TDA-02"],
    "WildFire Decrypted Content Not Forwarded":      ["TDA-02"],
    "WildFire Session Information Incomplete":       ["TDA-02"],
    "WildFire Malicious File Alerts Disabled":       ["TDA-02"],
    "WildFire Inline ML Not Enabled":                ["TDA-02"],
    "Antivirus Profile Not Blocking":                ["TDA-02", "VPM-01"],
    "Antivirus Profile Not Applied":                 ["TDA-02"],
    "Anti-Spyware Profile Not Blocking":             ["TDA-02", "NET-14"],
    "Anti-Spyware Profile Not Applied":              ["TDA-02"],
    "Vulnerability Profile Not Applied":             ["VPM-01"],
    "URL Filtering Not Configured":                  ["NET-04"],
    "URL Category Action Not Blocking":              ["NET-04"],
    "URL Filtering Not Logging":                     ["MON-06"],
    "URL Filtering Profile Not Applied":             ["NET-04"],
    "Data Filtering Not Applied":                    ["DAM-01"],
    "Reconnaissance Protection Disabled":            ["NET-04"],
    "Packet-Based Attack Protection Disabled":       ["NET-04"],
    "User Credential Submission Uncontrolled":       ["IAC-06", "NET-04"],
    "Advanced Threat Prevention Not Enabled":        ["TDA-02", "NET-14"],
    "Threat Intelligence Blocking Not Configured":   ["NET-04", "TDA-02"],
    "Default Policy Logging Disabled":               ["MON-06"],
    "Management Interface Not Restricted":           ["NET-04", "IAC-10"],
    "SSL Decryption Not Configured":                 ["CRY-03", "NET-14"],
}


def _scf_label(ctrl_ids: list[str]) -> str:
    return " · ".join(ctrl_ids)


# ── Vuln reference mapping ────────────────────────────────────────────────────
def _load_vulns(path: str) -> dict[int, str]:
    """Parse vulns.txt (one description per line, line number = vuln ID)."""
    vulns: dict[int, str] = {}
    try:
        with open(path, encoding="utf-8") as fh:
            for vuln_id, line in enumerate(fh, 1):
                text = line.strip()
                if text:
                    vulns[vuln_id] = text
        return vulns
    except FileNotFoundError:
        return vulns


def _find_vulns_file() -> dict[int, str]:
    """Try to load vulns.txt from the script directory, then CWD."""
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "vulns.txt"),
        os.path.join(os.getcwd(), "vulns.txt"),
    ]
    for path in candidates:
        vulns = _load_vulns(path)
        if vulns:
            return vulns
    print("[!] Warning: vulns.txt not found — Vuln column will be empty. "
          "Place vulns.txt alongside pa_analyzer.py.")
    return {}


# Maps each check category to its vuln ID in vulns.txt
CATEGORY_VULN_ID: dict[str, int] = {
    # ── Firewall rule checks ───────────────────────────────────────────────────
    "Any/Any/Any Allow Rule":                        9,
    "Missing Security Profiles":                     6,
    "No Logging Configured":                         28,
    "Allow Rule Not Logging Session End":            28,
    "Unrestricted Source Address":                   9,
    "Unrestricted Destination Address":              9,
    "Exposed RDP from Any Source":                   9,
    "Cleartext Telnet Allowed":                      4,
    "SSH Exposed from Any Source":                   9,
    "SMB Exposed from Any Source":                   9,
    "VNC Exposed from Any Source":                   9,
    "Risky Application Allowed Without User-ID Restriction": 27,
    "Disabled Rule":                                 2,
    "Missing Rule Description":                      5,
    "Negated Source Address":                        9,
    "Negated Destination Address":                   9,
    "Zone Missing Protection Profile":               6,
    "Potential Shadow Rule":                         30,
    "Application+Service Both Any":                  9,
    "Inbound Allow Without Inspection":              6,
    "Service=Any with Specific Application":         9,
    # ── Crypto checks ─────────────────────────────────────────────────────────
    "Weak IKE Encryption":                           29,
    "Weak IKE Hash/PRF":                             29,
    "Weak IKE DH Group":                             29,
    "Weak IPSec Encryption":                         29,
    "Weak IPSec Authentication":                     29,
    "Weak IPSec DH Group (PFS)":                     29,
    "IPSec PFS Disabled":                            31,
    "Weak Minimum TLS Version":                      29,
    "IKEv1 in Use":                                  29,
    "IKE Pre-Shared Key Authentication":             27,
    # ── Management / system checks ────────────────────────────────────────────
    "HTTP Management Enabled":                       4,
    "Telnet Management Enabled":                     4,
    "No Management IP Restrictions":                 14,
    "NTP Not Configured":                            18,
    "No Login Banner":                               21,
    "DNS Not Configured":                            17,
    "Admin Without Authentication Profile":          27,
    "Admin Account Has No Password":                 22,
    "Excessive Superuser Accounts":                  3,
    "SNMPv1 Enabled":                                15,
    "SNMPv2c Enabled":                               15,
    "Default/Weak SNMP Community String":            22,
    "SNMP Enabled Without Source Restrictions":      14,
    "No Syslog Servers Configured":                  28,
    "Syslog Transmitted Over UDP":                   10,
    # ── Password & session policy ─────────────────────────────────────────────
    "Password Complexity Not Enforced":              27,
    "Weak Password Minimum Length":                  27,
    "No Account Lockout Policy":                     27,
    "Long or No Management Session Timeout":         27,
    "Password Expiry Not Configured":                3,
    "Insufficient Password History":                 3,
    # ── Content updates ───────────────────────────────────────────────────────
    "AV/Threat Content Updates Not Automatic":       1,
    "WildFire Updates Not Automatic":                1,
    # ── Security profile quality ──────────────────────────────────────────────
    "Vulnerability Profile Allows Critical/High Threats": 6,
    "WildFire Profile Missing Rules":                6,
    "WildFire Profile Incomplete Coverage":          6,
    "File Blocking Not Applied":                     6,
    # ── Zone / User-ID ────────────────────────────────────────────────────────
    "User-ID Enabled on Untrusted Zone":             3,
    # ── NTP ───────────────────────────────────────────────────────────────────
    "Only One NTP Server":                           18,
    "NTP Authentication Not Configured":             7,
    # ── Insecure protocols / certificates ────────────────────────────────────
    "Insecure Protocol Allowed in Rule":             4,
    "TLS Profile Using Default Certificate":         38,
    # ── Rule completeness ─────────────────────────────────────────────────────
    "No Default Deny Rule":                          9,
    # ── Policy Optimizer / rule-usage (CSV-sourced rules) ─────────────────────
    "Unused Security Rule (Policy Optimizer)":       25,
    # ── CIS L2 benchmark checks ───────────────────────────────────────────────
    "Admin Interface Default Certificate":           38,
    "WMI Probing Enabled":                           3,
    "Zone Flood Protection Disabled":                9,
    "Decryption Certificate Untrusted":              38,
    "SNMPv3 Trap Not Configured":                    28,
    # ── CIS L1 benchmark checks ───────────────────────────────────────────────
    "High DP Load Logging Disabled":                 28,
    "SNMPv3 Polling Not Configured":                 37,
    "Password Profile Bypass Configured":            3,
    "User-ID Network Exclusions Not Configured":     3,
    "User-ID Agent Traffic Unrestricted":            9,
    "High Availability Not Configured":              11,
    "WildFire File Size Limits Not Maximized":       6,
    "WildFire Profile Not Applied":                  6,
    "WildFire Decrypted Content Not Forwarded":      6,
    "WildFire Session Information Incomplete":       6,
    "WildFire Malicious File Alerts Disabled":       6,
    "WildFire Inline ML Not Enabled":                6,
    "Antivirus Profile Not Blocking":                6,
    "Antivirus Profile Not Applied":                 6,
    "Anti-Spyware Profile Not Blocking":             6,
    "Anti-Spyware Profile Not Applied":              6,
    "Vulnerability Profile Not Applied":             6,
    "URL Filtering Not Configured":                  6,
    "URL Category Action Not Blocking":              6,
    "URL Filtering Not Logging":                     28,
    "URL Filtering Profile Not Applied":             6,
    "Data Filtering Not Applied":                    6,
    "User Credential Submission Uncontrolled":       6,
    "Advanced Threat Prevention Not Enabled":        6,
    "Default Policy Logging Disabled":               28,
    "Management Interface Not Restricted":           14,
}


# ── CIS audit file registries ────────────────────────────────────────────────
_CIS_L1_AUDIT_IN_TAR: dict[int, str] = {
    6:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_6_Benchmark_L1_v1.0.0.audit",
    7:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_7_Benchmark_L1_v1.0.0.audit",
    8:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_8_Benchmark_L1_v1.0.0.audit",
    9:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_9_Benchmark_v1.1.0_L1.audit",
    10: "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_10_Benchmark_v1.3.0_L1.audit",
    11: "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_11_Benchmark_v1.2.0_L1.audit",
}

_CIS_L2_AUDIT_IN_TAR: dict[int, str] = {
    6:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_6_Benchmark_L2_v1.0.0.audit",
    7:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_7_Benchmark_L2_v1.0.0.audit",
    8:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_8_Benchmark_L2_v1.0.0.audit",
    9:  "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_9_Benchmark_v1.1.0_L2.audit",
    10: "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_10_Benchmark_v1.3.0_L2.audit",
    11: "portal_audits/palo_alto/CIS_Palo_Alto_Firewall_11_Benchmark_v1.2.0_L2.audit",
}


_TAR_SEARCH_SKIP_DIRS = {".git", "__pycache__", "node_modules", ".venv", "venv", ".tox", "dist", "build"}


def _find_audits_tar() -> str | None:
    """Return path to audits.tar.gz.

    Checks alongside the script and in the CWD first (fast path, preserves prior
    behavior/priority), then falls back to a recursive walk under both of those
    directories so a bundle placed in a subfolder is still picked up. Common
    noise directories (.git, __pycache__, venvs, node_modules, etc.) are skipped.
    """
    roots = list(dict.fromkeys([os.path.dirname(os.path.abspath(__file__)), os.getcwd()]))

    for base in roots:
        direct = os.path.join(base, "audits.tar.gz")
        if os.path.isfile(direct):
            return direct

    for base in roots:
        for root, dirs, files in os.walk(base):
            dirs[:] = [d for d in dirs if d not in _TAR_SEARCH_SKIP_DIRS and not d.startswith(".")]
            if "audits.tar.gz" in files:
                return os.path.join(root, "audits.tar.gz")
    return None


# ── XSLT-based audit execution helpers ───────────────────────────────────────

# XPath substitutions: adapt Panorama/template paths to standalone device layout
# (In a standalone config export, deviceconfig and network are top-level under
#  <config>, not nested under devices/entry as Panorama templates assume.)
_XSLT_PATH_SUBS: list[tuple[str, str]] = [
    ('/response/result/config/devices/entry/deviceconfig/', '//deviceconfig/'),
    ('/response/result/config/devices/entry/network/',      '//network/'),
    ('//devices/entry/network/',                            '//network/'),
]

# Maps vague sub-check descriptions to (category, description) tuples.
# Keys are the raw check description strings from the audit file.
_AUDIT_SUBCHECK_META: dict[str, tuple[str, str]] = {
    # CIS 1.1.1.1 – syslog server
    "host":          ("No Syslog Servers Configured",
                      "No syslog server host is configured for log forwarding. Security events "
                      "will not be forwarded to an external log aggregation platform."),
    # CIS 1.1.1.2 – SNMP trap forwarding (L1 + L2)
    "system":        ("SNMPv3 Trap Not Configured",
                      "SNMP v3 traps are not configured to forward system log events. "
                      "Security events in the system log will not reach the SNMP monitoring platform."),
    "configuration": ("SNMPv3 Trap Not Configured",
                      "SNMP v3 traps are not configured to forward configuration change log events."),
    "user-id":       ("SNMPv3 Trap Not Configured",
                      "SNMP v3 traps are not configured to forward User-ID log events."),
    "hip match":     ("SNMPv3 Trap Not Configured",
                      "SNMP v3 traps are not configured to forward HIP match log events."),
    "ip-tag":        ("SNMPv3 Trap Not Configured",
                      "SNMP v3 traps are not configured to forward IP tag log events."),
    # CIS 1.2.x – management interface services
    "SSH":           ("Management Interface Not Restricted",
                      "SSH management access is not restricted to permitted source IP addresses. "
                      "Unrestricted SSH access exposes the management plane to brute-force and lateral movement."),
    "HTTPS":         ("Management Interface Not Restricted",
                      "HTTPS management access is not restricted to permitted source IP addresses."),
    "SNMP":          ("SNMPv3 Polling Not Configured",
                      "SNMP management polling is not using SNMPv3. Earlier SNMP versions lack "
                      "cryptographic authentication and privacy, exposing device metrics and MIB data."),
    "HTTP":          ("HTTP Management Enabled",
                      "HTTP is enabled on the management interface. Management credentials and "
                      "session tokens are transmitted in cleartext over HTTP."),
    "Telnet":        ("Telnet Management Enabled",
                      "Telnet is enabled on the management interface. All management traffic, "
                      "including credentials, is transmitted in cleartext."),
    # CIS 1.4.1 – lockout settings
    "Lockout Time":  ("No Account Lockout Policy",
                      "Admin account lockout duration is not configured or is too short to deter "
                      "brute-force login attempts."),
    "Failed Attempts": ("No Account Lockout Policy",
                        "Admin account lockout threshold (number of failed attempts before lockout) "
                        "is not configured."),
    # CIS 1.x – NTP sub-checks
    "Ensure primary-ntp-server is configured": ("NTP Not Configured",
                      "No primary NTP server is configured. Without accurate time synchronization, "
                      "log timestamps cannot be trusted for incident correlation or audit purposes."),
    "Ensure secondary-ntp-server is configured": ("Only One NTP Server",
                      "No secondary NTP server is configured. A single NTP server is a single point "
                      "of failure for time synchronization."),
    # CIS 2.7 – User-ID
    "Check that User ID is enabled on any profile": ("User-ID Not Configured",
                      "User-ID is not enabled on any security profile. Traffic policies can only "
                      "enforce access by IP address, not user identity."),
    # CIS 3.1 – HA sub-checks
    "Path Monitoring Failure Condition": ("High Availability Not Configured",
                      "High availability path monitoring failure condition is not configured. "
                      "Failover may not trigger when critical network paths fail."),
    "Link Monitoring Failure Condition": ("High Availability Not Configured",
                      "High availability link monitoring failure condition is not configured."),
    "Passive Link State": ("High Availability Not Configured",
                      "High availability passive link state is not set to shutdown on failover, "
                      "which may cause traffic to be misrouted through the passive unit."),
    "Election Setings": ("High Availability Not Configured",
                      "High availability election settings are not properly configured."),
    # CIS 6.3 / 6.24 – Anti-spyware DNS sub-checks
    "Anti-Spyware Profile DNS Signature Source Sinkhole": ("Anti-Spyware Profile Not Blocking",
                      "Anti-spyware profile is not configured with DNS sinkhole to redirect "
                      "malicious DNS queries. Infected hosts can continue resolving C2 domains."),
    "Anti-Spyware Profile DNS Security Policies": ("Anti-Spyware Profile Not Blocking",
                      "Anti-spyware profile does not have DNS security policies configured to "
                      "block known malicious domain categories."),
    "Anti-Spyware Profile DNS Sinkhole": ("Anti-Spyware Profile Not Blocking",
                      "Anti-spyware profile DNS sinkhole configuration is missing or incomplete."),
    "Anti-Spyware Profile DNS Security Command and Control Domains": ("Anti-Spyware Profile Not Blocking",
                      "Anti-spyware profile is not configured to block DNS queries to "
                      "command-and-control domains."),
    # CIS 6.10 – URL filtering HTTP header logging
    "Log Container Page": ("URL Filtering Not Logging",
                      "URL filtering profile is not logging container page requests, reducing "
                      "visibility into nested web content and iframe-based threats."),
    "User-Agent":    ("URL Filtering Not Logging",
                      "URL filtering profile is not logging the HTTP User-Agent header. "
                      "User-Agent data is essential for detecting malicious tools and bots."),
    "Referer":       ("URL Filtering Not Logging",
                      "URL filtering profile is not logging the HTTP Referer header."),
    "X-Forwarded-For": ("URL Filtering Not Logging",
                      "URL filtering profile is not logging the X-Forwarded-For header. "
                      "Source IP attribution is impaired for proxied or NAT'd clients."),
    # CIS 6.12/6.14 – data filtering sub-checks
    "Data Object":   ("Data Filtering Not Applied",
                      "No data objects are defined for data filtering. Sensitive content patterns "
                      "such as credit card numbers or SSNs are not being matched."),
    "Data Filtering Profile": ("Data Filtering Not Applied",
                      "No data filtering profile is configured on the firewall."),
    # CIS 7.4 – default policy logging sub-checks
    "Policies":      ("SSL Decryption Not Configured",
                      "No SSL Forward Proxy decryption policies are configured. Encrypted traffic "
                      "from internal hosts is not being inspected, masking threats in HTTPS."),
    "Invalid Categories": ("SSL Decryption Not Configured",
                      "SSL decryption profile category configuration is invalid or incomplete."),
}

# CIS check number prefix → category name (drives CIS/PCI/SCF/Vuln column lookups)
_CIS_NUM_TO_CATEGORY: dict[str, str] = {
    # L2 checks
    "1.1.1.2": "SNMPv3 Trap Not Configured",
    "1.2.5":   "Admin Interface Default Certificate",
    "2.2":     "WMI Probing Enabled",
    "6.16":    "Zone Flood Protection Disabled",
    "8.3":     "Decryption Certificate Untrusted",
    # L1 checks
    "1.1.2":   "No Login Banner",
    "1.1.3":   "High DP Load Logging Disabled",
    "1.2.1":   "Management Interface Not Restricted",    # PAN-OS 11+
    "1.2.3":   "HTTP Management Enabled",
    "1.3.1":   "Password Complexity Not Enforced",
    "1.3.2":   "Weak Password Minimum Length",
    "1.3.3":   "Password Complexity Not Enforced",
    "1.3.4":   "Password Complexity Not Enforced",
    "1.3.5":   "Password Complexity Not Enforced",
    "1.3.6":   "Password Complexity Not Enforced",
    "1.3.7":   "Password Expiry Not Configured",
    "1.3.8":   "Insufficient Password History",
    "1.3.9":   "Insufficient Password History",
    "1.3.10":  "Password Profile Bypass Configured",
    "1.4.1":   "Long or No Management Session Timeout",
    "1.5.1":   "SNMPv3 Polling Not Configured",
    "1.6.1":   "Update Server Verification Disabled",
    "1.6.2":   "Only One NTP Server",                   # PAN-OS 11+
    "2.4":     "User-ID Network Exclusions Not Configured",  # PAN-OS 11+
    "2.8":     "User-ID Agent Traffic Unrestricted",
    "3.1":     "High Availability Not Configured",
    "4.1":     "AV/Threat Content Updates Not Automatic",
    "4.2":     "AV/Threat Content Updates Not Automatic",
    "5.1":     "WildFire File Size Limits Not Maximized",
    "5.2":     "WildFire Profile Not Applied",
    "5.3":     "WildFire Decrypted Content Not Forwarded",
    "5.4":     "WildFire Session Information Incomplete",
    "5.5":     "WildFire Malicious File Alerts Disabled",
    "5.6":     "WildFire Updates Not Automatic",
    "5.8":     "WildFire Inline ML Not Enabled",
    "6.1":     "Antivirus Profile Not Blocking",
    "6.2":     "Antivirus Profile Not Applied",
    "6.3":     "Anti-Spyware Profile Not Blocking",
    "6.5":     "Anti-Spyware Profile Not Applied",
    "6.6":     "Vulnerability Profile Allows Critical/High Threats",
    "6.7":     "Vulnerability Profile Not Applied",
    "6.8":     "URL Filtering Not Configured",
    "6.9":     "URL Category Action Not Blocking",
    "6.10":    "URL Filtering Not Logging",
    "6.12":    "URL Filtering Profile Not Applied",
    "6.14":    "Data Filtering Not Applied",
    "6.15":    "Zone Flood Protection Disabled",
    "6.17":    "Reconnaissance Protection Disabled",
    "6.18":    "Packet-Based Attack Protection Disabled",
    "6.19":    "User Credential Submission Uncontrolled",
    "6.20":    "Antivirus Profile Not Blocking",
    "6.21":    "Antivirus Profile Not Blocking",
    "6.22":    "Advanced Threat Prevention Not Enabled",
    "6.23":    "Advanced Threat Prevention Not Enabled",
    "6.24":    "Anti-Spyware Profile Not Blocking",
    "7.2":     "Service=Any with Specific Application",
    "7.3":     "Threat Intelligence Blocking Not Configured",
    "7.4":     "Default Policy Logging Disabled",
}

# Categories fully covered by the always-running Python checks.
# XSLT findings in these categories are suppressed to prevent duplicates.
_PYTHON_COVERED_CATEGORIES: frozenset[str] = frozenset({
    "HTTP Management Enabled",
    "Telnet Management Enabled",
    "No Management IP Restrictions",
    "Management Interface Not Restricted",
    "NTP Not Configured",
    "Only One NTP Server",
    "NTP Authentication Not Configured",
    "No Login Banner",
    "DNS Not Configured",
    "No Syslog Servers Configured",
    "Syslog Transmitted Over UDP",
    "SNMPv1 Enabled",
    "SNMPv2c Enabled",
    "Default/Weak SNMP Community String",
    "SNMP Enabled Without Source Restrictions",
    "Password Complexity Not Enforced",
    "Weak Password Minimum Length",
    "No Account Lockout Policy",
    "Long or No Management Session Timeout",
    "Password Expiry Not Configured",
    "Insufficient Password History",
    "Admin Without Authentication Profile",
    "Admin Account Has No Password",
    "Excessive Superuser Accounts",
    "AV/Threat Content Updates Not Automatic",
    "WildFire Updates Not Automatic",
    "Vulnerability Profile Allows Critical/High Threats",
    "WildFire Profile Missing Rules",
    "WildFire Profile Incomplete Coverage",
    "No Default Deny Rule",
    "File Blocking Not Applied",
    "Missing Security Profiles",
    "No Logging Configured",
    "Allow Rule Not Logging Session End",
    "Service=Any with Specific Application",
    "TLS Profile Using Default Certificate",
    "Weak Minimum TLS Version",
})

_FW_REF_RE = re.compile(
    r'\b(?:CIS\s+[\d\.]+|PCI[\s\-]*DSS\s+[\d\.]+)',
    re.IGNORECASE,
)


def _strip_fw_refs(text: str) -> str:
    return re.sub(r'\s{2,}', ' ', _FW_REF_RE.sub('', text)).strip()


def _first_para(text: str) -> str:
    """Return the first substantive paragraph before any Rationale/Impact section."""
    for marker in ("\nRationale:", "\n\nRationale:", "\nImpact:", "\n\nImpact:"):
        if marker in text:
            text = text[:text.index(marker)]
    paras = [p.strip() for p in text.split('\n\n') if p.strip()]
    return paras[0] if paras else text.strip()


def _parse_audit_op_checks(content: str) -> list[dict]:
    """Extract actionable op-type custom_item checks from a .audit file."""
    blocks = re.findall(r'<custom_item>(.*?)</custom_item>', content, re.DOTALL)
    checks = []
    _SKIP_VARS  = ('@PLATFORM_VERSION@', '@SNMP_SERVER@', '@LOG_SERVER@')
    _SKIP_DESCS = {'panorama model', 'panorama system-mode'}
    for b in blocks:
        if 'api_request_type : "op"' not in b or 'xsl_stmt' not in b:
            continue
        if any(v in b for v in _SKIP_VARS):
            continue
        desc_m = re.search(r'description\s*:\s*"([^"]*)"', b)
        desc   = desc_m.group(1).strip() if desc_m else ""
        if desc.lower() in _SKIP_DESCS:
            continue
        if re.search(r'(?:expect|not_expect)\s*:\s*"[^"]*Manual Review', b):
            continue
        info_m  = re.search(r'\binfo\s*:\s*"((?:[^"\\]|\\.)*)"',     b, re.DOTALL)
        sol_m   = re.search(r'solution\s*:\s*"((?:[^"\\]|\\.)*)"',    b, re.DOTALL)
        exp_m   = re.search(r'(?m)^\s*expect\s*:\s*"([^"]*)"',        b)
        nexp_m  = re.search(r'not_expect\s*:\s*"([^"]*)"',            b)
        sev_m   = re.search(r'severity\s*:\s*(\w+)',                   b)
        stmts   = re.findall(r'xsl_stmt\s*:\s*"((?:[^"\\]|\\.)*)"',   b)
        checks.append({
            "description": desc,
            "info":        info_m.group(1).replace('\\"', '"').strip() if info_m  else "",
            "solution":    sol_m.group(1).replace('\\"', '"').strip()  if sol_m   else "",
            "expect":      exp_m.group(1)                              if exp_m   else "",
            "not_expect":  nexp_m.group(1)                            if nexp_m  else "",
            "severity":    sev_m.group(1).upper()                     if sev_m   else "MEDIUM",
            "xsl_stmts":   [s.replace('\\"', '"') for s in stmts],
        })
    return checks


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "hdr_bg":   "1F3864", "hdr_fg":   "FFFFFF",
    "critical": "C00000", "critical_l": "FFB3B3",
    "high":     "FF0000", "high_l":     "FFD9B3",
    "medium":   "FF8C00", "medium_l":   "FFF2CC",
    "low":      "0070C0", "low_l":      "BDD7EE",
    "info":     "595959", "info_l":     "F2F2F2",
    "allow":    "375623", "allow_l":    "E2EFDA",
    "deny":     "C00000", "deny_l":     "FFB3B3",
    "disabled": "A6A6A6",
    "alt_row":  "F5F5F5",
}

_thin_side = Side(style="thin", color="CCCCCC")
THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=size)


def _align(h="left", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Config parser ─────────────────────────────────────────────────────────────
class PaloAltoParser:
    def __init__(self, config_file: "str | None", csv_rules_path: "str | None" = None,
                 panos_version_override: "int | None" = None):
        self.config_file = config_file  # None => rules-only mode (CSV, no device config)
        self.csv_rules_path = csv_rules_path
        # Basename to show in reports — falls back to the CSV when there's no device config.
        self.source_label = os.path.basename(config_file) if config_file else \
            (os.path.basename(csv_rules_path) if csv_rules_path else "(none)")
        self.panos_version_override = panos_version_override
        self.config_format = "xml"  # set to "curly" in parse() when detected
        self.root: ET.Element | None = None

        self.security_rules: list[dict] = []
        self.nat_rules: list[dict] = []
        self.address_objects: list[dict] = []
        self.address_groups: list[dict] = []
        self.service_objects: list[dict] = []
        self.service_groups: list[dict] = []
        self.zones: list[dict] = []
        self.profile_groups: dict[str, dict] = {}
        self.issues: list[dict] = []
        self._seen_issues: set[tuple[str, str]] = set()

        # Crypto / system data
        self.ike_crypto_profiles: list[dict] = []
        self.ipsec_crypto_profiles: list[dict] = []
        self.ssl_tls_profiles: list[dict] = []
        self.ike_gateways: list[dict] = []
        self.admin_accounts: list[dict] = []
        self.mgmt_settings: dict = {}
        self.log_syslog_servers: list[dict] = []
        self.password_policy:    dict            = {}
        self.update_schedule:    dict            = {}
        self.vuln_profiles:      dict[str, dict] = {}
        self.wildfire_profiles:  dict[str, dict] = {}

        # CIS benchmark / .audit selection (set by _run_cis_checks once parsed)
        self.panos_version_str: str = ""
        self.audits_tar_path:   str = ""
        self.cis_l1_audit_used: str = ""
        self.cis_l2_audit_used: str = ""

    # ── Parse entry point ─────────────────────────────────────────────────────
    def parse(self):
        if self.config_file is None:
            # Rules-only mode: a CSV was given with no device config at all.
            # Device/system checks and the CIS Benchmark need real config data
            # to mean anything, so _run_checks() skips them for this format —
            # an empty <config/> root just lets every _parse_* helper below
            # find nothing, same as a config file with no such section.
            self.config_format = "csv-only"
            self.root = ET.Element("config")
            self._linemap = {}
            if not self.csv_rules_path:
                sys.exit("No config file and no --rules-csv given — nothing to analyze.")
        else:
            try:
                with open(self.config_file, "r", encoding="utf-8", errors="replace") as fh:
                    sniff = fh.read(65536)
            except FileNotFoundError:
                sys.exit(f"File not found: {self.config_file}")

            if _looks_like_curly_config(sniff):
                self.config_format = "curly"
                try:
                    self.root, self._linemap = _load_curly_config(self.config_file)
                except ValueError as exc:
                    sys.exit(f"Could not parse 'show config merged' capture: {exc}")
                print(f"[+] Detected 'show config merged' CLI text format "
                      f"({os.path.basename(self.config_file)})")
                if self.panos_version_override:
                    self.root.set("version", str(self.panos_version_override))
            else:
                try:
                    self.root, self._linemap = _parse_xml_with_linenos(self.config_file)
                except ET.ParseError as exc:
                    sys.exit(f"XML parse error: {exc}")

        self._parse_objects()
        self._parse_profile_groups()
        self._parse_security_rules()
        self._parse_nat_rules()
        self._parse_zones()
        self._parse_ike_crypto_profiles()
        self._parse_ipsec_crypto_profiles()
        self._parse_ssl_tls_profiles()
        self._parse_ike_gateways()
        self._parse_admin_accounts()
        self._parse_mgmt_settings()
        self._parse_syslog_servers()
        self._parse_password_policy()
        self._parse_update_schedule()
        self._parse_security_profiles()
        if self.csv_rules_path:
            self._parse_security_rules_from_csv(self.csv_rules_path)
        self._run_checks()

    # ── Low-level helpers ─────────────────────────────────────────────────────
    @staticmethod
    def _members(el: ET.Element, path: str) -> list[str]:
        container = el.find(path)
        if container is None:
            return []
        return [m.text.strip() for m in container.findall("member") if m.text]

    @staticmethod
    def _text(el: ET.Element, path: str, default: str = "") -> str:
        found = el.find(path)
        return found.text.strip() if (found is not None and found.text) else default

    def _lineno(self, el: ET.Element) -> int:
        """Return the source line number for an element, or 0 if unknown."""
        return self._linemap.get(id(el), 0)

    def _lineno_str(self, el: ET.Element, child_paths: list[str] | None = None) -> str:
        """Return a human-readable line reference like '42' or '42-51'."""
        start = self._lineno(el)
        if not start:
            return ""
        if child_paths:
            end = start
            for path in child_paths:
                child = el.find(path)
                if child is not None:
                    ln = self._lineno(child)
                    if ln and ln > end:
                        end = ln
            return f"{start}-{end}" if end != start else str(start)
        return str(start)

    # ── Object parsing ────────────────────────────────────────────────────────
    def _parse_objects(self):
        """Collect address / service objects from all scopes."""
        scopes = []
        for vsys in self.root.findall(".//vsys/entry"):
            scopes.append((vsys.get("name", "vsys1"), vsys))
        shared = self.root.find(".//shared")
        if shared is not None:
            scopes.append(("shared", shared))
        for dg in self.root.findall(".//device-group/entry"):
            scopes.append((dg.get("name", "device-group"), dg))
        if not scopes:
            scopes.append(("global", self.root))

        for scope_name, scope_el in scopes:
            for addr in scope_el.findall(".//address/entry"):
                ip = self._text(addr, "ip-netmask")
                ip_range = self._text(addr, "ip-range")
                fqdn = self._text(addr, "fqdn")
                if ip:
                    atype, val = "ip-netmask", ip
                elif ip_range:
                    atype, val = "ip-range", ip_range
                elif fqdn:
                    atype, val = "fqdn", fqdn
                else:
                    atype, val = "unknown", ""
                self.address_objects.append({
                    "scope": scope_name,
                    "name": addr.get("name", ""),
                    "type": atype, "value": val,
                    "description": self._text(addr, "description"),
                    "tags": ", ".join(self._members(addr, "tag")),
                })

            for grp in scope_el.findall(".//address-group/entry"):
                dyn = self._text(grp, "dynamic/filter")
                self.address_groups.append({
                    "scope": scope_name,
                    "name": grp.get("name", ""),
                    "type": "dynamic" if dyn else "static",
                    "members": ", ".join(self._members(grp, "static")),
                    "dynamic_filter": dyn,
                    "description": self._text(grp, "description"),
                    "tags": ", ".join(self._members(grp, "tag")),
                })

            for svc in scope_el.findall(".//service/entry"):
                proto, port = "", ""
                if svc.find("protocol/tcp") is not None:
                    proto = "tcp"
                    port = self._text(svc.find("protocol/tcp"), "port")
                elif svc.find("protocol/udp") is not None:
                    proto = "udp"
                    port = self._text(svc.find("protocol/udp"), "port")
                self.service_objects.append({
                    "scope": scope_name,
                    "name": svc.get("name", ""),
                    "protocol": proto, "port": port,
                    "description": self._text(svc, "description"),
                })

            for grp in scope_el.findall(".//service-group/entry"):
                self.service_groups.append({
                    "scope": scope_name,
                    "name": grp.get("name", ""),
                    "members": ", ".join(self._members(grp, "members")),
                    "tags": ", ".join(self._members(grp, "tag")),
                })

    def _parse_profile_groups(self):
        for grp in self.root.findall(".//profile-group/entry"):
            name = grp.get("name", "")
            self.profile_groups[name] = {
                "virus":            self._text(grp, "virus/member"),
                "vulnerability":    self._text(grp, "vulnerability/member"),
                "spyware":          self._text(grp, "spyware/member"),
                "url-filtering":    self._text(grp, "url-filtering/member"),
                "file-blocking":    self._text(grp, "file-blocking/member"),
                "wildfire-analysis": self._text(grp, "wildfire-analysis/member"),
            }

    # ── Security rules ────────────────────────────────────────────────────────
    def _parse_security_rules(self):
        paths = [
            ".//rulebase/security/rules/entry",
            ".//pre-rulebase/security/rules/entry",
            ".//post-rulebase/security/rules/entry",
        ]
        rule_num = 1
        for path in paths:
            rb = "security"
            if "pre-rulebase" in path:
                rb = "pre-security"
            elif "post-rulebase" in path:
                rb = "post-security"

            for rule in self.root.findall(path):
                ps = rule.find("profile-setting")
                profile_type, profiles = "", {}
                if ps is not None:
                    gm = ps.find("group/member")
                    if gm is not None and gm.text:
                        profile_type = "group"
                        profiles = {"group": gm.text.strip()}
                    else:
                        profile_type = "profiles"
                        profiles = {
                            k: self._text(ps, f"profiles/{k}/member")
                            for k in ("virus", "vulnerability", "spyware",
                                      "url-filtering", "file-blocking", "wildfire-analysis")
                        }

                self.security_rules.append({
                    "num": rule_num,
                    "name":          rule.get("name", f"rule-{rule_num}"),
                    "line":          self._lineno_str(rule, ["action", "description"]),
                    "rulebase":      rb,
                    "disabled":      self._text(rule, "disabled", "no"),
                    "src_zones":     ", ".join(self._members(rule, "from")),
                    "dst_zones":     ", ".join(self._members(rule, "to")),
                    "sources":       ", ".join(self._members(rule, "source")),
                    "destinations":  ", ".join(self._members(rule, "destination")),
                    "negate_src":    self._text(rule, "negate-source", "no"),
                    "negate_dst":    self._text(rule, "negate-destination", "no"),
                    "applications":  ", ".join(self._members(rule, "application")),
                    "services":      ", ".join(self._members(rule, "service")),
                    "categories":    ", ".join(self._members(rule, "category")),
                    "action":        self._text(rule, "action", "allow"),
                    "profile_type":  profile_type,
                    "profiles":      profiles,
                    "log_start":     self._text(rule, "log-start", "no"),
                    "log_end":       self._text(rule, "log-end", "yes"),
                    "log_setting":   self._text(rule, "log-setting"),
                    "hip_profiles":  ", ".join(self._members(rule, "hip-profiles")),
                    "schedule":      self._text(rule, "schedule"),
                    "tags":          ", ".join(self._members(rule, "tag")),
                    "description":   self._text(rule, "description"),
                    "source_user":   ", ".join(self._members(rule, "source-user")),
                })
                rule_num += 1

    # ── Security rules from a Policy Optimizer / rulebase CSV export ─────────
    @staticmethod
    def _csv_cell(v: str) -> str:
        """Normalize a CSV cell: strip whitespace and PAN-OS UI placeholder
        dashes ('-') for "not applicable" down to an empty string."""
        v = (v or "").strip()
        return "" if v in ("-", "--", "n/a", "N/A") else v

    def _parse_security_rules_from_csv(self, csv_path: str):
        """Load the Security Rulebase from a Palo Alto rule-usage CSV export.

        Two real export shapes are supported for the usage columns (the rest
        of the header is identical between them):
          - a single status column: "Rule Usage Rule Usage" (e.g. "Unused")
            plus "Rule Usage Apps Seen" / "Days With No New Apps".
          - separate "Rule Usage Hit Count" / "Rule Usage Last Hit" /
            "Rule Usage First Hit" columns (a 0 hit count means unused);
            "Location"/"Target" may also be absent in this shape.
        Whichever is present is normalized into the same internal fields so
        the Policy Optimizer checks below don't need to care which was used.

        This REPLACES any rules already parsed from the config file: the CSV is
        the ground truth for rule content when supplied — a 'show config merged'
        capture can have an empty rulebase when rules are pushed from Panorama
        and weren't captured locally, which is exactly the case a CSV export
        is meant to fill in. Fields the CSV doesn't carry (per-rule negate
        flags, URL categories, description) are left at PAN-OS defaults
        rather than guessed; per-rule logging IS inferred where possible from
        the free-text "Options" column (see below).
        """
        import csv as _csv

        try:
            fh = open(csv_path, "r", encoding="utf-8-sig", newline="")
        except FileNotFoundError:
            sys.exit(f"Rules CSV not found: {csv_path}")

        with fh:
            reader = _csv.DictReader(fh)
            if reader.fieldnames is None:
                sys.exit(f"Rules CSV has no header row: {csv_path}")

            loaded: list[dict] = []
            for rule_num, raw_row in enumerate(reader, 1):
                row = {(k or "").strip(): self._csv_cell(v) for k, v in raw_row.items()}
                name = row.get("Name", "")
                if not name:
                    continue

                # "Profile" is either a bare name, "none"/"disabled", or
                # "Profile Group: <name>" — strip that prefix when present.
                profile_raw = row.get("Profile", "")
                grp_m = re.match(r'^\s*Profile\s+Group\s*:\s*(.+)$', profile_raw, re.IGNORECASE)
                profile_name = grp_m.group(1).strip() if grp_m else profile_raw
                if profile_name and profile_name.lower() not in ("none", "disabled"):
                    profile_type, profiles = "group", {"group": profile_name}
                else:
                    profile_type, profiles = "", {}

                # "Options" is a ';'-joined free-text list, e.g.:
                #   "Traffic log sent at session start, and at session end;
                #    Log Forwarding Profile setting: PRIMARY-FORWARDING"
                # An explicit "Traffic log sent..." clause overrides the
                # PAN-OS default (log-end only); its total absence from a
                # non-empty Options value is read as logging being off.
                options = row.get("Options", "")
                traffic_log_m = re.search(r'Traffic log sent[^;]*', options, re.IGNORECASE)
                if traffic_log_m:
                    clause = traffic_log_m.group(0)
                    log_start = "yes" if re.search(r'session start', clause, re.IGNORECASE) else "no"
                    log_end   = "yes" if re.search(r'session end',   clause, re.IGNORECASE) else "no"
                elif options:
                    log_start, log_end = "no", "no"
                else:
                    log_start, log_end = "no", "yes"  # PAN-OS default, nothing to infer from
                log_setting_m = re.search(r'Log Forwarding(?:\s+Profile\s+setting)?\s*:\s*([^;]+)',
                                           options, re.IGNORECASE)
                schedule_m = re.search(r'Schedule\s*:\s*([^;]+)', options, re.IGNORECASE)

                action = (row.get("Action", "allow") or "allow").strip().lower()

                # Usage-status normalization across the two export shapes.
                usage_status = row.get("Rule Usage Rule Usage", "")
                hit_count_raw = row.get("Rule Usage Hit Count", "")
                if not usage_status and hit_count_raw:
                    digits = re.sub(r"[^\d]", "", hit_count_raw)
                    if digits:
                        usage_status = "Unused" if int(digits) == 0 else f"Used ({digits} hits)"

                loaded.append({
                    "num": rule_num,
                    "name": name,
                    "line": f"CSV row {rule_num + 1}",
                    "rulebase": "security",
                    "disabled": "no",   # not present as a column in either export shape
                    "src_zones":    row.get("Source Zone", ""),
                    "dst_zones":    row.get("Destination Zone", ""),
                    "sources":      row.get("Source Address", "") or "any",
                    "destinations": row.get("Destination Address", "") or "any",
                    "negate_src": "no",
                    "negate_dst": "no",
                    "applications": row.get("Application", "") or "any",
                    "services":     row.get("Service", "") or "any",
                    "categories": "",
                    "action": action,
                    "profile_type": profile_type,
                    "profiles": profiles,
                    "log_start": log_start,
                    "log_end":   log_end,
                    "log_setting": log_setting_m.group(1).strip() if log_setting_m else "",
                    "hip_profiles": "",
                    "schedule": schedule_m.group(1).strip() if schedule_m else "",
                    "tags": row.get("Tags", ""),
                    "description": "",
                    # Policy Optimizer / usage-report extras — consumed only by
                    # the CSV-specific checks below; every other check ignores
                    # unknown dict keys.
                    "rule_type":        row.get("Type", ""),
                    "source_user":      row.get("Source User", ""),
                    "source_device":    row.get("Source Device", ""),
                    "dest_device":      row.get("Destination Device", ""),
                    "target":           row.get("Target", ""),
                    "csv_location":     row.get("Location", ""),
                    "usage_status":     usage_status,
                    "hit_count":        row.get("Rule Usage Hit Count", ""),
                    "last_hit":         row.get("Rule Usage Last Hit", ""),
                    "first_hit":        row.get("Rule Usage First Hit", ""),
                    "apps_seen":        row.get("Rule Usage Apps Seen", ""),
                    "days_no_new_apps": row.get("Days With No New Apps", ""),
                    "modified": row.get("Modified", ""),
                    "created":  row.get("Created", ""),
                })

        if self.security_rules:
            print(f"[!] Discarding {len(self.security_rules)} security rule(s) parsed from "
                  f"{self.config_format} config — CSV rules take precedence.")
        self.security_rules = loaded
        print(f"[+] Loaded {len(loaded)} security rule(s) from CSV: {csv_path}")

    # ── NAT rules ─────────────────────────────────────────────────────────────
    def _parse_nat_rules(self):
        paths = [
            ".//rulebase/nat/rules/entry",
            ".//pre-rulebase/nat/rules/entry",
            ".//post-rulebase/nat/rules/entry",
        ]
        rule_num = 1
        for path in paths:
            for rule in self.root.findall(path):
                # Source translation
                st = rule.find("source-translation")
                src_type, src_val = "", ""
                if st is not None:
                    for st_kind in ("dynamic-ip-and-port", "dynamic-ip", "static-ip"):
                        el = st.find(st_kind)
                        if el is not None:
                            src_type = st_kind
                            addrs = self._members(el, "translated-address")
                            iface_el = el.find("interface-address")
                            if addrs:
                                src_val = ", ".join(addrs)
                            elif iface_el is not None:
                                src_val = f"interface:{self._text(iface_el, 'interface')}"
                            break

                # Destination translation
                dt = rule.find("destination-translation")
                dst_addr = self._text(dt, "translated-address") if dt is not None else ""
                dst_port = self._text(dt, "translated-port") if dt is not None else ""

                self.nat_rules.append({
                    "num": rule_num,
                    "name":        rule.get("name", f"nat-{rule_num}"),
                    "line":        self._lineno_str(rule),
                    "disabled":    self._text(rule, "disabled", "no"),
                    "nat_type":    self._text(rule, "nat-type", "ipv4"),
                    "src_zones":   ", ".join(self._members(rule, "from")),
                    "dst_zones":   ", ".join(self._members(rule, "to")),
                    "sources":     ", ".join(self._members(rule, "source")),
                    "destinations": ", ".join(self._members(rule, "destination")),
                    "services":    ", ".join(self._members(rule, "service")),
                    "src_trans_type":  src_type,
                    "src_trans_value": src_val,
                    "dst_trans_addr":  dst_addr,
                    "dst_trans_port":  dst_port,
                    "description": self._text(rule, "description"),
                    "tags":        ", ".join(self._members(rule, "tag")),
                })
                rule_num += 1

    # ── Zones ─────────────────────────────────────────────────────────────────
    def _parse_zones(self):
        for zone in self.root.findall(".//zone/entry"):
            zone_type, ifaces = "", []
            for zt in ("layer3", "layer2", "virtual-wire", "tap", "external", "tunnel", "loopback"):
                el = zone.find(zt)
                if el is not None:
                    zone_type = zt
                    ifaces = self._members(el, "member")
                    break
            self.zones.append({
                "name":     zone.get("name", ""),
                "line":     self._lineno_str(zone),
                "type":     zone_type,
                "interfaces": ", ".join(ifaces),
                "zone_protection_profile": self._text(zone, "network/zone-protection-profile"),
                "user_id":  self._text(zone, "enable-user-identification", "no"),
                "log_setting": self._text(zone, "log-setting"),
            })

    # ── Security checks ───────────────────────────────────────────────────────
    def _issue(self, severity, category, rule_name, description, recommendation,
               details="", line="", bench_override: "list[str] | None" = None):
        key = (category, rule_name)
        if key in self._seen_issues:
            return
        self._seen_issues.add(key)
        cis_ids   = CIS_CONTROL_MAP.get(category, [])
        pci_ids   = PCI_DSS_MAP.get(category, [])
        scf_ids   = SCF_MAP.get(category, [])
        bench_ids = bench_override if bench_override is not None else CIS_BENCHMARK_MAP.get(category, [])
        self.issues.append({
            "severity":        severity,
            "category":        category,
            "rule_name":       rule_name,
            "line":            str(line) if line else "",
            "description":     description,
            "recommendation":  recommendation,
            "details":         details,
            "cis_controls":    _cis_label(cis_ids),
            "cis_ids":         cis_ids,
            "cis_benchmark":   _cis_benchmark_label(bench_ids),
            "pci_dss":         _pci_label(pci_ids),
            "pci_ids":         pci_ids,
            "scf":             _scf_label(scf_ids),
        })

    def _run_checks(self):
        self._chk_any_any_any()
        self._chk_missing_profiles()
        self._chk_logging()
        self._chk_overly_permissive()
        self._chk_risky_services_from_any()
        self._chk_risky_apps_no_user_restriction()
        self._chk_insecure_cleartext_apps()
        self._chk_disabled_rules()
        self._chk_missing_descriptions()
        self._chk_negate_rules()
        self._chk_zones_no_protection()
        self._chk_shadow_rules()
        self._chk_app_any_svc_any_allow()
        self._chk_inbound_no_inspection()
        self._chk_service_any_allow()
        self._chk_user_id_untrust()
        self._chk_default_deny_rule()
        self._chk_file_blocking_inbound()
        # Policy Optimizer / rule-usage checks (only fire on CSV-sourced rules)
        self._chk_policy_optimizer_unused()
        self._chk_policy_optimizer_stale_apps()

        if self.config_format == "csv-only":
            # No device config was supplied — skip device/system checks and
            # the CIS Benchmark entirely rather than report a wall of "not
            # configured" findings for settings we simply have no data on.
            return

        # Crypto / system checks — need a real device config
        self._chk_weak_ike_crypto()
        self._chk_weak_ipsec_crypto()
        self._chk_weak_ssl_tls()
        self._chk_tls_default_cert()
        self._chk_ike_gateways()
        self._chk_management_access()
        self._chk_admin_accounts()
        self._chk_snmp()
        self._chk_no_syslog()
        self._chk_password_policy()
        self._chk_update_schedule()
        self._chk_security_profile_settings()
        # CIS Benchmark L1 + L2 checks (XSLT-based)
        self._run_cis_checks()

    def _active_allow(self):
        return [r for r in self.security_rules if r["disabled"] != "yes" and r["action"] == "allow"]

    def _active_rules(self):
        return [r for r in self.security_rules if r["disabled"] != "yes"]

    @staticmethod
    def _has_any(field: str) -> bool:
        return "any" in (x.strip().lower() for x in field.split(","))

    @staticmethod
    def _user_suffix(r: dict) -> str:
        """'  Source User: X' for an excessive-traffic finding's Details, when
        the rule has a Source User value at all (blank omits it entirely —
        most XML-derived rules never had this field before Source User
        support was added, so there's nothing to report either way)."""
        su = (r.get("source_user", "") or "").strip()
        return f"  Source User: {su}" if su else ""

    def _rule_has_security_profiles(self, rule: dict) -> bool:
        pt = rule["profile_type"]
        if pt == "group":
            return bool(rule["profiles"].get("group"))
        if pt == "profiles":
            p = rule["profiles"]
            return bool(p.get("virus") or p.get("vulnerability") or p.get("spyware"))
        return False

    # Check 1: any/any/any allow ───────────────────────────────────────────────
    def _chk_any_any_any(self):
        for r in self._active_allow():
            if self._has_any(r["sources"]) and self._has_any(r["destinations"]) and self._has_any(r["applications"]):
                self._issue(
                    "CRITICAL", "Any/Any/Any Allow Rule", r["name"],
                    "Rule allows ALL applications from ANY source to ANY destination.",
                    "Apply least-privilege: restrict source, destination, and application.",
                    f"Zones: {r['src_zones']} → {r['dst_zones']}{self._user_suffix(r)}",
                    line=r["line"],
                )

    # Check 2: missing security profiles ─────────────────────────────────────
    def _chk_missing_profiles(self):
        for r in self._active_allow():
            if not self._rule_has_security_profiles(r):
                self._issue(
                    "HIGH", "Missing Security Profiles", r["name"],
                    "Allow rule has no Antivirus / IPS / Anti-Spyware profile.",
                    "Attach a security profile group with AV, Vulnerability, and Spyware profiles.",
                    f"App: {r['applications']}  Svc: {r['services']}",
                    line=r["line"],
                )

    # Check 3: logging ────────────────────────────────────────────────────────
    def _chk_logging(self):
        for r in self._active_rules():
            if r["log_end"] == "no" and r["log_start"] == "no":
                self._issue(
                    "HIGH", "No Logging Configured", r["name"],
                    "Both log-at-session-start and log-at-session-end are disabled.",
                    "Enable at minimum log-at-session-end on all rules for audit visibility.",
                    f"Action: {r['action']}",
                    line=r["line"],
                )
            elif r["action"] == "allow" and r["log_end"] == "no":
                self._issue(
                    "MEDIUM", "Allow Rule Not Logging Session End", r["name"],
                    "Allow rule does not log at session end; traffic details won't be captured.",
                    "Enable log-at-session-end to record bytes transferred, duration, and threat data.",
                    line=r["line"],
                )

    # Check 4: overly permissive (partial any) ────────────────────────────────
    def _chk_overly_permissive(self):
        for r in self._active_allow():
            src_any = self._has_any(r["sources"])
            dst_any = self._has_any(r["destinations"])
            app_any = self._has_any(r["applications"])
            if src_any and dst_any and app_any:
                continue
            if src_any and not dst_any:
                self._issue(
                    "HIGH", "Unrestricted Source Address", r["name"],
                    "Allow rule permits traffic from any source IP.",
                    "Restrict the source to specific trusted IP ranges.",
                    f"Dest: {r['destinations']}  App: {r['applications']}{self._user_suffix(r)}",
                    line=r["line"],
                )
            if dst_any and not src_any:
                self._issue(
                    "HIGH", "Unrestricted Destination Address", r["name"],
                    "Allow rule permits traffic to any destination IP.",
                    "Restrict the destination to required hosts/subnets only.",
                    f"Src: {r['sources']}  App: {r['applications']}{self._user_suffix(r)}",
                    line=r["line"],
                )

    # Check 5: risky services from any ────────────────────────────────────────
    def _chk_risky_services_from_any(self):
        risky = {
            "rdp":    ("Exposed RDP from Any Source",    "HIGH",
                       "RDP allowed from any source — high brute-force/ransomware risk.",
                       "Restrict to management hosts or require VPN before RDP."),
            "telnet": ("Cleartext Telnet Allowed",        "HIGH",
                       "Telnet transmits credentials in cleartext.",
                       "Replace with SSH. Block all Telnet."),
            "ssh":    ("SSH Exposed from Any Source",     "MEDIUM",
                       "SSH allowed from any source — brute-force risk.",
                       "Restrict SSH to dedicated management network."),
            "smb":    ("SMB Exposed from Any Source",     "HIGH",
                       "SMB from any source enables lateral movement and ransomware spread.",
                       "Block SMB from untrusted zones entirely."),
            "vnc":    ("VNC Exposed from Any Source",     "HIGH",
                       "VNC often lacks strong authentication and uses cleartext.",
                       "Replace with a properly authenticated remote-access solution."),
        }
        for r in self._active_allow():
            if not self._has_any(r["sources"]):
                continue
            apps_lower = r["applications"].lower()
            for keyword, (category, sev, desc, rec) in risky.items():
                if keyword in apps_lower:
                    self._issue(sev, category, r["name"], desc, rec,
                                f"Src zones: {r['src_zones']}{self._user_suffix(r)}",
                                line=r["line"])

    # Check: risky/administrative apps allowed without a Source User restriction
    def _chk_risky_apps_no_user_restriction(self):
        """User-ID note: allow rules for admin/high-risk protocols with no
        Source User restriction lose an important compensating control —
        anyone who can reach the rule's source/zone can use it, authenticated
        or not."""
        risky_apps = ("rdp", "ssh", "telnet", "vnc", "smb")
        for r in self._active_allow():
            apps_lower = r["applications"].lower()
            if not any(app in apps_lower for app in risky_apps):
                continue
            src_user = (r.get("source_user", "") or "").strip()
            if src_user and src_user.lower() != "any":
                continue
            self._issue(
                "MEDIUM", "Risky Application Allowed Without User-ID Restriction", r["name"],
                "Rule allows a high-risk administrative protocol (RDP/SSH/Telnet/VNC/SMB) "
                "without restricting Source User to a specific user or group.",
                "Add a Source User restriction (User-ID) so only authorized accounts can "
                "use this access, as a compensating control alongside address/zone limits.",
                f"Source User: {src_user or 'any'}  Apps: {r['applications']}  "
                f"Src zones: {r['src_zones']}",
                line=r["line"],
            )

    # Check 6: disabled rules ─────────────────────────────────────────────────
    def _chk_disabled_rules(self):
        for r in self.security_rules:
            if r["disabled"] == "yes":
                self._issue(
                    "LOW", "Disabled Rule", r["name"],
                    "Rule is disabled, adding clutter and potential confusion.",
                    "Remove disabled rules that are no longer needed.",
                    line=r["line"],
                )

    # Check 7: missing descriptions ───────────────────────────────────────────
    def _chk_missing_descriptions(self):
        for r in self.security_rules:
            if not r["description"]:
                self._issue(
                    "LOW", "Missing Rule Description", r["name"],
                    "No description is set; future reviewers cannot determine the rule's purpose.",
                    "Add a description stating the business owner and intent of the rule.",
                    line=r["line"],
                )

    # Check 8: negate rules ───────────────────────────────────────────────────
    def _chk_negate_rules(self):
        for r in self._active_rules():
            if r["negate_src"] == "yes":
                self._issue(
                    "MEDIUM", "Negated Source Address", r["name"],
                    "Rule uses a negated source — all IPs *except* the listed ones match.",
                    "Confirm this is intentional; consider rewriting as explicit allow/deny pairs.",
                    f"NOT ({r['sources']})",
                    line=r["line"],
                )
            if r["negate_dst"] == "yes":
                self._issue(
                    "MEDIUM", "Negated Destination Address", r["name"],
                    "Rule uses a negated destination — all IPs *except* the listed ones match.",
                    "Confirm this is intentional; consider rewriting as explicit allow/deny pairs.",
                    f"NOT ({r['destinations']})",
                    line=r["line"],
                )

    # Check 9: zones without protection profile ───────────────────────────────
    def _chk_zones_no_protection(self):
        for z in self.zones:
            if not z["zone_protection_profile"]:
                self._issue(
                    "MEDIUM", "Zone Missing Protection Profile", f"Zone: {z['name']}",
                    f"Zone '{z['name']}' has no Zone Protection Profile (DoS, flood, recon protection).",
                    "Assign a Zone Protection Profile — especially critical for external/untrust zones.",
                    f"Zone type: {z['type']}  Interfaces: {z['interfaces']}",
                    line=z["line"],
                )

    # Check 10: shadow rules ──────────────────────────────────────────────────
    def _chk_shadow_rules(self):
        active = self._active_rules()

        def covers(a_val: str, b_val: str) -> bool:
            a_set = {x.strip().lower() for x in a_val.split(",")}
            b_set = {x.strip().lower() for x in b_val.split(",")}
            if "any" in a_set:
                return True
            return b_set.issubset(a_set)

        for i, rule in enumerate(active):
            for prior in active[:i]:
                if prior["action"] != rule["action"]:
                    continue
                if (covers(prior["src_zones"],    rule["src_zones"]) and
                        covers(prior["dst_zones"],    rule["dst_zones"]) and
                        covers(prior["sources"],      rule["sources"]) and
                        covers(prior["destinations"], rule["destinations"]) and
                        covers(prior["applications"], rule["applications"]) and
                        covers(prior["services"],     rule["services"])):
                    self._issue(
                        "MEDIUM", "Potential Shadow Rule", rule["name"],
                        f"Rule may never be matched because rule '{prior['name']}' (#{prior['num']}) already covers its traffic.",
                        "Review rule order and remove or reorder as appropriate.",
                        f"Shadowed by: {prior['name']} (rule #{prior['num']})",
                        line=rule["line"],
                    )
                    break

    # Check 11: application=any + service=any on allow ────────────────────────
    def _chk_app_any_svc_any_allow(self):
        for r in self._active_allow():
            app_any = self._has_any(r["applications"])
            svc_any = self._has_any(r["services"])
            src_any = self._has_any(r["sources"])
            dst_any = self._has_any(r["destinations"])
            if app_any and svc_any and not (src_any and dst_any):
                self._issue(
                    "HIGH", "Application+Service Both Any", r["name"],
                    "Allow rule uses application=any AND service=any, completely bypassing App-ID.",
                    "Specify explicit applications; use service=application-default to enforce App-ID.",
                    f"Src: {r['sources']}  Dst: {r['destinations']}{self._user_suffix(r)}",
                    line=r["line"],
                )

    # Check 12: inbound allow without security inspection ─────────────────────
    def _chk_inbound_no_inspection(self):
        ext_hints = {"untrust", "external", "internet", "outside", "wan", "dmz", "public"}
        int_hints = {"trust", "internal", "inside", "lan", "servers", "server", "corporate"}

        for r in self._active_allow():
            src_z = {z.strip().lower() for z in r["src_zones"].split(",")}
            dst_z = {z.strip().lower() for z in r["dst_zones"].split(",")}

            is_inbound = any(any(h in z for h in ext_hints) for z in src_z)
            is_to_int  = any(any(h in z for h in int_hints) for z in dst_z)

            if is_inbound and is_to_int and not self._rule_has_security_profiles(r):
                self._issue(
                    "HIGH", "Inbound Allow Without Inspection", r["name"],
                    "Inbound rule (external→internal) allows traffic with no AV/IPS/Spyware profiles.",
                    "Apply a security profile group with AV, Vulnerability, and Spyware to all inbound rules.",
                    f"Zones: {r['src_zones']} → {r['dst_zones']}{self._user_suffix(r)}",
                    line=r["line"],
                )

    # Check 13: service=any on allow (not application-default) ────────────────
    def _chk_service_any_allow(self):
        for r in self._active_allow():
            svcs = [s.strip().lower() for s in r["services"].split(",")]
            apps = r["applications"].lower()
            if "any" in svcs and "any" not in apps:
                self._issue(
                    "MEDIUM", "Service=Any with Specific Application", r["name"],
                    "Rule uses service=any; application traffic is allowed on non-standard ports too.",
                    "Change service to 'application-default' to enforce standard port usage.",
                    f"App: {r['applications']}{self._user_suffix(r)}",
                    line=r["line"],
                )

    # ── Crypto / system parsers ───────────────────────────────────────────────

    def _parse_ike_crypto_profiles(self):
        for prof in self.root.findall(".//ike-crypto-profiles/entry"):
            lt_h = self._text(prof, "lifetime/hours", "")
            lt_m = self._text(prof, "lifetime/minutes", "")
            lt_s = self._text(prof, "lifetime/seconds", "")
            lt_d = self._text(prof, "lifetime/days", "")
            lifetime = (lt_d + "d" if lt_d else "") or (lt_h + "h" if lt_h else "") or \
                       (lt_m + "m" if lt_m else "") or (lt_s + "s" if lt_s else "")
            self.ike_crypto_profiles.append({
                "name":       prof.get("name", ""),
                "line":       self._lineno_str(prof),
                "encryption": ", ".join(self._members(prof, "encryption")),
                "hash":       ", ".join(self._members(prof, "hash")),
                "dh_group":   ", ".join(self._members(prof, "dh-group")),
                "lifetime":   lifetime,
            })

    def _parse_ipsec_crypto_profiles(self):
        for prof in self.root.findall(".//ipsec-crypto-profiles/entry"):
            # ESP or AH
            esp = prof.find("esp")
            ah  = prof.find("ah")
            if esp is not None:
                enc  = ", ".join(self._members(esp, "encryption"))
                auth = ", ".join(self._members(esp, "authentication"))
                proto = "esp"
            elif ah is not None:
                enc  = ""
                auth = ", ".join(self._members(ah, "authentication"))
                proto = "ah"
            else:
                enc, auth, proto = "", "", ""

            lt_h = self._text(prof, "lifetime/hours", "")
            lt_m = self._text(prof, "lifetime/minutes", "")
            lt_s = self._text(prof, "lifetime/seconds", "")
            lt_d = self._text(prof, "lifetime/days", "")
            lifetime = (lt_d + "d" if lt_d else "") or (lt_h + "h" if lt_h else "") or \
                       (lt_m + "m" if lt_m else "") or (lt_s + "s" if lt_s else "")

            self.ipsec_crypto_profiles.append({
                "name":       prof.get("name", ""),
                "line":       self._lineno_str(prof),
                "protocol":   proto,
                "encryption": enc,
                "auth":       auth,
                "dh_group":   self._text(prof, "dh-group"),
                "lifetime":   lifetime,
            })

    def _parse_ssl_tls_profiles(self):
        for prof in self.root.findall(".//ssl-tls-service-profile/entry"):
            self.ssl_tls_profiles.append({
                "name":        prof.get("name", ""),
                "line":        self._lineno_str(prof),
                "min_version": self._text(prof, "protocol-settings/min-version", "tls1-0"),
                "max_version": self._text(prof, "protocol-settings/max-version", "max"),
                "certificate": self._text(prof, "certificate"),
            })

    def _parse_ike_gateways(self):
        for gw in self.root.findall(".//ike-gateways/entry"):
            # IKEv1 vs IKEv2
            proto_ver = "ikev1"
            if gw.find("protocol/ikev2") is not None:
                proto_ver = "ikev2"
            if gw.find("protocol/version") is not None:
                ver_text = self._text(gw, "protocol/version")
                if ver_text:
                    proto_ver = ver_text

            self.ike_gateways.append({
                "name":            gw.get("name", ""),
                "line":            self._lineno_str(gw),
                "version":         proto_ver,
                "peer_ip":         self._text(gw, "peer-ip-value"),
                "local_ip":        self._text(gw, "local-ip-value"),
                "crypto_profile":  self._text(gw, "ikev1/ike-crypto-profile") or
                                   self._text(gw, "ikev2/ike-crypto-profile"),
                "auth_type":       self._text(gw, "authentication/pre-shared-key") and "psk" or
                                   (self._text(gw, "authentication/certificate") and "cert" or ""),
                "psk_set":         "yes" if gw.find("authentication/pre-shared-key") is not None else "no",
                "nat_traversal":   self._text(gw, "protocol-common/nat-traversal/enable", "no"),
            })

    def _parse_admin_accounts(self):
        for admin in self.root.findall(".//administrator/entry"):
            name = admin.get("name", "")
            role_type = "superuser" if admin.find("permissions/role-based/superuser") is not None else ""
            if not role_type:
                if admin.find("permissions/role-based/superreader") is not None:
                    role_type = "superreader"
                elif admin.find("permissions/role-based/vsysadmin") is not None:
                    role_type = "vsysadmin"
                elif admin.find("permissions/role-based/custom") is not None:
                    role_type = "custom"
                else:
                    role_type = "unknown"

            self.admin_accounts.append({
                "name":             name,
                "line":             self._lineno_str(admin),
                "role":             role_type,
                "auth_profile":     self._text(admin, "authentication-profile"),
                "public_key":       "yes" if admin.find("public-key") is not None else "no",
                "password_hash":    "set" if admin.find("phash") is not None else "none",
            })

    def _parse_mgmt_settings(self):
        sys_el = self.root.find(".//deviceconfig/system")
        if sys_el is None:
            self.mgmt_settings = {}
            return

        # Permitted IPs for management
        permitted = [e.get("name", "") for e in sys_el.findall("permitted-ip/entry")]

        # Services enabled/disabled
        svc = sys_el.find("service")
        def svc_enabled(tag: str) -> bool:
            if svc is None:
                return True
            val = self._text(svc, f"disable-{tag}", "no")
            return val.lower() != "yes"

        # SNMP
        snmp_v1  = sys_el.find(".//snmp-setting/access-setting/version/v1")  is not None
        snmp_v2c = sys_el.find(".//snmp-setting/access-setting/version/v2c") is not None
        snmp_v3  = sys_el.find(".//snmp-setting/access-setting/version/v3")  is not None
        community = self._text(sys_el, ".//snmp-setting/access-setting/version/v2c/snmp-community-string")

        # NTP
        ntp_primary   = self._text(sys_el, "ntp-servers/primary-ntp-server/ntp-server-address")
        ntp_secondary = self._text(sys_el, "ntp-servers/secondary-ntp-server/ntp-server-address")

        # DNS
        dns_primary = self._text(sys_el, "dns-setting/servers/primary")
        dns_secondary = self._text(sys_el, "dns-setting/servers/secondary")

        # Login banner
        login_banner = self._text(sys_el, "login-banner")

        self.mgmt_settings = {
            "permitted_ips":   permitted,
            "http_enabled":    svc_enabled("http"),
            "https_enabled":   svc_enabled("https"),
            "telnet_enabled":  svc_enabled("telnet"),
            "ssh_enabled":     svc_enabled("ssh"),
            "snmp_enabled":    svc_enabled("snmp"),
            "snmp_v1":         snmp_v1,
            "snmp_v2c":        snmp_v2c,
            "snmp_v3":         snmp_v3,
            "snmp_community":  community,
            "ntp_primary":     ntp_primary,
            "ntp_secondary":   ntp_secondary,
            "dns_primary":     dns_primary,
            "dns_secondary":   dns_secondary,
            "login_banner":    login_banner,
            "hostname":        self._text(sys_el, "hostname"),
        }

    def _parse_syslog_servers(self):
        for srv in self.root.findall(".//syslog/entry"):
            for server in srv.findall("server/entry"):
                self.log_syslog_servers.append({
                    "profile": srv.get("name", ""),
                    "name":    server.get("name", ""),
                    "server":  self._text(server, "server"),
                    "port":    self._text(server, "port", "514"),
                    "transport": self._text(server, "transport", "UDP"),
                    "format":  self._text(server, "format", "BSD"),
                    "facility": self._text(server, "facility", "LOG_USER"),
                })

    def _parse_password_policy(self):
        """Parse password complexity, account lockout, session timeout, and NTP auth."""
        pp: dict = {
            "complexity_enabled": False,
            "min_length": 0,
            "min_upper": 0, "min_lower": 0, "min_numeric": 0, "min_special": 0,
            "password_age": 0, "password_history": 0,
            "lockout_attempts": 0, "lockout_time": 0,
            "idle_timeout": 0,
            "ntp_auth_primary": "", "ntp_auth_secondary": "",
            "line_pc": "", "line_lockout": "",
        }
        pc_el = self.root.find(".//deviceconfig/system/password-complexity")
        if pc_el is not None:
            pp["complexity_enabled"] = self._text(pc_el, "enabled", "no").lower() == "yes"
            for key, tag in [
                ("min_length",       "minimum-length"),
                ("min_upper",        "minimum-uppercase-letters"),
                ("min_lower",        "minimum-lowercase-letters"),
                ("min_numeric",      "minimum-numeric-letters"),
                ("min_special",      "minimum-special-characters"),
                ("password_age",     "password-age-enforcement-period"),
                ("password_history", "password-history-count"),
            ]:
                try:
                    pp[key] = int(self._text(pc_el, tag, "0"))
                except ValueError:
                    pass
            pp["line_pc"] = self._lineno_str(pc_el)

        mgmt_el = self.root.find(".//deviceconfig/setting/management")
        if mgmt_el is not None:
            lockout_el = mgmt_el.find("admin-lockout")
            if lockout_el is not None:
                try:
                    pp["lockout_attempts"] = int(self._text(lockout_el, "failed-attempts", "0"))
                    pp["lockout_time"]     = int(self._text(lockout_el, "lockout-time", "0"))
                except ValueError:
                    pass
                pp["line_lockout"] = self._lineno_str(lockout_el)
            try:
                pp["idle_timeout"] = int(self._text(mgmt_el, "idle-timeout", "0"))
            except ValueError:
                pass

        sys_el = self.root.find(".//deviceconfig/system")
        if sys_el is not None:
            pp["ntp_auth_primary"]   = self._text(
                sys_el, "ntp-servers/primary-ntp-server/authentication-type", "")
            pp["ntp_auth_secondary"] = self._text(
                sys_el, "ntp-servers/secondary-ntp-server/authentication-type", "")
        self.password_policy = pp

    def _parse_update_schedule(self):
        """Parse AV/threat/WildFire content update schedule from deviceconfig/system."""
        us: dict = {
            "av_action": "", "av_freq": "",
            "threats_action": "", "threats_freq": "",
            "wildfire_action": "", "wildfire_freq": "",
            "line": "",
        }
        us_el = self.root.find(".//deviceconfig/system/update-schedule")
        if us_el is not None:
            us["line"] = self._lineno_str(us_el)
            for section, a_key, f_key in [
                ("anti-virus", "av_action",      "av_freq"),
                ("threats",    "threats_action",  "threats_freq"),
                ("wildfire",   "wildfire_action", "wildfire_freq"),
            ]:
                sec = us_el.find(section)
                if sec is None:
                    continue
                rec = sec.find("recurring")
                if rec is None:
                    continue
                for interval_el in rec:
                    action = self._text(interval_el, "action")
                    if action:
                        us[a_key] = action
                        us[f_key] = interval_el.tag
                        break
        self.update_schedule = us

    def _parse_security_profiles(self):
        """Parse vulnerability and WildFire analysis profiles for depth checks."""
        for vp in self.root.findall(".//profiles/vulnerability/entry"):
            name  = vp.get("name", "")
            rules = []
            for r in vp.findall("rules/entry"):
                sevs = self._members(r, "severity")
                action_el = r.find("action")
                action = "allow"
                if action_el is not None:
                    for act in ("drop", "reset-both", "reset-client", "reset-server",
                                "block-ip", "alert", "allow"):
                        if action_el.find(act) is not None:
                            action = act
                            break
                rules.append({"severities": sevs, "action": action})
            self.vuln_profiles[name] = {"rules": rules, "line": self._lineno_str(vp)}

        for wfp in self.root.findall(".//profiles/wildfire-analysis/entry"):
            name  = wfp.get("name", "")
            rules = []
            for r in wfp.findall("rules/entry"):
                rules.append({
                    "file_types": self._members(r, "file-types"),
                    "apps":       self._members(r, "applications"),
                    "direction":  self._text(r, "direction", ""),
                    "analysis":   self._text(r, "analysis", ""),
                })
            self.wildfire_profiles[name] = {"rules": rules, "line": self._lineno_str(wfp)}

    # ── Crypto / system checks ────────────────────────────────────────────────

    # Weak values reference tables
    _WEAK_ENC  = {"des", "3des", "null"}
    _WEAK_HASH = {"md5", "sha1"}
    _WEAK_DH   = {"group1", "group2", "group5"}      # 768, 1024, 1536-bit
    _CRIT_DH   = {"group1"}                           # 768-bit
    _WEAK_TLS  = {"ssl3-0": "CRITICAL", "tls1-0": "HIGH", "tls1-1": "MEDIUM"}

    def _chk_weak_ike_crypto(self):
        for p in self.ike_crypto_profiles:
            name = f"IKE Profile: {p['name']}"
            encs = {e.strip().lower() for e in p["encryption"].split(",")}
            hashes = {h.strip().lower() for h in p["hash"].split(",")}
            dh_groups = {g.strip().lower() for g in p["dh_group"].split(",")}

            for enc in encs & self._WEAK_ENC:
                sev = "CRITICAL" if enc in ("des", "null") else "HIGH"
                self._issue(sev, "Weak IKE Encryption", name,
                    f"IKE crypto profile '{p['name']}' uses {enc.upper()} encryption.",
                    f"Replace {enc.upper()} with AES-256-GCM or AES-256-CBC.",
                    f"Encryption list: {p['encryption']}", line=p["line"])

            for h in hashes & self._WEAK_HASH:
                sev = "HIGH" if h == "md5" else "MEDIUM"
                self._issue(sev, "Weak IKE Hash/PRF", name,
                    f"IKE crypto profile '{p['name']}' uses {h.upper()} for integrity/PRF.",
                    f"Replace {h.upper()} with SHA-256, SHA-384, or SHA-512.",
                    f"Hash list: {p['hash']}", line=p["line"])

            for dh in dh_groups & self._WEAK_DH:
                sev = "CRITICAL" if dh in self._CRIT_DH else "HIGH"
                self._issue(sev, "Weak IKE DH Group", name,
                    f"IKE crypto profile '{p['name']}' includes {dh} (insufficient key size).",
                    "Use DH Group 14 (2048-bit) minimum; prefer Group 19/20 (ECDH-256/384).",
                    f"DH groups: {p['dh_group']}", line=p["line"])

    def _chk_weak_ipsec_crypto(self):
        for p in self.ipsec_crypto_profiles:
            name = f"IPSec Profile: {p['name']}"
            encs  = {e.strip().lower() for e in p["encryption"].split(",")} if p["encryption"] else set()
            auths = {a.strip().lower() for a in p["auth"].split(",")}       if p["auth"]       else set()
            dh    = p["dh_group"].strip().lower()

            for enc in encs & self._WEAK_ENC:
                sev = "CRITICAL" if enc in ("des", "null") else "HIGH"
                self._issue(sev, "Weak IPSec Encryption", name,
                    f"IPSec crypto profile '{p['name']}' uses {enc.upper()} for ESP encryption.",
                    f"Replace {enc.upper()} with AES-256-GCM (preferred) or AES-256-CBC.",
                    f"Encryption list: {p['encryption']}", line=p["line"])

            for a in auths & self._WEAK_HASH:
                sev = "HIGH" if a == "md5" else "MEDIUM"
                self._issue(sev, "Weak IPSec Authentication", name,
                    f"IPSec crypto profile '{p['name']}' uses {a.upper()} for ESP authentication.",
                    f"Replace {a.upper()} with SHA-256 or higher.",
                    f"Auth list: {p['auth']}", line=p["line"])

            if dh in self._WEAK_DH:
                sev = "CRITICAL" if dh in self._CRIT_DH else "HIGH"
                self._issue(sev, "Weak IPSec DH Group (PFS)", name,
                    f"IPSec profile '{p['name']}' uses {dh} for Perfect Forward Secrecy.",
                    "Use DH Group 14 minimum for PFS; prefer Group 19/20.",
                    f"DH group: {p['dh_group']}", line=p["line"])

            if dh in ("no-pfs", "") and p["dh_group"].lower() in ("no-pfs", ""):
                self._issue("MEDIUM", "IPSec PFS Disabled", name,
                    f"IPSec profile '{p['name']}' has Perfect Forward Secrecy disabled.",
                    "Enable PFS with at least DH Group 14 to limit session key exposure.",
                    "", line=p["line"])

    def _chk_weak_ssl_tls(self):
        weak_map = {"ssl3-0": "CRITICAL", "tls1-0": "HIGH", "tls1-1": "MEDIUM"}
        version_order = ["ssl3-0", "tls1-0", "tls1-1", "tls1-2", "tls1-3", "max"]

        for p in self.ssl_tls_profiles:
            name = f"SSL/TLS Profile: {p['name']}"
            min_ver = p["min_version"].strip().lower()

            if min_ver in weak_map:
                sev = weak_map[min_ver]
                self._issue(sev, "Weak Minimum TLS Version", name,
                    f"SSL/TLS profile '{p['name']}' allows {min_ver.upper().replace('-', '.')} connections.",
                    "Set minimum version to TLS 1.2; prefer TLS 1.3 where supported.",
                    f"min-version: {p['min_version']}  max-version: {p['max_version']}",
                    line=p["line"])

        # Also check GlobalProtect and management SSL profiles embedded elsewhere
        for prof in self.root.findall(".//ssl-exclude-cert/entry"):
            pass  # placeholder for future cert checks

    def _chk_ike_gateways(self):
        for gw in self.ike_gateways:
            name = f"IKE Gateway: {gw['name']}"
            if gw["version"].lower() == "ikev1":
                self._issue("MEDIUM", "IKEv1 in Use", name,
                    f"IKE gateway '{gw['name']}' uses IKEv1, which lacks several security improvements.",
                    "Migrate to IKEv2 for stronger authentication, built-in NAT-T, and DoS resistance.",
                    f"Peer: {gw['peer_ip']}", line=gw["line"])

            if gw["psk_set"] == "yes":
                self._issue("LOW", "IKE Pre-Shared Key Authentication", name,
                    f"IKE gateway '{gw['name']}' uses PSK authentication.",
                    "Certificate-based authentication is preferred over PSK for scalability and security.",
                    f"Peer: {gw['peer_ip']}", line=gw["line"])

    def _chk_management_access(self):
        m = self.mgmt_settings
        if not m:
            return

        sys_el = self.root.find(".//deviceconfig/system")
        svc    = sys_el.find("service") if sys_el is not None else None

        def _svc_rule(tag: str) -> str:
            return f"service/disable-{tag}"

        def _svc_line(tag: str) -> str:
            # Only return a line when the element is explicitly present (bad value set);
            # if absent, the issue is that nothing is there — no line to point to.
            if svc is None:
                return ""
            el = svc.find(f"disable-{tag}")
            return self._lineno_str(el) if el is not None else ""

        def _svc_details(tag: str) -> str:
            if svc is None:
                return f"disable-{tag}: not set (no <service> block — defaults to enabled)"
            el = svc.find(f"disable-{tag}")
            if el is None:
                return f"disable-{tag}: not set (defaults to enabled)"
            return f"disable-{tag}: {el.text or 'no'}"

        if m.get("http_enabled"):
            self._issue("HIGH", "HTTP Management Enabled", _svc_rule("http"),
                "HTTP access to the management interface is enabled (cleartext).",
                "Disable HTTP management: set service/disable-http to yes.",
                _svc_details("http"),
                line=_svc_line("http"))

        if m.get("telnet_enabled"):
            self._issue("HIGH", "Telnet Management Enabled", _svc_rule("telnet"),
                "Telnet access to the management interface is enabled (cleartext).",
                "Disable Telnet management: set service/disable-telnet to yes. Use SSH instead.",
                _svc_details("telnet"),
                line=_svc_line("telnet"))

        if not m.get("permitted_ips"):
            self._issue("MEDIUM", "No Management IP Restrictions", "permitted-ip",
                "No permitted-ip entries restrict which hosts can reach the management interface.",
                "Add permitted-ip entries to restrict management access to known admin hosts/subnets.",
                "permitted-ip: not configured — any host can attempt to reach the management interface")

        if not m.get("ntp_primary"):
            self._issue("MEDIUM", "NTP Not Configured", "ntp-servers",
                "No primary NTP server is configured.",
                "Configure at least two NTP servers for accurate timestamps in logs and certificates.",
                "ntp-servers/primary-ntp-server/ntp-server-address: not set")
        elif not m.get("ntp_secondary"):
            ntp_el   = sys_el.find("ntp-servers") if sys_el is not None else None
            ntp_line = self._lineno_str(ntp_el) if ntp_el is not None else ""
            self._issue("LOW", "Only One NTP Server", "ntp-servers",
                "Only one NTP server is configured. Loss of this server leaves the firewall without time sync.",
                "Add a secondary NTP server for redundancy.",
                f"primary-ntp-server: {m.get('ntp_primary')}\n"
                "secondary-ntp-server/ntp-server-address: not set",
                line=ntp_line)

        if not m.get("login_banner"):
            self._issue("LOW", "No Login Banner", "login-banner",
                "No login banner is configured on the management interface.",
                "Add a legal warning banner (login-banner) to satisfy compliance requirements and "
                "establish notice of unauthorized access.",
                "login-banner: not set")

        if not m.get("dns_primary"):
            self._issue("LOW", "DNS Not Configured", "dns-setting/servers",
                "No primary DNS server is configured in device settings.",
                "Configure DNS for FQDN resolution used by URL filtering, FQDN objects, and updates.",
                "dns-setting/servers/primary: not set")

    def _chk_admin_accounts(self):
        superusers = [a for a in self.admin_accounts if a["role"] == "superuser"]

        for admin in self.admin_accounts:
            name = f"Admin: {admin['name']}"
            if not admin["auth_profile"]:
                self._issue("HIGH", "Admin Without Authentication Profile", name,
                    f"Administrator '{admin['name']}' has no authentication profile — "
                    "local password only, no MFA.",
                    "Assign an authentication profile with MFA to all admin accounts.",
                    f"Role: {admin['role']}", line=admin["line"])

            if admin["password_hash"] == "none" and admin["public_key"] == "no":
                self._issue("CRITICAL", "Admin Account Has No Password", name,
                    f"Administrator '{admin['name']}' has no password hash or SSH key set.",
                    "Set a strong password or SSH key for this account immediately.",
                    f"Role: {admin['role']}", line=admin["line"])

        if len(superusers) > 2:
            names = ", ".join(a["name"] for a in superusers)
            self._issue("MEDIUM", "Excessive Superuser Accounts", "Admin Accounts",
                f"{len(superusers)} accounts have the superuser role.",
                "Limit superuser accounts to the minimum required; use role-based access for others.",
                f"Superusers: {names}")

    def _chk_snmp(self):
        m = self.mgmt_settings
        if not m:
            return

        sys_el    = self.root.find(".//deviceconfig/system")
        snmp_base = "deviceconfig/system/snmp-setting/access-setting/version"

        def _snmp_ver_line(ver: str) -> str:
            if sys_el is None:
                return ""
            el = sys_el.find(f".//snmp-setting/access-setting/version/{ver}")
            return self._lineno_str(el) if el is not None else (
                   self._lineno_str(sys_el.find(".//snmp-setting")) or
                   self._lineno_str(sys_el))

        if m.get("snmp_v1"):
            self._issue("HIGH", "SNMPv1 Enabled", "snmp-setting/version/v1",
                "SNMPv1 is enabled; it uses cleartext community strings and has no authentication.",
                "Disable SNMPv1. Use SNMPv3 with authPriv (auth + encryption).",
                "snmp-setting/access-setting/version/v1: present",
                line=_snmp_ver_line("v1"))

        if m.get("snmp_v2c"):
            self._issue("MEDIUM", "SNMPv2c Enabled", "snmp-setting/version/v2c",
                "SNMPv2c is enabled; community strings are transmitted in cleartext.",
                "Migrate to SNMPv3 with authPriv. If SNMPv2c is required, restrict source IPs.",
                "snmp-setting/access-setting/version/v2c: present",
                line=_snmp_ver_line("v2c"))

        community = m.get("snmp_community", "").lower()
        if community in ("public", "private", "cisco", "community", "snmp"):
            self._issue("CRITICAL", "Default/Weak SNMP Community String",
                "snmp-setting/version/v2c/community-string",
                f"SNMP community string is set to the well-known default value '{community}'.",
                "Change the community string to a long random value and restrict allowed SNMP hosts.",
                f"snmp-community-string: {community}",
                line=_snmp_ver_line("v2c"))

        if m.get("snmp_v1") or m.get("snmp_v2c"):
            if not m.get("permitted_ips"):
                self._issue("HIGH", "SNMP Enabled Without Source Restrictions", "permitted-ip",
                    "SNMP is enabled and no management permitted-ip list is configured.",
                    "Restrict SNMP access to specific NMS hosts via permitted-ip or ACL.",
                    "permitted-ip: not configured — any host can poll SNMP")

    def _chk_no_syslog(self):
        if not self.log_syslog_servers:
            self._issue("MEDIUM", "No Syslog Servers Configured", "Logging",
                "No syslog servers are defined; firewall logs may only reside on the local device.",
                "Configure at least one remote syslog server (preferably two) for log retention, "
                "correlation, and incident response.",
                "")
        else:
            # Check for unencrypted syslog
            for srv in self.log_syslog_servers:
                if srv["transport"].upper() == "UDP":
                    self._issue("LOW", "Syslog Transmitted Over UDP", f"Syslog: {srv['server']}",
                        f"Syslog server {srv['server']} uses UDP — logs can be lost or spoofed in transit.",
                        "Switch syslog transport to TCP or SSL for reliable, tamper-evident log delivery.",
                        f"Profile: {srv['profile']}  Port: {srv['port']}")

    def _chk_password_policy(self):
        pp = self.password_policy
        if not pp:
            return
        if not pp.get("complexity_enabled"):
            self._issue("HIGH", "Password Complexity Not Enforced", "Password Policy",
                "Password complexity requirements are disabled in deviceconfig/system/password-complexity.",
                "Enable password-complexity and require uppercase, lowercase, numeric, and special characters.",
                line=pp.get("line_pc", ""))
        elif pp.get("min_length", 0) < 12:
            self._issue("MEDIUM", "Weak Password Minimum Length", "Password Policy",
                f"Minimum password length is {pp.get('min_length', 0)} characters (recommended ≥ 12).",
                "Increase minimum-length to 12 or more in deviceconfig/system/password-complexity.",
                line=pp.get("line_pc", ""))

        if not pp.get("lockout_attempts"):
            self._issue("HIGH", "No Account Lockout Policy", "Password Policy",
                "No admin account lockout policy is configured.",
                "Set admin-lockout in deviceconfig/setting/management: "
                "failed-attempts ≤ 5 and lockout-time ≥ 5 minutes.",
                line=pp.get("line_lockout", ""))
        elif pp.get("lockout_attempts", 99) > 5:
            self._issue("MEDIUM", "No Account Lockout Policy", "Password Policy",
                f"Admin lockout threshold is {pp['lockout_attempts']} failed attempts (recommended ≤ 5).",
                "Reduce failed-attempts to 5 or fewer.",
                line=pp.get("line_lockout", ""))

        timeout = pp.get("idle_timeout", 0)
        if timeout == 0:
            self._issue("HIGH", "Long or No Management Session Timeout", "Password Policy",
                "Management session idle timeout is 0 (disabled). Idle sessions never expire.",
                "Set idle-timeout to 15 or 30 minutes in deviceconfig/setting/management.",
                line=pp.get("line_lockout", ""))
        elif timeout > 30:
            self._issue("MEDIUM", "Long or No Management Session Timeout", "Password Policy",
                f"Management session idle timeout is {timeout} minutes (recommended ≤ 30).",
                "Reduce idle-timeout to 15 or 30 minutes.",
                line=pp.get("line_lockout", ""))

        age = pp.get("password_age", 0)
        if age == 0:
            self._issue("MEDIUM", "Password Expiry Not Configured", "Password Policy",
                "Password expiry is not enforced (password-age-enforcement-period = 0). "
                "Accounts with non-expiring passwords increase the window of exposure "
                "if credentials are compromised.",
                "Set password-age-enforcement-period to 90 or fewer days.",
                line=pp.get("line_pc", ""))
        elif age > 90:
            self._issue("MEDIUM", "Password Expiry Not Configured", "Password Policy",
                f"Password expiry period is {age} days, which exceeds the recommended 90-day maximum. "
                "Long-lived passwords remain valid longer after a compromise.",
                "Reduce password-age-enforcement-period to 90 days or fewer.",
                line=pp.get("line_pc", ""))

        history = pp.get("password_history", 0)
        if history < 4:
            self._issue("LOW", "Insufficient Password History", "Password Policy",
                f"Password history count is {history}. Without a sufficient history, users can "
                "immediately reuse old passwords, undermining password rotation policies.",
                "Set password-history-count to at least 4.",
                line=pp.get("line_pc", ""))

        m = self.mgmt_settings
        for label, ntp_key, auth_key in [
            ("Primary",   "ntp_primary",   "ntp_auth_primary"),
            ("Secondary", "ntp_secondary", "ntp_auth_secondary"),
        ]:
            if m.get(ntp_key) and not pp.get(auth_key, "").strip():
                self._issue("LOW", "NTP Authentication Not Configured", "Management",
                    f"{label} NTP server '{m[ntp_key]}' has no NTP authentication configured.",
                    "Enable NTP authentication: set authentication-type to symmetric-key "
                    "and configure a matching key ID on the NTP server.",
                    f"NTP server: {m[ntp_key]}")

    def _chk_update_schedule(self):
        us = self.update_schedule
        for label, a_key, f_key in [
            ("Anti-Virus",     "av_action",      "av_freq"),
            ("Threat content", "threats_action",  "threats_freq"),
        ]:
            action = us.get(a_key, "")
            if not action:
                self._issue("MEDIUM", "AV/Threat Content Updates Not Automatic", f"Update: {label}",
                    f"{label} content updates are not scheduled.",
                    "Configure automatic updates with 'download-and-install' action "
                    "in Device > Dynamic Updates.",
                    "Without scheduled updates the device will miss new threat signatures.",
                    line=us.get("line", ""))
            elif action != "download-and-install":
                self._issue("LOW", "AV/Threat Content Updates Not Automatic", f"Update: {label}",
                    f"{label} update action is '{action}' (not 'download-and-install').",
                    "Change the update action to 'download-and-install' for fully automatic updates.",
                    f"Frequency: {us.get(f_key, '')}",
                    line=us.get("line", ""))

        wf_action = us.get("wildfire_action", "")
        if not wf_action:
            self._issue("MEDIUM", "WildFire Updates Not Automatic", "Update: WildFire",
                "WildFire content updates are not scheduled.",
                "Configure WildFire updates (every-min or every-15-min) with 'download-and-install'. "
                "WildFire delivers near-real-time protection against novel malware.",
                line=us.get("line", ""))
        elif wf_action != "download-and-install":
            self._issue("LOW", "WildFire Updates Not Automatic", "Update: WildFire",
                f"WildFire update action is '{wf_action}' (not 'download-and-install').",
                "Change WildFire update action to 'download-and-install'.",
                f"Frequency: {us.get('wildfire_freq', '')}",
                line=us.get("line", ""))

    def _chk_security_profile_settings(self):
        block_actions = {"drop", "reset-both", "reset-client", "reset-server", "block-ip"}
        for name, vp in self.vuln_profiles.items():
            if not vp["rules"]:
                self._issue("HIGH", "Vulnerability Profile Allows Critical/High Threats",
                    f"Vuln Profile: {name}",
                    f"Vulnerability profile '{name}' has no rules — all threats pass through.",
                    "Add rules to block (reset-both or drop) at minimum critical and high severity threats.",
                    line=vp["line"])
                continue
            blocks_critical = any(
                any(s in ("critical", "high", "any") for s in r["severities"])
                and r["action"] in block_actions
                for r in vp["rules"]
            )
            if not blocks_critical:
                self._issue("HIGH", "Vulnerability Profile Allows Critical/High Threats",
                    f"Vuln Profile: {name}",
                    f"Vulnerability profile '{name}' does not block critical or high severity threats.",
                    "Add a rule with action reset-both (or drop) for critical and high severity threats.",
                    line=vp["line"])

        for name, wfp in self.wildfire_profiles.items():
            if not wfp["rules"]:
                self._issue("HIGH", "WildFire Profile Missing Rules", f"WildFire Profile: {name}",
                    f"WildFire analysis profile '{name}' has no submission rules.",
                    "Add rules to submit 'any' file type in 'both' directions to WildFire.",
                    line=wfp["line"])
                continue
            has_broad = any(
                "any" in r["file_types"]
                and r.get("direction", "") in ("both", "upload", "download", "")
                for r in wfp["rules"]
            )
            if not has_broad:
                self._issue("MEDIUM", "WildFire Profile Incomplete Coverage",
                    f"WildFire Profile: {name}",
                    f"WildFire profile '{name}' does not submit all file types for analysis.",
                    "Add a rule covering file-types='any' and direction='both' "
                    "to maximise unknown-threat detection.",
                    line=wfp["line"])

    def _chk_user_id_untrust(self):
        untrusted_kw = {"untrust", "external", "outside", "internet", "wan", "public"}
        for z in self.zones:
            if z.get("user_id", "no").lower() == "yes":
                if any(kw in z["name"].lower() for kw in untrusted_kw):
                    self._issue("HIGH", "User-ID Enabled on Untrusted Zone", f"Zone: {z['name']}",
                        f"User-ID is enabled on zone '{z['name']}'. Untrusted hosts can inject "
                        "forged User-ID mappings, bypassing identity-based policy.",
                        "Disable User-ID on all untrusted and DMZ zones. "
                        "Enable only on internal trusted zones.",
                        line=z["line"])

    def _chk_tls_default_cert(self):
        for prof in self.ssl_tls_profiles:
            cert = prof.get("certificate", "")
            if cert.lower() in ("default", ""):
                sev = "HIGH" if cert.lower() == "default" else "MEDIUM"
                self._issue(sev, "TLS Profile Using Default Certificate", f"TLS Profile: {prof['name']}",
                    f"SSL/TLS service profile '{prof['name']}' uses "
                    f"{'the factory default self-signed certificate' if cert else 'no certificate'}. "
                    "Clients cannot validate the server identity.",
                    "Replace with a CA-signed certificate (Device > Certificate Management).",
                    line=prof.get("line", ""))

    def _chk_insecure_cleartext_apps(self):
        """Flag allow rules using cleartext protocols not already caught by _chk_risky_services_from_any."""
        ALWAYS_INSECURE: dict[str, tuple[str, str]] = {
            "ftp":    ("MEDIUM", "FTP transmits credentials and data in cleartext."),
            "tftp":   ("MEDIUM", "TFTP has no authentication or encryption."),
            "rsh":    ("HIGH",   "Remote Shell (RSH) sends commands over cleartext."),
            "rlogin": ("HIGH",   "rlogin is unauthenticated and cleartext."),
            "finger": ("LOW",    "Finger discloses user accounts; obsolete protocol."),
        }
        UNTRUSTED_KW = {"untrust", "external", "outside", "internet", "wan", "public"}

        for r in self._active_allow():
            apps_lower = {a.strip().lower() for a in r["applications"].split(",")}
            src_zones  = {z.strip().lower() for z in r["src_zones"].split(",")}
            is_any_src = self._has_any(r["sources"])
            from_external = any(kw in z for z in src_zones for kw in UNTRUSTED_KW) \
                            or "any" in src_zones

            for app, (sev, desc) in ALWAYS_INSECURE.items():
                if app in apps_lower:
                    self._issue(sev, "Insecure Protocol Allowed in Rule", r["name"],
                        f"Rule allows '{app}': {desc}",
                        f"Replace '{app}' with an encrypted equivalent (SSH/SFTP/SCP). "
                        "Block this application entirely if not required.",
                        f"Src zones: {r['src_zones']}  Destinations: {r['destinations']}",
                        line=r["line"])

            # Telnet with specific source — any-source case already handled by _chk_risky_services_from_any
            if "telnet" in apps_lower and not is_any_src:
                self._issue("HIGH", "Insecure Protocol Allowed in Rule", r["name"],
                    "Telnet transmits credentials in cleartext.",
                    "Replace Telnet with SSH. Block all Telnet traffic.",
                    f"Src zones: {r['src_zones']}  Sources: {r['sources']}",
                    line=r["line"])

            # Unencrypted HTTP from an external zone into internal/DMZ
            if from_external and ("http" in apps_lower or "web-browsing" in apps_lower):
                dst_zones = {z.strip().lower() for z in r["dst_zones"].split(",")}
                internal  = dst_zones - UNTRUSTED_KW - {"any"}
                if internal:
                    self._issue("LOW", "Insecure Protocol Allowed in Rule", r["name"],
                        "Rule allows unencrypted HTTP from an external zone into an internal zone.",
                        "Replace HTTP with HTTPS. If hosting a web service, enforce HTTP→HTTPS redirect.",
                        f"Src zones: {r['src_zones']}  Dst zones: {r['dst_zones']}",
                        line=r["line"])

    def _chk_default_deny_rule(self):
        """PCI DSS 1.2.4: Every rulebase should end with an explicit deny-all."""
        from collections import defaultdict
        rb_rules: dict[str, list] = defaultdict(list)
        for r in self.security_rules:
            rb_rules[r["rulebase"]].append(r)
        for rb_name, rules in rb_rules.items():
            if not rules:
                continue
            last = rules[-1]
            if last["action"] not in ("deny", "drop"):
                self._issue("MEDIUM", "No Default Deny Rule", f"Rulebase: {rb_name}",
                    f"The last rule in rulebase '{rb_name}' is '{last['action']}' — "
                    "not an explicit deny-all. Without a terminating deny rule, "
                    "traffic that does not match any rule falls through on the implicit default, "
                    "which may not be logged or may differ from the intended policy.",
                    "Add an explicit deny-all rule as the last rule in every rulebase "
                    "(action=deny, all zones, sources, destinations, applications).",
                    f"Last rule: '{last['name']}'",
                    line=last.get("line", ""))

    def _chk_file_blocking_inbound(self):
        """PCI DSS 5.3.1: File-blocking profile should be attached to inbound allow rules."""
        UNTRUSTED_KW = {"untrust", "external", "outside", "internet", "wan", "public"}
        for r in self._active_allow():
            src_zones = {z.strip().lower() for z in r["src_zones"].split(",")}
            from_external = any(kw in z for z in src_zones for kw in UNTRUSTED_KW) \
                            or "any" in src_zones
            if not from_external:
                continue
            pt = r["profile_type"]
            if pt == "group":
                grp_data = self.profile_groups.get(r["profiles"].get("group", ""), {})
                has_fb = bool(grp_data.get("file-blocking"))
            elif pt == "profiles":
                has_fb = bool(r["profiles"].get("file-blocking"))
            else:
                has_fb = False
            if not has_fb:
                self._issue("MEDIUM", "File Blocking Not Applied", r["name"],
                    f"Inbound allow rule '{r['name']}' from external zone(s) has no "
                    "file-blocking profile attached. Without it, malicious files "
                    "(executables, scripts, documents with macros) can pass into the network unchecked.",
                    "Assign a file-blocking security profile (or profile group) to this rule. "
                    "Block executable file types at minimum.",
                    f"Src zones: {r['src_zones']}  Apps: {r['applications']}",
                    line=r["line"])

    # ── Policy Optimizer / rule-usage checks (CSV-sourced rules only) ────────
    def _chk_policy_optimizer_unused(self):
        """Flag rules Policy Optimizer reports as unused (no matching traffic).

        `usage_status` is already normalized in _parse_security_rules_from_csv
        to "Unused" for both real export shapes: a "Rule Usage Rule Usage"
        status column, or a "Rule Usage Hit Count" of 0.
        """
        for r in self.security_rules:
            status = r.get("usage_status", "")
            if not status or "unused" not in status.lower():
                continue
            evidence = f"Usage: {status}"
            if r.get("last_hit"):
                evidence += f"  Last hit: {r['last_hit']}"
            evidence += f"  Created: {r.get('created', '')}  Modified: {r.get('modified', '')}"
            self._issue(
                "MEDIUM", "Unused Security Rule (Policy Optimizer)", r["name"],
                "Policy Optimizer reports this rule has not matched any traffic.",
                "Confirm the rule is no longer needed and remove it, or investigate why "
                "the traffic it was written for isn't hitting it.",
                evidence,
                line=r["line"],
            )

    def _chk_policy_optimizer_stale_apps(self):
        """Flag rules still allowing Application=any despite long-standing app
        visibility that could be used to narrow it (classic App-ID cleanup).

        "Rule Usage Apps Seen" varies by export shape: sometimes a delimited
        list of actual application names, sometimes just a count of distinct
        apps observed — handle both, and phrase the recommendation to match
        what's actually known.
        """
        for r in self.security_rules:
            days_raw = r.get("days_no_new_apps", "")
            if not days_raw:
                continue
            digits = re.sub(r"[^\d]", "", days_raw)
            if not digits or int(digits) < 90:
                continue
            if not self._has_any(r["applications"]):
                continue
            apps_seen = r.get("apps_seen", "")
            if not apps_seen or apps_seen.strip().lower() in ("any", "unknown"):
                continue
            if apps_seen.strip().isdigit():
                seen_count = int(apps_seen.strip())
                if seen_count <= 0:
                    continue  # nothing observed — that's an unused rule, not an unrefined one
                recommendation = (
                    f"Policy Optimizer has observed {seen_count} distinct application(s) on "
                    "this rule — use Policy Optimizer's app-usage view to see which ones and "
                    "restrict Application to that set."
                )
            else:
                recommendation = f"Restrict Application to the observed set: {apps_seen}."
            self._issue(
                "MEDIUM", "Application Not Refined To Observed Traffic", r["name"],
                f"Rule still allows Application=any after {digits} day(s) with no new "
                "applications observed — enough visibility exists to narrow it to the "
                "apps actually seen.",
                recommendation,
                f"Apps seen: {apps_seen}  Days with no new apps: {digits}",
                line=r["line"],
            )

    # ── CIS L2 benchmark checks ───────────────────────────────────────────────
    def _panfw_major_version(self) -> int:
        """Return the major PAN-OS version from the config root version attribute."""
        ver_str = (self.root.get("version", "") or "") if self.root is not None else ""
        try:
            return int(ver_str.split(".")[0])
        except (ValueError, IndexError):
            return 0

    def _run_cis_checks(self):
        """Run CIS Benchmark L1 and L2 checks via XSLT execution against the audit files.

        The benchmark/.audit pair is picked from the PAN-OS major version found in the
        config's own <config version="..."> attribute (e.g. "11.1.13" -> major 11), so a
        device on any 11.x train is automatically checked against the CIS Palo Alto
        Firewall 11 Benchmark rather than a fixed/hardcoded version.
        """
        major   = self._panfw_major_version()
        clamped = max(6, min(11, major)) if 6 <= major <= 11 else 11
        tar_path = _find_audits_tar()

        self.panos_version_str = (self.root.get("version", "") or "unknown") \
            if self.root is not None else "unknown"

        if not tar_path:
            print("[!] audits.tar.gz not found — running built-in CIS L2 checks.")
            self.audits_tar_path   = ""
            self.cis_l1_audit_used = self.cis_l2_audit_used = \
                "audits.tar.gz not found — built-in checks only"
            self._chk_cis125_admin_cert()
            self._chk_cis22_wmi_probing()
            self._chk_cis616_zone_flood()
            return

        self.audits_tar_path = tar_path
        print(f"[+] audits.tar.gz found: {tar_path}")

        try:
            from lxml import etree as _letree  # type: ignore[import]
        except ImportError:
            print("[!] lxml not installed — running built-in CIS L2 checks. pip install lxml")
            self.cis_l1_audit_used = self.cis_l2_audit_used = \
                "lxml not installed — built-in checks only"
            self._chk_cis125_admin_cert()
            self._chk_cis22_wmi_probing()
            self._chk_cis616_zone_flood()
            return

        # Wrap config in the envelope the audit XSLTs expect
        try:
            config_xml = ET.tostring(self.root, encoding="unicode")
            wrapped    = f"<response><result>{config_xml}</result></response>"
            xml_doc    = _letree.fromstring(wrapped.encode("utf-8"))
        except Exception as exc:
            print(f"[!] Could not prepare XML for XSLT execution: {exc}")
            return

        ver_str = major if major else "unknown"

        # Run L1 first (comprehensive baseline hardening)
        l1_name = _CIS_L1_AUDIT_IN_TAR.get(clamped, _CIS_L1_AUDIT_IN_TAR[11])
        try:
            with tarfile.open(tar_path, "r:gz") as tf:
                if l1_name in tf.getnames():
                    l1_bytes = tf.extractfile(l1_name).read()  # type: ignore[union-attr]
                    l1_content = l1_bytes.decode("utf-8", errors="replace")
                    l1_checks  = _parse_audit_op_checks(l1_content)
                    ran = found = 0
                    for chk in l1_checks:
                        result = self._run_audit_xslt_check(chk, xml_doc, _letree)
                        if result is not None:
                            ran += 1
                            if result:
                                found += 1
                    print(f"[+] CIS L1 audit: {os.path.basename(l1_name)} "
                          f"(PAN-OS {ver_str}): {ran} checks, {found} finding(s)")
                    self.cis_l1_audit_used = os.path.basename(l1_name)
                else:
                    print(f"[!] CIS L1 audit not found in tarball: {l1_name}")
                    self.cis_l1_audit_used = f"{os.path.basename(l1_name)} (missing from tarball)"
        except Exception as exc:
            print(f"[!] Could not run L1 audit: {exc}")

        # Run L2 next (additional hardening; deduplication via _seen_issues)
        l2_name = _CIS_L2_AUDIT_IN_TAR.get(clamped, _CIS_L2_AUDIT_IN_TAR[11])
        try:
            with tarfile.open(tar_path, "r:gz") as tf:
                if l2_name in tf.getnames():
                    l2_bytes = tf.extractfile(l2_name).read()  # type: ignore[union-attr]
                    l2_content = l2_bytes.decode("utf-8", errors="replace")
                    l2_checks  = _parse_audit_op_checks(l2_content)
                    ran = found = 0
                    for chk in l2_checks:
                        result = self._run_audit_xslt_check(chk, xml_doc, _letree)
                        if result is not None:
                            ran += 1
                            if result:
                                found += 1
                    print(f"[+] CIS L2 audit: {os.path.basename(l2_name)} "
                          f"(PAN-OS {ver_str}): {ran} checks, {found} finding(s)")
                    self.cis_l2_audit_used = os.path.basename(l2_name)
                else:
                    print(f"[!] CIS L2 audit not found in tarball: {l2_name}")
                    self.cis_l2_audit_used = f"{os.path.basename(l2_name)} (missing from tarball)"
                    self._chk_cis125_admin_cert()
                    self._chk_cis22_wmi_probing()
                    self._chk_cis616_zone_flood()
        except Exception as exc:
            print(f"[!] Could not run L2 audit: {exc}")

    def _run_audit_xslt_check(self, check: dict, xml_doc, etree) -> "bool | None":
        """Execute one XSLT check. Returns True=issue found, False=passed, None=error/skip."""
        xsl_body = "\n".join(check["xsl_stmts"])
        for old, new in _XSLT_PATH_SUBS:
            xsl_body = xsl_body.replace(old, new)
        if "</xsl:template>" not in xsl_body:
            xsl_body += "\n</xsl:template>"
        xslt_src = (
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<xsl:stylesheet version="1.0" xmlns:xsl="http://www.w3.org/1999/XSL/Transform">'
            '<xsl:output method="text"/>'
            + xsl_body
            + '</xsl:stylesheet>'
        )
        try:
            xslt_doc  = etree.fromstring(xslt_src.encode("utf-8"))
            transform = etree.XSLT(xslt_doc)
            output    = str(transform(xml_doc)).strip()
        except Exception:
            return None

        expect    = check["expect"]
        not_exp   = check["not_expect"]
        # Contradictory criteria (expect == not_expect) can never be satisfied; skip.
        if expect and not_exp and expect == not_exp:
            return None
        failed    = False
        if expect and not re.search(expect, output, re.MULTILINE):
            failed = True
        if not_exp and re.search(not_exp, output, re.MULTILINE):
            failed = True
        if not expect and not not_exp:
            return None  # no pass/fail criterion
        if not failed:
            return False

        # ── Map to category ──────────────────────────────────────────────────
        desc_raw = check["description"]
        cis_m    = re.match(r'^(\d+\.\d+(?:\.\d+)?(?:\.\d+)?)\s+', desc_raw)
        cis_num  = cis_m.group(1) if cis_m else ""

        if cis_num in _CIS_NUM_TO_CATEGORY:
            category = _CIS_NUM_TO_CATEGORY[cis_num]
        elif desc_raw in _AUDIT_SUBCHECK_META:
            category = _AUDIT_SUBCHECK_META[desc_raw][0]
        else:
            category = desc_raw

        # Skip categories already covered by always-running Python checks
        if category in _PYTHON_COVERED_CATEGORIES:
            return False

        # ── Build plain description (no framework refs) ───────────────────
        if check["info"]:
            description = _strip_fw_refs(_first_para(check["info"]))[:500]
        elif desc_raw in _AUDIT_SUBCHECK_META:
            description = _AUDIT_SUBCHECK_META[desc_raw][1]
        else:
            description = _strip_fw_refs(desc_raw)

        # ── Recommendation ────────────────────────────────────────────────
        sol = check["solution"]
        recommendation = _first_para(sol)[:500] if sol else ""

        # ── Severity ──────────────────────────────────────────────────────
        severity = check["severity"] if check["severity"] in ("CRITICAL", "HIGH", "MEDIUM", "LOW") else "MEDIUM"

        # ── Rule / object name ────────────────────────────────────────────
        if desc_raw in _AUDIT_SUBCHECK_META:
            rule_obj = desc_raw
        elif cis_m:
            rule_obj = desc_raw[cis_m.end():].strip()
        else:
            rule_obj = category

        # ── Details (evidence) ────────────────────────────────────────────
        details = f"Audit check: {desc_raw}\nXSLT output: {output[:500]}"

        # Append full audit solution and extracted xpath to details so both
        # surface in the Output column.
        full_sol = check["solution"]
        if full_sol:
            details += f"\nAudit solution: {full_sol[:1000]}"
        xpath_m = re.search(r'select="([^"]+)"', xsl_body)
        if xpath_m:
            raw_xp = xpath_m.group(1).strip()
            # Unwrap single XSLT function wrapper e.g. string(...) or normalize-space(...)
            fn_m   = re.match(r'^[\w-]+\((.*)\)$', raw_xp, re.DOTALL)
            audit_xpath = fn_m.group(1).strip() if fn_m else raw_xp
            details += f"\nxpath: {audit_xpath}"

        bench_override = [cis_num] if cis_num else None
        self._issue(severity, category, rule_obj, description, recommendation, details,
                    bench_override=bench_override)
        return True

    def _chk_cis125_admin_cert(self):
        """CIS 1.2.5: Ensure a valid certificate is set for the browser-based admin interface."""
        sys_el = self.root.find("./deviceconfig/system") if self.root is not None else None
        prof_el = sys_el.find("ssl-tls-service-profile") if sys_el is not None else None
        if prof_el is None or not (prof_el.text and prof_el.text.strip()):
            self._issue(
                "HIGH", "Admin Interface Default Certificate", "deviceconfig/system",
                "No SSL/TLS service profile is assigned to the browser-based management "
                "interface. The factory-default self-signed certificate is in use; "
                "administrators cannot detect a man-in-the-middle attack on their admin session.",
                "Navigate to Device > Certificate Management > SSL/TLS Service Profile, configure "
                "a profile using a CA-signed certificate, then assign it at "
                "Device > Setup > Management > General Settings > SSL/TLS Service Profile.",
                "deviceconfig/system/ssl-tls-service-profile: (not configured)")

    def _chk_cis22_wmi_probing(self):
        """CIS 2.2: Ensure WMI probing is disabled."""
        if self.root is None:
            return
        for vsys_entry in self.root.findall(".//vsys/entry"):
            vsys_name = vsys_entry.get("name", "vsys1")
            probe_el = vsys_entry.find("user-id-collector/setting/enable-probing")
            if probe_el is not None and (probe_el.text or "").strip() == "yes":
                self._issue(
                    "MEDIUM", "WMI Probing Enabled", f"vsys: {vsys_name}",
                    "WMI probing is enabled. This exposes a domain administrator credential "
                    "whose NTLM hash can be captured by a hostile host and cracked offline "
                    "or used in relay attacks.",
                    "Disable WMI probing unless explicitly required for User-ID: "
                    "Device > User Identification > User Mapping > Palo Alto Networks "
                    "User ID Agent Setup > uncheck Enable Probing.",
                    f"user-id-collector/setting/enable-probing: yes  (vsys: {vsys_name})")

    def _chk_cis616_zone_flood(self):
        """CIS 6.16: Zone protection profiles must have all flood-protection types enabled."""
        if self.root is None:
            return
        flood_types = [
            ("icmp",     "ICMP"),
            ("icmpv6",   "ICMPv6"),
            ("udp",      "UDP"),
            ("other-ip", "Other-IP"),
        ]
        for prof in self.root.findall(".//network/profiles/zone-protection-profile/entry"):
            prof_name = prof.get("name", "")
            for xml_tag, label in flood_types:
                el = prof.find(f"flood/{xml_tag}/enable")
                if el is None or (el.text or "").strip() != "yes":
                    self._issue(
                        "MEDIUM", "Zone Flood Protection Disabled",
                        f"Zone Profile: {prof_name}",
                        f"{label} flood protection is not enabled on zone protection profile "
                        f"'{prof_name}'. Attackers can overwhelm network resources with "
                        "targeted flood traffic through this zone.",
                        "Enable all flood protection types (ICMP, ICMPv6, UDP, Other-IP) in "
                        "each Zone Protection Profile: Network > Network Profiles > "
                        "Zone Protection > Flood Protection.",
                        f"Profile: {prof_name}  |  flood/{xml_tag}/enable: (not set or 'no')",
                        line=self._lineno_str(prof))


# ── Excel report writer ───────────────────────────────────────────────────────
class ExcelReporter:
    SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    SEV_COLORS = {
        "CRITICAL": (C["critical"],   C["critical_l"]),
        "HIGH":     (C["high"],       C["high_l"]),
        "MEDIUM":   (C["medium"],     C["medium_l"]),
        "LOW":      (C["low"],        C["low_l"]),
        "INFO":     (C["info"],       C["info_l"]),
    }
    ACTION_COLORS = {
        "allow":        (C["allow"],  C["allow_l"]),
        "deny":         (C["deny"],   C["deny_l"]),
        "drop":         (C["deny"],   C["deny_l"]),
        "reset-client": (C["deny"],   C["deny_l"]),
        "reset-server": (C["deny"],   C["deny_l"]),
        "reset-both":   (C["deny"],   C["deny_l"]),
    }

    def __init__(self, parser: PaloAltoParser, output_file: str):
        self.p = parser
        self.out = output_file
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self._vulns = _find_vulns_file()

    # ── Sheet helpers ─────────────────────────────────────────────────────────
    def _hdr(self, ws, headers: list[str], row: int = 1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill  = _fill(C["hdr_bg"])
            c.font  = _font(bold=True, color=C["hdr_fg"])
            c.alignment = _align("center", wrap=False)
            c.border = THIN
        ws.row_dimensions[row].height = 28

    def _row_fill(self, row_idx: int, disabled: bool = False) -> str | None:
        if disabled:
            return C["info_l"]
        return C["alt_row"] if row_idx % 2 == 0 else None

    def _set_widths(self, ws, widths: list[int | float]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_cell(self, ws, row, col, value, bold=False, bg=None,
                    fg="000000", h="left", border=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(bold=bold, color=fg)
        c.alignment = _align(h)
        if bg:
            c.fill = _fill(bg)
        if border:
            c.border = THIN
        return c

    # ── Summary ───────────────────────────────────────────────────────────────
    def _sheet_summary(self):
        ws = self.wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        # Title banner
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = "Palo Alto Firewall — Configuration Security Report"
        c.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        c.fill  = _fill(C["hdr_bg"])
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 42

        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    "
                   f"Source file: {self.p.source_label}")
        c.font  = _font(italic=True, color="595959", size=9)
        c.fill  = _fill("F2F2F2")
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 18

        row = 4
        p = self.p

        def section_header(label: str):
            nonlocal row
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = _font(bold=True, color=C["hdr_bg"], size=12)
            row += 1

        def kv(label: str, value):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = _font()
            c1.alignment = _align()
            c2 = ws.cell(row=row, column=2, value=value)
            c2.font = _font(bold=True)
            c2.alignment = _align("center")
            if row % 2 == 0:
                c1.fill = c2.fill = _fill(C["alt_row"])
            row += 1

        section_header("CONFIGURATION OVERVIEW")
        fmt_label = {"curly": "show config merged (CLI text)",
                     "xml": "PAN-OS XML export",
                     "csv-only": "Rules CSV only (no device config)"}[p.config_format]
        kv("Config format",   fmt_label)
        if p.csv_rules_path:
            kv("Rules CSV",   os.path.basename(p.csv_rules_path))
        if p.config_format == "csv-only":
            kv("Device/system checks", "skipped — no device config supplied")
        else:
            kv("PAN-OS Version (detected)", p.panos_version_str or "unknown")
            kv("audits.tar.gz used",       p.audits_tar_path or "not found")
            kv("CIS Benchmark — L1 .audit", p.cis_l1_audit_used or "n/a")
            kv("CIS Benchmark — L2 .audit", p.cis_l2_audit_used or "n/a")
        kv("Security Rules (total)",    len(p.security_rules))
        kv("  Active",  sum(1 for r in p.security_rules if r["disabled"] != "yes"))
        kv("  Disabled", sum(1 for r in p.security_rules if r["disabled"] == "yes"))
        kv("  Allow",   sum(1 for r in p.security_rules if r["action"] == "allow"))
        kv("  Deny/Drop/Reset", sum(1 for r in p.security_rules
                                   if r["action"] not in ("allow",)))
        kv("NAT Rules", len(p.nat_rules))
        kv("Address Objects", len(p.address_objects))
        kv("Address Groups",  len(p.address_groups))
        kv("Service Objects", len(p.service_objects))
        kv("Service Groups",  len(p.service_groups))
        kv("Zones", len(p.zones))
        row += 1

        section_header("SECURITY FINDINGS BY SEVERITY")
        self._hdr(ws, ["Severity", "Count"], row=row)
        row += 1
        total = 0
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
            cnt = sum(1 for i in p.issues if i["severity"] == sev)
            total += cnt
            fg, bg = self.SEV_COLORS[sev]
            c1 = ws.cell(row=row, column=1, value=sev)
            c1.fill  = _fill(bg); c1.font = _font(bold=True, color=fg)
            c1.alignment = _align("center"); c1.border = THIN
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.fill  = _fill(bg); c2.font = _font(bold=True, color=fg)
            c2.alignment = _align("center"); c2.border = THIN
            row += 1
        c1 = ws.cell(row=row, column=1, value="TOTAL")
        c1.font = _font(bold=True); c1.border = THIN
        c2 = ws.cell(row=row, column=2, value=total)
        c2.font = _font(bold=True); c2.border = THIN
        c2.alignment = _align("center")
        row += 2

        section_header("FINDINGS BY CATEGORY")
        self._hdr(ws, ["Category", "Count"], row=row)
        row += 1
        cat_counts: dict[str, int] = defaultdict(int)
        for iss in p.issues:
            cat_counts[iss["category"]] += 1
        for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat).font = _font()
            c2 = ws.cell(row=row, column=2, value=cnt)
            c2.font = _font(bold=True)
            c2.alignment = _align("center")
            if row % 2 == 0:
                ws.cell(row=row, column=1).fill = _fill(C["alt_row"])
                c2.fill = _fill(C["alt_row"])
            row += 1

        self._set_widths(ws, [40, 15, 15, 15, 15, 15, 15])

    # ── Security rules ────────────────────────────────────────────────────────
    def _sheet_security_rules(self):
        ws = self.wb.create_sheet("Security Rules")
        ws.sheet_view.showGridLines = False
        headers = [
            "#", "Rule Name", "Line #", "Status", "Rulebase",
            "Src Zone", "Dst Zone", "Source Address", "Destination Address",
            "Negate Src", "Negate Dst", "Application", "Service", "URL Category",
            "Action",
            "Profile/Group", "AV", "Vulnerability", "Spyware",
            "URL Filter", "File Blocking", "WildFire",
            "Log Start", "Log End", "Log Profile",
            "HIP Profiles", "Schedule", "Tags", "Description",
        ]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for rule in self.p.security_rules:
            row = rule["num"] + 1
            disabled = rule["disabled"] == "yes"
            row_bg = self._row_fill(row, disabled)

            # Expand profile group
            pt = rule["profile_type"]
            if pt == "group":
                pg = rule["profiles"].get("group", "")
                prof_display = f"Group: {pg}"
                grp_data = self.p.profile_groups.get(pg, {})
                av   = grp_data.get("virus", "")
                vuln = grp_data.get("vulnerability", "")
                spy  = grp_data.get("spyware", "")
                url  = grp_data.get("url-filtering", "")
                fb   = grp_data.get("file-blocking", "")
                wf   = grp_data.get("wildfire-analysis", "")
            else:
                prof_display = ""
                av   = rule["profiles"].get("virus", "")
                vuln = rule["profiles"].get("vulnerability", "")
                spy  = rule["profiles"].get("spyware", "")
                url  = rule["profiles"].get("url-filtering", "")
                fb   = rule["profiles"].get("file-blocking", "")
                wf   = rule["profiles"].get("wildfire-analysis", "")

            values = [
                rule["num"], rule["name"], rule.get("line", ""),
                "Disabled" if disabled else "Active", rule["rulebase"],
                rule["src_zones"], rule["dst_zones"],
                rule["sources"], rule["destinations"],
                rule["negate_src"], rule["negate_dst"],
                rule["applications"], rule["services"], rule["categories"],
                rule["action"].upper(),
                prof_display, av, vuln, spy, url, fb, wf,
                rule["log_start"], rule["log_end"], rule["log_setting"],
                rule["hip_profiles"], rule["schedule"],
                rule["tags"], rule["description"],
            ]

            action = rule["action"].lower()
            act_fg, act_bg = self.ACTION_COLORS.get(action, (C["info"], C["info_l"]))

            for col, val in enumerate(values, 1):
                fg = C["disabled"] if disabled else "000000"
                bg = row_bg
                bold = False

                if col == 15:  # Action (shifted +1 for Line # column)
                    fg, bg, bold = act_fg, act_bg, True
                elif col == 3:  # Line # — center
                    c2 = ws.cell(row=row, column=col, value=val)
                    c2.font = _font(color=C["info"] if not disabled else C["disabled"])
                    c2.alignment = _align("center")
                    c2.border = THIN
                    if bg:
                        c2.fill = _fill(bg)
                    continue

                c = ws.cell(row=row, column=col, value=val)
                c.font      = _font(bold=bold, color=fg)
                c.alignment = _align()
                c.border    = THIN
                if bg:
                    c.fill = _fill(bg)

        widths = [4, 30, 10, 9, 13, 20, 20, 28, 28, 10, 10,
                  30, 20, 18, 10,
                  22, 15, 15, 15, 15, 15, 15,
                  10, 10, 20, 20, 15, 20, 40]
        self._set_widths(ws, widths)

    # ── NAT rules ─────────────────────────────────────────────────────────────
    def _sheet_nat_rules(self):
        ws = self.wb.create_sheet("NAT Rules")
        ws.sheet_view.showGridLines = False
        headers = [
            "#", "Rule Name", "Line #", "Status", "Type",
            "Src Zone", "Dst Zone", "Source", "Destination", "Service",
            "Src Trans Type", "Src Trans Value",
            "Dst Trans Address", "Dst Trans Port",
            "Description", "Tags",
        ]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for rule in self.p.nat_rules:
            row = rule["num"] + 1
            disabled = rule["disabled"] == "yes"
            bg = self._row_fill(row, disabled)
            values = [
                rule["num"], rule["name"], rule.get("line", ""),
                "Disabled" if disabled else "Active", rule["nat_type"],
                rule["src_zones"], rule["dst_zones"],
                rule["sources"], rule["destinations"], rule["services"],
                rule["src_trans_type"], rule["src_trans_value"],
                rule["dst_trans_addr"], rule["dst_trans_port"],
                rule["description"], rule["tags"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = _font(color=C["disabled"] if disabled else "000000")
                c.alignment = _align()
                c.border    = THIN
                if bg:
                    c.fill = _fill(bg)

        widths = [4, 32, 10, 9, 9, 20, 20, 28, 28, 18, 22, 28, 22, 16, 38, 20]
        self._set_widths(ws, widths)

    # ── Address objects ───────────────────────────────────────────────────────
    def _sheet_address_objects(self):
        ws = self.wb.create_sheet("Address Objects")
        ws.sheet_view.showGridLines = False
        headers = ["Scope", "Name", "Type", "Value / Address", "Description", "Tags"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for i, obj in enumerate(self.p.address_objects, 2):
            bg = self._row_fill(i)
            for col, val in enumerate([
                obj["scope"], obj["name"], obj["type"],
                obj["value"], obj["description"], obj["tags"]
            ], 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)

        self._set_widths(ws, [16, 36, 14, 32, 42, 26])

    # ── Address groups ────────────────────────────────────────────────────────
    def _sheet_address_groups(self):
        ws = self.wb.create_sheet("Address Groups")
        ws.sheet_view.showGridLines = False
        headers = ["Scope", "Name", "Type", "Static Members / Dynamic Filter", "Description", "Tags"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for i, grp in enumerate(self.p.address_groups, 2):
            bg = self._row_fill(i)
            val = grp["dynamic_filter"] if grp["type"] == "dynamic" else grp["members"]
            for col, v in enumerate([
                grp["scope"], grp["name"], grp["type"], val,
                grp["description"], grp["tags"]
            ], 1):
                c = ws.cell(row=i, column=col, value=v)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)

        self._set_widths(ws, [16, 36, 12, 65, 42, 26])

    # ── Service objects ───────────────────────────────────────────────────────
    def _sheet_service_objects(self):
        ws = self.wb.create_sheet("Service Objects")
        ws.sheet_view.showGridLines = False

        # Service objects section
        ws.cell(row=1, column=1).value = "SERVICE OBJECTS"
        ws.cell(row=1, column=1).font  = _font(bold=True, color=C["hdr_bg"], size=11)
        ws.row_dimensions[1].height = 22

        headers = ["Scope", "Name", "Protocol", "Port(s)", "Description"]
        self._hdr(ws, headers, row=2)
        ws.freeze_panes = "A3"

        for i, svc in enumerate(self.p.service_objects, 3):
            bg = self._row_fill(i)
            for col, v in enumerate([
                svc["scope"], svc["name"], svc["protocol"],
                svc["port"], svc["description"]
            ], 1):
                c = ws.cell(row=i, column=col, value=v)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)

        sep = len(self.p.service_objects) + 4
        if self.p.service_groups:
            ws.cell(row=sep, column=1).value = "SERVICE GROUPS"
            ws.cell(row=sep, column=1).font = _font(bold=True, color=C["hdr_bg"], size=11)
            self._hdr(ws, ["Scope", "Name", "Members", "", "Tags"], row=sep + 1)
            for i, grp in enumerate(self.p.service_groups, sep + 2):
                bg = self._row_fill(i)
                for col, v in enumerate([
                    grp["scope"], grp["name"], grp["members"], "", grp["tags"]
                ], 1):
                    c = ws.cell(row=i, column=col, value=v)
                    c.font = _font(); c.alignment = _align(); c.border = THIN
                    if bg:
                        c.fill = _fill(bg)

        ws.auto_filter.ref = f"A2:{get_column_letter(5)}{sep - 1}" if sep > 3 else None
        self._set_widths(ws, [16, 36, 12, 30, 44])

    # ── Zones ─────────────────────────────────────────────────────────────────
    def _sheet_zones(self):
        ws = self.wb.create_sheet("Zones")
        ws.sheet_view.showGridLines = False
        headers = ["Zone Name", "Type", "Interfaces",
                   "Zone Protection Profile", "User-ID", "Log Setting"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for i, z in enumerate(self.p.zones, 2):
            bg = self._row_fill(i)
            has_prot = bool(z["zone_protection_profile"])
            values = [
                z["name"], z["type"], z["interfaces"],
                z["zone_protection_profile"] or "NONE",
                z["user_id"], z["log_setting"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if col == 4 and not has_prot:
                    c.fill  = _fill(C["medium_l"])
                    c.font  = _font(bold=True, color=C["medium"])
                elif bg:
                    c.fill = _fill(bg)

        self._set_widths(ws, [30, 18, 44, 36, 12, 26])

    # ── Security issues ───────────────────────────────────────────────────────
    def _sheet_issues(self):
        ws = self.wb.create_sheet("Security Issues")
        ws.sheet_view.showGridLines = False
        headers = ["#", "Validated", "Severity", "Residual Risk", "Residual Risk Note",
                   "Category", "Rule / Object", "Config Line(s)", "CIS v8", "CIS Benchmark",
                   "PCI DSS", "SCF",
                   "Description", "Recommendation", "Details",
                   "Asset", "Target", "Vuln", "Output", "Source"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        hostname = self.p.mgmt_settings.get("hostname", "") or ""
        source = self.p.source_label

        sorted_issues = sorted(
            self.p.issues,
            key=lambda x: self.SEV_ORDER.get(x["severity"], 9),
        )
        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            sev = iss["severity"]
            fg, bg = self.SEV_COLORS[sev]
            row_bg = self._row_fill(row)

            rule_name = iss["rule_name"]
            line = iss.get("line", "")
            target = f"{rule_name} ({line})" if line else rule_name
            details = iss.get("details", "")
            output = f"{iss['description']}\n{details}" if details else iss["description"]
            vuln_id = CATEGORY_VULN_ID.get(iss["category"])
            vuln = self._vulns[vuln_id] if vuln_id and vuln_id in self._vulns else ""

            values = [idx, "Y", sev, "", "",
                      iss["category"], rule_name,
                      line, iss.get("cis_controls", ""), iss.get("cis_benchmark", ""),
                      iss.get("pci_dss", ""), iss.get("scf", ""),
                      iss["description"], iss["recommendation"], details,
                      hostname, target, vuln, output, source]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                if col == 3:  # Severity
                    c.fill  = _fill(bg)
                    c.font  = _font(bold=True, color=fg)
                    c.alignment = _align("center")
                elif col in (1, 8):  # # and Config Line(s)
                    c.font = _font(bold=(col == 1))
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 9:  # CIS v8
                    c.font = _font(bold=True, color="17375E", size=9)
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 10:  # CIS Benchmark
                    c.font = _font(bold=True, color="1F618D", size=9)
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 11:  # PCI DSS
                    c.font = _font(bold=True, color="7B2D8B", size=9)
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 12:  # SCF
                    c.font = _font(bold=True, color="1A5C3A", size=9)
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                elif col == 2:  # Validated
                    c.font = _font(bold=True)
                    c.alignment = _align("center")
                    if row_bg:
                        c.fill = _fill(row_bg)
                else:
                    c.font = _font()
                    c.alignment = _align()
                    if row_bg:
                        c.fill = _fill(row_bg)
            ws.row_dimensions[row].height = 40

        self._set_widths(ws, [4, 12, 12, 18, 28, 32, 36, 14, 20, 20, 18, 18, 60, 60, 36, 24, 44, 16, 70, 30])

    # ── Ticketing Export ──────────────────────────────────────────────────────
    def _sheet_export(self):
        ws = self.wb.create_sheet("Export")
        ws.sheet_view.showGridLines = False
        headers = ["Hostname", "line#", "protocol", "port", "output"]
        self._hdr(ws, headers)
        ws.freeze_panes = "A2"

        hostname = self.p.mgmt_settings.get("hostname", "") or ""
        sorted_issues = sorted(
            self.p.issues,
            key=lambda x: self.SEV_ORDER.get(x["severity"], 9),
        )
        for idx, iss in enumerate(sorted_issues, 1):
            row = idx + 1
            output_text = (
                f"{iss['rule_name']}\n"
                f"{iss['description']}\n"
                f"{iss['recommendation']}"
            )
            values = [hostname, iss.get("line", ""), "tcp", 0, output_text]
            row_bg = self._row_fill(row)
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = THIN
                c.font = _font()
                c.alignment = _align("left" if col == 5 else "center", wrap=(col == 5))
                if row_bg:
                    c.fill = _fill(row_bg)
            ws.row_dimensions[row].height = 60

        self._set_widths(ws, [24, 10, 10, 8, 80])

    # ── Crypto & System Config ────────────────────────────────────────────────
    def _sheet_crypto_system(self):
        ws = self.wb.create_sheet("Crypto & System")
        ws.sheet_view.showGridLines = False
        p = self.p

        WEAK_ENC  = {"des", "3des", "null"}
        WEAK_HASH = {"md5", "sha1"}
        WEAK_DH   = {"group1", "group2", "group5"}

        def flag_cell(cell, value: str, bad_set: set, crit_set: set | None = None):
            lv = value.strip().lower()
            if lv in (crit_set or set()):
                cell.fill  = _fill(C["critical_l"])
                cell.font  = _font(bold=True, color=C["critical"])
            elif lv in bad_set:
                cell.fill  = _fill(C["high_l"])
                cell.font  = _font(bold=True, color=C["high"])
            else:
                cell.font  = _font()

        def flag_multi(cell, csv_value: str, bad_set: set, crit_set: set | None = None):
            """Highlight cell if any token in a CSV value is in the bad set."""
            tokens = {t.strip().lower() for t in csv_value.split(",") if t.strip()}
            if tokens & (crit_set or set()):
                cell.fill = _fill(C["critical_l"])
                cell.font = _font(bold=True, color=C["critical"])
            elif tokens & bad_set:
                cell.fill = _fill(C["high_l"])
                cell.font = _font(bold=True, color=C["high"])
            else:
                cell.font = _font()

        def flag_tls(cell, version: str):
            ver = version.strip().lower()
            if ver == "ssl3-0":
                cell.fill = _fill(C["critical_l"]); cell.font = _font(bold=True, color=C["critical"])
            elif ver == "tls1-0":
                cell.fill = _fill(C["high_l"]);     cell.font = _font(bold=True, color=C["high"])
            elif ver == "tls1-1":
                cell.fill = _fill(C["medium_l"]);   cell.font = _font(bold=True, color=C["medium"])
            else:
                cell.font = _font()

        row = 1

        def section(title: str, headers: list[str]) -> int:
            nonlocal row
            ws.cell(row=row, column=1).value = title
            ws.cell(row=row, column=1).font  = _font(bold=True, color=C["hdr_bg"], size=12)
            ws.row_dimensions[row].height = 22
            row += 1
            self._hdr(ws, headers, row=row)
            row += 1
            return row  # first data row

        # ── IKE Crypto Profiles ───────────────────────────────────────────────
        section("IKE CRYPTO PROFILES", ["Name", "Encryption", "Hash / PRF", "DH Group", "Lifetime"])
        for prof in p.ike_crypto_profiles:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [prof["name"], prof["encryption"], prof["hash"], prof["dh_group"], prof["lifetime"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            flag_multi(ws.cell(row=row, column=2), prof["encryption"], WEAK_ENC, {"des", "null"})
            flag_multi(ws.cell(row=row, column=3), prof["hash"],       WEAK_HASH, {"md5"})
            flag_multi(ws.cell(row=row, column=4), prof["dh_group"],   WEAK_DH,  {"group1"})
            row += 1
        if not p.ike_crypto_profiles:
            ws.cell(row=row, column=1).value = "(none found)"
            ws.cell(row=row, column=1).font = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── IPSec Crypto Profiles ─────────────────────────────────────────────
        section("IPSEC CRYPTO PROFILES",
                ["Name", "Protocol", "Encryption", "Authentication", "DH Group (PFS)", "Lifetime"])
        for prof in p.ipsec_crypto_profiles:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [prof["name"], prof["protocol"], prof["encryption"],
                    prof["auth"], prof["dh_group"], prof["lifetime"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            flag_multi(ws.cell(row=row, column=3), prof["encryption"], WEAK_ENC, {"des", "null"})
            flag_multi(ws.cell(row=row, column=4), prof["auth"],       WEAK_HASH, {"md5"})
            dh_val = prof["dh_group"].strip().lower()
            dh_cell = ws.cell(row=row, column=5)
            if dh_val in ("no-pfs", ""):
                dh_cell.fill = _fill(C["medium_l"]); dh_cell.font = _font(bold=True, color=C["medium"])
            else:
                flag_cell(dh_cell, dh_val, WEAK_DH, {"group1"})
            row += 1
        if not p.ipsec_crypto_profiles:
            ws.cell(row=row, column=1).value = "(none found)"
            ws.cell(row=row, column=1).font = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── SSL/TLS Profiles ──────────────────────────────────────────────────
        section("SSL/TLS SERVICE PROFILES", ["Name", "Min TLS Version", "Max TLS Version", "Certificate"])
        for prof in p.ssl_tls_profiles:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [prof["name"], prof["min_version"], prof["max_version"], prof["certificate"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            flag_tls(ws.cell(row=row, column=2), prof["min_version"])
            row += 1
        if not p.ssl_tls_profiles:
            ws.cell(row=row, column=1).value = "(none found)"
            ws.cell(row=row, column=1).font = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── IKE Gateways ──────────────────────────────────────────────────────
        section("IKE GATEWAYS",
                ["Name", "IKE Version", "Peer IP", "Local IP",
                 "Crypto Profile", "Auth Type", "NAT-T"])
        for gw in p.ike_gateways:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [gw["name"], gw["version"], gw["peer_ip"], gw["local_ip"],
                    gw["crypto_profile"], gw["auth_type"], gw["nat_traversal"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            ver_cell = ws.cell(row=row, column=2)
            if gw["version"].lower() == "ikev1":
                ver_cell.fill = _fill(C["medium_l"])
                ver_cell.font = _font(bold=True, color=C["medium"])
            row += 1
        if not p.ike_gateways:
            ws.cell(row=row, column=1).value = "(none found)"
            ws.cell(row=row, column=1).font = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── Management Settings ───────────────────────────────────────────────
        section("MANAGEMENT SETTINGS", ["Setting", "Value", "Status"])
        m = p.mgmt_settings
        if m:
            def mgmt_row(label: str, value: str, good: bool | None = None):
                nonlocal row
                bg = C["alt_row"] if row % 2 == 0 else None
                for col, v in enumerate([label, value], 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = _font(); c.alignment = _align(); c.border = THIN
                    if bg:
                        c.fill = _fill(bg)
                status_cell = ws.cell(row=row, column=3)
                status_cell.border = THIN; status_cell.alignment = _align("center")
                if good is True:
                    status_cell.value = "OK"
                    status_cell.font  = _font(bold=True, color=C["allow"])
                    status_cell.fill  = _fill(C["allow_l"])
                elif good is False:
                    status_cell.value = "ISSUE"
                    status_cell.font  = _font(bold=True, color=C["high"])
                    status_cell.fill  = _fill(C["high_l"])
                else:
                    status_cell.value = "INFO"
                    status_cell.font  = _font(color=C["info"])
                row += 1

            mgmt_row("Hostname",              m.get("hostname", "(not set)"), None)
            mgmt_row("HTTP Management",       "Enabled" if m.get("http_enabled") else "Disabled",
                                               not m.get("http_enabled"))
            mgmt_row("HTTPS Management",      "Enabled" if m.get("https_enabled") else "Disabled",
                                               m.get("https_enabled"))
            mgmt_row("Telnet Management",     "Enabled" if m.get("telnet_enabled") else "Disabled",
                                               not m.get("telnet_enabled"))
            mgmt_row("SSH Management",        "Enabled" if m.get("ssh_enabled") else "Disabled",
                                               m.get("ssh_enabled"))
            mgmt_row("SNMP Enabled",          "Yes" if m.get("snmp_enabled") else "No", None)
            mgmt_row("SNMP v1",               "Enabled" if m.get("snmp_v1") else "Disabled",
                                               not m.get("snmp_v1"))
            mgmt_row("SNMP v2c",              "Enabled" if m.get("snmp_v2c") else "Disabled",
                                               not m.get("snmp_v2c"))
            mgmt_row("SNMP v3",               "Enabled" if m.get("snmp_v3") else "Disabled",
                                               m.get("snmp_v3"))
            mgmt_row("SNMP Community String", m.get("snmp_community") or "(not set)", None)
            mgmt_row("Permitted Mgmt IPs",
                     ", ".join(m["permitted_ips"]) if m.get("permitted_ips") else "NONE (unrestricted)",
                     bool(m.get("permitted_ips")))
            mgmt_row("NTP Primary",           m.get("ntp_primary") or "(not configured)",
                                               bool(m.get("ntp_primary")))
            mgmt_row("NTP Secondary",         m.get("ntp_secondary") or "(not configured)", None)
            mgmt_row("DNS Primary",           m.get("dns_primary") or "(not configured)",
                                               bool(m.get("dns_primary")))
            mgmt_row("Login Banner",          "Set" if m.get("login_banner") else "Not configured",
                                               bool(m.get("login_banner")))
        else:
            ws.cell(row=row, column=1).value = "(deviceconfig/system not found in this export)"
            ws.cell(row=row, column=1).font  = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── Admin Accounts ────────────────────────────────────────────────────
        section("ADMINISTRATOR ACCOUNTS",
                ["Username", "Role", "Auth Profile (MFA)", "SSH Key", "Password"])
        for admin in p.admin_accounts:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [admin["name"], admin["role"], admin["auth_profile"] or "(none)",
                    admin["public_key"], admin["password_hash"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            if not admin["auth_profile"]:
                ws.cell(row=row, column=3).fill = _fill(C["high_l"])
                ws.cell(row=row, column=3).font = _font(bold=True, color=C["high"])
            if admin["password_hash"] == "none" and admin["public_key"] == "no":
                ws.cell(row=row, column=5).fill = _fill(C["critical_l"])
                ws.cell(row=row, column=5).font = _font(bold=True, color=C["critical"])
            row += 1
        if not p.admin_accounts:
            ws.cell(row=row, column=1).value = "(none found)"
            ws.cell(row=row, column=1).font = _font(italic=True, color=C["info"])
            row += 1
        row += 1

        # ── Syslog Servers ────────────────────────────────────────────────────
        section("SYSLOG SERVERS", ["Profile", "Server Name", "Address", "Port", "Transport", "Format"])
        for srv in p.log_syslog_servers:
            bg = C["alt_row"] if row % 2 == 0 else None
            vals = [srv["profile"], srv["name"], srv["server"],
                    srv["port"], srv["transport"], srv["format"]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = _font(); c.alignment = _align(); c.border = THIN
                if bg:
                    c.fill = _fill(bg)
            if srv["transport"].upper() == "UDP":
                ws.cell(row=row, column=5).fill = _fill(C["medium_l"])
                ws.cell(row=row, column=5).font = _font(bold=True, color=C["medium"])
            row += 1
        if not p.log_syslog_servers:
            ws.cell(row=row, column=1).value = "(no syslog servers configured)"
            ws.cell(row=row, column=1).fill  = _fill(C["high_l"])
            ws.cell(row=row, column=1).font  = _font(bold=True, color=C["high"])
            row += 1

        self._set_widths(ws, [30, 22, 30, 25, 20, 22, 18])

    # ── CIS v8 Mapping ────────────────────────────────────────────────────────
    def _sheet_cis_mapping(self):
        ws = self.wb.create_sheet("CIS v8 Mapping")
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "CIS Controls v8 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(C["hdr_bg"])
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = ("Each CIS safeguard lists all findings from this config that map to it, "
                   "with count and severity breakdown.")
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        # Build reverse map: ctrl_id → list of issues
        from collections import defaultdict as _dd
        ctrl_issues: dict[str, list[dict]] = _dd(list)
        for iss in self.p.issues:
            for cid in iss.get("cis_ids", []):
                ctrl_issues[cid].append(iss)

        SEV_ORDER = self.SEV_ORDER
        SEV_COLORS = self.SEV_COLORS

        row = 4
        # Sort controls numerically (e.g. 3.10 < 4.2 < 12.2)
        def _sort_key(k):
            parts = k.split(".")
            return (int(parts[0]), float("0." + parts[1]) if len(parts) > 1 else 0)

        all_ctrl_ids = sorted(CIS_CTRL_DESC.keys(), key=_sort_key)

        for ctrl_id in all_ctrl_ids:
            ctrl_desc = CIS_CTRL_DESC[ctrl_id]
            issues_for_ctrl = sorted(
                ctrl_issues.get(ctrl_id, []),
                key=lambda x: SEV_ORDER.get(x["severity"], 9),
            )

            # Control header row
            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"CIS {ctrl_id} — {ctrl_desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill("17375E")
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 22
            row += 1

            if not issues_for_ctrl:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this control")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9")
                nc.alignment = _align()
                nc.border = THIN
                row += 1
            else:
                # Column sub-headers
                sub_hdrs = ["Severity", "Category", "Rule / Object",
                            "Config Line(s)", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font  = _font(bold=True, color="FFFFFF")
                    c.fill  = _fill("2E4057")
                    c.alignment = _align("center", wrap=False)
                    c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1

                for iss in issues_for_ctrl:
                    sev = iss["severity"]
                    fg, bg = SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["rule_name"],
                            iss.get("line", ""),
                            iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill  = _fill(bg)
                            c.font  = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb:
                                c.fill = _fill(rb)
                        else:
                            c.font = _font()
                            c.alignment = _align()
                            if rb:
                                c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1

            row += 1  # gap between controls

        self._set_widths(ws, [12, 34, 36, 14, 60, 60])

    # ── PCI DSS v4.0 Mapping ──────────────────────────────────────────────────
    def _sheet_pci_mapping(self):
        ws = self.wb.create_sheet("PCI DSS Mapping")
        ws.sheet_view.showGridLines = False
        PCI_HDR = "5C1A8C"  # purple for PCI DSS

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = "PCI DSS v4.0 — Finding Cross-Reference"
        t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        t.fill  = _fill(PCI_HDR)
        t.alignment = _align("center", wrap=False)
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:F2")
        s = ws["A2"]
        s.value = ("Each PCI DSS v4.0 requirement lists all findings from this config that map to it.")
        s.font  = _font(italic=True, color=C["info"], size=9)
        s.fill  = _fill("F2F2F2")
        s.alignment = _align("center", wrap=False)
        ws.row_dimensions[2].height = 16

        from collections import defaultdict as _dd
        req_issues: dict[str, list[dict]] = _dd(list)
        for iss in self.p.issues:
            for pid in iss.get("pci_ids", []):
                req_issues[pid].append(iss)

        SEV_ORDER  = self.SEV_ORDER
        SEV_COLORS = self.SEV_COLORS
        row = 3
        for req_id in sorted(PCI_DSS_DESC.keys(),
                              key=lambda x: [int(p) for p in x.split(".")]):
            desc = PCI_DSS_DESC[req_id]
            issues_for_req = sorted(
                req_issues.get(req_id, []),
                key=lambda x: SEV_ORDER.get(x["severity"], 9),
            )
            count = len(issues_for_req)
            # Requirement header row
            ws.merge_cells(f"A{row}:F{row}")
            hc = ws.cell(row=row, column=1,
                         value=f"PCI DSS {req_id}  [{count} finding{'s' if count != 1 else ''}]  {desc}")
            hc.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            hc.fill  = _fill(PCI_HDR)
            hc.alignment = _align("left", wrap=False)
            hc.border = THIN
            ws.row_dimensions[row].height = 28
            row += 1

            if not issues_for_req:
                ws.merge_cells(f"A{row}:F{row}")
                nc = ws.cell(row=row, column=1, value="No findings for this requirement")
                nc.font = _font(italic=True, color=C["info"])
                nc.fill = _fill("F9F9F9")
                nc.alignment = _align()
                nc.border = THIN
                row += 1
            else:
                sub_hdrs = ["Severity", "Category", "Rule / Object",
                            "Config Line(s)", "Description", "Recommendation"]
                for col, h in enumerate(sub_hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font  = _font(bold=True, color="FFFFFF")
                    c.fill  = _fill("2E4057")
                    c.alignment = _align("center", wrap=False)
                    c.border = THIN
                ws.row_dimensions[row].height = 20
                row += 1
                for iss in issues_for_req:
                    sev = iss["severity"]
                    fg, bg = SEV_COLORS[sev]
                    rb = C["alt_row"] if row % 2 == 0 else None
                    vals = [sev, iss["category"], iss["rule_name"],
                            iss.get("line", ""),
                            iss["description"], iss["recommendation"]]
                    for col, val in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = THIN
                        if col == 1:
                            c.fill  = _fill(bg)
                            c.font  = _font(bold=True, color=fg)
                            c.alignment = _align("center")
                        elif col == 4:
                            c.font = _font(color=C["info"], size=9)
                            c.alignment = _align("center")
                            if rb:
                                c.fill = _fill(rb)
                        else:
                            c.font = _font()
                            c.alignment = _align()
                            if rb:
                                c.fill = _fill(rb)
                    ws.row_dimensions[row].height = 36
                    row += 1
            row += 1

        self._set_widths(ws, [12, 34, 36, 14, 60, 60])

    # ── Save ──────────────────────────────────────────────────────────────────
    def save(self):
        self._sheet_summary()
        self._sheet_security_rules()
        self._sheet_nat_rules()
        self._sheet_address_objects()
        self._sheet_address_groups()
        self._sheet_service_objects()
        self._sheet_zones()
        self._sheet_crypto_system()
        self._sheet_issues()
        self._sheet_export()
        self._sheet_cis_mapping()
        self._sheet_pci_mapping()
        self.wb.save(self.out)


# ── CLI entry point ───────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Palo Alto Firewall Config Analyzer — outputs Excel report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pa_analyzer.py running-config.xml
  python pa_analyzer.py running-config.xml -o audit-$(date +%%Y%%m%%d).xlsx
  python pa_analyzer.py panorama-export.xml -o panorama-audit.xlsx
  python pa_analyzer.py "show merged combined.txt" --rules-csv rules.csv
  python pa_analyzer.py rules.csv
""",
    )
    ap.add_argument("config",
                     help="PAN-OS configuration file: an XML export (device or Panorama), a "
                          "'show config merged'/'show config running' CLI text capture (e.g. "
                          "a PuTTY session log), or — on its own, with no device config at "
                          "all — a Policy Optimizer/rulebase CSV export (a '.csv' file is "
                          "routed to --rules-csv automatically and runs rule-based checks "
                          "only; device/system checks and the CIS Benchmark need a config).")
    ap.add_argument("-o", "--output", default=None,
                    help="Output Excel file (default: <config-stem>_analysis.xlsx)")
    ap.add_argument("--rules-csv", default=None,
                    help="Security policy / Policy Optimizer CSV export (Name, Source/"
                         "Destination Zone+Address, Application, Service, Action, Profile, "
                         "Rule Usage, Rule Usage Apps Seen, Days With No New Apps, ...). "
                         "Replaces any rulebase parsed from `config` and adds Policy "
                         "Optimizer usage checks — use this when the CLI capture's "
                         "rulebase is empty (rules pushed from Panorama, not captured "
                         "locally).")
    ap.add_argument("--panos-version", type=int, default=None,
                    help="PAN-OS major version (e.g. 10, 11) to select the CIS Benchmark "
                         "against. Only needed for a CLI text capture, which — unlike an "
                         "XML export — doesn't embed its own version string; ignored for "
                         "XML input.")
    args = ap.parse_args()

    # A bare CSV positional (no device config at all) is rules-only mode.
    config_file = args.config
    csv_rules_path = args.rules_csv
    if config_file.lower().endswith(".csv"):
        if csv_rules_path and os.path.abspath(csv_rules_path) != os.path.abspath(config_file):
            sys.exit("Both a .csv positional and --rules-csv were given — pass the device "
                      "config (XML or CLI text) as the positional and the CSV via --rules-csv.")
        csv_rules_path, config_file = config_file, None

    if not args.output:
        stem = os.path.splitext(os.path.basename(args.config))[0]
        args.output = f"{stem}_analysis.xlsx"

    print(f"[*] Parsing:  {args.config}")
    parser = PaloAltoParser(config_file, csv_rules_path=csv_rules_path,
                             panos_version_override=args.panos_version)
    parser.parse()

    sev_counts: dict[str, int] = defaultdict(int)
    for iss in parser.issues:
        sev_counts[iss["severity"]] += 1

    print(f"[*] Parsed:")
    print(f"      Security rules : {len(parser.security_rules)}")
    print(f"      NAT rules      : {len(parser.nat_rules)}")
    print(f"      Address objects: {len(parser.address_objects)}")
    print(f"      Address groups : {len(parser.address_groups)}")
    print(f"      Service objects: {len(parser.service_objects)}")
    print(f"      Zones          : {len(parser.zones)}")
    print(f"[*] Security findings: {len(parser.issues)}")
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        if sev_counts[sev]:
            print(f"      {sev:<10}: {sev_counts[sev]}")

    print(f"[*] Writing Excel report...")
    reporter = ExcelReporter(parser, args.output)
    reporter.save()
    print(f"[+] Saved: {args.output}")


if __name__ == "__main__":
    main()
